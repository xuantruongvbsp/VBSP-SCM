"""Đọc và tổng hợp dữ liệu Chấm điểm Tổ TK&VV (CDTOTKVV)."""
import re
import os
from pathlib import Path

import pandas as pd
import streamlit as st

from config import BASE_DIR, CDTOTKVV_DIR, CDTOTKVV_COLS, CDTOTKVV_DATA_ROW_START
from config import PGD_DATA_DIR
from data.core import ts_file
from data.pgd import duong_dan_pgd

_COLS_FLOAT = ["du_no", "so_du_tk", "tong_diem"]
_COLS_STR   = ["ma_dv", "ma_xa", "ma_to"]
_XEP_LOAI_TOT = "T\u1ed1t"
_XEP_LOAI_KHA = "Kh\u00e1"
_XEP_LOAI_TB  = "Trung b\u00ecnh"
_XEP_LOAI_YEU = "Y\u1ebfu"
_TINH_TRANG_A = "A"
_TINH_TRANG_B = "B"
_TINH_TRANG_C = "C"


def _ten_file(thang_nam: str) -> Path:
    return CDTOTKVV_DIR / f"CDTOTKVV_{thang_nam}.xlsx"


def doc_thang_nam_tu_file(file_bytes: bytes) -> str | None:
    """Trích kỳ chấm điểm mm/yyyy từ phần đầu file CDTOTKVV (tiêu đề / ngày báo cáo)."""
    from datetime import date, datetime
    from io import BytesIO

    import openpyxl

    pat_ngay_vn = re.compile(
        r"ngày\s+(\d{1,2})\s+tháng\s+(\d{1,2})\s+năm\s+(\d{4})",
        re.IGNORECASE,
    )
    pat_ddmmyyyy = re.compile(
        r"\b([0-2]?\d|3[0-1])[/\-]([0]?\d|1[0-2])[/\-](\d{4})\b"
    )
    pat_thang_slash_nam = re.compile(r"\b(0[1-9]|1[0-2])/\s*(\d{4})\b")
    pat_thang_word = re.compile(
        r"tháng\s*(\d{1,2})\s*/\s*(\d{4})",
        re.IGNORECASE,
    )
    try:
        wb = openpyxl.load_workbook(
            BytesIO(file_bytes), read_only=True, data_only=True
        )
        ws = wb.active
        cap = min(ws.max_row or 0, 25)
        for row in ws.iter_rows(min_row=1, max_row=cap, values_only=True):
            for cell in row:
                if cell is None:
                    continue
                if isinstance(cell, (datetime, date)):
                    return cell.strftime("%m/%Y")
                text = str(cell).strip()
                if not text or text.lower() == "nan":
                    continue
                m = pat_ngay_vn.search(text)
                if m:
                    mm = m.group(2).zfill(2)
                    yyyy = m.group(3)
                    return f"{mm}/{yyyy}"
                m = pat_thang_word.search(text)
                if m:
                    return f"{m.group(1).zfill(2)}/{m.group(2)}"
                m = pat_thang_slash_nam.search(text)
                if m:
                    return f"{m.group(1)}/{m.group(2)}"
                m = pat_ddmmyyyy.search(text)
                if m:
                    mm = m.group(2).zfill(2)
                    yyyy = m.group(3)
                    return f"{mm}/{yyyy}"
    except Exception:
        return None
    return None


@st.cache_data(show_spinner=False)
def doc_cdtotkvv_path(duong_dan: str, _ts) -> pd.DataFrame | None:
    if not duong_dan or not os.path.exists(duong_dan):
        return None
    df = pd.read_excel(
        duong_dan,
        engine="openpyxl",
        header=None,
        skiprows=CDTOTKVV_DATA_ROW_START,
    )
    df = df.iloc[:, : len(CDTOTKVV_COLS)].copy()
    df.columns = CDTOTKVV_COLS
    df = df[pd.to_numeric(df["stt"], errors="coerce").notna()].reset_index(drop=True)
    for col in _COLS_FLOAT:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    for col in _COLS_STR:
        if col in df.columns:
            df[col] = (
                df[col]
                .astype(str)
                .str.strip()
                .str.replace(r"\.0$", "", regex=True)
                .replace("nan", pd.NA)
            )
    return df


@st.cache_data(show_spinner=False)
def doc_cdtotkvv(thang_nam: str) -> pd.DataFrame | None:
    try:
        mm, yyyy = thang_nam.split("/")
        suffix = f"{yyyy}_{mm}"
    except ValueError:
        return None
    pgd_data = BASE_DIR / "pgd_data"
    frames = []
    for f in pgd_data.rglob(f"cdtotkvv_{suffix}.xlsx"):
        df = doc_cdtotkvv_path(str(f), ts_file(str(f)))
        if df is not None and not df.empty:
            frames.append(df)
    if not frames:
        for f in pgd_data.rglob("cdtotkvv_latest.xlsx"):
            try:
                thang = doc_thang_nam_tu_file(f.read_bytes())
                if thang == thang_nam:
                    df = doc_cdtotkvv_path(str(f), ts_file(str(f)))
                    if df is not None and not df.empty:
                        frames.append(df)
            except Exception:
                pass
    if not frames:
        return None
    return pd.concat(frames, ignore_index=True)


@st.cache_data(show_spinner=False)
def doc_cdtotkvv_pgd(ten_pgd: str, _ts) -> pd.DataFrame | None:
    path = duong_dan_pgd(ten_pgd, "cdtotkvv")
    return doc_cdtotkvv_path(path, _ts)


def doc_cdtotkvv_toan_cn_pgd() -> pd.DataFrame | None:
    if not os.path.exists(PGD_DATA_DIR):
        return None
    frames: list[pd.DataFrame] = []
    for d in sorted(Path(PGD_DATA_DIR).iterdir()):
        if not d.is_dir():
            continue
        path = d / "cdtotkvv_latest.xlsx"
        if not path.exists():
            continue
        try:
            df_p = doc_cdtotkvv_path(str(path), ts_file(str(path)))
            if df_p is None or df_p.empty:
                continue
            frames.append(df_p)
        except Exception:
            continue
    if not frames:
        return None
    return pd.concat(frames, ignore_index=True)


@st.cache_data(show_spinner=False)
def tong_hop_tu_pgd_data() -> pd.DataFrame | None:
    """Tổng hợp CDTOTKVV từ pgd_data/*/cdtotkvv_latest.xlsx (hệ thống tập trung)."""
    return doc_cdtotkvv_toan_cn_pgd()


@st.cache_data(show_spinner=False, ttl=60)
def ds_thang_nam() -> list[str]:
    pat = re.compile(r"cdtotkvv_(\d{4})_(\d{2})\.xlsx$", re.IGNORECASE)
    thang_set = set()
    pgd_data = BASE_DIR / "pgd_data"
    if not pgd_data.exists():
        return []
    for f in pgd_data.rglob("cdtotkvv_*.xlsx"):
        m = pat.search(f.name)
        if m:
            yyyy, mm = m.group(1), m.group(2)
            thang_set.add(f"{mm}/{yyyy}")
    for f in pgd_data.rglob("cdtotkvv_latest.xlsx"):
        try:
            thang = doc_thang_nam_tu_file(f.read_bytes())
            if thang:
                thang_set.add(thang)
        except Exception:
            pass
    return sorted(thang_set, key=lambda s: (s[3:], s[:2]), reverse=True)


@st.cache_data(show_spinner=False)
def tong_hop_theo_pgd(df: pd.DataFrame) -> pd.DataFrame:
    """
    Nhóm theo ma_dv + ten_dv, tính:
    - tong_to           : tổng số tổ
    - tong_diem_tb      : trung bình tổng_diem (làm tròn 2 chữ số)
    - to_tot/kha/tb/yeu : số tổ theo xep_loai
    - to_tinh_trang_a/b/c : số tổ theo tình trạng A/B/C
    Trả về DataFrame sort theo ma_dv.
    """
    nhom = df.groupby(["ma_dv", "ten_dv"], as_index=False)
    tong_hop = nhom.agg(
        tong_to=("stt", "count"),
        tong_diem_tb=("tong_diem", "mean"),
    )
    for ten_col, gia_tri in [
        ("to_tot", _XEP_LOAI_TOT),
        ("to_kha", _XEP_LOAI_KHA),
        ("to_tb",  _XEP_LOAI_TB),
        ("to_yeu", _XEP_LOAI_YEU),
    ]:
        dem = (
            df[df["xep_loai"] == gia_tri]
            .groupby(["ma_dv", "ten_dv"], as_index=False)
            .agg(**{ten_col: ("stt", "count")})
        )
        tong_hop = tong_hop.merge(dem, on=["ma_dv", "ten_dv"], how="left")
    for ten_col, gia_tri in [
        ("to_tinh_trang_a", _TINH_TRANG_A),
        ("to_tinh_trang_b", _TINH_TRANG_B),
        ("to_tinh_trang_c", _TINH_TRANG_C),
    ]:
        dem = (
            df[df["tinh_trang"] == gia_tri]
            .groupby(["ma_dv", "ten_dv"], as_index=False)
            .agg(**{ten_col: ("stt", "count")})
        )
        tong_hop = tong_hop.merge(dem, on=["ma_dv", "ten_dv"], how="left")
    cols_dem = [
        "to_tot", "to_kha", "to_tb", "to_yeu",
        "to_tinh_trang_a", "to_tinh_trang_b", "to_tinh_trang_c",
    ]
    tong_hop[cols_dem] = tong_hop[cols_dem].fillna(0).astype(int)
    tong_hop["tong_diem_tb"] = tong_hop["tong_diem_tb"].round(2)
    return tong_hop.sort_values("ma_dv").reset_index(drop=True)
