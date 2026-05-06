from __future__ import annotations

from datetime import datetime
from io import BytesIO
import shutil
from pathlib import Path
import sys

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from openpyxl import Workbook

from config import PGD_DATA_DIR
from data.cdtotkvv import doc_thang_nam_tu_file
from data.pgd import luu_file_pgd_voi_lich_su, pgd_slug


def tao_file_cdtotkvv_gia_lap(ngay: int, thang: int, nam: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A5"] = "CHI TIẾT CHẤM ĐIỂM TỔ TIẾT KIỆM VÀ VAY VỐN"
    ws["A6"] = f"Ngày {ngay} tháng {thang} năm {nam}"
    ws["B11"] = "004604"
    ws["C11"] = "PGD Long Khánh"
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def main() -> int:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    ten_pgd_test = f"PGD SMOKE TEST {ts}"
    slug = pgd_slug(ten_pgd_test)
    dir_test = Path(PGD_DATA_DIR) / slug

    b1 = tao_file_cdtotkvv_gia_lap(31, 3, 2026)
    b2 = tao_file_cdtotkvv_gia_lap(30, 4, 2026)

    try:
        t1 = doc_thang_nam_tu_file(b1)
        t2 = doc_thang_nam_tu_file(b2)
        assert t1 == "03/2026", f"Parse tháng 1 sai: {t1}"
        assert t2 == "04/2026", f"Parse tháng 2 sai: {t2}"

        luu_file_pgd_voi_lich_su(ten_pgd_test, "cdtotkvv", b1, t1)
        luu_file_pgd_voi_lich_su(ten_pgd_test, "cdtotkvv", b2, t2)

        f_latest = dir_test / "cdtotkvv_latest.xlsx"
        f_032026 = dir_test / "cdtotkvv_2026_03.xlsx"
        f_042026 = dir_test / "cdtotkvv_2026_04.xlsx"

        assert f_latest.exists(), "Thiếu latest.xlsx"
        assert f_032026.exists(), "Thiếu file lịch sử tháng 03/2026"
        assert f_042026.exists(), "Thiếu file lịch sử tháng 04/2026"
        assert f_latest.read_bytes() == b2, "latest không phải file tháng mới nhất"

        print("PASS: parse month + save 2 monthly history files + latest updated.")
        print(f"Test folder: {dir_test}")
        return 0
    finally:
        if dir_test.exists():
            shutil.rmtree(dir_test, ignore_errors=True)


if __name__ == "__main__":
    raise SystemExit(main())
