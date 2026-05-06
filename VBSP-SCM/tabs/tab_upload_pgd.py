"""
Tab Upload Dữ liệu — Hỗ trợ địa bàn (PGD tự upload file của mình).
──────────────────────────────────────────────────────────────────────
Quyền:
  role == "user"              → không có selectbox, chỉ thấy đơn vị pgd_user
  role in ("admin","manager") → có selectbox chọn đơn vị

Giao diện: giống tab_upload_khnv nhưng không có nút merge thủ công trên tab.
Lưu file qua luu_pgd_file — HSTD/NQ11/GQVL có gộp toàn CN tự động theo service.

Bảng trạng thái:
  user   → 1 hàng (đơn vị của mình)
  admin/manager → 22 hàng (toàn Chi nhánh)
"""
from io import BytesIO

import pandas as pd
import streamlit as st

import db
from config import DS_PGD, DON_VI_CHI_NHANH, MA_PGD_MAP
from pathlib import Path

from data.pgd import doc_trang_thai_file, duong_dan_pgd
from services.upload_service import (
    KetQuaUpload,
    kiem_tra_file,
    lay_meta_chat_luong,
    luu_pgd_file,
)
# data_priority_service và _render_upload_hang_loat đã được tách ra
# theo kiến trúc 2 luồng độc lập (xem HUONG_DAN_NGUON_DU_LIEU.md)


# ── Danh sách 22 đơn vị ──────────────────────────────────────────────────────
DS_DON_VI: list[str] = [DON_VI_CHI_NHANH] + DS_PGD

# Nhãn ngắn cho thông báo sau upload (không kèm icon cột)
NHAN_LOAI: dict[str, str] = {
    "hstd": "HSTD",
    "nq11": "NQ11",
    "gqvl": "GQVL",
    "cdtotkvv": "CDTOTKVV",
}


# ── Đọc tên đơn vị từ file upload ────────────────────────────────────────────

def _lay_ten_don_vi_trong_file(file_bytes: bytes, loai: str) -> str | None:
    """
    Đọc tên đơn vị từ file Excel theo từng loại.

    HSTD/GQVL : header=4, cột "Tên PGD"
    NQ11       : header=4, cột "Mã PGD" → tra MA_PGD_MAP
    CDTOTKVV   : header=7, dropna("Mã đơn vị"), cột "Tên đơn vị"
    """
    try:
        buf = BytesIO(file_bytes)
        if loai == "hstd":
            df = pd.read_excel(buf, sheet_name="BCQUERY", header=4, nrows=5)
            df = df.iloc[:, 1:]
            cot = next((c for c in df.columns if "tên pgd" in str(c).lower()), None)
            if cot is None:
                return None
            vals = df[cot].dropna().astype(str).unique()
            return str(vals[0]).strip() if len(vals) > 0 else None

        elif loai == "nq11":
            df = pd.read_excel(buf, sheet_name="BCQUERY", header=4, nrows=5)
            df = df.iloc[:, 1:]
            cot = next((c for c in df.columns if "mã pgd" in str(c).lower()), None)
            if cot is None:
                return None
            ma_pgd = str(df[cot].dropna().iloc[0]).strip().zfill(6) if not df[cot].dropna().empty else None
            if ma_pgd is None:
                return None
            return MA_PGD_MAP.get(ma_pgd)

        elif loai == "gqvl":
            # GQVL dùng sheet="Sheet1", header=7 (khác HSTD/NQ11)
            df = pd.read_excel(buf, sheet_name="Sheet1", header=7, nrows=10)
            df = df.iloc[:, 1:].dropna(how="all")
            cot = next(
                (c for c in df.columns
                 if any(k in str(c).lower()
                        for k in ["tên pgd", "ten pgd", "tên đơn vị", "đơn vị"])),
                None
            )
            if cot is not None:
                vals = df[cot].dropna().astype(str).unique()
                return str(vals[0]).strip() if len(vals) > 0 else None
            return None

        elif loai == "cdtotkvv":
            df = pd.read_excel(buf, header=7, nrows=10)
            cot_ma = next((c for c in df.columns if "mã đơn vị" in str(c).lower()), None)
            cot_ten = next((c for c in df.columns if "tên đơn vị" in str(c).lower()), None)
            if cot_ten is None:
                return None
            if cot_ma is not None:
                df = df.dropna(subset=[cot_ma])
            vals = df[cot_ten].dropna().astype(str).unique()
            return str(vals[0]).strip() if len(vals) > 0 else None

    except Exception:
        return None


def _kiem_tra_loai_file(file_bytes: bytes, loai: str) -> tuple[bool, str]:
    """
    Kiểm tra file có đúng loại khai báo không (HSTD/NQ11/GQVL/CDTOTKVV).
    Dựa vào sheet name đặc trưng.
    Trả về (ok, thong_bao).
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(
            BytesIO(file_bytes), read_only=True, data_only=True
        )
        sheet_names = [s.lower() for s in wb.sheetnames]

        if loai in ("hstd", "nq11"):
            # HSTD và NQ11 bắt buộc có sheet "bcquery"
            if "bcquery" not in sheet_names:
                return False, (
                    f"❌ File không phải **{loai.upper()}**! "
                    f"Thiếu sheet ‘BCQUERY’. "
                    f"Sheets có trong file: {', '.join(wb.sheetnames)}"
                )

        elif loai == "gqvl":
            # GQVL bắt buộc có sheet "sheet1", KHÔNG có sheet "bcquery"
            if "bcquery" in sheet_names:
                return False, (
                    "❌ Bạn đang upload file **HSTD/NQ11** vào ô **GQVL**! "
                    "Vui lòng chọn đúng file Sao kê GQVL."
                )
            if "sheet1" not in sheet_names:
                return False, (
                    f"❌ File không phải **GQVL**! "
                    f"Thiếu sheet ‘Sheet1’. "
                    f"Sheets có trong file: {', '.join(wb.sheetnames)}"
                )

        elif loai == "cdtotkvv":
            # CDTOTKVV không có sheet bcquery — chỉ kiểm tra KHÔNG phải HSTD
            if "bcquery" in sheet_names:
                return False, (
                    "❌ Bạn đang upload file **HSTD/NQ11** vào ô **CDTOTKVV**! "
                    "Vui lòng chọn đúng file Chấm điểm Tổ TK&VV."
                )

        return True, "OK"
    except Exception:
        # Không đọc được → fail-open, không chặn upload
        return True, "Bỏ qua kiểm tra loại file."


def _kiem_tra_don_vi(file_bytes: bytes, loai: str, ten_dv_chon: str) -> tuple[bool, str]:
    """
    Kiểm tra tên đơn vị trong file có khớp với đơn vị đang chọn.
    Trả về (khop: bool, thong_bao: str).
    """
    ten_trong_file = _lay_ten_don_vi_trong_file(file_bytes, loai)
    if ten_trong_file is None:
        return True, (
            f"⚠️ Không đọc được tên đơn vị từ file **{loai.upper()}**"
            f" — hệ thống tin tưởng bạn đang upload đúng file của "
            f"**{ten_dv_chon}**."
        )
    
    # Chuẩn hóa tên để so sánh (loại bỏ khoảng trắng, chuyển thường, loại bỏ ký tự đặc biệt)
    def chuan_hoa_ten(ten: str) -> str:
        import re
        ten = ten.strip().lower()
        # Loại bỏ các từ không quan trọng
        ten = re.sub(r'\b(pgd|phòng|giao|dịch|hội|sở|chi|nhánh|tỉnh)\b', '', ten)
        # Loại bỏ ký tự đặc biệt và khoảng trắng thừa
        ten = re.sub(r'[^\w\s]', ' ', ten)
        ten = re.sub(r'\s+', ' ', ten).strip()
        return ten
    
    ten_file_chuan = chuan_hoa_ten(ten_trong_file)
    ten_chon_chuan = chuan_hoa_ten(ten_dv_chon)
    
    # Kiểm tra khớp chính xác
    if ten_trong_file.strip() == ten_dv_chon.strip():
        return True, f"✅ Đơn vị khớp: **{ten_trong_file}**"
    
    # Kiểm tra khớp sau chuẩn hóa
    if ten_file_chuan == ten_chon_chuan:
        return True, f"✅ Đơn vị khớp (chuẩn hóa): **{ten_trong_file}**"
    
    # Kiểm tra chứa từ khóa chính
    if ten_file_chuan and ten_chon_chuan and len(ten_file_chuan) > 2:
        if ten_file_chuan in ten_chon_chuan or ten_chon_chuan in ten_file_chuan:
            return True, f"✅ Đơn vị tương tự: **{ten_trong_file}**"
    
    return False, (
        f"⚠️ Upload nhầm đơn vị! File chứa: **{ten_trong_file}** | Đang chọn: **{ten_dv_chon}**"
    )


# ── Form upload ───────────────────────────────────────────────────────────────

def _render_upload_form(ten_dv: str, prefix: str, username: str) -> None:
    """Form upload 4 file cho đơn vị. Không merge toàn CN (việc của KH-NV)."""
    st.markdown(f"##### 📤 Upload file cho: **{ten_dv}**")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.caption("📊 HSTD Chi tiết")
        f_hstd = st.file_uploader("HSTD", type=["xlsx", "xls"],
                                   key=f"{prefix}_hstd", label_visibility="collapsed")
    with c2:
        st.caption("📑 Sao kê NQ11")
        f_nq11 = st.file_uploader("NQ11", type=["xlsx", "xls"],
                                   key=f"{prefix}_nq11", label_visibility="collapsed")
    with c3:
        st.caption("📋 Sao kê GQVL")
        f_gqvl = st.file_uploader("GQVL", type=["xlsx", "xls"],
                                   key=f"{prefix}_gqvl", label_visibility="collapsed")
    with c4:
        st.caption("🏆 Chấm điểm Tổ TK&VV")
        f_cdtotkvv = st.file_uploader("CDTOTKVV", type=["xlsx", "xls"],
                                       key=f"{prefix}_cdtotkvv", label_visibility="collapsed")

    co_file = any(f is not None for f in [f_hstd, f_nq11, f_gqvl, f_cdtotkvv])
    if not co_file:
        st.info("Chọn ít nhất 1 file để bắt đầu upload.")
        return

    if st.button("📤 Upload", type="primary", key=f"{prefix}_btn_upload"):
        _xu_ly_upload(ten_dv, username,
                      f_hstd, f_nq11, f_gqvl, f_cdtotkvv)


def _xu_ly_upload(
    ten_dv: str,
    username: str,
    f_hstd, f_nq11, f_gqvl, f_cdtotkvv,
) -> None:
    """Xử lý upload, kiểm tra đơn vị, lưu file."""
    danh_sach_file = [
        ("hstd",     f_hstd),
        ("nq11",     f_nq11),
        ("gqvl",     f_gqvl),
        ("cdtotkvv", f_cdtotkvv),
    ]

    co_luu_thanh_cong = False

    for loai, f_obj in danh_sach_file:
        if f_obj is None:
            continue

        file_bytes = f_obj.read()

        # Kiểm tra cơ bản (định dạng, kích thước)
        ok_kt, msg_kt = kiem_tra_file(f_obj.name, file_bytes)
        if not ok_kt:
            nhan_kt = NHAN_LOAI.get(loai, loai.upper())
            st.error(f"❌ {nhan_kt}: {msg_kt}")
            continue

        # Kiểm tra loại file (sheet name)
        ok_loai, msg_loai = _kiem_tra_loai_file(file_bytes, loai)
        if not ok_loai:
            st.error(msg_loai)
            continue

        # Kiểm tra tên đơn vị trong file
        khop, msg_khop = _kiem_tra_don_vi(file_bytes, loai, ten_dv)
        if not khop:
            st.warning(msg_khop)
            continue

        path_excel = duong_dan_pgd(ten_dv, loai)
        pq = Path(path_excel).with_suffix(".parquet")
        if pq.exists():
            pq.unlink()

        try:
            kq = luu_pgd_file(ten_dv, loai, file_bytes)
        except Exception as e:
            kq = KetQuaUpload(False, f"Lỗi lưu: {e}")

        nhan = NHAN_LOAI.get(loai, loai.upper())

        if kq.thanh_cong:
            co_luu_thanh_cong = True
            mb = len(file_bytes) / 1024 / 1024
            db.ghi_audit(
                username,
                "upload_pgd_dia_ban",
                f"{loai.upper()} — {ten_dv} ({mb:.1f} MB)",
            )
            st.success(f"✅ Đã lưu **{nhan}** — {ten_dv}")

            meta_dq = lay_meta_chat_luong(loai)
            bao_cao = meta_dq.get("bao_cao", []) if meta_dq else []
            don_vi_loi = [
                r
                for r in bao_cao
                if r.get("don_vi") == ten_dv and r.get("so_loi", 0) > 0
            ]

            if don_vi_loi:
                r = don_vi_loi[0]
                st.warning(
                    f"⚠️ **Dữ liệu đã lưu nhưng có {r['so_loi']} vấn đề chất lượng "
                    f"cần kiểm tra:**\n\n"
                    + "\n".join(f"• {e}" for e in r.get("errors", []))
                    + "\n\n_Dữ liệu vẫn được hệ thống sử dụng bình thường._"
                )
        else:
            st.error(
                f"❌ **Không thể lưu {nhan}** — {ten_dv}\n\n"
                f"{kq.thong_bao}\n\n"
                "_Vui lòng kiểm tra lại file và thử upload lại._"
            )

    if co_luu_thanh_cong:
        st.success(
            "✅ Đã upload thành công. "
            "Phòng KH-NV sẽ tổng hợp dữ liệu toàn Chi nhánh."
        )

    st.cache_data.clear()
    st.rerun()


def _bang_trang_thai_don_vi(ten_dv: str) -> None:
    """Hiển thị trạng thái 4 loại file của một đơn vị."""
    LOAI_MAP = {
        "hstd":     "📊 HSTD",
        "nq11":     "📑 NQ11",
        "gqvl":     "📋 GQVL",
        "cdtotkvv": "🏆 CDTOTKVV",
    }
    cols = st.columns(4)
    for col, (loai, nhan) in zip(cols, LOAI_MAP.items()):
        info = doc_trang_thai_file(ten_dv, loai)
        with col:
            if not info["co_file"]:
                st.markdown(
                    f"**{nhan}**  \n❌ Chưa có file",
                    help="Chưa upload lần nào",
                )
            else:
                ngay = info["ngay_upload"].strftime("%d/%m/%Y %H:%M")
                so_ngay = info["so_ngay_cu"]
                canh_bao = info.get("canh_bao", "ok")
                icon = "✅" if canh_bao == "ok" else "⚠️"
                st.markdown(
                    f"**{nhan}**  \n{icon} {ngay}  \n"
                    f"<span style='font-size:0.82rem;color:#6b7280'>"
                    f"{so_ngay} ngày trước</span>",
                    unsafe_allow_html=True,
                )


# ── Entry point ───────────────────────────────────────────────────────────────

def render(tab=None, **kwargs) -> None:
    """
    Render tab Upload Dữ liệu cho CBTD địa bàn.
    Nhận tab (st.tab object) hoặc render trực tiếp trong context hiện tại.
    """
    role     = kwargs.get("role", "user")
    username = kwargs.get("username", "unknown")
    pgd_user = kwargs.get("pgd_user", "")

    ctx = tab if tab is not None else st

    with ctx:
        st.markdown("### 📤 Upload Dữ liệu — Phòng Giao Dịch")
        st.caption(
            "Upload file dữ liệu của PGD — dùng riêng cho Hỗ trợ địa bàn"
        )

        st.info(
            "📌 Dữ liệu upload tại đây chỉ dùng cho **Hỗ trợ địa bàn** "
            "(tra cứu, báo cáo giao ban, mẫu biểu). "
            "Không ảnh hưởng số liệu tổng hợp của **Phòng KH-NV**."
        )

        # ── Trạng thái theo đơn vị (không còn upload thủ công từng ô) ───────
        st.markdown("**📋 Trạng thái dữ liệu hiện tại**")
        if role == "user":
            ten_dv = pgd_user or (DS_PGD[0] if DS_PGD else DON_VI_CHI_NHANH)
            st.info(f"🏢 Đơn vị của bạn: **{ten_dv}**")
            _bang_trang_thai_don_vi(ten_dv)
        else:
            ten_xem = st.selectbox(
                "🏢 Đơn vị (xem trạng thái)",
                DS_PGD,
                key="pgd_trang_thai_chon_dv",
            )
            _bang_trang_thai_don_vi(ten_xem)

        st.divider()
        # Form upload riêng cho phân hệ Hỗ trợ địa bàn
        # Không dùng form của KH-NV — 2 luồng độc lập
        if role == "user":
            ten_dv = pgd_user or ""
            if ten_dv:
                _render_upload_form(ten_dv, "pgd_op", username)
            else:
                st.warning("⚠️ Không xác định được PGD. Liên hệ Admin.")
        else:
            ten_dv = st.selectbox(
                "🏢 Chọn PGD cần upload",
                DS_PGD,
                key="pgd_upload_op_chon_dv",
            )
            st.caption(
                "⚠️ Dữ liệu Hội sở Chi nhánh upload tại **Phòng KH-NV → Upload KH-NV**"
            )
            _render_upload_form(ten_dv, f"pgd_op_{ten_dv[:8]}", username)

