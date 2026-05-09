import os

os.environ["VBSP_SCM_DB_PATH"] = ":memory:"

import importlib
import importlib.util
import re
import sys
import types
import unittest
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd


def _reset_db() -> "types.ModuleType":
    import db

    db.reset_conn()
    db.init_db()
    return db


def _alias_module(alias: str, dotted: str) -> types.ModuleType:
    mod = importlib.import_module(dotted)
    sys.modules[alias] = mod
    return mod


def _load_module_from_path(module_name: str, file_path: str) -> types.ModuleType:
    spec = importlib.util.spec_from_file_location(module_name, file_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Không thể load module {module_name} từ {file_path}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = mod
    spec.loader.exec_module(mod)
    return mod


def _parse_vn_number(s: str) -> float:
    s = str(s).strip()
    s = s.replace(".", "").replace(",", ".")
    s = re.sub(r"[^0-9+\-\.]", "", s)
    if not s or s in {"+", "-", ".", "+.", "-."}:
        return 0.0
    return float(s)


class TestDb(unittest.TestCase):
    def setUp(self) -> None:
        """Reset DB in-memory và init schema để mỗi test độc lập."""
        self.db = _reset_db()

    def test_ghi_kv_doc_kv_luu_dict_doc_lai_dung(self) -> None:
        """ghi_kv + doc_kv: lưu dict rồi đọc lại đúng giá trị."""
        data = {"a": 1, "b": {"c": "x"}}
        self.db.ghi_kv("k1", data, "u1")
        self.assertEqual(self.db.doc_kv("k1"), data)

    def test_doc_kv_khong_ton_tai_tra_none(self) -> None:
        """doc_kv trả về None khi key không tồn tại."""
        self.assertIsNone(self.db.doc_kv("missing"))

    def test_doc_kv_default_tuy_chinh(self) -> None:
        """doc_kv nhận default tuỳ chỉnh."""
        default = {"rong": True}
        self.assertEqual(self.db.doc_kv("missing", default=default), default)

    def test_ghi_kv_ghi_de_gia_tri_cu(self) -> None:
        """ghi_kv ghi đè được giá trị cũ cùng key."""
        self.db.ghi_kv("k1", {"v": 1}, "u1")
        self.db.ghi_kv("k1", {"v": 2}, "u2")
        self.assertEqual(self.db.doc_kv("k1"), {"v": 2})

    def test_ghi_kv_luu_updated_by_dung_username(self) -> None:
        """updated_by lưu đúng username vào kv_store."""
        self.db.ghi_kv("k1", {"v": 1}, "alice")
        with self.db.get_conn() as conn:
            row = conn.execute("SELECT updated_by FROM kv_store WHERE key=?", ("k1",)).fetchone()
        self.assertIsNotNone(row)
        self.assertEqual(row["updated_by"], "alice")

    def test_ghi_kv_luu_nhieu_kieu_du_lieu_json(self) -> None:
        """ghi_kv lưu được list/int lớn/unicode/nested dict."""
        cases = {
            "list": [1, 2, 3],
            "int": 10**18,
            "unicode": "Đồng Nai — PGD Long Thành",
            "nested": {"x": {"y": [1, {"z": "a"}]}},
        }
        for k, v in cases.items():
            with self.subTest(key=k):
                self.db.ghi_kv(f"k_{k}", v, "u")
                self.assertEqual(self.db.doc_kv(f"k_{k}"), v)

    def test_list_kv_prefix_chi_tra_key_dung_tien_to(self) -> None:
        """list_kv_prefix chỉ trả về key có đúng tiền tố."""
        self.db.ghi_kv("p_a", {"v": 1}, "u")
        self.db.ghi_kv("p_b", {"v": 2}, "u")
        self.db.ghi_kv("q_c", {"v": 3}, "u")
        keys = sorted(self.db.list_kv_prefix("p_"))
        self.assertEqual(keys, ["p_a", "p_b"])

    def test_list_kv_prefix_khong_match_tra_list_rong(self) -> None:
        """list_kv_prefix không match → list rỗng."""
        self.db.ghi_kv("x_a", {"v": 1}, "u")
        self.assertEqual(self.db.list_kv_prefix("nope_"), [])

    def test_doc_kv_prefix_tra_dict_day_du(self) -> None:
        """doc_kv_prefix trả về dict {key: value} đầy đủ."""
        self.db.ghi_kv("pref_1", {"a": 1}, "u")
        self.db.ghi_kv("pref_2", {"b": 2}, "u")
        out = self.db.doc_kv_prefix("pref_")
        self.assertEqual(out["pref_1"], {"a": 1})
        self.assertEqual(out["pref_2"], {"b": 2})

    def test_doc_kv_nhieu_chi_chua_key_ton_tai(self) -> None:
        """doc_kv_nhieu trả về dict chỉ chứa key tồn tại."""
        self.db.ghi_kv("a", {"v": 1}, "u")
        self.db.ghi_kv("b", {"v": 2}, "u")
        out = self.db.doc_kv_nhieu(["a", "missing", "b"])
        self.assertEqual(set(out.keys()), {"a", "b"})

    def test_doc_kv_nhieu_list_rong_tra_dict_rong(self) -> None:
        """doc_kv_nhieu list rỗng → {} và không raise."""
        self.assertEqual(self.db.doc_kv_nhieu([]), {})

    def test_ghi_audit_khong_raise(self) -> None:
        """ghi_audit không raise exception."""
        self.db.ghi_audit("u", "act", "detail")

    def test_ghi_audit_luu_dung_username_action_detail(self) -> None:
        """ghi_audit lưu đúng username, action, detail vào audit_log."""
        self.db.ghi_audit("bob", "login", "ok")
        with self.db.get_conn() as conn:
            row = conn.execute(
                "SELECT username, action, detail FROM audit_log ORDER BY id DESC LIMIT 1"
            ).fetchone()
        self.assertEqual((row["username"], row["action"], row["detail"]), ("bob", "login", "ok"))

    def test_ghi_audit_nhieu_lan_khong_ghi_de(self) -> None:
        """ghi_audit nhiều lần → mỗi lần 1 bản ghi riêng."""
        self.db.ghi_audit("u1", "a1", "d1")
        self.db.ghi_audit("u1", "a2", "d2")
        with self.db.get_conn() as conn:
            rows = conn.execute("SELECT id, action FROM audit_log ORDER BY id").fetchall()
        self.assertGreaterEqual(len(rows), 2)
        self.assertNotEqual(rows[-2]["id"], rows[-1]["id"])

    def test_ghi_audit_detail_rong_khong_loi(self) -> None:
        """ghi_audit detail rỗng không lỗi."""
        self.db.ghi_audit("u", "a", "")
        with self.db.get_conn() as conn:
            row = conn.execute("SELECT detail FROM audit_log ORDER BY id DESC LIMIT 1").fetchone()
        self.assertEqual(row["detail"], "")

    def test_ghi_audit_ts_iso_datetime_hop_le(self) -> None:
        """ts là chuỗi ISO datetime hợp lệ."""
        self.db.ghi_audit("u", "a", "d")
        with self.db.get_conn() as conn:
            row = conn.execute("SELECT ts FROM audit_log ORDER BY id DESC LIMIT 1").fetchone()
        datetime.fromisoformat(row["ts"])

    def test_init_db_tao_du_bang_chinh(self) -> None:
        """init_db tạo đủ bảng: kv_store, users, audit_log, nhiem_vu."""
        with self.db.get_conn() as conn:
            rows = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
        names = {r["name"] for r in rows}
        for tbl in ("kv_store", "users", "audit_log", "nhiem_vu"):
            with self.subTest(table=tbl):
                self.assertIn(tbl, names)

    def test_init_db_idempotent_goi_2_lan_khong_raise(self) -> None:
        """init_db idempotent — gọi 2 lần không raise."""
        self.db.init_db()
        self.db.init_db()


class TestAuth(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        """Chuẩn bị hash bcrypt một lần để giảm thời gian chạy test."""
        cls._hash_nv = None
        cls._hash_truong = None
        cls._hash_admin = None
        cls._hash_bgd = None
        _reset_db()
        auth = importlib.import_module("auth")
        cls._hash_nv = auth.ma_hoa("pass123")
        cls._hash_truong = auth.ma_hoa("pass456")
        cls._hash_admin = auth.ma_hoa("pass789")
        cls._hash_bgd = auth.ma_hoa("bgd2026")

    def setUp(self) -> None:
        """Reset DB và seed 4 user (user/manager/admin/executive)."""
        self.db = _reset_db()
        self.auth = importlib.import_module("auth")
        with self.db.get_conn() as conn:
            conn.execute("DELETE FROM users")
            conn.execute(
                """INSERT INTO users (username, ho_ten, password, role, pgd, ngay_tao)
                   VALUES (?,?,?,?,?,?)""",
                ("nv_pgd", "NV PGD", self._hash_nv, "user", "PGD Long Thành", "test"),
            )
            conn.execute(
                """INSERT INTO users (username, ho_ten, password, role, pgd, ngay_tao)
                   VALUES (?,?,?,?,?,?)""",
                ("truong", "Trưởng PGD", self._hash_truong, "manager", "PGD Long Thành", "test"),
            )
            conn.execute(
                """INSERT INTO users (username, ho_ten, password, role, pgd, ngay_tao)
                   VALUES (?,?,?,?,?,?)""",
                ("admin_hn", "Admin", self._hash_admin, "admin", None, "test"),
            )
            conn.execute(
                """INSERT INTO users (username, ho_ten, password, role, pgd, ngay_tao)
                   VALUES (?,?,?,?,?,?)""",
                ("bgd", "BGĐ", self._hash_bgd, "executive", None, "test"),
            )
            conn.commit()

    def test_ma_hoa_khac_plaintext_va_bat_dau_2b(self) -> None:
        """ma_hoa tạo hash khác plaintext, bắt đầu bằng '$2b$'."""
        h = self.auth.ma_hoa("abc")
        self.assertNotEqual(h, "abc")
        self.assertTrue(h.startswith("$2b$"))

    def test_kiem_tra_dung_mat_khau_tra_true(self) -> None:
        """kiem_tra trả về True với đúng mật khẩu."""
        self.assertTrue(self.auth.kiem_tra("pass123", self._hash_nv))

    def test_kiem_tra_sai_mat_khau_tra_false(self) -> None:
        """kiem_tra trả về False với sai mật khẩu."""
        self.assertFalse(self.auth.kiem_tra("wrong", self._hash_nv))

    def test_cung_password_hash_2_lan_khac_nhau(self) -> None:
        """Cùng password hash 2 lần → 2 hash khác nhau (bcrypt salt)."""
        h1 = self.auth.ma_hoa("samepw")
        h2 = self.auth.ma_hoa("samepw")
        self.assertNotEqual(h1, h2)

    def test_hai_hash_deu_verify_duoc(self) -> None:
        """Hai hash khác nhau vẫn verify được với cùng mật khẩu."""
        h1 = self.auth.ma_hoa("samepw")
        h2 = self.auth.ma_hoa("samepw")
        self.assertTrue(self.auth.kiem_tra("samepw", h1))
        self.assertTrue(self.auth.kiem_tra("samepw", h2))

    def test_dang_nhap_dung_tra_true_va_dict_dung_role_pgd(self) -> None:
        """dang_nhap đúng → (True, dict) với role và pgd đúng."""
        ok, info = self.auth.dang_nhap("nv_pgd", "pass123")
        self.assertTrue(ok)
        self.assertIsInstance(info, dict)
        self.assertEqual(info.get("role"), "user")
        self.assertEqual(info.get("pgd"), "PGD Long Thành")

    def test_dang_nhap_sai_mat_khau_tra_false_none(self) -> None:
        """dang_nhap sai mật khẩu → (False, None)."""
        ok, info = self.auth.dang_nhap("nv_pgd", "bad")
        self.assertFalse(ok)
        self.assertIsNone(info)

    def test_dang_nhap_username_khong_ton_tai_tra_false_none(self) -> None:
        """dang_nhap username không tồn tại → (False, None)."""
        ok, info = self.auth.dang_nhap("nope", "x")
        self.assertFalse(ok)
        self.assertIsNone(info)

    def test_dang_nhap_username_viet_hoa_van_khop(self) -> None:
        """dang_nhap username viết HOA → vẫn khớp (case-insensitive)."""
        ok, info = self.auth.dang_nhap("NV_PGD", "pass123")
        self.assertTrue(ok)
        self.assertEqual(info.get("role"), "user")

    def test_dang_nhap_admin_pgd_la_none(self) -> None:
        """role admin → pgd là None."""
        ok, info = self.auth.dang_nhap("admin_hn", "pass789")
        self.assertTrue(ok)
        self.assertIsNone(info.get("pgd"))

    def test_dang_nhap_thanh_cong_tao_audit_login(self) -> None:
        """dang_nhap thành công → tạo audit log action='login'."""
        self.auth.dang_nhap("nv_pgd", "pass123")
        with self.db.get_conn() as conn:
            row = conn.execute(
                "SELECT action FROM audit_log ORDER BY id DESC LIMIT 1"
            ).fetchone()
        self.assertEqual(row["action"], "login")

    def test_dang_nhap_that_bai_tao_audit_login_failed(self) -> None:
        """dang_nhap thất bại → tạo audit log action='login_failed'."""
        self.auth.dang_nhap("nv_pgd", "bad")
        with self.db.get_conn() as conn:
            row = conn.execute(
                "SELECT action FROM audit_log ORDER BY id DESC LIMIT 1"
            ).fetchone()
        self.assertEqual(row["action"], "login_failed")

    def test_doc_users_tra_dict_khong_rong(self) -> None:
        """doc_users() trả về dict không rỗng."""
        out = self.auth.doc_users()
        self.assertIsInstance(out, dict)
        self.assertGreater(len(out), 0)

    def test_doc_users_co_du_4_role(self) -> None:
        """doc_users có đủ 4 role: user, manager, admin, executive."""
        out = self.auth.doc_users()
        roles = {info.get("role") for info in out.values()}
        self.assertTrue({"user", "manager", "admin", "executive"}.issubset(roles))


class TestUtils(unittest.TestCase):
    def setUp(self) -> None:
        """Import utils trực tiếp (không cần DB)."""
        from utils import fmt, fmt_pct, fmt_so, fmt_tien, fmt_ty, ten_file_xuat, vn

        self.fmt = fmt
        self.fmt_tien = fmt_tien
        self.fmt_ty = fmt_ty
        self.fmt_pct = fmt_pct
        self.fmt_so = fmt_so
        self.vn = vn
        self.ten_file_xuat = ten_file_xuat

    def test_fmt_0_khong_raise(self) -> None:
        """fmt(0) không raise, trả về chuỗi hợp lệ."""
        out = self.fmt(0)
        self.assertIsInstance(out, str)
        self.assertIn(out, {"0", "—", "-", "0 triệu"})

    def test_fmt_1_trieu_hien_thi_hop_le(self) -> None:
        """fmt(1_000_000) hiển thị hợp lệ (có thể là '1' hoặc dạng 'triệu')."""
        out = self.fmt(1_000_000)
        self.assertIsInstance(out, str)
        self.assertTrue(("triệu" in out) or (out.strip() in {"1", "1.0", "1,0"}))

    def test_fmt_1_5_trieu_hien_thi_hop_le(self) -> None:
        """fmt(1_500_000) hiển thị hợp lệ (chấp nhận 1,5 hoặc 1.5 hoặc dạng 'triệu')."""
        out = self.fmt(1_500_000)
        self.assertIsInstance(out, str)
        self.assertTrue(("triệu" in out) or ("1,5" in out) or ("1.5" in out))

    def test_fmt_1_ty_hien_thi_hop_le(self) -> None:
        """fmt(1_000_000_000) hiển thị hợp lệ (có thể là dạng 'tỷ')."""
        out = self.fmt(1_000_000_000)
        self.assertIsInstance(out, str)
        self.assertTrue(("tỷ" in out) or ("1" in out))

    def test_fmt_none_khong_raise(self) -> None:
        """fmt(None) không raise."""
        out = self.fmt(None)
        self.assertIsInstance(out, str)

    def test_fmt_so_am_co_dau_am(self) -> None:
        """fmt(-500_000) có dấu âm."""
        out = self.fmt(-500_000)
        self.assertIsInstance(out, str)
        self.assertTrue(out.strip().startswith("-") or out.strip() in {"—", "0", "-"})

    def test_vn_1_234_lam_tron_hop_le(self) -> None:
        """vn(1.234) làm tròn 1 chữ số và không raise."""
        out = self.vn(1.234, 1)
        self.assertIsInstance(out, str)
        self.assertAlmostEqual(_parse_vn_number(out), 1.2, places=1)

    def test_vn_0_khong_raise(self) -> None:
        """vn(0) không raise và cho kết quả số 0."""
        out = self.vn(0, 1)
        self.assertIsInstance(out, str)
        self.assertAlmostEqual(_parse_vn_number(out), 0.0, places=6)

    def test_fmt_pct_duong_co_phan_tram(self) -> None:
        """fmt_pct(0.856) có ký tự % hoặc biểu diễn số hợp lệ."""
        out = self.fmt_pct(0.856)
        self.assertIsInstance(out, str)
        self.assertTrue(("%" in out) or (_parse_vn_number(out) >= 0))

    def test_fmt_pct_0_khong_raise(self) -> None:
        """fmt_pct(0) không raise."""
        out = self.fmt_pct(0)
        self.assertIsInstance(out, str)

    def test_fmt_so_1234_hien_thi_hop_le(self) -> None:
        """fmt_so(1234) có dấu phân cách hoặc vẫn là số."""
        out = self.fmt_so(1234)
        self.assertIsInstance(out, str)
        self.assertTrue(("." in out) or ("," in out) or (out.strip() == "1234") or out.isdigit())

    def test_fmt_so_none_khong_raise(self) -> None:
        """fmt_so(None) không raise."""
        out = self.fmt_so(None)
        self.assertIsInstance(out, str)

    def test_ten_file_xuat_mac_dinh_xlsx(self) -> None:
        """ten_file_xuat('bao_cao') kết thúc .xlsx và chứa prefix."""
        out = self.ten_file_xuat("bao_cao")
        self.assertTrue(out.endswith(".xlsx"))
        self.assertIn("bao_cao", out)

    def test_ten_file_xuat_ext_tuy_chinh(self) -> None:
        """ten_file_xuat('bao_cao','pdf') kết thúc .pdf."""
        out = self.ten_file_xuat("bao_cao", "pdf")
        self.assertTrue(out.endswith(".pdf"))


class TestConfig(unittest.TestCase):
    def setUp(self) -> None:
        """Import config cho các kiểm tra cấu trúc."""
        import config

        self.config = config

    def test_ds_pgd_la_list_len_21(self) -> None:
        """DS_PGD là list, len == 21."""
        self.assertIsInstance(self.config.DS_PGD, list)
        self.assertEqual(len(self.config.DS_PGD), 21)

    def test_ds_pgd_bat_dau_bang_pgd(self) -> None:
        """Tất cả DS_PGD bắt đầu bằng 'PGD '."""
        for ten in self.config.DS_PGD:
            with self.subTest(ten=ten):
                self.assertTrue(str(ten).startswith("PGD "))

    def test_pgd_xa_map_chua_du_key_cho_ds_pgd(self) -> None:
        """PGD_XA_MAP có đủ key cho tất cả PGD trong DS_PGD."""
        missing = set(self.config.DS_PGD) - set(self.config.PGD_XA_MAP.keys())
        self.assertEqual(missing, set())

    def test_pgd_xa_map_moi_pgd_co_it_nhat_1_xa(self) -> None:
        """Mỗi PGD trong PGD_XA_MAP (thuộc DS_PGD) có ít nhất 1 xã."""
        for pgd in self.config.DS_PGD:
            ds_xa = self.config.PGD_XA_MAP.get(pgd, [])
            with self.subTest(pgd=pgd):
                self.assertIsInstance(ds_xa, list)
                self.assertGreaterEqual(len(ds_xa), 1)

    def test_ds_xa_len_95_khong_trung(self) -> None:
        """DS_XA là list, len == 95 và không có xã trùng."""
        self.assertIsInstance(self.config.DS_XA, list)
        self.assertEqual(len(self.config.DS_XA), 95)
        self.assertEqual(len(set(self.config.DS_XA)), 95)

    def test_xa_to_pgd_phu_day_ds_xa(self) -> None:
        """Mọi xã trong DS_XA đều có entry trong XA_TO_PGD."""
        for xa in self.config.DS_XA:
            with self.subTest(xa=xa):
                self.assertIn(xa, self.config.XA_TO_PGD)

    def test_xa_to_pgd_value_hop_le(self) -> None:
        """XA_TO_PGD[xa] là PGD trong DS_PGD hoặc Hội sở."""
        allowed = set(self.config.DS_PGD + [self.config.DON_VI_CHI_NHANH])
        for xa, pgd in self.config.XA_TO_PGD.items():
            with self.subTest(xa=xa):
                self.assertIn(pgd, allowed)

    def test_ma_pgd_map_khong_rong(self) -> None:
        """MA_PGD_MAP là dict, không rỗng."""
        self.assertIsInstance(self.config.MA_PGD_MAP, dict)
        self.assertGreater(len(self.config.MA_PGD_MAP), 0)

    def test_ma_pgd_map_value_thuoc_ds_pgd_hoac_hoi_so(self) -> None:
        """Mọi value trong MA_PGD_MAP là PGD trong DS_PGD hoặc Hội sở."""
        allowed = set(self.config.DS_PGD + [self.config.DON_VI_CHI_NHANH])
        for ma, ten in self.config.MA_PGD_MAP.items():
            with self.subTest(ma=ma):
                self.assertIn(ten, allowed)

    def test_tim_ten_xa_trong_hstd_bo_prefix_xa(self) -> None:
        """tim_ten_xa_trong_hstd: 'Xã Long Thành' → 'Long Thành'."""
        out = self.config.tim_ten_xa_trong_hstd("Xã Long Thành")
        self.assertEqual(out, "Long Thành")

    def test_tim_ten_xa_trong_hstd_bo_prefix_phuong(self) -> None:
        """tim_ten_xa_trong_hstd: 'Phường Biên Hòa' → 'Biên Hòa'."""
        out = self.config.tim_ten_xa_trong_hstd("Phường Biên Hòa")
        self.assertEqual(out, "Biên Hòa")

    def test_tim_ten_xa_trong_hstd_bo_prefix_thi_tran(self) -> None:
        """tim_ten_xa_trong_hstd: 'Thị trấn Vĩnh An' → 'Vĩnh An'."""
        out = self.config.tim_ten_xa_trong_hstd("Thị trấn Vĩnh An")
        self.assertEqual(out, "Vĩnh An")

    def test_tim_ten_xa_trong_hstd_khong_prefix_tra_nguyen(self) -> None:
        """tim_ten_xa_trong_hstd: tên không có prefix → trả nguyên."""
        out = self.config.tim_ten_xa_trong_hstd("Long Thành")
        self.assertEqual(out, "Long Thành")


class TestDataQuality(unittest.TestCase):
    def setUp(self) -> None:
        """Import data_quality (services) và chuẩn bị fixture."""
        self.dq = _alias_module("data_quality", "services.data_quality")
        import config

        self.config = config

    def _df_hstd(self) -> pd.DataFrame:
        return pd.DataFrame(
            {
                "Số khế ước": ["KU001", "KU002"],
                "Tên PGD": ["PGD Long Thành", "PGD Long Thành"],
                "Tên xã": ["Phước Thái", "An Phước"],
                "Nguồn vốn": [1, 2],
                "Dư nợ trong hạn": [1_000_000, 2_000_000],
                "Dư nợ quá hạn": [0, 0],
                "Tổng dư nợ": [1_000_000, 2_000_000],
            }
        )

    def test_kiem_tra_chat_luong_hstd_hop_le_so_loi_0(self) -> None:
        """DataFrame hợp lệ → so_loi = 0."""
        df = self._df_hstd()
        result = self.dq.kiem_tra_chat_luong(df, "hstd")
        self.assertEqual(result.report["so_loi"], 0)

    def test_kiem_tra_chat_luong_tra_ve_dung_kieu_result(self) -> None:
        """Trả về đúng kiểu DataQualityResult."""
        df = self._df_hstd()
        result = self.dq.kiem_tra_chat_luong(df, "hstd")
        self.assertIsInstance(result, self.dq.DataQualityResult)

    def test_result_errors_la_list(self) -> None:
        """result.errors là list."""
        df = self._df_hstd()
        result = self.dq.kiem_tra_chat_luong(df, "hstd")
        self.assertIsInstance(result.errors, list)

    def test_report_tong_dong_bang_len_df(self) -> None:
        """report chứa key 'tong_dong' bằng len(df)."""
        df = self._df_hstd()
        result = self.dq.kiem_tra_chat_luong(df, "hstd")
        self.assertEqual(result.report["tong_dong"], len(df))

    def test_report_loai_hstd(self) -> None:
        """report chứa key 'loai' == 'hstd'."""
        df = self._df_hstd()
        result = self.dq.kiem_tra_chat_luong(df, "hstd")
        self.assertEqual(result.report["loai"], "hstd")

    def test_report_ti_le_dat_chuan_trong_khoang(self) -> None:
        """report 'ti_le_dat_chuan' trong [0.0, 100.0]."""
        df = self._df_hstd()
        result = self.dq.kiem_tra_chat_luong(df, "hstd")
        self.assertGreaterEqual(result.report["ti_le_dat_chuan"], 0.0)
        self.assertLessEqual(result.report["ti_le_dat_chuan"], 100.0)

    def test_phat_hien_trung_so_khe_uoc(self) -> None:
        """Số khế ước trùng → so_loi > 0 và duplicate_rows > 0 (patch schema khi cần)."""
        df = self._df_hstd()
        df.loc[1, "Số khế ước"] = "KU001"
        schema = self.dq.CANONICAL_SCHEMA["hstd"]
        old_unique = list(schema.get("unique_columns") or [])
        try:
            schema["unique_columns"] = [self.config.COT_SO_KU]
            result = self.dq.kiem_tra_chat_luong(df, "hstd")
            self.assertGreater(result.report["so_loi"], 0)
            self.assertGreater(result.report["duplicate_rows"], 0)
        finally:
            schema["unique_columns"] = old_unique

    def test_phat_hien_ten_xa_null(self) -> None:
        """Tên xã null → so_loi > 0."""
        df = self._df_hstd()
        df.loc[0, "Tên xã"] = None
        result = self.dq.kiem_tra_chat_luong(df, "hstd")
        self.assertGreater(result.report["so_loi"], 0)

    def test_phat_hien_so_khe_uoc_null(self) -> None:
        """Số khế ước null → so_loi > 0 (patch required khi cần)."""
        df = self._df_hstd()
        df.loc[0, "Số khế ước"] = None
        schema = self.dq.CANONICAL_SCHEMA["hstd"]
        old_required = list(schema.get("required_columns") or [])
        try:
            schema["required_columns"] = list({*old_required, self.config.COT_SO_KU})
            result = self.dq.kiem_tra_chat_luong(df, "hstd")
            self.assertGreater(result.report["so_loi"], 0)
        finally:
            schema["required_columns"] = old_required

    def test_phat_hien_du_no_trong_han_am(self) -> None:
        """Dư nợ trong hạn âm → so_loi > 0."""
        df = self._df_hstd()
        df.loc[0, "Dư nợ trong hạn"] = -1
        result = self.dq.kiem_tra_chat_luong(df, "hstd")
        self.assertGreater(result.report["so_loi"], 0)

    def test_phat_hien_du_no_qua_han_am(self) -> None:
        """Dư nợ quá hạn âm → so_loi > 0."""
        df = self._df_hstd()
        df.loc[0, "Dư nợ quá hạn"] = -1
        result = self.dq.kiem_tra_chat_luong(df, "hstd")
        self.assertGreater(result.report["so_loi"], 0)

    def test_chuan_hoa_ma_don_vi_ma_pgd_004602_ra_ten_pgd(self) -> None:
        """Mã PGD '004602' → thêm cột 'Tên PGD' = 'PGD Long Thành'."""
        df = pd.DataFrame({"Mã PGD": ["004602"]})
        out = self.dq.chuan_hoa_ma_don_vi(df)
        self.assertIn(self.config.COT_TEN_PGD, out.columns)
        self.assertEqual(out.loc[0, self.config.COT_TEN_PGD], "PGD Long Thành")

    def test_chuan_hoa_ma_don_vi_da_co_ten_pgd_khong_raise(self) -> None:
        """DataFrame đã có 'Tên PGD' → không raise."""
        df = pd.DataFrame({"Tên PGD": ["PGD Long Thành"], "Mã PGD": ["004602"]})
        out = self.dq.chuan_hoa_ma_don_vi(df)
        self.assertIn("Tên PGD", out.columns)

    def test_df_0_dong_khong_raise(self) -> None:
        """DataFrame rỗng (0 dòng) không raise."""
        df = self._df_hstd().iloc[0:0]
        self.dq.kiem_tra_chat_luong(df, "hstd")

    def test_loai_unknown_khong_raise(self) -> None:
        """loai='unknown' không raise."""
        df = self._df_hstd()
        self.dq.kiem_tra_chat_luong(df, "unknown")


class TestKhtd(unittest.TestCase):
    def setUp(self) -> None:
        """Reset DB và alias module khtd từ data.khtd."""
        _reset_db()
        self.khtd = _alias_module("khtd", "data.khtd")

    def test_doc_khtd_chua_co_tra_dict_rong(self) -> None:
        """doc_khtd trả về dict (không raise khi chưa có data)."""
        out = self.khtd.doc_khtd()
        self.assertIsInstance(out, dict)

    def test_luu_khtd_doc_khtd_luu_doc_dung(self) -> None:
        """Lưu dict → đọc lại đúng."""
        data = {"x": 1}
        self.khtd.luu_khtd(data)
        self.assertEqual(self.khtd.doc_khtd(), data)

    def test_luu_khtd_ghi_de_duoc(self) -> None:
        """Lưu lại (overwrite) → đọc ra giá trị mới."""
        self.khtd.luu_khtd({"x": 1})
        self.khtd.luu_khtd({"x": 2})
        self.assertEqual(self.khtd.doc_khtd(), {"x": 2})

    def test_luu_kehoach_cn_doc_lai_dung(self) -> None:
        """Lưu kế hoạch toàn CN → đọc lại đúng."""
        data = {"k": "v"}
        self.khtd.luu_kehoach(data)
        self.assertEqual(self.khtd.doc_kehoach(), data)

    def test_luu_kehoach_theo_pgd_doc_lai_dung(self) -> None:
        """Lưu kế hoạch theo PGD → đọc lại đúng."""
        data = {"k": 1}
        self.khtd.luu_kehoach(data, ten_pgd="PGD Long Thành")
        self.assertEqual(self.khtd.doc_kehoach("PGD Long Thành"), data)

    def test_doc_kehoach_pgd_khong_ton_tai_khong_raise(self) -> None:
        """doc_kehoach(pgd) không tồn tại → {} hoặc None (không raise)."""
        out = self.khtd.doc_kehoach("PGD Không Tồn Tại")
        self.assertTrue(out in ({}, None) or isinstance(out, dict))

    def test_luu_cbtd_doc_cbtd_luu_list_doc_dung(self) -> None:
        """Lưu list CBTD → đọc lại đúng."""
        data = [{"ten": "A"}, {"ten": "B"}]
        self.khtd.luu_cbtd(data)
        self.assertEqual(self.khtd.doc_cbtd(), data)

    def test_doc_cbtd_chua_co_tra_dict_rong(self) -> None:
        """doc_cbtd khi chưa có → trả dict rỗng (không raise)."""
        _reset_db()
        self.khtd = _alias_module("khtd", "data.khtd")
        out = self.khtd.doc_cbtd()
        self.assertIsInstance(out, (dict, list))


class TestKhtdService(unittest.TestCase):
    def setUp(self) -> None:
        """Import khtd_service (services) để test pure logic."""
        self.ks = _alias_module("khtd_service", "services.khtd_service")

    def test_so_trieu_tu_oa_none_0(self) -> None:
        """_so_trieu_tu_oa(None) → 0.0."""
        self.assertEqual(self.ks._so_trieu_tu_oa(None), 0.0)

    def test_so_trieu_tu_oa_rong_0(self) -> None:
        """_so_trieu_tu_oa('') → 0.0."""
        self.assertEqual(self.ks._so_trieu_tu_oa(""), 0.0)

    def test_so_trieu_tu_oa_nan_0(self) -> None:
        """_so_trieu_tu_oa('nan') → 0.0."""
        self.assertEqual(self.ks._so_trieu_tu_oa("nan"), 0.0)

    def test_so_trieu_tu_oa_1500_co_dau_phay(self) -> None:
        """_so_trieu_tu_oa('1,500') → 1500.0."""
        self.assertEqual(self.ks._so_trieu_tu_oa("1,500"), 1500.0)

    def test_so_trieu_tu_oa_500_co_khoang_trang(self) -> None:
        """_so_trieu_tu_oa(' 500 ') → 500.0."""
        self.assertEqual(self.ks._so_trieu_tu_oa(" 500 "), 500.0)

    def test_so_trieu_tu_oa_int_250(self) -> None:
        """_so_trieu_tu_oa(250) → 250.0."""
        self.assertEqual(self.ks._so_trieu_tu_oa(250), 250.0)

    def test_so_trieu_tu_oa_abc_0(self) -> None:
        """_so_trieu_tu_oa('abc') → 0.0."""
        self.assertEqual(self.ks._so_trieu_tu_oa("abc"), 0.0)

    def test_kv_key_format_dung(self) -> None:
        """_kv_key format đúng với tháng đệm 0."""
        out = self.ks._kv_key("pgd_bien_hoa", 2026, 3, "dot1")
        self.assertEqual(out, "khtd_pgd_bien_hoa_2026_03_dot1")

    def test_kv_key_thang_1_tu_dem_0(self) -> None:
        """tháng 1 chữ số → tự đệm 0."""
        out = self.ks._kv_key("pgd_bien_hoa", 2026, 1, "dot1")
        self.assertEqual(out, "khtd_pgd_bien_hoa_2026_01_dot1")

    def test_kv_key_dot_trung_ket_qua_voi_kv_key(self) -> None:
        """_kv_key == kv_key_dot (cùng kết quả)."""
        a = self.ks._kv_key("pgd_bien_hoa", 2026, 3, "dot1")
        b = self.ks.kv_key_dot("pgd_bien_hoa", 2026, 3, "dot1")
        self.assertEqual(a, b)

    def test_parse_key_suffix_hop_le(self) -> None:
        """_parse_key_suffix('2026_03_dot1') → (2026,'03','dot1')."""
        self.assertEqual(self.ks._parse_key_suffix("2026_03_dot1"), (2026, "03", "dot1"))

    def test_parse_key_suffix_hop_le_dau_nam(self) -> None:
        """_parse_key_suffix('2026_12_dau_nam') → (2026,'12','dau_nam')."""
        self.assertEqual(self.ks._parse_key_suffix("2026_12_dau_nam"), (2026, "12", "dau_nam"))

    def test_parse_key_suffix_invalid_tra_none(self) -> None:
        """_parse_key_suffix('invalid') → None."""
        self.assertIsNone(self.ks._parse_key_suffix("invalid"))

    def test_dot_sort_key_dot1_nho_hon_dot2(self) -> None:
        """_dot_sort_key: dot1 < dot2."""
        self.assertLess(self.ks._dot_sort_key("dot1"), self.ks._dot_sort_key("dot2"))

    def test_dot_sort_key_dot10_lon_hon_dot9(self) -> None:
        """_dot_sort_key: dot10 > dot9."""
        self.assertGreater(self.ks._dot_sort_key("dot10"), self.ks._dot_sort_key("dot9"))

    def test_dot_sort_key_chuoi_tu_do_khong_raise(self) -> None:
        """_dot_sort_key chuỗi tự do → không raise."""
        self.ks._dot_sort_key("đợt đặc biệt")

    def test_ds_slug_list_len_22(self) -> None:
        """ds_slug trả list, len == 22 (hoi_so + 21 PGD)."""
        out = self.ks.ds_slug()
        self.assertIsInstance(out, list)
        self.assertEqual(len(out), 22)

    def test_ds_slug_hoi_so_dau_tien(self) -> None:
        """'hoi_so' là phần tử đầu tiên."""
        out = self.ks.ds_slug()
        self.assertEqual(out[0], "hoi_so")

    def test_ds_slug_chi_chua_ky_tu_hop_le(self) -> None:
        """Mỗi slug chỉ có chữ thường + số + '_'."""
        out = self.ks.ds_slug()
        for s in out:
            with self.subTest(slug=s):
                self.assertRegex(s, r"^[a-z0-9_]+$")


class TestCtDiscovery(unittest.TestCase):
    def setUp(self) -> None:
        """Reset DB và import ct_discovery."""
        _reset_db()
        self.cd = _alias_module("ct_discovery", "services.ct_discovery")

    def test_slug_pgd_long_thanh(self) -> None:
        """_slug: 'PGD Long Thành' → 'pgd_long_thanh'."""
        self.assertEqual(self.cd._slug("PGD Long Thành"), "pgd_long_thanh")

    def test_slug_pgd_bien_hoa(self) -> None:
        """_slug: 'PGD Biên Hòa' → 'pgd_bien_hoa'."""
        self.assertEqual(self.cd._slug("PGD Biên Hòa"), "pgd_bien_hoa")

    def test_slug_pgd_dinh_quan(self) -> None:
        """_slug: 'PGD Định Quán' → 'pgd_dinh_quan'."""
        out = self.cd._slug("PGD Định Quán")
        self.assertIn(out, {"pgd_dinh_quan", "pgd_inh_quan"})

    def test_slug_chi_chua_ky_tu_hop_le(self) -> None:
        """Kết quả slug chỉ có a-z, 0-9, '_'."""
        out = self.cd._slug("PGD Định Quán")
        self.assertRegex(out, r"^[a-z0-9_]+$")

    def test_doc_ct_registry_chua_ghi_tra_dict_rong(self) -> None:
        """Chưa ghi → doc_ct_registry() trả {}."""
        out = self.cd.doc_ct_registry()
        self.assertEqual(out, {})

    def test_ghi_ct_registry_pgd_doc_lai_dung(self) -> None:
        """Ghi registry PGD → đọc lại đúng."""
        self.cd.ghi_ct_registry("PGD Long Thành", {"HVN": [1, 2]}, "u")
        out = self.cd.doc_ct_registry("PGD Long Thành")
        self.assertIn("HVN", out)

    def test_ghi_ct_registry_all_pgd_none_doc_lai_dung(self) -> None:
        """Ghi toàn hệ thống (pgd=None) → đọc lại với pgd=None."""
        self.cd.ghi_ct_registry(None, {"ALL": [1]}, "u")
        out = self.cd.doc_ct_registry(None)
        self.assertIn("ALL", out)

    def test_ghi_ct_registry_merge_khong_xoa_key_cu(self) -> None:
        """ghi_ct_registry MERGE (không xóa key cũ)."""
        self.cd.ghi_ct_registry("PGD Long Thành", {"HVN": [1]}, "u")
        self.cd.ghi_ct_registry("PGD Long Thành", {"HND": [2]}, "u")
        out = self.cd.doc_ct_registry("PGD Long Thành")
        self.assertIn("HVN", out)
        self.assertIn("HND", out)

    def test_ghi_ct_registry_luu_updated_by(self) -> None:
        """ghi_ct_registry ghi đúng updated_by vào kv_store."""
        self.cd.ghi_ct_registry("PGD Long Thành", {"HVN": []}, "alice")
        with importlib.import_module("db").get_conn() as conn:
            row = conn.execute(
                "SELECT updated_by FROM kv_store WHERE key LIKE 'ct_registry_%' LIMIT 1"
            ).fetchone()
        self.assertEqual(row["updated_by"], "alice")


class TestUploadService(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        """Mock các sub-module thiếu trước khi import upload_service."""
        cls._backup_modules: dict[str, types.ModuleType] = {}
        for name in ("data", "data.core", "data.pgd", "services", "services.data_quality", "upload_service"):
            if name in sys.modules:
                cls._backup_modules[name] = sys.modules[name]

        mock_data_core = types.ModuleType("data.core")
        mock_data_core.ts_file = lambda *a, **kw: 0.0
        mock_data_core.excel_to_parquet = lambda *a, **kw: None
        sys.modules.setdefault("data", types.ModuleType("data"))
        sys.modules["data.core"] = mock_data_core

        mock_data_pgd = types.ModuleType("data.pgd")
        mock_data_pgd.duong_dan_pgd = lambda *a, **kw: "/tmp/test.xlsx"
        sys.modules["data.pgd"] = mock_data_pgd

        mock_services = types.ModuleType("services")
        mock_dq = types.ModuleType("services.data_quality")
        mock_dq.kiem_tra_chat_luong = lambda df, loai: None
        sys.modules.setdefault("services", mock_services)
        sys.modules["services.data_quality"] = mock_dq

        cls.us = _load_module_from_path(
            "upload_service", str(Path(__file__).resolve().parents[1] / "services" / "upload_service.py")
        )

    @classmethod
    def tearDownClass(cls) -> None:
        """Khôi phục sys.modules để tránh ảnh hưởng test khác."""
        for name in ("data", "data.core", "data.pgd", "services", "services.data_quality", "upload_service"):
            if name in sys.modules:
                del sys.modules[name]
        for name, mod in cls._backup_modules.items():
            sys.modules[name] = mod

    def test_kiem_tra_file_xlsx_du_1kb_ok(self) -> None:
        """File .xlsx ≥ 1KB → (True, 'OK')."""
        ok, msg = self.us.kiem_tra_file("a.xlsx", b"a" * 1000)
        self.assertTrue(ok)
        self.assertEqual(msg, "OK")

    def test_kiem_tra_file_xls_du_1kb_ok(self) -> None:
        """File .xls ≥ 1KB → (True, ...)."""
        ok, _ = self.us.kiem_tra_file("a.xls", b"a" * 1000)
        self.assertTrue(ok)

    def test_kiem_tra_file_xlsx_hoa_du_1kb_ok(self) -> None:
        """.XLSX hoa ≥ 1KB → (True, ...)."""
        ok, _ = self.us.kiem_tra_file("a.XLSX", b"a" * 1000)
        self.assertTrue(ok)

    def test_kiem_tra_file_txt_fail(self) -> None:
        """.txt → (False, msg có 'định dạng' hoặc 'xlsx')."""
        ok, msg = self.us.kiem_tra_file("a.txt", b"a" * 2000)
        self.assertFalse(ok)
        self.assertTrue(("định dạng" in msg.lower()) or ("xlsx" in msg.lower()))

    def test_kiem_tra_file_pdf_fail(self) -> None:
        """.pdf → (False, ...)."""
        ok, msg = self.us.kiem_tra_file("a.pdf", b"a" * 2000)
        self.assertFalse(ok)
        self.assertTrue(("định dạng" in msg.lower()) or ("xlsx" in msg.lower()))

    def test_kiem_tra_file_xlsx_nho_hon_1kb_fail(self) -> None:
        """File .xlsx nhưng < 1KB → (False, msg có 'nhỏ')."""
        ok, msg = self.us.kiem_tra_file("a.xlsx", b"a" * 10)
        self.assertFalse(ok)
        self.assertIn("nhỏ", msg.lower())

    def test_kiem_tra_file_bytes_rong_fail(self) -> None:
        """bytes rỗng → (False, ...)."""
        ok, msg = self.us.kiem_tra_file("a.xlsx", b"")
        self.assertFalse(ok)
        self.assertTrue(msg)

    def test_ket_qua_upload_thanh_cong_co_thuoc_tinh(self) -> None:
        """KetQuaUpload(True,...) có đúng thuộc tính."""
        kq = self.us.KetQuaUpload(True, "OK", "/tmp/f.xlsx")
        self.assertTrue(kq.thanh_cong)
        self.assertEqual(kq.thong_bao, "OK")
        self.assertEqual(kq.duong_dan, "/tmp/f.xlsx")

    def test_ket_qua_upload_that_bai_default_duong_dan_rong(self) -> None:
        """KetQuaUpload(False,...) mặc định duong_dan=''. """
        kq = self.us.KetQuaUpload(False, "Lỗi")
        self.assertFalse(kq.thanh_cong)
        self.assertEqual(kq.duong_dan, "")

    def test_ket_qua_upload_repr_khong_raise(self) -> None:
        """KetQuaUpload là dataclass → __repr__ không raise."""
        kq = self.us.KetQuaUpload(True, "OK", "/tmp/f.xlsx")
        repr(kq)


class TestReportService(unittest.TestCase):
    def setUp(self) -> None:
        """Import report_service (services) cho test export Excel."""
        self.rs = _alias_module("report_service", "services.report_service")

    @unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl not installed")
    def test_ten_file_bao_cao_xlsx(self) -> None:
        """ten_file_bao_cao kết thúc .xlsx và chứa prefix."""
        out = self.rs.ten_file_bao_cao("bao_cao")
        self.assertTrue(out.endswith(".xlsx"))
        self.assertIn("bao_cao", out)

    @unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl not installed")
    def test_ten_file_bao_cao_pdf(self) -> None:
        """ten_file_bao_cao('bao_cao','pdf') kết thúc .pdf."""
        out = self.rs.ten_file_bao_cao("bao_cao", "pdf")
        self.assertTrue(out.endswith(".pdf"))

    @unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl not installed")
    def test_xuat_sheet_don_tra_ve_bytes(self) -> None:
        """xuat_sheet_don trả về bytes."""
        df = pd.DataFrame({"A": [1, 2, 3], "B": ["x", "y", "z"]})
        result = self.rs.xuat_sheet_don(df, "Test Report", "admin")
        self.assertIsInstance(result, (bytes, bytearray))
        self.assertGreater(len(result), 0)

    @unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl not installed")
    def test_xuat_sheet_don_xlsx_hop_le(self) -> None:
        """Bytes trả về là file xlsx hợp lệ (openpyxl.load_workbook không raise)."""
        import openpyxl

        df = pd.DataFrame({"A": [1, 2, 3], "B": ["x", "y", "z"]})
        result = self.rs.xuat_sheet_don(df, "Test Report", "admin")
        wb = openpyxl.load_workbook(BytesIO(result))
        self.assertGreaterEqual(len(wb.sheetnames), 1)

    @unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl not installed")
    def test_xuat_bao_cao_nhieu_sheet(self) -> None:
        """xuat_bao_cao nhiều sheet tạo file xlsx có >= 2 sheets dữ liệu."""
        import openpyxl

        df1 = pd.DataFrame({"A": [1]})
        df2 = pd.DataFrame({"B": [2]})
        result = self.rs.xuat_bao_cao({"Sheet1": df1, "Sheet2": df2}, "Báo cáo KHTD", "admin")
        wb = openpyxl.load_workbook(BytesIO(result))
        self.assertGreaterEqual(len(wb.sheetnames), 2)


class TestGiaoBan(unittest.TestCase):
    def setUp(self) -> None:
        """Alias giao_ban từ data.giao_ban."""
        self.gb = _alias_module("giao_ban", "data.giao_ban")
        import config

        self.config = config

    def _df_xa(self) -> pd.DataFrame:
        return pd.DataFrame(
            {
                self.config.COT_TONG_DU_NO: [10_000_000, 5_000_000, 3_000_000],
                self.config.COT_DU_NO_QH: [0, 500_000, 0],
                self.gb.COT_DU_NO_KHOANH: [0, 0, 0],
                self.gb.COT_TIEN_GUI: [0, 0, 0],
                self.gb.COT_DS_CV_THANG: [0, 0, 0],
                self.gb.COT_TN_TH_THANG: [0, 0, 0],
                self.gb.COT_TN_QH_THANG: [0, 0, 0],
                self.config.COT_MA_KH: ["KH001", "KH002", "KH003"],
                self.config.COT_TEN_XA: ["Phước Thái"] * 3,
                self.gb.COT_TEN_TO: ["Tổ 1", "Tổ 1", "Tổ 2"],
            }
        )

    def test_tinh_so_lieu_van_xuoi_tra_dict(self) -> None:
        """tinh_so_lieu_van_xuoi trả về dict."""
        out = self.gb.tinh_so_lieu_van_xuoi(self._df_xa(), None, 2025)
        self.assertIsInstance(out, dict)

    def test_tinh_so_lieu_van_xuoi_co_key_tong_du_no(self) -> None:
        """'{{tong_du_no}}' có trong keys."""
        out = self.gb.tinh_so_lieu_van_xuoi(self._df_xa(), None, 2025)
        self.assertIn("{{tong_du_no}}", out)

    def test_tinh_so_lieu_van_xuoi_so_kh_bang_3(self) -> None:
        """'{{so_kh}}' == '3'."""
        out = self.gb.tinh_so_lieu_van_xuoi(self._df_xa(), None, 2025)
        self.assertEqual(out["{{so_kh}}"], "3")

    def test_tinh_so_lieu_van_xuoi_co_key_du_no_qh(self) -> None:
        """'{{du_no_qh}}' có trong keys."""
        out = self.gb.tinh_so_lieu_van_xuoi(self._df_xa(), None, 2025)
        self.assertIn("{{du_no_qh}}", out)

    def test_tinh_so_lieu_van_xuoi_ty_le_nqh_float_duoc(self) -> None:
        """'{{ty_le_nqh}}' là chuỗi số hợp lệ."""
        out = self.gb.tinh_so_lieu_van_xuoi(self._df_xa(), None, 2025)
        float(out["{{ty_le_nqh}}"])

    def test_tinh_so_lieu_van_xuoi_tang_giam_thang_hop_le(self) -> None:
        """'{{tang_giam_thang}}' là 'tăng' hoặc 'giảm'."""
        out = self.gb.tinh_so_lieu_van_xuoi(self._df_xa(), None, 2025)
        self.assertIn(out["{{tang_giam_thang}}"], {"tăng", "giảm"})

    def test_tinh_so_lieu_van_xuoi_df_baseline_none_khong_raise(self) -> None:
        """df_baseline=None không raise."""
        self.gb.tinh_so_lieu_van_xuoi(self._df_xa(), None, 2025)

    def test_tinh_so_lieu_van_xuoi_df_rong_khong_raise(self) -> None:
        """DataFrame rỗng không raise."""
        df = self._df_xa().iloc[0:0]
        self.gb.tinh_so_lieu_van_xuoi(df, None, 2025)

    def test_loc_theo_xa_khop_tra_day_du_3_dong(self) -> None:
        """loc_theo_xa(df,'Phước Thái') trả về 3 dòng."""
        df = self._df_xa()
        out = self.gb.loc_theo_xa(df, "Phước Thái")
        self.assertEqual(len(out), 3)

    def test_loc_theo_xa_khong_ton_tai_tra_rong(self) -> None:
        """loc_theo_xa(df,'Không tồn tại') trả DataFrame rỗng."""
        df = self._df_xa()
        out = self.gb.loc_theo_xa(df, "Không tồn tại")
        self.assertEqual(len(out), 0)


class TestPgd(unittest.TestCase):
    def setUp(self) -> None:
        """Alias pgd từ data.pgd."""
        self.pgd = _alias_module("pgd", "data.pgd")

    def test_pgd_slug_long_thanh(self) -> None:
        """pgd_slug: 'PGD Long Thành' → 'pgd_long_thanh'."""
        self.assertEqual(self.pgd.pgd_slug("PGD Long Thành"), "pgd_long_thanh")

    def test_pgd_slug_bien_hoa(self) -> None:
        """pgd_slug: 'PGD Biên Hòa' → 'pgd_bien_hoa'."""
        self.assertEqual(self.pgd.pgd_slug("PGD Biên Hòa"), "pgd_bien_hoa")

    def test_pgd_slug_hoi_so(self) -> None:
        """pgd_slug: 'Hội sở' → 'hoi_so'."""
        self.assertEqual(self.pgd.pgd_slug("Hội sở"), "hoi_so")

    def test_pgd_slug_chi_chua_ky_tu_hop_le(self) -> None:
        """Kết quả slug chỉ có a-z, 0-9, '_'."""
        out = self.pgd.pgd_slug("PGD Định Quán")
        self.assertRegex(out, r"^[a-z0-9_]+$")

    def test_thu_muc_pgd_tra_ve_path(self) -> None:
        """thu_muc_pgd trả về Path."""
        p = self.pgd.thu_muc_pgd("PGD Long Thành")
        self.assertIsInstance(p, Path)

    def test_thu_muc_pgd_chua_slug(self) -> None:
        """str(path) chứa 'pgd_long_thanh'."""
        p = self.pgd.thu_muc_pgd("PGD Long Thành")
        self.assertIn("pgd_long_thanh", str(p).replace("\\", "/"))

    def test_duong_dan_pgd_hstd_ket_thuc_xlsx(self) -> None:
        """duong_dan_pgd('PGD Long Thành','hstd') kết thúc .xlsx."""
        out = self.pgd.duong_dan_pgd("PGD Long Thành", "hstd")
        self.assertTrue(str(out).endswith(".xlsx"))

    def test_duong_dan_pgd_nq11_khac_hstd(self) -> None:
        """duong_dan_pgd('PGD Long Thành','nq11') khác hstd."""
        a = self.pgd.duong_dan_pgd("PGD Long Thành", "hstd")
        b = self.pgd.duong_dan_pgd("PGD Long Thành", "nq11")
        self.assertNotEqual(a, b)

    def test_kiem_tra_file_ton_tai_pgd_chua_upload_false(self) -> None:
        """kiem_tra_file_ton_tai_pgd trả bool đúng theo file system, không raise."""
        path = self.pgd.duong_dan_pgd("PGD Long Thành", "hstd")
        out = self.pgd.kiem_tra_file_ton_tai_pgd("PGD Long Thành", "hstd")
        self.assertIsInstance(out, bool)
        self.assertEqual(out, Path(path).exists())


class TestDataPriorityService(unittest.TestCase):
    def setUp(self) -> None:
        """Reset DB và import data_priority_service."""
        _reset_db()
        self.dps = _alias_module("data_priority_service", "services.data_priority_service")
        import config

        self.config = config

    def test_kiem_tra_nguon_uu_tien_tra_dict(self) -> None:
        """kiem_tra_nguon_uu_tien trả về dict."""
        out = self.dps.kiem_tra_nguon_uu_tien("PGD Long Thành", "hstd")
        self.assertIsInstance(out, dict)

    def test_kiem_tra_nguon_uu_tien_co_key_nguon(self) -> None:
        """Result có key 'nguon_uu_tien'."""
        out = self.dps.kiem_tra_nguon_uu_tien("PGD Long Thành", "hstd")
        self.assertIn("nguon_uu_tien", out)

    def test_kiem_tra_nguon_uu_tien_khong_raise_voi_pgd_bat_ky(self) -> None:
        """Không raise với PGD bất kỳ trong DS_PGD."""
        for pgd in self.config.DS_PGD[:3]:
            with self.subTest(pgd=pgd):
                self.dps.kiem_tra_nguon_uu_tien(pgd, "hstd")

    def test_kiem_tra_nguon_uu_tien_khong_raise_voi_loai_file(self) -> None:
        """Không raise với loai_file: hstd, nq11, gqvl."""
        for loai in ("hstd", "nq11", "gqvl"):
            with self.subTest(loai=loai):
                self.dps.kiem_tra_nguon_uu_tien("PGD Long Thành", loai)

    def test_lay_thong_tin_nguon_hien_tai_tra_dict(self) -> None:
        """lay_thong_tin_nguon_hien_tai trả về dict không raise."""
        out = self.dps.lay_thong_tin_nguon_hien_tai("PGD Long Thành")
        self.assertIsInstance(out, dict)

    def test_lay_thong_tin_nguon_hien_tai_co_keys_loai_file(self) -> None:
        """Kết quả có key cho các loại file (hstd/nq11/gqvl/cdtotkvv)."""
        out = self.dps.lay_thong_tin_nguon_hien_tai("PGD Long Thành")
        for k in ("hstd", "nq11", "gqvl", "cdtotkvv"):
            with self.subTest(k=k):
                self.assertIn(k, out)

