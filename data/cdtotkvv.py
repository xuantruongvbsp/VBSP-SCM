"""Đọc và tổng hợp dữ liệu Chấm điểm Tổ TK&VV (CDTOTKVV)."""
import os
import re
import unicodedata
from pathlib import Path

import pandas as pd
import streamlit as st

from config import (
    BASE_DIR,
    CDTOTKVV_DIR,
    CDTOTKVV_COLS,
    CDTOTKVV_DATA_ROW_START,
    COT_MA_PGD,
    COT_MA_TO,
    COT_HINH_THUC_VAY,
    COT_TEN_PGD,
    COT_TEN_TO,
    COT_TONG_DU_NO,
    TEN_PGD_TO_MA,
)
from config import PGD_DATA_DIR
from data.core import ts_file
from data.pgd import duong_dan_pgd

_COLS_FLOAT = ["du_no", "so_du_tk", "tong_diem"]
_COLS_STR   = ["ma_dv", "ma_xa", "ma_to"]
_CODE_WIDTHS = {"ma_dv": 6, "ma_xa": 6, "ma_to": 7}
_XEP_LOAI_TOT = "T\u1ed1t"
_XEP_LOAI_KHA = "Kh\u00e1"
_XEP_LOAI_TB  = "Trung b\u00ecnh"
_XEP_LOAI_YEU = "Y\u1ebfu"
_TINH_TRANG_A = "A"
_TINH_TRANG_B = "B"
_TINH_TRANG_C = "C"


def _norm_text_key(val) -> str:
    """Chuẩn hóa text header Excel để dò cột linh hoạt hơn."""
    if val is None:
        return ""
    text = unicodedata.normalize("NFKD", str(val))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("đ", "d").replace("Đ", "D").lower().strip()
    return "".join(ch for ch in text if ch.isalnum())


def _normalize_code_value(val, width: int | None = None):
    """Chuẩn hóa mã đọc từ Excel, giữ số 0 đầu cho các cột định danh."""
    if val is None or pd.isna(val):
        return pd.NA
    text = str(val).strip()
    if not text or text.lower() in {"nan", "none", "nat"}:
        return pd.NA
    text = re.sub(r"\.0+$", "", text)
    if width and re.fullmatch(r"\d+", text):
        return text.zfill(width)
    return text


def chuan_hoa_cdtotkvv_phan_tich(df: pd.DataFrame | None) -> pd.DataFrame:
    """Chuẩn hóa tập CDTO dùng cho KPI/báo cáo, không sửa DataFrame đầu vào.

    - Chỉ giữ Tổ còn dư nợ dương.
    - Quy đổi các biến thể ``Yếu kém``/``Yếu`` về nhãn chuẩn ``Yếu``.

    Nếu thiếu cột ``du_no`` thì giữ nguyên dữ liệu để tương thích với các tập
    tối giản trong kiểm thử và các file lịch sử cũ chưa đủ schema.
    """
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df.copy()

    out = df.copy()
    if "du_no" in out.columns:
        out["du_no"] = pd.to_numeric(out["du_no"], errors="coerce")
        out = out[out["du_no"].fillna(0) > 0].copy()

    if "xep_loai" in out.columns:
        def _xep_loai_chuan(value):
            if value is None or pd.isna(value):
                return pd.NA
            text = str(value).strip()
            if _norm_text_key(text) in {"yeu", "yeukem"}:
                return _XEP_LOAI_YEU
            return text

        out["xep_loai"] = out["xep_loai"].map(_xep_loai_chuan)

    return out.reset_index(drop=True)


def doi_chieu_cdtotkvv_hstd(
    df_cdto: pd.DataFrame | None,
    df_hstd: pd.DataFrame | None,
) -> dict:
    """Đối chiếu CDTO/HSTD bằng khóa chuẩn ``Mã PGD + Mã tổ``.

    HSTD chỉ tính mã Tổ có tổng dư nợ dương và loại ``0000000`` (dư nợ
    không qua Tổ). Kết quả giữ hai danh sách ngoại lệ để UI cho phép đối chiếu
    dữ liệu nguồn, không tự ghép theo tên Tổ trưởng.
    """
    empty = {
        "tong_cdto": 0,
        "tong_hstd": 0,
        "so_khop": 0,
        "chi_hstd": pd.DataFrame(),
        "chi_cdto": pd.DataFrame(),
        "cho_vay_truc_tiep": pd.DataFrame(),
    }
    if df_cdto is None or df_cdto.empty or df_hstd is None or df_hstd.empty:
        return empty

    required_cdto = {"ma_dv", "ma_to"}
    required_hstd = {COT_MA_PGD, COT_MA_TO, COT_TONG_DU_NO}
    if not required_cdto.issubset(df_cdto.columns) or not required_hstd.issubset(df_hstd.columns):
        return empty

    cdto = chuan_hoa_cdtotkvv_phan_tich(df_cdto)
    cdto["ma_dv_chuan"] = cdto["ma_dv"].map(
        lambda value: _normalize_code_value(value, 6)
    )
    cdto["ma_to_chuan"] = cdto["ma_to"].map(
        lambda value: _normalize_code_value(value, 7)
    )
    cdto = cdto.dropna(subset=["ma_dv_chuan", "ma_to_chuan"]).copy()
    cdto = cdto[cdto["ma_to_chuan"] != "0000000"].copy()
    cdto["_key"] = cdto["ma_dv_chuan"].astype(str) + "|" + cdto["ma_to_chuan"].astype(str)
    cdto = cdto.drop_duplicates("_key", keep="last")

    hstd = df_hstd.copy()
    hstd["ma_dv_chuan"] = hstd[COT_MA_PGD].map(
        lambda value: _normalize_code_value(value, 6)
    )
    hstd["ma_to_chuan"] = hstd[COT_MA_TO].map(
        lambda value: _normalize_code_value(value, 7)
    )
    hstd[COT_TONG_DU_NO] = pd.to_numeric(hstd[COT_TONG_DU_NO], errors="coerce").fillna(0)
    if COT_HINH_THUC_VAY in hstd.columns:
        hinh_thuc = pd.to_numeric(hstd[COT_HINH_THUC_VAY], errors="coerce")
        hstd["_du_no_truc_tiep"] = hstd[COT_TONG_DU_NO].where(hinh_thuc == 1, 0)
    else:
        hstd["_du_no_truc_tiep"] = 0
    hstd = hstd.dropna(subset=["ma_dv_chuan", "ma_to_chuan"]).copy()
    hstd = hstd[hstd["ma_to_chuan"] != "0000000"].copy()

    agg_kwargs: dict[str, tuple[str, str]] = {
        "du_no": (COT_TONG_DU_NO, "sum"),
        "du_no_truc_tiep": ("_du_no_truc_tiep", "sum"),
    }
    if COT_TEN_PGD in hstd.columns:
        agg_kwargs["ten_dv"] = (COT_TEN_PGD, "first")
    if COT_TEN_TO in hstd.columns:
        agg_kwargs["ten_to"] = (COT_TEN_TO, "first")
    hstd_to = (
        hstd.groupby(["ma_dv_chuan", "ma_to_chuan"], as_index=False)
        .agg(**agg_kwargs)
    )
    hstd_to = hstd_to[hstd_to["du_no"] > 0].copy()
    cho_vay_truc_tiep = hstd_to[
        (hstd_to["du_no_truc_tiep"] > 0)
        & (hstd_to["du_no_truc_tiep"] == hstd_to["du_no"])
    ].copy()
    hstd_to = hstd_to.drop(index=cho_vay_truc_tiep.index).copy()
    hstd_to["_key"] = hstd_to["ma_dv_chuan"].astype(str) + "|" + hstd_to["ma_to_chuan"].astype(str)

    keys_cdto = set(cdto["_key"])
    keys_hstd = set(hstd_to["_key"])
    chi_hstd = hstd_to[~hstd_to["_key"].isin(keys_cdto)].copy()
    chi_cdto = cdto[~cdto["_key"].isin(keys_hstd)].copy()

    return {
        "tong_cdto": len(keys_cdto),
        "tong_hstd": len(keys_hstd),
        "so_khop": len(keys_cdto & keys_hstd),
        "chi_hstd": chi_hstd.reset_index(drop=True),
        "chi_cdto": chi_cdto.reset_index(drop=True),
        "cho_vay_truc_tiep": cho_vay_truc_tiep.reset_index(drop=True),
    }


def _tim_header_cdto_toan_cn(all_rows: list[list]) -> tuple[int | None, dict[str, int]]:
    """Tìm dòng header và map tên trường -> index cột trong file toàn CN."""
    alias_map = {
        "stt": {"stt"},
        "ma_dv": {"mapgd", "madonvi", "maphonggiaodich", "mapgd6so"},
        "ten_dv": {"tenpgd", "tendonvi", "tendv"},
        "ma_xa": {"maxa"},
        "ten_xa": {"tenxa"},
        "ma_to": {"mato"},
        "ten_to_truong": {"tento", "tentotruong"},
        "loai_to": {"loaito"},
        "dvut": {"dvut", "madvut", "tendvut"},
        "du_no": {"duno"},
        "tong_diem": {"tongdiem"},
        "xep_loai": {"xeploai"},
        "ngaybc": {"ngaybc", "ngaybaocao"},
    }
    for row_idx, row in enumerate(all_rows[:20]):
        idx_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            key = _norm_text_key(cell)
            if not key:
                continue
            for field, aliases in alias_map.items():
                if key in aliases and field not in idx_map:
                    idx_map[field] = col_idx
                    break
        has_unit_col = "ma_dv" in idx_map or "ten_dv" in idx_map
        has_data_col = any(
            field in idx_map
            for field in ("ma_xa", "ten_xa", "ma_to", "ten_to_truong", "tong_diem", "xep_loai")
        )
        if "stt" in idx_map and has_unit_col and has_data_col:
            return row_idx, idx_map
    return None, {}


def _chon_cot_ma_pgd_tot_nhat(
    all_rows: list[list],
    valid_codes: set[str],
    preferred_idx: int | None = None,
    start_row: int = 0,
) -> int | None:
    """Chọn cột có nhiều mã PGD hợp lệ/đa dạng nhất trong phần data."""

    def _norm_ma_local(val) -> str | None:
        if val is None:
            return None
        try:
            return str(int(float(str(val).strip()))).zfill(6)
        except (ValueError, TypeError):
            return None

    sample_rows = all_rows[start_row : start_row + 300]
    max_cols = max((len(r) for r in sample_rows), default=0)
    best_idx = None
    best_score = (-1, -1, -1)
    for col_idx in range(max_cols):
        hits = 0
        unique_codes: set[str] = set()
        for row in sample_rows:
            if len(row) <= col_idx:
                continue
            ma = _norm_ma_local(row[col_idx])
            if ma and ma in valid_codes:
                hits += 1
                unique_codes.add(ma)
        if not hits:
            continue
        score = (
            len(unique_codes),
            hits,
            1 if preferred_idx is not None and col_idx == preferred_idx else 0,
        )
        if score > best_score:
            best_score = score
            best_idx = col_idx
    return best_idx


def _chon_cot_ten_dv_tot_nhat(
    all_rows: list[list],
    preferred_idx: int | None = None,
    start_row: int = 0,
) -> int | None:
    """Chọn cột có nhiều tên đơn vị chuẩn hóa hợp lệ nhất trong phần data."""
    from services.file_detection_service import ten_doc_ve_don_vi_chuan

    sample_rows = all_rows[start_row : start_row + 300]
    max_cols = max((len(r) for r in sample_rows), default=0)
    best_idx = None
    best_score = (-1, -1, -1)
    for col_idx in range(max_cols):
        hits = 0
        unique_names: set[str] = set()
        for row in sample_rows:
            if len(row) <= col_idx:
                continue
            ten = ten_doc_ve_don_vi_chuan(row[col_idx])
            if ten:
                hits += 1
                unique_names.add(ten)
        if not hits:
            continue
        score = (
            len(unique_names),
            hits,
            1 if preferred_idx is not None and col_idx == preferred_idx else 0,
        )
        if score > best_score:
            best_score = score
            best_idx = col_idx
    return best_idx


def _ten_file(thang_nam: str) -> Path:
    return CDTOTKVV_DIR / f"CDTOTKVV_{thang_nam}.xlsx"


def doc_thang_nam_tu_file(file_bytes: bytes) -> str | None:
    """Trích kỳ chấm điểm mm/yyyy từ phần đầu file CDTOTKVV (tiêu đề / ngày báo cáo)."""
    from datetime import date, datetime
    from io import BytesIO

    import openpyxl

    pat_ngay_vn = re.compile(
        r"ngày\s+(\d{1,2})\s+tháng\s+(\d{1,2})\s+năm\s+(\d{4})",
        re.IGNORECASE,
    )
    pat_ddmmyyyy = re.compile(
        r"\b([0-2]?\d|3[0-1])[/\-]([0]?\d|1[0-2])[/\-](\d{4})\b"
    )
    pat_thang_slash_nam = re.compile(r"\b(0[1-9]|1[0-2])/\s*(\d{4})\b")
    pat_thang_word = re.compile(
        r"tháng\s*(\d{1,2})\s*/\s*(\d{4})",
        re.IGNORECASE,
    )
    try:
        wb = openpyxl.load_workbook(
            BytesIO(file_bytes), read_only=True, data_only=True
        )
        ws = wb.active
        cap = min(ws.max_row or 0, 25)
        for row in ws.iter_rows(min_row=1, max_row=cap, values_only=True):
            for cell in row:
                if cell is None:
                    continue
                if isinstance(cell, (datetime, date)):
                    return cell.strftime("%m/%Y")
                text = str(cell).strip()
                if not text or text.lower() == "nan":
                    continue
                m = pat_ngay_vn.search(text)
                if m:
                    mm = m.group(2).zfill(2)
                    yyyy = m.group(3)
                    return f"{mm}/{yyyy}"
                m = pat_thang_word.search(text)
                if m:
                    return f"{m.group(1).zfill(2)}/{m.group(2)}"
                m = pat_thang_slash_nam.search(text)
                if m:
                    return f"{m.group(1)}/{m.group(2)}"
                m = pat_ddmmyyyy.search(text)
                if m:
                    mm = m.group(2).zfill(2)
                    yyyy = m.group(3)
                    return f"{mm}/{yyyy}"
    except Exception:
        return None
    return None


@st.cache_data(show_spinner=False)
def doc_cdtotkvv_path(duong_dan: str, _ts) -> pd.DataFrame | None:
    if not duong_dan or not os.path.exists(duong_dan):
        return None
    df = pd.read_excel(
        duong_dan,
        engine="openpyxl",
        header=None,
        skiprows=CDTOTKVV_DATA_ROW_START,
    )
    if df.empty:
        return None
    df = df.iloc[:, : len(CDTOTKVV_COLS)].copy()
    # File tách từ toàn CN: openpyxl bỏ trailing None → có thể thiếu cột cuối (tinh_trang)
    for col in CDTOTKVV_COLS[len(df.columns):]:
        df[col] = pd.NA
    df.columns = CDTOTKVV_COLS
    df = df[pd.to_numeric(df["stt"], errors="coerce").notna()].reset_index(drop=True)
    for col in _COLS_FLOAT:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    for col in _COLS_STR:
        if col in df.columns:
            width = _CODE_WIDTHS.get(col)
            df[col] = df[col].map(lambda v, w=width: _normalize_code_value(v, w))
    return chuan_hoa_cdtotkvv_phan_tich(df)


@st.cache_data(show_spinner=False)
def doc_cdtotkvv(thang_nam: str) -> pd.DataFrame | None:
    try:
        mm, yyyy = thang_nam.split("/")
        suffix = f"{yyyy}_{mm}"
    except ValueError:
        return None
    pgd_data = BASE_DIR / "pgd_data"
    frames = []
    # Đọc file cũ trước, file được cập nhật sau đọc sau để khi loại trùng sẽ
    # ưu tiên bản sửa mới nhất của cùng một Tổ.
    files = sorted(
        pgd_data.rglob(f"cdtotkvv_{suffix}.xlsx"),
        key=lambda path: path.stat().st_mtime,
    )
    for f in files:
        df = doc_cdtotkvv_path(str(f), ts_file(str(f)))
        if df is not None and not df.empty:
            frames.append(df)
    if not frames:
        for f in pgd_data.rglob("cdtotkvv_latest.xlsx"):
            try:
                thang = doc_thang_nam_tu_file(f.read_bytes())
                if thang == thang_nam:
                    df = doc_cdtotkvv_path(str(f), ts_file(str(f)))
                    if df is not None and not df.empty:
                        frames.append(df)
            except Exception:
                pass
    if not frames:
        return None
    result = pd.concat(frames, ignore_index=True)

    # Phòng vệ dữ liệu lịch sử: một file toàn CN từng có thể bị đặt nhầm trong
    # thư mục một đơn vị, khiến cùng Tổ xuất hiện thêm lần nữa khi concat.
    # Chỉ loại trùng các dòng có đủ định danh để không làm mất dòng lỗi cần rà soát.
    key_cols = [col for col in ("ma_dv", "ma_to") if col in result.columns]
    if len(key_cols) == 2:
        complete_key = result[key_cols].notna().all(axis=1)
        rows_valid = result.loc[complete_key].drop_duplicates(
            subset=key_cols,
            keep="last",
        )
        rows_incomplete = result.loc[~complete_key]
        result = pd.concat([rows_valid, rows_incomplete], ignore_index=True)
    return result


@st.cache_data(show_spinner=False)
def doc_cdtotkvv_pgd(ten_pgd: str, _ts) -> pd.DataFrame | None:
    path = duong_dan_pgd(ten_pgd, "cdtotkvv")
    return doc_cdtotkvv_path(path, _ts)


def doc_cdtotkvv_toan_cn_pgd() -> pd.DataFrame | None:
    if not os.path.exists(PGD_DATA_DIR):
        return None
    frames: list[pd.DataFrame] = []
    for d in sorted(Path(PGD_DATA_DIR).iterdir()):
        if not d.is_dir():
            continue
        path = d / "cdtotkvv_latest.xlsx"
        if not path.exists():
            continue
        try:
            df_p = doc_cdtotkvv_path(str(path), ts_file(str(path)))
            if df_p is None or df_p.empty:
                continue
            frames.append(df_p)
        except Exception:
            continue
    if not frames:
        return None
    return pd.concat(frames, ignore_index=True)


@st.cache_data(show_spinner=False)
def tong_hop_tu_pgd_data() -> pd.DataFrame | None:
    """Tổng hợp CDTOTKVV từ pgd_data/*/cdtotkvv_latest.xlsx (hệ thống tập trung)."""
    return doc_cdtotkvv_toan_cn_pgd()


@st.cache_data(show_spinner=False)
def ds_thang_nam() -> list[str]:
    pat = re.compile(r"cdtotkvv_(\d{4})_(\d{2})\.xlsx$", re.IGNORECASE)
    thang_set = set()
    pgd_data = BASE_DIR / "pgd_data"
    if not pgd_data.exists():
        return []
    for f in pgd_data.rglob("cdtotkvv_*.xlsx"):
        m = pat.search(f.name)
        if m:
            yyyy, mm = m.group(1), m.group(2)
            thang_set.add(f"{mm}/{yyyy}")
    for f in pgd_data.rglob("cdtotkvv_latest.xlsx"):
        try:
            thang = doc_thang_nam_tu_file(f.read_bytes())
            if thang:
                thang_set.add(thang)
        except Exception:
            pass
    return sorted(thang_set, key=lambda s: (s[3:], s[:2]), reverse=True)


def doc_thang_tu_cdto_toan_cn(file_bytes: bytes) -> str | None:
    """
    Đọc tháng báo cáo từ cột NGAYBC (cột S, index 18) của file CDTOTKVV toàn CN.
    Đáng tin hơn doc_thang_nam_tu_file() vì đọc từ dữ liệu thực (không bị
    ảnh hưởng bởi ngày tạo/export file trong header).
    Trả về "MM/YYYY" hoặc None nếu không đọc được.
    """
    from io import BytesIO
    from datetime import datetime, date as _date

    import openpyxl

    from config import MA_PGD_MAP

    def _norm_ma(val) -> str | None:
        try:
            return str(int(float(str(val).strip()))).zfill(6)
        except (ValueError, TypeError):
            return None

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        from services.file_detection_service import ten_doc_ve_don_vi_chuan

        _header_row, idx_map = _tim_header_cdto_toan_cn(all_rows)
        col_mapgd = _chon_cot_ma_pgd_tot_nhat(
            all_rows,
            set(MA_PGD_MAP),
            preferred_idx=idx_map.get("ma_dv"),
            start_row=(_header_row or 0) + 1,
        )
        col_tendv = _chon_cot_ten_dv_tot_nhat(
            all_rows,
            preferred_idx=idx_map.get("ten_dv"),
            start_row=(_header_row or 0) + 1,
        )
        col_ngaybc = idx_map.get("ngaybc")
        current_unit = None

        for row in all_rows:
            row = list(row)
            if col_mapgd is None:
                ma_dv = next(
                    (
                        ma
                        for idx in (2, 1)
                        if len(row) > idx
                        for ma in [_norm_ma(row[idx])]
                        if ma and ma in MA_PGD_MAP
                    ),
                    None,
                )
            else:
                if len(row) <= col_mapgd:
                    ma_dv = None
                else:
                    ma_dv = _norm_ma(row[col_mapgd])
            ten_dv = None
            if col_tendv is not None and len(row) > col_tendv:
                ten_dv = ten_doc_ve_don_vi_chuan(row[col_tendv])
            unit = MA_PGD_MAP.get(ma_dv) if ma_dv in MA_PGD_MAP else ten_dv or current_unit
            if not unit:
                continue
            current_unit = unit
            if col_ngaybc is None:
                val = next((row[idx] for idx in (18, 17) if len(row) > idx), None)
            else:
                if len(row) <= col_ngaybc:
                    break
                val = row[col_ngaybc]
            if val is None:
                break
            if isinstance(val, (datetime, _date)):
                return val.strftime("%m/%Y")
            if val:
                m = re.search(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})", str(val).strip())
                if m:
                    return f"{m.group(2).zfill(2)}/{m.group(3)}"
            break
    except Exception:
        pass
    return None


def tach_file_cdto_toan_cn(file_bytes: bytes) -> dict[str, bytes]:
    """
    Đọc file CDTOTKVV toàn CN, tách thành dict {ten_pgd: excel_bytes}.
    Mỗi file con được chuẩn hóa theo đúng CDTOTKVV_COLS (20 cột per-PGD)
    để doc_cdtotkvv_path() đọc được chính xác.

    Format toàn CN (18 cột):
      0:STT 1:MAPGD 2:TEN_PGD 3:MAXA 4:TENXA 5:MATO 6:TENTO
      7:LOAITO 8:DVUT 9:DUNO 10:Tham gia GDX 11:TL thu nợ gốc
      12:TL thu lãi 13:TG Tổ TKVV 14:TL nợ quá hạn
      15:TONGDIEM 16:XEPLOAI 17:NGAYBC

    Format per-PGD / CDTOTKVV_COLS (20 cột):
      0:stt 1:ma_dv 2:ten_dv 3:ma_xa 4:ten_xa 5:ma_to
      6:ten_to_truong 7:dvut 8:loai_to 9:du_no 10:so_du_tk
      11-16:diem_* 17:tong_diem 18:xep_loai 19:tinh_trang

    Raises ValueError nếu không đọc được hoặc không tìm thấy đơn vị hợp lệ.
    """
    from io import BytesIO
    from collections import defaultdict

    import openpyxl

    from config import MA_PGD_MAP, CDTOTKVV_DATA_ROW_START

    # Số cột đầu ra cố định theo CDTOTKVV_COLS
    N_COLS_OUT = 20
    FALLBACK_IDX = {
        "stt": 1,
        "ma_dv": 2,
        "ten_dv": 3,
        "ma_xa": 4,
        "ten_xa": 5,
        "ma_to": 6,
        "ten_to_truong": 7,
        "loai_to": 8,
        "dvut": 9,
        "du_no": 10,
        "tong_diem": 16,
        "xep_loai": 17,
    }

    def _norm_ma(val) -> str | None:
        if val is None:
            return None
        try:
            return str(int(float(str(val).strip()))).zfill(6)
        except (ValueError, TypeError):
            return None

    def _get(row: list, idx: int):
        return row[idx] if idx < len(row) else None

    def _src(row: list, field: str):
        idx = idx_map.get(field, FALLBACK_IDX.get(field))
        if idx is None:
            return None
        return _get(row, idx)

    def _map_row(src: list) -> list:
        """Chuyển 1 data row từ format toàn CN → CDTOTKVV_COLS (20 cột)."""
        return [
            _src(src, "stt"),
            _src(src, "ma_dv"),
            _src(src, "ten_dv"),
            _src(src, "ma_xa"),
            _src(src, "ten_xa"),
            _src(src, "ma_to"),
            _src(src, "ten_to_truong"),
            _src(src, "dvut"),
            _src(src, "loai_to"),
            _src(src, "du_no"),
            None,           # so_du_tk     (không có trong toàn CN)
            None,           # diem_gdtx    (không có)
            None,           # diem_nqh     (không có)
            None,           # diem_thu_no  (không có)
            None,           # diem_thu_lai (không có)
            None,           # diem_tv_tiengui (không có)
            None,           # diem_ds_tg   (không có)
            _src(src, "tong_diem"),
            _src(src, "xep_loai"),
            None,           # tinh_trang   (không có trong toàn CN)
        ]

    # Tự phát hiện dòng bắt đầu dữ liệu: dòng đầu tiên có MAPGD hợp lệ
    # (đáng tin hơn STT vì MAPGD là cột ta cần, header sẽ không có mã 6 số)
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    from services.file_detection_service import ten_doc_ve_don_vi_chuan

    _header_row, idx_map = _tim_header_cdto_toan_cn(all_rows)
    col_ma_dv_in = _chon_cot_ma_pgd_tot_nhat(
        all_rows,
        set(MA_PGD_MAP),
        preferred_idx=idx_map.get("ma_dv", FALLBACK_IDX["ma_dv"]),
        start_row=(_header_row or 0) + 1,
    )
    if col_ma_dv_in is None:
        col_ma_dv_in = idx_map.get("ma_dv")
    if col_ma_dv_in is None and not idx_map:
        col_ma_dv_in = FALLBACK_IDX["ma_dv"]
    col_ten_dv_in = _chon_cot_ten_dv_tot_nhat(
        all_rows,
        preferred_idx=idx_map.get("ten_dv", FALLBACK_IDX["ten_dv"]),
        start_row=(_header_row or 0) + 1,
    )
    if col_ten_dv_in is None:
        col_ten_dv_in = idx_map.get("ten_dv")
    if col_ten_dv_in is None and not idx_map:
        col_ten_dv_in = FALLBACK_IDX["ten_dv"]

    def _resolve_unit(row: list, current_unit: str | None = None) -> tuple[str | None, str | None]:
        ma_dv = None
        if col_ma_dv_in is not None and len(row) > col_ma_dv_in:
            ma_dv = _norm_ma(row[col_ma_dv_in])
        ten_dv = None
        if col_ten_dv_in is not None and len(row) > col_ten_dv_in:
            ten_dv = ten_doc_ve_don_vi_chuan(row[col_ten_dv_in])
        unit = MA_PGD_MAP.get(ma_dv) if ma_dv in MA_PGD_MAP else ten_dv or current_unit
        ma_std = ma_dv if ma_dv in MA_PGD_MAP else (TEN_PGD_TO_MA.get(unit) if unit else None)
        return unit, ma_std
    idx_map["ma_dv"] = col_ma_dv_in

    current_unit = None
    data_start = None
    for i, row in enumerate(all_rows):
        if not row:
            continue
        unit, ma_dv = _resolve_unit(row, current_unit)
        if unit and ma_dv:
            current_unit = unit
            data_start = i
            break

    if data_start is None:
        raise ValueError(
            "Không tìm thấy dòng dữ liệu hợp lệ trong file "
            "(cần có 'Mã PGD' hợp lệ hoặc 'Tên PGD/Tên đơn vị' nhận diện được)"
        )

    raw_headers = all_rows[:data_start]
    data_rows   = all_rows[data_start:]

    # Đảm bảo đúng CDTOTKVV_DATA_ROW_START dòng header để
    # doc_cdtotkvv_path(skiprows=CDTOTKVV_DATA_ROW_START) đọc không lệch.
    empty_row = [None] * N_COLS_OUT
    header_rows = raw_headers[:CDTOTKVV_DATA_ROW_START]
    while len(header_rows) < CDTOTKVV_DATA_ROW_START:
        header_rows.append(empty_row)

    groups: dict[str, list] = defaultdict(list)
    current_unit = None
    for row in data_rows:
        if not row:
            continue
        unit, ma_dv = _resolve_unit(row, current_unit)
        if unit and ma_dv:
            current_unit = unit
            groups[ma_dv].append(row)

    if not groups:
        raise ValueError(
            "Không tìm thấy mã đơn vị hợp lệ trong file "
            "(kiểm tra cột 'Mã PGD' hoặc 'Tên PGD/Tên đơn vị')"
        )

    result: dict[str, bytes] = {}
    for ma_dv, rows in groups.items():
        ten_pgd = MA_PGD_MAP[ma_dv]
        last_unit = ten_pgd
        wb_out  = openpyxl.Workbook(write_only=True)
        ws_out  = wb_out.create_sheet()

        # Ghi header gốc (metadata, bị skip khi đọc)
        for hrow in header_rows:
            ws_out.append(hrow + [None] * max(0, N_COLS_OUT - len(hrow)))

        # Ghi data rows đã chuẩn hóa theo CDTOTKVV_COLS
        for drow in rows:
            unit_row, ma_row = _resolve_unit(drow, last_unit)
            last_unit = unit_row or last_unit
            mapped = _map_row(drow)
            mapped[1] = ma_row or mapped[1]
            mapped[2] = last_unit or mapped[2]
            ws_out.append(mapped)

        buf = BytesIO()
        wb_out.save(buf)
        result[ten_pgd] = buf.getvalue()

    return result


@st.cache_data(show_spinner=False)
def tong_hop_theo_pgd(df: pd.DataFrame) -> pd.DataFrame:
    """
    Nhóm theo ma_dv + ten_dv, tính:
    - tong_to           : tổng số tổ
    - tong_diem_tb      : trung bình tổng_diem (làm tròn 2 chữ số)
    - to_tot/kha/tb/yeu : số tổ theo xep_loai
    - to_tinh_trang_a/b/c : số tổ theo tình trạng A/B/C
    Trả về DataFrame sort theo ma_dv.
    """
    df = chuan_hoa_cdtotkvv_phan_tich(df)
    if df.empty:
        return pd.DataFrame(
            columns=[
                "ma_dv", "ten_dv", "tong_to", "tong_diem_tb",
                "to_tot", "to_kha", "to_tb", "to_yeu",
                "to_tinh_trang_a", "to_tinh_trang_b", "to_tinh_trang_c",
            ]
        )
    nhom = df.groupby(["ma_dv", "ten_dv"], as_index=False)
    tong_hop = nhom.agg(
        tong_to=("stt", "count"),
        tong_diem_tb=("tong_diem", "mean"),
    )
    for ten_col, gia_tri in [
        ("to_tot", _XEP_LOAI_TOT),
        ("to_kha", _XEP_LOAI_KHA),
        ("to_tb",  _XEP_LOAI_TB),
        ("to_yeu", _XEP_LOAI_YEU),
    ]:
        dem = (
            df[df["xep_loai"] == gia_tri]
            .groupby(["ma_dv", "ten_dv"], as_index=False)
            .agg(**{ten_col: ("stt", "count")})
        )
        tong_hop = tong_hop.merge(dem, on=["ma_dv", "ten_dv"], how="left")
    for ten_col, gia_tri in [
        ("to_tinh_trang_a", _TINH_TRANG_A),
        ("to_tinh_trang_b", _TINH_TRANG_B),
        ("to_tinh_trang_c", _TINH_TRANG_C),
    ]:
        dem = (
            df[df["tinh_trang"] == gia_tri]
            .groupby(["ma_dv", "ten_dv"], as_index=False)
            .agg(**{ten_col: ("stt", "count")})
        )
        tong_hop = tong_hop.merge(dem, on=["ma_dv", "ten_dv"], how="left")
    cols_dem = [
        "to_tot", "to_kha", "to_tb", "to_yeu",
        "to_tinh_trang_a", "to_tinh_trang_b", "to_tinh_trang_c",
    ]
    tong_hop[cols_dem] = tong_hop[cols_dem].fillna(0).astype(int)
    tong_hop["tong_diem_tb"] = tong_hop["tong_diem_tb"].round(2)
    return tong_hop.sort_values("ma_dv").reset_index(drop=True)
