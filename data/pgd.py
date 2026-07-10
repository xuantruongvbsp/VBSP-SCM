"""Quản lý file dữ liệu riêng từng PGD (upload, đọc, gộp)."""
import logging
import os
import re
import unicodedata
import html
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Literal

import pandas as pd
import streamlit as st

_log = logging.getLogger(__name__)
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.datetime import from_excel

from data.core import ts_file
from config import PGD_DATA_DIR, GQVL_PGD_DIR, UPLOAD_CANH_BAO_NGAY


# ── Slug helper ───────────────────────────────────────────────────────────────
def pgd_slug(ten_pgd: str) -> str:
    """'PGD Biên Hòa 1' → 'pgd_bien_hoa_1'
    Lưu ý: "đ/Đ" (U+0111/U+0110) không decompose được qua NFD
    → phải replace thủ công trước khi normalize.
    """
    s = ten_pgd.strip().lower()
    s = s.replace("đ", "d").replace("Đ", "D")   # xử lý trước khi NFD
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]+", "_", s).strip("_")


def thu_muc_pgd(ten_pgd: str) -> Path:
    """Trả về Path thư mục riêng của PGD: pgd_data/{slug}/. Tạo nếu chưa có."""
    d = Path(PGD_DATA_DIR) / pgd_slug(ten_pgd)
    d.mkdir(parents=True, exist_ok=True)
    return d


def duong_dan_pgd(ten_pgd: str, loai: str) -> str:
    """
    Đường dẫn file PGD theo cấu trúc mới.
    loai: "hstd" | "nq11" | "gqvl" | "cdtotkvv" → pgd_data/{slug}/{loai}_latest.xlsx
          "hstd_khnv"  → pgd_data/{slug}/hstd_khnv.xlsx  (riêng Phòng KH-NV, tránh ghi đè bởi PGD support)
          "dienbao_ht" | "dienbao_prev" → pgd_data/{slug}/dienbao_ht.xlsx | dienbao_prev.xlsx
    """
    slug_dir = Path(PGD_DATA_DIR) / pgd_slug(ten_pgd)
    if loai == "dienbao_ht":
        return str(slug_dir / "dienbao_ht.xlsx")
    if loai == "dienbao_prev":
        return str(slug_dir / "dienbao_prev.xlsx")
    if loai == "hstd_khnv":
        return str(thu_muc_pgd(ten_pgd) / "hstd_khnv.xlsx")
    return str(thu_muc_pgd(ten_pgd) / f"{loai}_latest.xlsx")


def kiem_tra_file_ton_tai_pgd(ten_pgd: str, loai: str) -> bool:
    """Kiểm tra xem PGD đã upload file loại {loai} chưa.
    HSTD: kiểm tra cả hstd_latest.xlsx lẫn hstd_khnv.xlsx (KH-NV upload)."""
    if Path(duong_dan_pgd(ten_pgd, loai)).exists():
        return True
    if loai == "hstd" and Path(duong_dan_pgd(ten_pgd, "hstd_khnv")).exists():
        return True
    return False


def duong_dan_hstd_hien_hanh(ten_pgd: str) -> str:
    """
    Chọn file HSTD hiện hành theo mtime giữa PGD upload và KH-NV upload.
    Nếu chưa có file nào, trả về đường dẫn chuẩn hstd_latest.xlsx.
    """
    path_latest = Path(duong_dan_pgd(ten_pgd, "hstd"))
    path_khnv = Path(duong_dan_pgd(ten_pgd, "hstd_khnv"))
    candidates = [p for p in (path_latest, path_khnv) if p.exists()]
    if not candidates:
        return str(path_latest)
    return str(max(candidates, key=lambda p: p.stat().st_mtime))


# ── Trạng thái chi tiết từng file PGD ────────────────────────────────────────

LoaiFile = Literal["hstd", "nq11", "gqvl", "cdtotkvv"]

_RE_NGAY_SO_LIEU_VN = re.compile(r"Ngày\s+(\d+)\s+tháng\s+(\d+)\s+năm\s+(\d+)")
_HSTD_NGAY_SO_LIEU_LABEL = "Ng\u00e0y s\u1ed1 li\u1ec7u"
_RE_XLSX_CELL = re.compile(rb"<c[^>]* r=\"([A-Z]+)(\d+)\"[^>]*>.*?</c>", re.S)
_RE_XLSX_TEXT = re.compile(rb"<t[^>]*>(.*?)</t>", re.S)
_RE_XLSX_VALUE = re.compile(rb"<v[^>]*>(.*?)</v>", re.S)


def _decode_xlsx_xml_text(raw: bytes) -> str:
    return html.unescape(raw.decode("utf-8", errors="ignore")).strip()


def _xlsx_cell_text(cell_xml: bytes, shared_strings: list[str]) -> str:
    texts = [_decode_xlsx_xml_text(m.group(1)) for m in _RE_XLSX_TEXT.finditer(cell_xml)]
    if texts:
        return "".join(texts).strip()

    m = _RE_XLSX_VALUE.search(cell_xml)
    if not m:
        return ""

    val = _decode_xlsx_xml_text(m.group(1))
    if b't="s"' in cell_xml:
        try:
            return shared_strings[int(val)]
        except Exception:
            return ""
    return val


def _xlsx_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    try:
        data = zf.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    items: list[str] = []
    for si in re.finditer(rb"<si[^>]*>.*?</si>", data, re.S):
        parts = [_decode_xlsx_xml_text(m.group(1)) for m in _RE_XLSX_TEXT.finditer(si.group(0))]
        items.append("".join(parts).strip())
    return items


def _doc_ngay_so_lieu_hstd_xml(path: Path) -> datetime | None:
    """Đọc nhanh ngày số liệu HSTD từ sheet XML, tự tìm cột theo header dòng 5."""
    try:
        with zipfile.ZipFile(path) as zf:
            shared_strings = _xlsx_shared_strings(zf)
            sheet_name = next(
                name for name in zf.namelist()
                if name.startswith("xl/worksheets/") and name.endswith(".xml")
            )
            sheet = zf.read(sheet_name)
    except Exception:
        return None

    col_ngay = ""
    for cell in _RE_XLSX_CELL.finditer(sheet):
        col, row_s = cell.group(1).decode("ascii"), cell.group(2).decode("ascii")
        if row_s != "5":
            continue
        if _xlsx_cell_text(cell.group(0), shared_strings) == _HSTD_NGAY_SO_LIEU_LABEL:
            col_ngay = col
            break
    if not col_ngay:
        return None

    for cell in _RE_XLSX_CELL.finditer(sheet):
        col, row_s = cell.group(1).decode("ascii"), cell.group(2).decode("ascii")
        if col != col_ngay:
            continue
        try:
            if int(row_s) < 6:
                continue
        except ValueError:
            continue
        dt = _xlsx_val_to_datetime(_xlsx_cell_text(cell.group(0), shared_strings))
        if dt is not None:
            return dt
    return None


def _xlsx_val_to_datetime(val) -> datetime | None:
    """Chuẩn hóa giá trị ô Excel (datetime, date, serial, chuỗi) → datetime."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime.combine(val, datetime.min.time())
    if isinstance(val, (int, float)):
        try:
            return from_excel(float(val))
        except Exception:
            return None
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        head = s.split()[0]
        for part in (head, s[:10], s[:19]):
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(part, fmt)
                except ValueError:
                    continue
    return None


@st.cache_data(ttl=7200, show_spinner=False)
def _doc_ngay_so_lieu(path: Path, loai: str, mtime: float = 0.0) -> datetime | None:
    """Đọc ngày số liệu từ trong file xlsx (read_only, tối thiểu ô cần thiết). Lỗi → None.
    mtime: os.path.getmtime(path) — nằm trong cache key, tự invalidate khi file thay đổi."""
    if loai == "hstd":
        return _doc_ngay_so_lieu_hstd_xml(path)

    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            if loai == "nq11":
                ws = wb["BCQUERY"]
                cidx = column_index_from_string("BA")
                val = None
                for row in ws.iter_rows(min_row=6, max_row=6, min_col=cidx, max_col=cidx):
                    for cell in row:
                        val = cell.value
                        break
                    break
                return _xlsx_val_to_datetime(val)
            if loai in ("gqvl", "cdtotkvv"):
                ws = wb.worksheets[0]
                for row in ws.iter_rows(min_row=6, max_row=6):
                    for cell in row:
                        v = cell.value
                        if not isinstance(v, str):
                            continue
                        m = _RE_NGAY_SO_LIEU_VN.search(v)
                        if m:
                            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
                            return datetime(y, mo, d)
                return None
        finally:
            wb.close()
    except Exception:
        return None


@st.cache_data(ttl=7200, show_spinner=False)
def doc_trang_thai_file(ten_don_vi: str, loai: LoaiFile, mtime: float = 0.0) -> dict:
    """
    Đọc trạng thái file upload của một đơn vị theo loại.
    mtime: os.path.getmtime của file — nằm trong cache key, tự invalidate khi file thay đổi.
    Truyền mtime=0 nếu file chưa tồn tại (cache sẽ trả {"co_file": False}).
    Khi file được upload mới, mtime thay đổi → Streamlit tạo cache entry mới.

    Trả về dict:
      {
        "co_file":    bool,
        "ngay_upload": datetime | None,   # từ os.path.getmtime
        "so_ngay_cu":  int | None,
        "canh_bao":     "ok" | "cu" | "khong_co",
        "ngay_so_lieu": datetime | None  # từ nội dung file; None nếu không đọc được
      }
    Phân loại:
      ok       — file tồn tại và còn trong ngưỡng cảnh báo
      cu       — file tồn tại nhưng đã quá ngưỡng ngày
      khong_co — chưa có file
    """
    path = Path(duong_dan_pgd(ten_don_vi, loai))
    # Với HSTD: ưu tiên file mới hơn giữa PGD upload (`hstd_latest.xlsx`)
    # và KH-NV upload (`hstd_khnv.xlsx`).
    if loai == "hstd":
        path = Path(duong_dan_hstd_hien_hanh(ten_don_vi))
    if not path.exists():
        return {
            "co_file": False,
            "ngay_upload": None,
            "so_ngay_cu": None,
            "canh_bao": "khong_co",
            "ngay_so_lieu": None,
        }

    ngay_upload = datetime.fromtimestamp(os.path.getmtime(path))
    so_ngay_cu  = (datetime.now() - ngay_upload).days
    nguong      = UPLOAD_CANH_BAO_NGAY.get(loai, 3)
    canh_bao    = "ok" if so_ngay_cu <= nguong else "cu"

    try:
        ngay_so_lieu = _doc_ngay_so_lieu(path, loai, mtime=os.path.getmtime(path))
    except Exception:
        ngay_so_lieu = None

    return {
        "co_file":    True,
        "ngay_upload": ngay_upload,
        "so_ngay_cu":  so_ngay_cu,
        "canh_bao":   canh_bao,
        "ngay_so_lieu": ngay_so_lieu,
    }


def _format_badge(tt: dict) -> str:
    """Chuyển dict trạng thái → chuỗi badge hiển thị trong bảng."""
    if not tt["co_file"]:
        return "❌ Chưa có"
    if tt.get("ngay_so_lieu"):
        ngay_str = tt["ngay_so_lieu"].strftime("%d/%m")
        date_part = f"SL: {ngay_str}"
    else:
        ngay_str = tt["ngay_upload"].strftime("%d/%m")
        date_part = ngay_str
    if tt["canh_bao"] == "ok":
        return f"✅ {date_part}"
    return f"⚠️ {date_part} ({tt['so_ngay_cu']} ngày)"


def lay_trang_thai_upload_pgd(ds_don_vi: list[str]) -> pd.DataFrame:
    """
    Trả về DataFrame trạng thái upload của tất cả đơn vị.
    Columns: Đơn vị | HSTD | NQ11 | GQVL | CDTOTKVV
    Giá trị ô: "✅ dd/mm" | "⚠️ dd/mm (N ngày)" | "❌ Chưa có"

    Truyền mtime thực vào doc_trang_thai_file để cache tự invalidate khi file thay đổi.
    Khi chưa có file: mtime=0.0 → cache entry riêng, không lẫn với entry sau khi upload.
    """
    def _mt(ten_pgd: str, loai_: str) -> float:
        """Lấy mtime của file PGD (0.0 nếu chưa tồn tại)."""
        try:
            p = Path(duong_dan_pgd(ten_pgd, loai_))
            return p.stat().st_mtime if p.exists() else 0.0
        except Exception:
            return 0.0

    rows = []
    for ten_dv in ds_don_vi:
        # HSTD: lấy mtime lớn hơn giữa hstd_latest.xlsx và hstd_khnv.xlsx
        hstd_mt = max(_mt(ten_dv, "hstd"), _mt(ten_dv, "hstd_khnv"))
        rows.append({
            "Đơn vị":   ten_dv,
            "HSTD":     _format_badge(doc_trang_thai_file(ten_dv, "hstd",     hstd_mt)),
            "NQ11":     _format_badge(doc_trang_thai_file(ten_dv, "nq11",     _mt(ten_dv, "nq11"))),
            "GQVL":     _format_badge(doc_trang_thai_file(ten_dv, "gqvl",     _mt(ten_dv, "gqvl"))),
            "CDTOTKVV": _format_badge(doc_trang_thai_file(ten_dv, "cdtotkvv", _mt(ten_dv, "cdtotkvv"))),
        })
    return pd.DataFrame(rows)


# ── Upload / Lưu ─────────────────────────────────────────────────────────────
def luu_file_pgd(ten_pgd: str, loai: str, file_bytes: bytes) -> str:
    """
    Lưu file upload cho PGD vào pgd_data/{slug}/{loai}_latest.xlsx.
    Trả về đường dẫn tuyệt đối.
    """
    path = duong_dan_pgd(ten_pgd, loai)
    with open(path, "wb") as f:
        f.write(file_bytes)
    # Xóa cache parquet liên quan
    cache = Path(path).with_suffix(".parquet")
    if cache.exists():
        cache.unlink()
    return path


def luu_file_pgd_voi_lich_su(
    ten_pgd: str,
    loai: str,
    file_bytes: bytes,
    thang_nam: str,          # "MM/YYYY"
) -> str:
    """
    Lưu song song 2 bản:
      1. {loai}_latest.xlsx      — ghi đè (realtime)
      2. {loai}_{YYYY}_{MM}.xlsx — lưu lịch sử, KHÔNG ghi đè
    Trả về đường dẫn latest.
    """
    thu_muc = thu_muc_pgd(ten_pgd)

    try:
        dt = datetime.strptime(thang_nam, "%m/%Y")
        suffix = dt.strftime("%Y_%m")
    except ValueError:
        suffix = thang_nam.replace("/", "_")

    path_latest = thu_muc / f"{loai}_latest.xlsx"
    path_version = thu_muc / f"{loai}_{suffix}.xlsx"

    # Ghi latest (luôn ghi đè)
    path_latest.write_bytes(file_bytes)

    # Ghi version (chỉ ghi nếu chưa có)
    if not path_version.exists():
        path_version.write_bytes(file_bytes)

    # Xóa parquet cache
    for p in [path_latest, path_version]:
        cache = p.with_suffix(".parquet")
        if cache.exists():
            cache.unlink()

    return str(path_latest)


# ── Legacy: đường dẫn cũ (gqvl_pgd/) ────────────────────────────────────────
def duong_dan_gqvl_pgd(ten_pgd: str) -> str:
    """Đường dẫn file GQVL cũ trong gqvl_pgd/ (legacy — fallback khi cần)."""
    os.makedirs(GQVL_PGD_DIR, exist_ok=True)
    slug = "gqvl_" + pgd_slug(ten_pgd).replace("pgd_", "", 1) \
           if pgd_slug(ten_pgd).startswith("pgd_") else f"gqvl_{pgd_slug(ten_pgd)}"
    return os.path.join(GQVL_PGD_DIR, f"{slug}.xlsx")


def luu_gqvl_pgd(ten_pgd: str, file_bytes: bytes) -> str:
    """Lưu file GQVL — dùng luu_file_pgd thay thế."""
    return luu_file_pgd(ten_pgd, "gqvl", file_bytes)


def ds_pgd_co_file(loai: str) -> list:
    """Danh sách tên PGD đã upload file theo loại (hstd/nq11/gqvl).
    Với HSTD: kiểm tra cả hstd_latest.xlsx (PGD tự upload) và hstd_khnv.xlsx (Phòng KH-NV upload).
    """
    if not os.path.exists(PGD_DATA_DIR):
        return []
    result = []
    for d in sorted(Path(PGD_DATA_DIR).iterdir()):
        if not d.is_dir():
            continue
        co_file = (d / f"{loai}_latest.xlsx").exists()
        if not co_file and loai == "hstd":
            co_file = (d / "hstd_khnv.xlsx").exists()
        if co_file:
            # Khôi phục tên hiển thị từ slug (gần đúng)
            result.append(d.name.replace("_", " ").title())
    return result


def ds_pgd_co_gqvl() -> list:
    """Danh sách PGD đã có file GQVL."""
    return ds_pgd_co_file("gqvl")


# ── Đọc file PGD đơn lẻ ──────────────────────────────────────────────────────
@st.cache_data(ttl=7200, show_spinner=False)
def doc_hstd_pgd(ten_pgd: str, file_mtime: float = 0.0) -> pd.DataFrame | None:
    """
    Đọc HSTD của một PGD từ file Parquet riêng.

    Tham số `file_mtime` (timestamp) là cache buster — khi file được upload
    mới, truyền vào os.path.getmtime(path) để cache khớp key mới (invalidate
    theo PGD) mà không cần st.cache_data.clear() toàn cục.

    Cách gọi khi đọc bình thường:
        df = doc_hstd_pgd("PGD Long Thành")

    Cách gọi sau khi upload file mới:
        ts = os.path.getmtime(duong_dan_pgd(ten_pgd, "hstd"))
        df = doc_hstd_pgd(ten_pgd, file_mtime=ts)

    (Tham số không dùng tiền tố `_` để Streamlit đưa vào cache key.)
    """
    from data.hstd import doc_file

    path = duong_dan_hstd_hien_hanh(ten_pgd)
    return doc_file(path, file_mtime) if os.path.exists(path) else None


@st.cache_data(ttl=7200, show_spinner=False)
def doc_nq11_pgd(ten_pgd: str, _ts):
    from data.hstd import doc_file_nq11

    path = duong_dan_pgd(ten_pgd, "nq11")
    return doc_file_nq11(path, _ts) if os.path.exists(path) else None


@st.cache_data(ttl=7200, show_spinner=False)
def doc_gqvl_pgd_v2(ten_pgd: str, _ts):
    """Đọc GQVL PGD — ưu tiên pgd_data/{slug}/, fallback gqvl_pgd/ (legacy)."""
    from data.hstd import doc_file_gqvl

    path = duong_dan_pgd(ten_pgd, "gqvl")
    if not os.path.exists(path):
        path = duong_dan_gqvl_pgd(ten_pgd)
    return doc_file_gqvl(path, ts_file(path)) if os.path.exists(path) else None


@st.cache_data(ttl=7200, show_spinner=False)
def doc_gqvl_pgd(ten_pgd: str, _ts):
    return doc_gqvl_pgd_v2(ten_pgd, _ts)


# ── Gộp toàn CN ──────────────────────────────────────────────────────────────
@st.cache_data(ttl=7200, show_spinner=False)
def doc_hstd_toan_cn_pgd(pgd_dir_mtime: float = 0.0) -> pd.DataFrame | None:
    """Gộp HSTD tất cả PGD — ws_operation, độc lập với CACHE_HSTD.

    Chọn file mới hơn giữa hstd_latest.xlsx (PGD upload) và hstd_khnv.xlsx (KH-NV).
    `pgd_dir_mtime` = max mtime của các file nguồn → cache key.
    (Không dùng tiền tố `_` để Streamlit đưa vào cache key.)
    """
    from data.hstd import doc_file

    if not os.path.exists(PGD_DATA_DIR):
        return None
    frames = []
    for d in sorted(Path(PGD_DATA_DIR).iterdir()):
        if not d.is_dir():
            continue
        ten_pgd = d.name.replace("_", " ").title()
        path = Path(duong_dan_hstd_hien_hanh(ten_pgd))
        if path.exists():
            try:
                frames.append(doc_file(str(path), ts_file(str(path))))
            except Exception:
                pass
    return pd.concat(frames, ignore_index=True) if frames else None


@st.cache_data(ttl=7200, show_spinner=False)
def doc_gqvl_toan_cn_pgd() -> pd.DataFrame | None:
    """Gộp GQVL tất cả PGD đã upload."""
    from data.hstd import doc_file_gqvl

    if not os.path.exists(PGD_DATA_DIR):
        return _doc_gqvl_toan_cn_legacy()
    frames = []
    for d in sorted(Path(PGD_DATA_DIR).iterdir()):
        path = d / "gqvl_latest.xlsx"
        if d.is_dir() and path.exists():
            try:
                df_p = doc_file_gqvl(str(path), ts_file(str(path)))
                df_p["_pgd"] = d.name.replace("_", " ").title()
                frames.append(df_p)
            except Exception:
                pass
    if not frames:
        return _doc_gqvl_toan_cn_legacy()
    return pd.concat(frames, ignore_index=True)


def doc_gqvl_toan_cn(_ts_sentinel=None) -> pd.DataFrame | None:
    """Alias — gộp GQVL toàn CN."""
    return doc_gqvl_toan_cn_pgd()


def _doc_gqvl_toan_cn_legacy() -> pd.DataFrame | None:
    """Fallback: gộp từ thư mục gqvl_pgd/ cũ."""
    from data.hstd import doc_file_gqvl

    if not os.path.exists(GQVL_PGD_DIR):
        return None
    frames = []
    for fname in os.listdir(GQVL_PGD_DIR):
        if not fname.endswith(".xlsx"):
            continue
        path = os.path.join(GQVL_PGD_DIR, fname)
        try:
            df_p = doc_file_gqvl(path, ts_file(path))
            df_p["_pgd"] = fname.replace("gqvl_", "").replace(".xlsx", "") \
                               .replace("_", " ").title()
            frames.append(df_p)
        except Exception:
            pass
    return pd.concat(frames, ignore_index=True) if frames else None


@st.cache_data(ttl=7200, show_spinner=False)
def doc_nq11_toan_cn_pgd() -> pd.DataFrame | None:
    """Gộp NQ11 tất cả PGD đã upload."""
    from data.hstd import doc_file_nq11

    if not os.path.exists(PGD_DATA_DIR):
        return None
    frames = []
    for d in sorted(Path(PGD_DATA_DIR).iterdir()):
        path = d / "nq11_latest.xlsx"
        if d.is_dir() and path.exists():
            try:
                frames.append(doc_file_nq11(str(path), ts_file(str(path))))
            except Exception:
                pass
    return pd.concat(frames, ignore_index=True) if frames else None
