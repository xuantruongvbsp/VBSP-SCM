#!/usr/bin/env python3
"""
daily_report.py — Tạo báo cáo Excel tóm tắt định kỳ hằng ngày.
Chạy độc lập: python scripts/daily_report.py
Có thể gọi từ Task Scheduler hoặc chạy thủ công.
"""
from __future__ import annotations

import os
import sys
import json
from datetime import datetime, date, timedelta
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import db
from data.core import _duckdb_query
from config import (
    CACHE_HSTD,
    COT_TEN_PGD, COT_TEN_XA, COT_TEN_CT, COT_TEN_KH,
    COT_SO_KU, COT_MA_KH, COT_TONG_DU_NO, COT_DU_NO_QH, COT_DU_NO_TH,
    COT_DU_NO_KHOANH, COT_LAI_TON, COT_NGAY_DH, COT_NGUON_VON,
    DS_PGD, COT_TEN_TO,
)
from data.pgd import pgd_slug
from utils import fmt_ty, fmt_so

REPORT_DIR = BASE_DIR / "cache" / "reports"
HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=10)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
AMBER_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")


def _style_header(ws, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _style_data(ws, nrows: int, ncols: int, tien_cols: list[int] | None = None):
    tien_cols = tien_cols or []
    for row in range(2, nrows + 1):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.font = NORMAL_FONT
            if col in tien_cols:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
        if row == nrows:
            for col in range(1, ncols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = BOLD_FONT


def _auto_width(ws, ncols: int):
    for col in range(1, ncols + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value or ""
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)


def _build_tong_quan_sheet(wb: Workbook, parquet_path: str):
    """Sheet 1: Tổng quan dư nợ theo PGD."""
    ws = wb.active
    ws.title = "Tổng quan PGD"

    if not os.path.exists(parquet_path):
        ws.cell(row=1, column=1, value="⚠️ Chưa có dữ liệu parquet").font = Font(bold=True, color="FF0000")
        return

    sql = f"""
        SELECT
            "{COT_TEN_PGD}"          AS "PGD",
            COUNT(DISTINCT "{COT_MA_KH}") AS "Số KH",
            COUNT("{COT_SO_KU}")      AS "Số món",
            SUM("{COT_TONG_DU_NO}")   AS "Dư nợ",
            SUM("{COT_DU_NO_QH}")     AS "Nợ QH",
            SUM("{COT_DU_NO_TH}")     AS "Nợ TH"
        FROM read_parquet(?)
        WHERE "{COT_TONG_DU_NO}" IS NOT NULL
        GROUP BY "{COT_TEN_PGD}"
        ORDER BY "Dư nợ" DESC
    """
    try:
        df = _duckdb_query(sql, [parquet_path])
    except Exception:
        ws.cell(row=1, column=1, value="⚠️ Lỗi truy vấn dữ liệu").font = Font(bold=True, color="FF0000")
        return

    total = pd.DataFrame([{
        "PGD": "TOÀN CHI NHÁNH",
        "Số KH": df["Số KH"].sum(),
        "Số món": df["Số món"].sum(),
        "Dư nợ": df["Dư nợ"].sum(),
        "Nợ QH": df["Nợ QH"].sum(),
        "Nợ TH": df["Nợ TH"].sum(),
    }])
    df["Tỷ lệ NQH"] = (df["Nợ QH"] / df["Dư nợ"].replace(0, pd.NA) * 100).round(2)
    total["Tỷ lệ NQH"] = round(total["Nợ QH"].iloc[0] / total["Dư nợ"].iloc[0] * 100, 2) if total["Dư nợ"].iloc[0] > 0 else 0
    df_out = pd.concat([df, total], ignore_index=True)

    for c in ["Dư nợ", "Nợ QH", "Nợ TH"]:
        df_out[c] = df_out[c].astype(int)

    cols = list(df_out.columns)
    for i, col in enumerate(cols, 1):
        ws.cell(row=1, column=i, value=col)

    for r, (_, row) in enumerate(df_out.iterrows(), 2):
        for c, col in enumerate(cols, 1):
            ws.cell(row=r, column=c, value=row[col])

    ncols = len(cols)
    _style_header(ws, ncols)
    tien_cols = [i + 1 for i, c in enumerate(cols) if c in ("Dư nợ", "Nợ QH", "Nợ TH")]
    _style_data(ws, len(df_out) + 1, ncols, tien_cols)

    for r in range(2, len(df_out) + 2):
        tl = ws.cell(row=r, column=ncols).value
        if isinstance(tl, (int, float)):
            if tl > 2:
                ws.cell(row=r, column=ncols).fill = RED_FILL
            elif tl > 1:
                ws.cell(row=r, column=ncols).fill = AMBER_FILL

    _auto_width(ws, ncols)


def _build_nqh_top_sheet(wb: Workbook, parquet_path: str):
    """Sheet 2: Top khoản vay NQH cao nhất."""
    ws = wb.create_sheet("Top NQH")

    if not os.path.exists(parquet_path):
        ws.cell(row=1, column=1, value="⚠️ Chưa có dữ liệu")
        return

    sql = f"""
        SELECT
            "{COT_TEN_PGD}"   AS "PGD",
            "{COT_TEN_KH}"    AS "Tên KH",
            "{COT_SO_KU}"     AS "Số KU",
            "{COT_TEN_CT}"    AS "CT",
            "{COT_TONG_DU_NO}" AS "Dư nợ",
            "{COT_DU_NO_QH}"  AS "Nợ QH",
            "{COT_LAI_TON}"   AS "Lãi tồn"
        FROM read_parquet(?)
        WHERE "{COT_DU_NO_QH}" > 0
        ORDER BY "{COT_DU_NO_QH}" DESC
        LIMIT 50
    """
    try:
        df = _duckdb_query(sql, [parquet_path])
    except Exception:
        ws.cell(row=1, column=1, value="⚠️ Lỗi truy vấn")
        return

    cols = list(df.columns)
    for i, col in enumerate(cols, 1):
        ws.cell(row=1, column=i, value=col)
    for r, (_, row) in enumerate(df.iterrows(), 2):
        for c, col in enumerate(cols, 1):
            ws.cell(row=r, column=c, value=row[col])

    ncols = len(cols)
    _style_header(ws, ncols)
    tien_cols = [i + 1 for i, c in enumerate(cols) if c in ("Dư nợ", "Nợ QH", "Lãi tồn")]
    _style_data(ws, len(df) + 1, ncols, tien_cols)
    _auto_width(ws, ncols)


def _build_den_han_sheet(wb: Workbook, parquet_path: str):
    """Sheet 3: Khoản vay đến hạn trong 30 ngày."""
    ws = wb.create_sheet("Đến hạn 30 ngày")

    if not os.path.exists(parquet_path):
        ws.cell(row=1, column=1, value="⚠️ Chưa có dữ liệu")
        return

    cutoff = (date.today() + timedelta(days=30)).isoformat()

    sql = f"""
        SELECT
            "{COT_TEN_PGD}"   AS "PGD",
            "{COT_TEN_KH}"    AS "Tên KH",
            "{COT_SO_KU}"     AS "Số KU",
            "{COT_NGAY_DH}"   AS "Ngày đến hạn",
            "{COT_TEN_CT}"    AS "CT",
            "{COT_TONG_DU_NO}" AS "Dư nợ",
            "{COT_TEN_TO}"    AS "Tổ TK&VV"
        FROM read_parquet(?)
        WHERE "{COT_NGAY_DH}" IS NOT NULL
          AND "{COT_NGAY_DH}" <= '{cutoff}'
          AND "{COT_NGAY_DH}" >= CURRENT_DATE
          AND "{COT_TONG_DU_NO}" > 0
        ORDER BY "{COT_NGAY_DH}", "{COT_TONG_DU_NO}" DESC
    """
    try:
        df = _duckdb_query(sql, [parquet_path])
    except Exception:
        ws.cell(row=1, column=1, value="⚠️ Lỗi truy vấn")
        return

    cols = list(df.columns)
    for i, col in enumerate(cols, 1):
        ws.cell(row=1, column=i, value=col)
    for r, (_, row) in enumerate(df.iterrows(), 2):
        for c, col in enumerate(cols, 1):
            ws.cell(row=r, column=c, value=row[col])

    ncols = len(cols)
    _style_header(ws, ncols)
    tien_cols = [i + 1 for i, c in enumerate(cols) if c in ("Dư nợ",)]
    _style_data(ws, len(df) + 1, ncols, tien_cols)
    _auto_width(ws, ncols)


def _build_khtd_sheet(wb: Workbook):
    """Sheet 4: KHTD tiến độ."""
    ws = wb.create_sheet("KHTD tiến độ")

    data = db.doc_kv("khtd_cn")
    if not data:
        ws.cell(row=1, column=1, value="⚠️ Chưa nhập KHTD Chi nhánh")
        return

    rows = []
    for ct, targets in sorted(data.items()):
        if not isinstance(targets, dict):
            continue
        for ten_pgd, val in targets.items():
            if isinstance(val, (int, float)):
                rows.append({"CT": ct, "PGD": ten_pgd, "KHTD (triệu đ)": round(val / 1_000_000, 1)})

    if not rows:
        ws.cell(row=1, column=1, value="⚠️ Chưa có dữ liệu KHTD")
        return

    df = pd.DataFrame(rows)
    cols = list(df.columns)
    for i, col in enumerate(cols, 1):
        ws.cell(row=1, column=i, value=col)
    for r, (_, row) in enumerate(df.iterrows(), 2):
        for c, col in enumerate(cols, 1):
            ws.cell(row=r, column=c, value=row[col])

    ncols = len(cols)
    _style_header(ws, ncols)
    tien_cols = [i + 1 for i, c in enumerate(cols) if "KHTD" in c]
    _style_data(ws, len(df) + 1, ncols, tien_cols)
    _auto_width(ws, ncols)


def _build_bia_sheet(wb: Workbook):
    """Sheet Bìa: thông tin báo cáo."""
    ws = wb.create_sheet("Bìa", 0)
    ws.merge_cells("A1:E1")
    title_cell = ws.cell(row=1, column=1, value="BÁO CÁO TÓM TẮT ĐỊNH KỲ")
    title_cell.font = Font(bold=True, size=16, color="2E7D32")
    title_cell.alignment = Alignment(horizontal="center")

    now_str = datetime.now().strftime("%H:%M %d/%m/%Y")
    ws.merge_cells("A3:E3")
    ws.cell(row=3, column=1, value=f"Thời gian xuất: {now_str}").font = Font(size=11)
    ws.cell(row=3, column=1).alignment = Alignment(horizontal="center")

    ws.merge_cells("A5:E5")
    ws.cell(row=5, column=1, value="NHCSXH Chi nhánh Đồng Nai").font = Font(bold=True, size=12)
    ws.cell(row=5, column=1).alignment = Alignment(horizontal="center")

    ws.merge_cells("A7:E7")
    ws.cell(row=7, column=1, value="Hệ thống Quản trị Tín dụng Nội bộ — VBSP-SCM").font = Font(size=10, italic=True)
    ws.cell(row=7, column=1).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15


def _vn(x: float, decimals: int = 0) -> str:
    """Format số kiểu Việt Nam: dấu . ngàn, dấu , thập phân."""
    s = f"{x:,.{decimals}f}"
    return s.replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def _trong_gio_gui(key: str) -> bool:
    """Kiểm tra giờ hiện tại có nằm trong cửa sổ ±15 phút so với giờ đã cấu hình không.
    Trả True nếu chưa cấu hình hoặc đang trong cửa sổ gửi.
    """
    try:
        cfg = db.doc_kv("telegram_schedule_config") or {}
        gio_cfg = cfg.get(key, "")
        if not gio_cfg or ":" not in gio_cfg:
            return True
        h, m = map(int, gio_cfg.split(":", 1))
        now = datetime.now()
        diff = abs(now.hour * 60 + now.minute - (h * 60 + m))
        return diff <= 15
    except Exception:
        return True


def generate_daily_report() -> str | None:
    """Tạo báo cáo Excel hằng ngày. Trả về đường dẫn file hoặc None nếu lỗi."""
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    parquet_path = str(CACHE_HSTD)
    if not os.path.exists(parquet_path):
        print(f"❌ Không tìm thấy {CACHE_HSTD}")
        return None

    wb = Workbook()
    _build_bia_sheet(wb)
    _build_tong_quan_sheet(wb, parquet_path)
    _build_nqh_top_sheet(wb, parquet_path)
    _build_den_han_sheet(wb, parquet_path)
    _build_khtd_sheet(wb)

    now = datetime.now()
    filename = f"BaoCao_Ngay_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = REPORT_DIR / filename

    wb.save(str(filepath))

    info = {
        "ts": now.isoformat(),
        "file": filename,
        "size": filepath.stat().st_size,
    }
    meta_path = REPORT_DIR / "latest.json"
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(info, f, ensure_ascii=False)

    print(f"✅ Đã tạo báo cáo: {filepath} ({info['size'] / 1024:.1f} KB)")

    # Gửi tóm tắt qua Telegram
    try:
        from services.telegram_service import gui_bao_cao_sang
        from config import COT_TONG_DU_NO, COT_DU_NO_QH

        if os.path.exists(parquet_path):
            sql_sum = f"""
                SELECT
                    SUM("{COT_TONG_DU_NO}") AS tong_dn,
                    SUM("{COT_DU_NO_QH}")   AS tong_qh
                FROM read_parquet(?)
                WHERE "{COT_TONG_DU_NO}" IS NOT NULL
            """
            df_sum = _duckdb_query(sql_sum, [parquet_path])
            tong_dn = int(df_sum["tong_dn"].iloc[0] or 0)
            tong_qh = int(df_sum["tong_qh"].iloc[0] or 0)
            ty_le_qh = f"{tong_qh / tong_dn * 100:.2f}%" if tong_dn > 0 else "—"

            merge_meta = db.doc_kv("merge_meta_hstd") or {}
            so_pgd = merge_meta.get("so_pgd", 0)

            from config import DS_PGD
            gui_bao_cao_sang(
                ngay=now.strftime("%d/%m/%Y"),
                tong_du_no=f"{_vn(tong_dn / 1e9, 1)} tỷ",
                tong_qh=f"{_vn(tong_qh / 1e6, 0)} triệu",
                ty_le_qh=ty_le_qh,
                so_pgd_da_upload=so_pgd,
                tong_pgd=len(DS_PGD),
            )
    except Exception as e:
        print(f"⚠️ Telegram: {e}")

    # Gửi nhắc khoản đến hạn trong tháng qua Telegram
    try:
        import calendar
        from services.telegram_service import gui_nhac_khoang_den_han

        last_day = calendar.monthrange(now.year, now.month)[1]
        end_of_month = date(now.year, now.month, last_day).isoformat()

        sql_dh = f"""
            SELECT
                "{COT_TEN_KH}"     AS ten_kh,
                "{COT_SO_KU}"      AS so_ku,
                "{COT_NGAY_DH}"    AS ngay_dh,
                "{COT_TONG_DU_NO}" AS du_no,
                "{COT_TEN_PGD}"    AS ten_pgd
            FROM read_parquet(?)
            WHERE "{COT_NGAY_DH}" IS NOT NULL
              AND "{COT_NGAY_DH}" >= CURRENT_DATE
              AND "{COT_NGAY_DH}" <= '{end_of_month}'
              AND "{COT_TONG_DU_NO}" > 0
            ORDER BY "{COT_NGAY_DH}", "{COT_TONG_DU_NO}" DESC
        """
        if os.path.exists(parquet_path):
            df_dh = _duckdb_query(sql_dh, [parquet_path])
            ds_khoang = [
                {
                    "ten_kh":  str(row["ten_kh"] or ""),
                    "so_ku":   str(row["so_ku"] or ""),
                    "ngay_dh": str(row["ngay_dh"] or ""),
                    "du_no":   f"{int(row['du_no'] or 0) / 1e6:,.0f} triệu".replace(",", "."),
                    "ten_pgd": str(row["ten_pgd"] or ""),
                }
                for _, row in df_dh.iterrows()
            ]
            gui_nhac_khoang_den_han(ds_khoang)
    except Exception as e:
        print(f"⚠️ Telegram nhắc đến hạn: {e}")

    try:
        _nhac_phan_ky_nxh()
    except Exception as e:
        print(f"⚠️ Telegram nhắc phân kỳ NXH: {e}")

    if _trong_gio_gui("qh_moi"):
        try:
            _canh_bao_qh_moi()
        except Exception as e:
            print(f"⚠️ Telegram cảnh báo NQH: {e}")

    # Thứ Sáu: báo cáo giải ngân tuần
    if date.today().weekday() == 4 and _trong_gio_gui("giai_ngan_tuan"):
        try:
            _giai_ngan_tuan()
        except Exception as e:
            print(f"⚠️ Telegram giải ngân tuần: {e}")

    # Mỗi ngày: cảnh báo nợ khoanh tăng
    if _trong_gio_gui("khoanh_tang"):
        try:
            _canh_bao_khoanh_tang()
        except Exception as e:
            print(f"⚠️ Telegram nợ khoanh: {e}")

    # Thứ Hai: báo cáo NQH tuần
    if date.today().weekday() == 0 and _trong_gio_gui("nqh_tuan"):
        try:
            _bao_cao_nqh_tuan()
        except Exception as e:
            print(f"⚠️ Telegram NQH tuần: {e}")

    # Mỗi ngày (nếu bật): tiến độ KHTD theo chương trình
    if _trong_gio_gui("khtd_ct"):
        try:
            _bao_cao_khtd_theo_ct()
        except Exception as e:
            print(f"⚠️ Telegram KHTD chương trình: {e}")

    # Ngày 25–31: tổng kết tháng
    if date.today().day >= 25 and _trong_gio_gui("tong_ket_thang"):
        try:
            _tong_ket_thang()
        except Exception as e:
            print(f"⚠️ Telegram tổng kết tháng: {e}")

    return str(filepath)


def _nhac_phan_ky_nxh() -> int:
    """Đọc parquet NXH, lọc từ hôm nay đến cuối tháng, gửi 1 tin/PGD qua Telegram."""
    from data.phan_ky_nxh import doc_phan_ky_nxh
    from services.telegram_service import gui_nhac_phan_ky_nxh

    df = doc_phan_ky_nxh()
    if df.empty:
        return 0

    today_ts  = pd.Timestamp.today().normalize()
    last_day  = today_ts.replace(day=1) + pd.offsets.MonthEnd(0)

    COL_NGAY = "Ngày đến hạn kỳ con"
    COL_TIEN = "Dư nợ kỳ con đến hạn"
    COL_TGK  = "Tổng TG, TK"
    COL_LAI  = "Lãi tồn"
    COL_PGD  = "Tên PGD"

    if COL_NGAY not in df.columns or COL_PGD not in df.columns:
        return 0

    ngay_du_lieu = today_ts.strftime("%d/%m/%Y")
    # Chỉ lấy khoản đến hạn TỪ HÔM NAY trở đi (bỏ các khoản đã qua)
    mask = (
        df[COL_NGAY].notna()
        & (df[COL_NGAY] >= today_ts)
        & (df[COL_NGAY] <= last_day)
    )
    df_thang = df[mask].sort_values(["Tên xã", COL_NGAY])

    if df_thang.empty:
        return 0

    sent = 0
    for ten_pgd, grp in df_thang.groupby(COL_PGD):
        ds = []
        for _, row in grp.iterrows():
            ngay_dh = ""
            try:
                if pd.notna(row[COL_NGAY]):
                    ngay_dh = pd.Timestamp(row[COL_NGAY]).strftime("%d/%m/%Y")
            except Exception:
                pass
            ds.append({
                "ten_kh":        str(row.get("Tên khách hàng") or ""),
                "so_ku":         str(row.get("Số khế ước") or ""),
                "ngay_dh":       ngay_dh,
                "du_no":         float(row.get(COL_TIEN) or 0),
                "lai_ton":       float(row.get(COL_LAI) or 0) if COL_LAI in grp.columns else 0.0,
                "tong_tgk":      float(row.get(COL_TGK) or 0) if COL_TGK in grp.columns else 0.0,
                "sdt":           str(row.get("Số điện thoại") or ""),
                "ten_xa":        str(row.get("Tên xã") or ""),
                "ten_to_truong": str(row.get("Tên tổ trưởng") or ""),
                "ghi_chu":       str(row.get("Ghi chú") or ""),
            })
        ok = gui_nhac_phan_ky_nxh(str(ten_pgd), ds, ngay_du_lieu=ngay_du_lieu)
        if ok:
            sent += 1
    return sent


def _canh_bao_qh_moi() -> int:
    """So sánh snapshot NQH hôm nay vs kỳ trước, gửi cảnh báo nếu tăng bất thường."""
    from services.telegram_service import gui_canh_bao_qh_moi
    try:
        from snapshot_service import doc_snapshot_range
        from config import COT_TEN_PGD, COT_TONG_DU_NO, COT_DU_NO_QH
        snapshots = doc_snapshot_range(tu_ky=None, den_ky=None, n_ky=2)
        if len(snapshots) < 2:
            return 0
        df_moi, df_cu = snapshots[0], snapshots[1]
        _NGUONG = 0.5  # tăng ≥ 0.5pp được coi là bất thường
        ds_tang = []
        for pgd in df_moi[COT_TEN_PGD].unique():
            row_m = df_moi[df_moi[COT_TEN_PGD] == pgd].iloc[0]
            row_c = df_cu[df_cu[COT_TEN_PGD] == pgd]
            if row_c.empty:
                continue
            row_c = row_c.iloc[0]
            dn_m = float(row_m.get(COT_TONG_DU_NO, 0) or 0)
            qh_m = float(row_m.get(COT_DU_NO_QH, 0) or 0)
            dn_c = float(row_c.get(COT_TONG_DU_NO, 0) or 0)
            qh_c = float(row_c.get(COT_DU_NO_QH, 0) or 0)
            tl_m = qh_m / dn_m * 100 if dn_m else 0.0
            tl_c = qh_c / dn_c * 100 if dn_c else 0.0
            if tl_m - tl_c >= _NGUONG:
                ds_tang.append({
                    "ten_pgd":  str(pgd),
                    "ty_le_cu": tl_c,
                    "ty_le_moi": tl_m,
                    "tang":     tl_m - tl_c,
                })
        if ds_tang:
            gui_canh_bao_qh_moi(ds_tang)
        return len(ds_tang)
    except Exception as e:
        print(f"⚠️ _canh_bao_qh_moi: {e}")
        return 0


def _giai_ngan_tuan() -> int:
    """Thứ Sáu: tổng hợp giải ngân (khoản vay mới) 7 ngày qua."""
    from services.telegram_service import gui_giai_ngan_tuan
    from config import DON_VI_CHI_NHANH
    try:
        if not Path(CACHE_HSTD).exists():
            return 0
        df = pd.read_parquet(CACHE_HSTD)
        if COT_NGAY_VAY not in df.columns:
            return 0
        today_ts = pd.Timestamp.today().normalize()
        t7 = today_ts - pd.Timedelta(days=7)
        mask = df[COT_NGAY_VAY].notna() & (df[COT_NGAY_VAY] >= t7) & (df[COT_NGAY_VAY] <= today_ts)
        df_gn = df[mask & (df[COT_TEN_PGD] != DON_VI_CHI_NHANH)]
        if df_gn.empty:
            return 0
        # t7_str → hiển thị khoảng tuần
        tuan_str = f"{t7.strftime('%d/%m')}–{today_ts.strftime('%d/%m/%Y')}"
        grp = df_gn.groupby(COT_TEN_PGD)[COT_TONG_DU_NO].agg(["sum", "count"]).reset_index()
        ds_pgd = [
            {
                "ten_pgd":  str(r[COT_TEN_PGD]),
                "so_khoan": int(r["count"]),
                "giai_ngan": float(r["sum"]),
            }
            for _, r in grp.iterrows()
        ]
        gui_giai_ngan_tuan(ds_pgd, tuan_str)
        return len(ds_pgd)
    except Exception as e:
        print(f"⚠️ _giai_ngan_tuan: {e}")
        return 0


def _canh_bao_khoanh_tang() -> int:
    """So sánh nợ khoanh snapshot hôm nay vs kỳ trước, cảnh báo nếu tăng ≥ 5%."""
    from services.telegram_service import gui_canh_bao_khoanh_tang
    try:
        from snapshot_service import doc_snapshot_range
        from config import COT_DU_NO_KHOANH
        snapshots = doc_snapshot_range(tu_ky=None, den_ky=None, n_ky=2)
        if len(snapshots) < 2 or COT_DU_NO_KHOANH not in snapshots[0].columns:
            return 0
        df_moi, df_cu = snapshots[0], snapshots[1]
        _NGUONG_PCT = 5.0  # tăng ≥ 5% giá trị tuyệt đối
        ds_tang = []
        for pgd in df_moi[COT_TEN_PGD].unique():
            row_m = df_moi[df_moi[COT_TEN_PGD] == pgd].iloc[0]
            row_c = df_cu[df_cu[COT_TEN_PGD] == pgd]
            if row_c.empty:
                continue
            row_c = row_c.iloc[0]
            kh_moi = float(row_m.get(COT_DU_NO_KHOANH, 0) or 0)
            kh_cu  = float(row_c.get(COT_DU_NO_KHOANH, 0) or 0)
            if kh_cu == 0 or kh_moi == 0:
                continue
            tang_pct = (kh_moi - kh_cu) / kh_cu * 100
            if tang_pct >= _NGUONG_PCT:
                ds_tang.append({
                    "ten_pgd":   str(pgd),
                    "khoanh_cu": kh_cu,
                    "khoanh_moi": kh_moi,
                    "tang_pct":  tang_pct,
                })
        if ds_tang:
            gui_canh_bao_khoanh_tang(ds_tang)
        return len(ds_tang)
    except Exception as e:
        print(f"⚠️ _canh_bao_khoanh_tang: {e}")
        return 0


def _bao_cao_nqh_tuan() -> int:
    """Thứ Hai: gửi báo cáo NQH từng đơn vị qua Telegram."""
    from services.telegram_service import gui_bao_cao_nqh_tuan
    from config import DON_VI_CHI_NHANH
    try:
        if not Path(CACHE_HSTD).exists():
            return 0
        df = pd.read_parquet(CACHE_HSTD, columns=[COT_TEN_PGD, COT_TONG_DU_NO, COT_DU_NO_QH])
        df = df[df[COT_TEN_PGD] != DON_VI_CHI_NHANH]
        grp = df.groupby(COT_TEN_PGD)[[COT_TONG_DU_NO, COT_DU_NO_QH]].sum().reset_index()
        meta = db.doc_kv("merge_meta_hstd") or {}
        ngay_sl = str(meta.get("ngay_sl", date.today().strftime("%d/%m/%Y")))
        ds_pgd = []
        for _, r in grp.iterrows():
            dn  = float(r[COT_TONG_DU_NO] or 0)
            qh  = float(r[COT_DU_NO_QH]   or 0)
            tl  = qh / dn * 100 if dn > 0 else 0.0
            ds_pgd.append({
                "ten_pgd":   str(r[COT_TEN_PGD]),
                "du_no":     dn,
                "nqh":       qh,
                "ty_le_nqh": round(tl, 2),
            })
        gui_bao_cao_nqh_tuan(ds_pgd, ngay_sl)
        return len(ds_pgd)
    except Exception as e:
        print(f"⚠️ _bao_cao_nqh_tuan: {e}")
        return 0


def _bao_cao_khtd_theo_ct() -> int:
    """Gửi tiến độ KHTD theo từng chương trình tín dụng."""
    from services.telegram_service import gui_khtd_theo_chuong_trinh
    try:
        from config import CHUONG_TRINH_KHTD
        khtd_cn = db.doc_kv("khtd_cn")
        if not khtd_cn:
            return 0
        if not Path(CACHE_HSTD).exists():
            return 0
        # Tính thực hiện theo từng chương trình
        df = pd.read_parquet(CACHE_HSTD)
        from tabs.tab_khtd_xuat import _tinh_thuc_hien_theo_ct
        df_th = _tinh_thuc_hien_theo_ct(df)
        meta = db.doc_kv("merge_meta_hstd") or {}
        ngay_sl = str(meta.get("ngay_sl", date.today().strftime("%d/%m/%Y")))
        ds_ct = []
        for ma_key, _ma_ct, ten_hien_thi, nguon_von, _tm in CHUONG_TRINH_KHTD:
            kh_ct  = float((khtd_cn.get(ma_key) or {}).get("_cn", 0) or 0)
            th_row = df_th[df_th["ma_key"] == ma_key] if not df_th.empty and "ma_key" in df_th.columns else pd.DataFrame()
            th_val = float(th_row["thuc_hien"].iloc[0]) if not th_row.empty else 0.0
            pct    = th_val / kh_ct * 100 if kh_ct > 0 else 0.0
            ds_ct.append({
                "ten_ct":    ten_hien_thi,
                "nguon_von": nguon_von,
                "ke_hoach":  kh_ct,
                "thuc_hien": th_val,
                "pct":       round(pct, 1),
            })
        gui_khtd_theo_chuong_trinh(ds_ct, ngay_sl)
        return len(ds_ct)
    except Exception as e:
        print(f"⚠️ _bao_cao_khtd_theo_ct: {e}")
        return 0


def _tong_ket_thang() -> None:
    """Gửi tổng kết tháng — chạy từ ngày 25 đến 31."""
    from services.telegram_service import gui_tong_ket_thang
    from config import DON_VI_CHI_NHANH
    try:
        if not Path(CACHE_HSTD).exists():
            return
        df = pd.read_parquet(CACHE_HSTD)
        khtd_cn = db.doc_kv("khtd_cn") or {}
        du_no  = float(df[COT_TONG_DU_NO].sum()) if COT_TONG_DU_NO in df.columns else 0.0
        nqh    = float(df[COT_DU_NO_QH].sum())   if COT_DU_NO_QH   in df.columns else 0.0
        # KH tổng CN
        ke_hoach = 0.0
        for _ct, targets in khtd_cn.items():
            if isinstance(targets, dict):
                ke_hoach += float(targets.get("_cn", 0) or 0)
        # Khoản đến hạn tháng sau
        so_dh, dn_dh = 0, 0.0
        if COT_NGAY_DH in df.columns:
            today_ts = pd.Timestamp.today().normalize()
            nm1    = today_ts.replace(day=1) + pd.offsets.MonthBegin(1)
            nm_end = nm1 + pd.offsets.MonthEnd(0)
            mask_dh = df[COT_NGAY_DH].notna() & (df[COT_NGAY_DH] >= nm1) & (df[COT_NGAY_DH] <= nm_end)
            so_dh   = int(mask_dh.sum())
            dn_dh   = float(df.loc[mask_dh, COT_TONG_DU_NO].sum()) if COT_TONG_DU_NO in df.columns else 0.0
        # Top/bottom PGD
        kh_pgd: dict[str, float] = {}
        for _ct, targets in khtd_cn.items():
            if isinstance(targets, dict):
                for pgd, val in targets.items():
                    if pgd != "_cn" and isinstance(val, (int, float)):
                        kh_pgd[pgd] = kh_pgd.get(pgd, 0) + float(val)
        ds_ranked = []
        if COT_TEN_PGD in df.columns:
            df_pgd = (
                df[df[COT_TEN_PGD] != DON_VI_CHI_NHANH]
                .groupby(COT_TEN_PGD)[COT_TONG_DU_NO].sum()
                .reset_index()
            )
            for _, r in df_pgd.iterrows():
                pgd = str(r[COT_TEN_PGD])
                kh  = kh_pgd.get(pgd, 0)
                th  = float(r[COT_TONG_DU_NO] or 0)
                pct = th / kh * 100 if kh > 0 else 0.0
                ds_ranked.append({"ten_pgd": pgd, "pct_kh": round(pct, 1)})
        ds_ranked.sort(key=lambda x: x["pct_kh"], reverse=True)
        top5 = ds_ranked[:5]
        bot5 = list(reversed(ds_ranked[-5:])) if len(ds_ranked) >= 5 else ds_ranked
        thang = date.today().month
        nam   = date.today().year
        gui_tong_ket_thang(
            thang, nam, du_no, ke_hoach, nqh,
            so_dh, dn_dh, top5, bot5,
        )
    except Exception as e:
        print(f"⚠️ _tong_ket_thang: {e}")


def _cleanup_old(n_days: int = 30):
    """Xóa báo cáo cũ hơn N ngày."""
    if not REPORT_DIR.exists():
        return
    cutoff = datetime.now() - timedelta(days=n_days)
    deleted = 0
    for f in REPORT_DIR.glob("BaoCao_Ngay_*.xlsx"):
        try:
            ts = datetime.fromtimestamp(f.stat().st_mtime)
            if ts < cutoff:
                f.unlink()
                deleted += 1
        except Exception:
            pass
    if deleted:
        print(f"🗑️ Đã xóa {deleted} báo cáo cũ (> {n_days} ngày)")


def list_reports() -> list[dict]:
    """Liệt kê tất cả báo cáo đã tạo."""
    if not REPORT_DIR.exists():
        return []
    reports = []
    for f in sorted(REPORT_DIR.glob("BaoCao_Ngay_*.xlsx"), reverse=True):
        ts = datetime.fromtimestamp(f.stat().st_mtime)
        reports.append({
            "file": f.name,
            "ts": ts.strftime("%d/%m/%Y %H:%M"),
            "size_kb": round(f.stat().st_size / 1024, 1),
            "path": str(f),
        })
    return reports


def get_latest_report() -> str | None:
    """Lấy đường dẫn báo cáo mới nhất."""
    meta_path = REPORT_DIR / "latest.json"
    if meta_path.exists():
        try:
            with open(meta_path) as f:
                info = json.load(f)
            fp = REPORT_DIR / info["file"]
            if fp.exists():
                return str(fp)
        except Exception:
            pass
    reports = list_reports()
    return reports[0]["path"] if reports else None


if __name__ == "__main__":
    print(f"VBSP-SCM Daily Report Generator — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    result = generate_daily_report()
    _cleanup_old(30)
    if not result:
        sys.exit(1)
