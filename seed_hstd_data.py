"""
seed_hstd_data.py
─────────────────
Sinh dữ liệu HSTD mẫu để chạy app trên máy mới sau khi pull từ GitHub.
Chạy 1 lần: python seed_hstd_data.py

Tạo:
  - data/HSTD_Du_lieu_tho.XLSX  (BCQUERY sheet, header dòng 4)
  - cache/hstd.parquet           (cache Parquet)
  -> App khởi động được ngay, không báo "Chưa có dữ liệu HSTD".

Có thể chạy lại để reset dữ liệu.
"""
import os
import sys
import random
from datetime import datetime, timedelta
from pathlib import Path

# ── Xử lý UTF-8 trên console Windows ─────────────────────────────────────────
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

import pandas as pd

# ── Đảm bảo import được từ project ───────────────────────────────────────────
BASE_DIR = Path(__file__).parent.resolve()
sys.path.insert(0, str(BASE_DIR))

from config import (
    DS_PGD, DON_VI_CHI_NHANH, PGD_XA_MAP,
    CACHE_HSTD, FILE_PATH, CACHE_DIR, THU_MUC_DATA,
    COT_TEN_PGD, COT_MA_KH, COT_TEN_KH, COT_SO_KU,
    COT_NGAY_VAY, COT_NGAY_DH, COT_NGAY_DH_HD,
    COT_THOI_HAN, COT_LAI_SUAT, COT_MUC_VAY,
    COT_DU_NO_TH, COT_DU_NO_QH, COT_TONG_DU_NO,
    COT_DU_NO_KHOANH, COT_TEN_CT, COT_TINH_TRANG,
    COT_DIA_CHI, COT_SDT, COT_NGAY_SL,
    COT_GOC_TRA, COT_CMND, COT_NGAY_SINH,
    COT_TEN_TO, COT_TEN_XA, COT_TEN_THON,
    COT_NGUON_VON, COT_MA_CHUONG_TRINH,
    COT_LAI_TON, COT_LAI_TON_QH, COT_LAI_THANG,
    COT_DVUT, COT_PHAN_LOAI,
    COT_NGAY_GDGN, COT_HINH_THUC_VAY,
    COT_NGAY_CAP_CMND, COT_NOI_CAP_CMND,
)


# ── Dữ liệu mẫu ──────────────────────────────────────────────────────────────

random.seed(42)

HO = [
    "Nguyễn", "Trần", "Lê", "Phạm", "Hoàng", "Huỳnh", "Phan", "Vũ",
    "Đặng", "Bùi", "Đỗ", "Hồ", "Ngô", "Dương", "Lý", "Vương",
]
DEM = [
    "Văn", "Hữu", "Đức", "Minh", "Quốc", "Thị", "Ngọc", "Thanh",
    "Công", "Xuân", "Đình", "Quang", "Trọng", "Huy", "Tấn", "Anh",
]
TEN = [
    "An", "Bình", "Cường", "Dũng", "Hà", "Hải", "Hiền", "Hoa",
    "Hùng", "Hương", "Lan", "Linh", "Long", "Mai", "Nam", "Ngọc",
    "Sơn", "Thắng", "Thảo", "Tùng", "Tuyết", "Vân", "Vinh", "Yến",
]
CHUONG_TRINH = [
    "Hộ nghèo", "Hộ cận nghèo", "Hộ mới thoát nghèo", "HSSV",
    "Nước sạch VSMT", "NS&VSTT NT", "Cho vay GQVL", "Cho vay SXKD",
    "Cho vay cư trú", "Cho vay xe ôm", "Cho vay nhà ở", "Cho vay đi lao động",
]
MA_CT_PREFIX = {ct: f"CT{100+i:03d}" for i, ct in enumerate(CHUONG_TRINH)}
DVUT_LIST = ["Hội Phụ nữ", "Hội Nông dân", "Hội CCB", "Đoàn Thanh niên", ""]
TINH_TRANG_LIST = ["Đang vay", "Đã trả", "Gia hạn", "Quá hạn"]
PHAN_LOAI_LIST = ["E", "D", "C", "B", "A"]
THON_LIST = ["Ấp 1", "Ấp 2", "Ấp 3", "Ấp 4", "Ấp 5", "Sóc Bến", "Sóc Lớn", "Xóm Mới"]
TO_LIST = ["Tổ 1", "Tổ 2", "Tổ 3", "Tổ 4", "Tổ 5"]


def sinh_ten():
    return f"{random.choice(HO)} {random.choice(DEM)} {random.choice(TEN)}"


def sinh_sdt():
    return f"09{random.randint(10000000, 99999999)}"


def sinh_cmnd():
    return f"{random.randint(100000000, 999999999)}"


def sinh_dia_chi(ten_xa):
    so_nha = random.randint(1, 999)
    duong = ["Đường 30/4", "Đường Hùng Vương", "Đường Lê Lợi", "QL 1A", "Tỉnh lộ 1",
             "Đường Nguyễn Văn Linh", "Đường Phạm Văn Thuận", "Hẻm 42"]
    return f"{so_nha} {random.choice(duong)}, {ten_xa}"


def sinh_ma_kh(pgd_idx, stt):
    return f"KH{pgd_idx:02d}{stt:04d}"


def sinh_so_ku(pgd_idx, stt):
    return f"KU{pgd_idx:02d}{stt:04d}"


def tao_du_lieu_mau(so_ho_moi_pgd: int = 30):
    """Tạo DataFrame HSTD mẫu với dữ liệu giả định."""
    tat_ca_dv = [DON_VI_CHI_NHANH] + DS_PGD
    rows = []
    ngay_sl = datetime.now().strftime("%d/%m/%Y")

    for pgd_idx, ten_pgd in enumerate(tat_ca_dv):
        ds_xa = PGD_XA_MAP.get(ten_pgd, ["Xã mẫu"])
        for stt in range(1, so_ho_moi_pgd + 1):
            ten_xa = random.choice(ds_xa)
            ten_thon = random.choice(THON_LIST)
            ten_to = random.choice(TO_LIST)
            ten_ct = random.choice(CHUONG_TRINH)
            ma_ct = MA_CT_PREFIX[ten_ct]
            ten_kh = sinh_ten()
            ma_kh = sinh_ma_kh(pgd_idx, stt)
            so_ku = sinh_so_ku(pgd_idx, stt)

            ngay_vay = datetime.now() - timedelta(days=random.randint(365, 1825))
            thoi_han = random.choice([12, 24, 36, 48, 60, 84, 120])
            ngay_dh = ngay_vay + timedelta(days=thoi_han * 30)
            ngay_dh_hd = ngay_dh - timedelta(days=random.randint(0, 90))

            muc_vay = random.choice([10, 15, 20, 30, 50, 80, 100, 200])
            tong_du_no = round(random.uniform(5, muc_vay), 2)
            du_no_th = round(tong_du_no * random.uniform(0.6, 1.0), 2)
            du_no_qh = round(max(0, tong_du_no - du_no_th), 2)
            du_no_khoanh = round(tong_du_no * 0.1, 2) if random.random() < 0.05 else 0
            goc_da_tra = round(random.uniform(0, muc_vay * 0.3), 2) if random.random() < 0.5 else 0
            lai_suat = round(random.uniform(0.65, 0.82), 2)
            lai_ton_th = round(du_no_th * lai_suat / 100 / 12, 0)
            lai_ton_qh = round(du_no_qh * lai_suat / 100 / 12 * 1.5, 0)
            lai_thang = round(tong_du_no * lai_suat / 100 / 12, 0)

            tinh_trang = random.choice(TINH_TRANG_LIST)
            phan_loai = random.choice(PHAN_LOAI_LIST)
            dvut = random.choice(DVUT_LIST)
            ngay_sinh = datetime.now() - timedelta(days=random.randint(6570, 29200))
            hinh_thuc = random.randint(1, 3)
            nguon_von = random.choice([1, 2])

            rows.append({
                COT_TEN_PGD: ten_pgd,
                COT_MA_KH: ma_kh,
                COT_TEN_KH: ten_kh,
                COT_SO_KU: so_ku,
                COT_NGAY_VAY: ngay_vay.strftime("%d/%m/%Y"),
                COT_NGAY_DH: ngay_dh.strftime("%d/%m/%Y"),
                COT_NGAY_DH_HD: ngay_dh_hd.strftime("%d/%m/%Y"),
                COT_THOI_HAN: thoi_han,
                COT_LAI_SUAT: lai_suat,
                COT_MUC_VAY: muc_vay,
                COT_DU_NO_TH: du_no_th,
                COT_DU_NO_QH: du_no_qh,
                COT_TONG_DU_NO: tong_du_no,
                COT_DU_NO_KHOANH: du_no_khoanh,
                COT_TEN_CT: ten_ct,
                COT_TINH_TRANG: tinh_trang,
                COT_DIA_CHI: sinh_dia_chi(ten_xa),
                COT_SDT: sinh_sdt(),
                COT_NGAY_SL: ngay_sl,
                COT_GOC_TRA: goc_da_tra,
                COT_CMND: sinh_cmnd(),
                COT_NGAY_SINH: ngay_sinh.strftime("%d/%m/%Y"),
                COT_NGAY_CAP_CMND: ngay_sinh.strftime("%d/%m/%Y"),
                COT_NOI_CAP_CMND: "CA Đồng Nai",
                COT_TEN_TO: ten_to,
                COT_TEN_XA: ten_xa,
                COT_TEN_THON: ten_thon,
                COT_NGUON_VON: nguon_von,
                COT_MA_CHUONG_TRINH: ma_ct,
                COT_LAI_TON: lai_ton_th,
                COT_LAI_TON_QH: lai_ton_qh,
                COT_LAI_THANG: lai_thang,
                COT_DVUT: dvut,
                COT_PHAN_LOAI: phan_loai,
                COT_NGAY_GDGN: (datetime.now() - timedelta(days=random.randint(0, 30))).strftime("%d/%m/%Y"),
                COT_HINH_THUC_VAY: hinh_thuc,
            })

    df = pd.DataFrame(rows)

    # Sắp xếp theo PGD
    df = df.sort_values([COT_TEN_PGD, COT_MA_KH]).reset_index(drop=True)
    return df


def main():
    print("=" * 60)
    print("  VBSP-SCM — Seed dữ liệu HSTD mẫu")
    print("=" * 60)

    # ── 1. Tạo thư mục ─────────────────────────────────────────────────────
    THU_MUC_DATA.mkdir(parents=True, exist_ok=True)
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    print(f"✅ Thư mục data/   → {THU_MUC_DATA}")
    print(f"✅ Thư mục cache/  → {CACHE_DIR}")

    # ── 2. Sinh dữ liệu mẫu ────────────────────────────────────────────────
    so_ho = 30
    print(f"\n⏳ Đang sinh {so_ho} hồ sơ x 22 đơn vị = {so_ho * 22} dòng...")
    df = tao_du_lieu_mau(so_ho_moi_pgd=so_ho)
    print(f"✅ Đã sinh {len(df)} dòng, {len(df.columns)} cột")

    # ── 3. Ghi Excel ───────────────────────────────────────────────────────
    excel_path = str(Path(FILE_PATH).with_suffix(".xlsx"))
    print(f"\n⏳ Đang ghi Excel → {excel_path}")
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    # Tạo sheet "BCQUERY" với 3 dòng trống phía trên (header=4 khi đọc)
    df.to_excel(excel_path, sheet_name="BCQUERY", startrow=3, index=False)

    # Ghi header dòng 1-3
    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["BCQUERY"]
    ws.cell(row=1, column=1, value="BÁO CÁO HOẠT ĐỘNG TÍN DỤNG")
    ws.cell(row=2, column=1, value=f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y')}")
    ws.cell(row=3, column=1, value="Dữ liệu mẫu - seed_hstd_data.py")
    wb.save(excel_path)
    wb.close()
    print(f"✅ Đã ghi Excel: {os.path.getsize(excel_path) / 1024:.0f} KB")

    # Cập nhật FILE_PATH nếu khác
    if excel_path != FILE_PATH:
        print(f"⚠️ Lưu ý: FILE_PATH trong config là .XLSX (hoa), file thực tế là .xlsx")
        print(f"   App sẽ tự động xử lý qua upload_service, không ảnh hưởng.")

    # ── 4. Ghi Parquet cache trực tiếp ─────────────────────────────────────
    print(f"\n⏳ Đang ghi Parquet cache → {CACHE_HSTD}")
    os.makedirs(os.path.dirname(CACHE_HSTD), exist_ok=True)
    df.to_parquet(CACHE_HSTD, index=False, engine="pyarrow", compression="zstd", compression_level=3)
    print(f"✅ Đã ghi Parquet: {os.path.getsize(CACHE_HSTD) / 1024:.0f} KB")

    # ── 5. Tóm tắt ─────────────────────────────────────────────────────────
    print(f"\n{'=' * 60}")
    print(f"  HOÀN THÀNH! Dữ liệu mẫu đã sẵn sàng.")
    print(f"  - Excel:  {excel_path}")
    print(f"  - Cache:  {CACHE_HSTD}")
    print(f"  - Config: {FILE_PATH}")
    print(f"  - Dòng:   {len(df):,}")
    print(f"  - Cột:    {len(df.columns)}")
    print(f"\n  Chạy app: streamlit run app.py")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
