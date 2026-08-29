"""Dịch vụ xuất báo cáo Excel với định dạng chuẩn VBSP."""

from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import Dict
import pandas as pd

from config import TEN_CHI_NHANH_HIEN_THI

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.worksheet import Worksheet
    # Import Image riêng biệt với try-except
    try:
        from openpyxl.drawing.image import Image as ExcelImage
    except ImportError:
        ExcelImage = None  # Sẽ skip logo nếu không import được
except ImportError:
    raise ImportError("Cần cài đặt openpyxl: pip install openpyxl")

NUMBER_FORMAT_TIEN = "#,##0"

def xuat_bao_cao(sheets: Dict[str, pd.DataFrame], tieu_de: str, nguoi_xuat: str) -> bytes:
    """
    Xuất báo cáo Excel với sheet bìa và các sheet dữ liệu đã định dạng chuẩn.
    
    Params
    ------
    sheets    : Dict[str, DataFrame] - Dictionary chứa tên sheet và DataFrame tương ứng
    tieu_de   : str - Tiêu đề báo cáo (hiển thị trong sheet Bìa)
    nguoi_xuat: str - Tên người xuất báo cáo
    
    Returns
    -------
    bytes - Nội dung file Excel để download
    """
    wb = Workbook()
    
    # Tạo sheet Bìa đầu tiên
    _tao_sheet_bia(wb, tieu_de, nguoi_xuat)
    
    # Thêm các sheet dữ liệu
    for sheet_name, df in sheets.items():
        if df is not None and not df.empty:
            _tao_sheet_du_lieu(wb, sheet_name, df)
    
    # Lưu vào BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _tao_sheet_bia(wb: Workbook, tieu_de: str, nguoi_xuat: str) -> None:
    """Tạo sheet Bìa với logo, thông tin báo cáo."""
    # Xóa sheet mặc định và tạo sheet Bìa
    if wb.worksheets:
        wb.remove(wb.active)
    ws = wb.create_sheet("Bìa", 0)
    
    # Thiết lập độ rộng cột
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    
    # Font styles
    font_title = Font(name='Times New Roman', size=18, bold=True, color='2E75B6')
    font_header = Font(name='Times New Roman', size=14, bold=True)
    font_normal = Font(name='Times New Roman', size=12)
    font_small = Font(name='Times New Roman', size=10)
    
    # Alignment
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    
    # Thêm logo nếu có
    logo_path = Path("assets/logo.png")
    if logo_path.exists() and ExcelImage is not None:
        try:
            logo = ExcelImage(str(logo_path))
            # Điều chỉnh kích thước logo
            logo.height = 80
            logo.width = 80
            ws.add_image(logo, 'B2')
        except Exception as e:  # conv: skip
            print(f"Không thể thêm logo: {e}")
    
    # Thông tin header
    row = 7 if logo_path.exists() else 2
    
    # Tên ngân hàng
    cell_bank = ws[f'B{row}']
    cell_bank.value = "NGÂN HÀNG CHÍNH SÁCH XÃ HỘI VIỆT NAM"
    cell_bank.font = font_header
    cell_bank.alignment = center
    
    # Chi nhánh
    row += 1
    cell_branch = ws[f'B{row}']
    cell_branch.value = TEN_CHI_NHANH_HIEN_THI.upper()
    cell_branch.font = font_normal
    cell_branch.alignment = center
    
    # Tiêu đề báo cáo
    row += 3
    cell_title = ws[f'B{row}']
    cell_title.value = tieu_de.upper()
    cell_title.font = font_title
    cell_title.alignment = center
    
    # Ngày xuất
    row += 3
    cell_date = ws[f'B{row}']
    ngay_xuat = datetime.now().strftime("%d/%m/%Y")
    cell_date.value = f"Ngày xuất: {ngay_xuat}"
    cell_date.font = font_normal
    cell_date.alignment = left
    
    # Người xuất
    row += 1
    cell_user = ws[f'B{row}']
    cell_user.value = f"Người xuất: {nguoi_xuat}"
    cell_user.font = font_normal
    cell_user.alignment = left
    
    # Thời gian xuất
    row += 1
    cell_time = ws[f'B{row}']
    gio_xuat = datetime.now().strftime("%H:%M:%S")
    cell_time.value = f"Thời gian: {gio_xuat}"
    cell_time.font = font_small
    cell_time.alignment = left
    
    # Ghi chú
    row += 3
    cell_note = ws[f'B{row}']
    cell_note.value = "Báo cáo được tạo tự động bởi hệ thống VBSP-SCM"
    cell_note.font = Font(name='Times New Roman', size=9, italic=True, color='666666')
    cell_note.alignment = center


def _tao_sheet_du_lieu(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Tạo sheet dữ liệu với định dạng chuẩn:
    - Freeze row 1 (header)
    - Header bold + background xanh nhạt
    - Auto column width (tối đa 50)
    - Format cột tiền bằng utils.fmt()
    """
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel giới hạn 31 ký tự

    # Tiêu đề sheet (in đậm)
    ws.cell(row=1, column=1, value=sheet_name).font = Font(
        name="Times New Roman", size=13, bold=True
    )
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(df.columns)))

    # Header tại dòng 2
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=2, column=col_idx, value=col_name)

    # Dữ liệu từ dòng 3
    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=False), start=3):
        for c_idx, v in enumerate(r, start=1):
            ws.cell(row=r_idx, column=c_idx, value=v)

    # Định dạng header (dòng 2)
    _dinh_dang_header(ws, header_row=2)

    # Freeze panes tại dòng 3 (giữ tiêu đề + header cố định)
    ws.freeze_panes = "A3"

    # Auto-filter on header row
    ws.auto_filter.ref = ws.dimensions

    # Auto-adjust column width
    _tu_dong_dieu_chinh_do_rong_cot(ws, df)
    
    # Định dạng cột số tiền
    _dinh_dang_cot_tien(ws, df)
    
    # Thêm border cho toàn bộ dữ liệu
    _them_border(ws, df)

    # Tô màu dòng NQH
    _dinh_dang_nqh(ws, df)


def _dinh_dang_header(ws: Worksheet, header_row: int = 1) -> None:
    """Định dạng header: bold + background xanh nhạt #DAEEF3."""
    header_font = Font(name='Times New Roman', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[header_row].height = 30

    # Áp dụng định dạng cho dòng header
    for cell in ws[header_row]:
        if cell.value:  # Chỉ format cell có giá trị
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment


def _tu_dong_dieu_chinh_do_rong_cot(ws: Worksheet, df: pd.DataFrame) -> None:
    """Tự động điều chỉnh độ rộng cột, tối đa 50 characters."""
    for idx, column in enumerate(df.columns, 1):
        # Lấy ký tự cột (A, B, C, ...) độc lập với merged cell ở dòng tiêu đề
        col_letter = get_column_letter(idx)
        
        # Tính độ rộng dựa trên tên cột và dữ liệu mẫu
        max_length = len(str(column))  # Độ dài tên cột
        
        # Kiểm tra độ dài dữ liệu trong 5 dòng đầu (để tránh xử lý quá chậm)
        sample_rows = min(5, len(df))
        for i in range(sample_rows):
            if i < len(df):
                cell_value = str(df.iloc[i, idx-1]) if pd.notna(df.iloc[i, idx-1]) else ""
                max_length = max(max_length, len(cell_value))
        
        # Điều chỉnh độ rộng (thêm padding và giới hạn tối đa)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = max(adjusted_width, 8)  # Tối thiểu 8


def _dinh_dang_cot_tien(ws: Worksheet, df: pd.DataFrame) -> None:
    """
    Định dạng các cột chứa số tiền bằng utils.fmt().
    Tự động nhận diện cột tiền dựa trên tên cột và kiểu dữ liệu.
    """
    # Danh sách từ khóa để nhận diện cột tiền
    tien_keywords = [
        'tiền', 'số tiền', 'thành tiền', 'dư nợ', 'du_no', 'so_tien',
        'doanh số', 'doanh_so', 'giá trị', 'gia_tri', 'tổng', 'tong',
        'phí', 'lãi', 'lai_suat', 'gốc', 'goc', 'nợ', 'no',
        'muc_vay', 'qua_han', 'trong_han', 'khoanh', 'giai_ngan',
        'lai_ton', 'von'
    ]
    
    for idx, column in enumerate(df.columns, 1):
        col_name = str(column).lower()
        
        # Kiểm tra xem có phải cột tiền không
        is_money_column = (
            any(keyword in col_name for keyword in tien_keywords) and
            pd.api.types.is_numeric_dtype(df.iloc[:, idx-1])
        )
        
        if is_money_column:
            # Định dạng các cell trong cột tiền (dữ liệu từ dòng 3)
            for row in range(3, len(df) + 3):
                cell = ws.cell(row=row, column=idx)
                if cell.value is not None:
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = NUMBER_FORMAT_TIEN
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    except (ValueError, TypeError):
                        pass


def _dinh_dang_nqh(ws: Worksheet, df: pd.DataFrame) -> None:
    """Tô màu đỏ nhạt các dòng có dư nợ quá hạn > 0."""
    # Tìm index cột NQH
    nqh_keywords = ["quá hạn", "nqh", "qh"]
    nqh_col_idx = None
    for idx, col in enumerate(df.columns, 1):
        if any(kw in str(col).lower() for kw in nqh_keywords):
            nqh_col_idx = idx
            break
    if nqh_col_idx is None:
        return
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    # Data starts at row 3 (row 1 = title, row 2 = header)
    for row_idx in range(3, len(df) + 3):
        cell_val = ws.cell(row=row_idx, column=nqh_col_idx).value
        try:
            val = float(str(cell_val).replace(",", "").replace(".", "")) if cell_val else 0
        except (ValueError, TypeError):
            val = 0
        if val > 0:
            for col_idx in range(1, len(df.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = red_fill


def _them_border(ws: Worksheet, df: pd.DataFrame) -> None:
    """Thêm border mảnh cho toàn bộ vùng dữ liệu."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Áp dụng border cho vùng bảng (header + data, bỏ dòng tiêu đề)
    for row in range(2, len(df) + 3):
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border


def xuat_bao_cao_nang_cao(
    sheets: Dict[str, pd.DataFrame],
    tieu_de: str,
    nguoi_xuat: str,
    pivot_config: Dict[str, list] | None = None,
) -> bytes:
    """Xuất báo cáo Excel nâng cao với conditional formatting + pivot summary.

    pivot_config: {sheet_name: [col_group_by, col_value]} — tạo thêm sheet pivot tổng hợp
    """
    wb = Workbook()
    _tao_sheet_bia(wb, tieu_de, nguoi_xuat)
    for sheet_name, df in sheets.items():
        if df is not None and not df.empty:
            _tao_sheet_du_lieu(wb, sheet_name, df)
    if pivot_config:
        for sheet_name, df in sheets.items():
            if df is None or df.empty:
                continue
            cfg = pivot_config.get(sheet_name)
            if cfg and len(cfg) >= 2:
                col_group, col_val = cfg[0], cfg[1]
                if col_group in df.columns and col_val in df.columns:
                    try:
                        piv = df.groupby(col_group, as_index=False)[col_val].sum()
                        _tao_sheet_du_lieu(wb, f"Pivot_{sheet_name[:20]}", piv)
                    except Exception:
                        pass
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def ten_file_bao_cao(prefix: str, ext: str = "xlsx") -> str:
    """
    Tạo tên file báo cáo với timestamp.
    
    Params
    ------
    prefix : str - Tiền tố tên file
    ext    : str - Phần mở rộng file (mặc định xlsx)
    
    Returns
    -------
    str - Tên file với format: {prefix}_DDMMYYYY.{ext}
    """
    timestamp = datetime.now().strftime("%d%m%Y")
    return f"{prefix}_{timestamp}.{ext}"


# Hàm tiện ích để xuất nhanh 1 sheet
def xuat_sheet_don(df: pd.DataFrame, tieu_de: str, nguoi_xuat: str) -> bytes:
    """
    Xuất nhanh 1 DataFrame thành Excel với sheet Bìa.
    
    Params
    ------
    df        : DataFrame cần xuất
    tieu_de   : Tiêu đề báo cáo
    nguoi_xuat: Tên người xuất
    
    Returns
    -------
    bytes - Nội dung file Excel
    """
    sheets = {"Dữ liệu": df}
    return xuat_bao_cao(sheets, tieu_de, nguoi_xuat)
