"""
Xuất Excel báo cáo Tiến độ công việc — pure openpyxl, không phụ thuộc Streamlit.
Dùng bởi: tabs/tab_tien_do.py (_render_xuat).
"""
from __future__ import annotations

from io import BytesIO

import pandas as pd


def xuat_excel_tien_do(df_tonghop: pd.DataFrame,
                       df_matran: pd.DataFrame,
                       df_ct: pd.DataFrame) -> bytes:
    """
    Tạo file Excel 3 sheet: Tổng hợp / Ma trận PGD / Chi tiết xã.
    Trả về bytes (dùng trực tiếp cho st.download_button).
    """
    import openpyxl  # noqa: F401 — import kiểm tra tại runtime
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_tonghop.to_excel(writer, sheet_name="Tổng hợp", index=False)
        df_matran.to_excel(writer, sheet_name="Ma trận PGD", index=False)
        df_ct.to_excel(writer, sheet_name="Chi tiết xã", index=False)

        header_fill = PatternFill("solid", fgColor="003D7A")
        header_font = Font(bold=True, color="FFFFFF", size=10, name="Times New Roman")
        body_font   = Font(size=10, name="Times New Roman")
        body_bold   = Font(bold=True, size=10, name="Times New Roman")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        right  = Alignment(horizontal="right",  vertical="center", wrap_text=True)
        thin   = Side(style="thin", color="BDBDBD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill = PatternFill("solid", fgColor="EEF4FB")

        def _style_sheet(ws, col_left, col_right, col_center, col_pct):
            col_names = [cell.value for cell in ws[1]]

            for cell in ws[1]:
                cell.fill   = header_fill
                cell.font   = header_font
                cell.alignment = center
                cell.border = border

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cn = col_names[cell.column - 1]
                    cell.border = border
                    cell.font   = body_font
                    if cn in col_left:
                        cell.alignment = left
                    elif cn in col_right:
                        cell.alignment = right
                        cell.number_format = "#,##0"
                    elif cn in col_center:
                        cell.alignment = center
                    elif cn in col_pct:
                        cell.alignment = center
                        cell.number_format = "0.0\"%\""

            for i, row in enumerate(
                ws.iter_rows(min_row=2, max_row=ws.max_row), start=1
            ):
                if i % 2 == 0:
                    for cell in row:
                        if not cell.fill or cell.fill.fgColor.rgb in (
                            "00000000", "003D7A"
                        ):
                            cell.fill = alt_fill

            for col_idx in range(1, len(col_names) + 1):
                max_len = max(
                    (
                        len(str(cell.value or ""))
                        for cell in ws[ws.min_row : ws.max_row][:, col_idx - 1]
                    ),
                    default=15,
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(
                    max_len + 4, 45
                )

        # Sheet 1: Tổng hợp
        ws_th = writer.sheets["Tổng hợp"]
        _style_sheet(
            ws_th,
            col_left={
                "Đầu việc", "Loại", "Người phụ trách",
                "CB KH-NV phụ trách", "Hội sở CN tỉnh",
            },
            col_right={
                "Số PGD", "Tổng xã", "Đã hoàn thành",
                "Chưa thực hiện", "Trễ hạn", "N/A",
            },
            col_center={
                "STT", "Ưu tiên", "Thời hạn",
                "Loại theo dõi", "Ngày bắt đầu",
            },
            col_pct={"Tỷ lệ HT%"},
        )

        # Sheet 2: Ma trận PGD
        ws_mt = writer.sheets["Ma trận PGD"]
        col_names_mt = [cell.value for cell in ws_mt[1]]
        for cell in ws_mt[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center
            cell.border    = border
        for row in ws_mt.iter_rows(min_row=2, max_row=ws_mt.max_row):
            for cell in row:
                cn = col_names_mt[cell.column - 1]
                cell.border = border
                cell.font   = body_font
                if cn == "Đơn vị":
                    cell.alignment = left
                    cell.font      = body_bold
                else:
                    cell.alignment = center
        for i, row in enumerate(
            ws_mt.iter_rows(min_row=2, max_row=ws_mt.max_row), start=1
        ):
            if i % 2 == 0:
                for cell in row:
                    if not cell.fill or cell.fill.fgColor.rgb in (
                        "00000000", "003D7A"
                    ):
                        cell.fill = alt_fill
        for col_idx in range(1, len(col_names_mt) + 1):
            max_len = max(
                (
                    len(str(cell.value or ""))
                    for cell in ws_mt[ws_mt.min_row : ws_mt.max_row][:, col_idx - 1]
                ),
                default=12,
            )
            ws_mt.column_dimensions[get_column_letter(col_idx)].width = min(
                max_len + 3, 28
            )

        # Sheet 3: Chi tiết xã
        ws_ct = writer.sheets["Chi tiết xã"]
        _style_sheet(
            ws_ct,
            col_left={"Đầu việc", "PGD", "Xã / Phường", "Ghi chú"},
            col_right=set(),
            col_center={"Task ID", "Thời hạn", "Trạng thái", "Ngày hoàn thành"},
            col_pct=set(),
        )

    buf.seek(0)
    return buf.getvalue()
