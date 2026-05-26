"""Service: các hàm thuần túy cho tab Tổng quan (không có st.* calls)."""
from __future__ import annotations

from logger import get_logger
logger = get_logger(__name__)

import logging
from io import BytesIO

import pandas as pd


def tinh_kpi_tongquan(
    df: pd.DataFrame,
    cot_tdn: str,
    cot_dth: str,
    cot_dqh: str,
    cot_nk: str,
    cot_ku: str,
    cot_ma_kh: str,
) -> dict:
    df_loc = df
    for c in [cot_tdn, cot_dth, cot_dqh, cot_nk]:
        if c and c in df_loc.columns:
            df_loc = df_loc.copy()
            break
    for c in [cot_tdn, cot_dth, cot_dqh, cot_nk]:
        if c and c in df_loc.columns:
            df_loc[c] = pd.to_numeric(df_loc[c], errors="coerce").fillna(0)
    tdn = df_loc[cot_tdn].sum() if cot_tdn in df_loc.columns else 0
    dth = df_loc[cot_dth].sum() if cot_dth in df_loc.columns else 0
    dqh = df_loc[cot_dqh].sum() if cot_dqh in df_loc.columns else 0
    dnk = df_loc[cot_nk].sum() if cot_nk in df_loc.columns else 0
    n_mon_vay = df_loc[cot_ku].nunique() if cot_ku in df_loc.columns else len(df_loc)
    n_kh = df_loc[cot_ma_kh].nunique() if cot_ma_kh in df_loc.columns else 0
    try:
        from data import danh_dau_khong_hd

        df_kh = danh_dau_khong_hd(df_loc)
        n_3m = int(df_kh["is_3m_inactive"].sum()) if "is_3m_inactive" in df_kh.columns else 0
        dn_3m = (
            df_kh.loc[df_kh["is_3m_inactive"], cot_tdn].sum()
            if ("is_3m_inactive" in df_kh.columns and cot_tdn in df_kh.columns)
            else 0
        )
    except Exception as e:
        logger.error("Lỗi trong khối except: %s", e, exc_info=True)
        n_3m = 0
        dn_3m = 0
        logging.warning("[tongquan_kpi] danh_dau_khong_hd lỗi: %s", e)
    return dict(
        tdn=tdn,
        dth=dth,
        dqh=dqh,
        dnk=dnk,
        n_mon_vay=n_mon_vay,
        n_kh=n_kh,
        n_3m=n_3m,
        dn_3m=dn_3m,
    )


def tinh_heatmap_pgd(
    df: pd.DataFrame,
    cot_pgd: str,
    cot_tdn: str,
    cot_ma_kh: str,
    cot_dqh: str,
) -> pd.DataFrame:
    df_loc = df
    for c in [cot_tdn, cot_dqh]:
        if c and c in df_loc.columns:
            df_loc = df_loc.copy()
            break
    for c in [cot_tdn, cot_dqh]:
        if c and c in df_loc.columns:
            df_loc[c] = pd.to_numeric(df_loc[c], errors="coerce").fillna(0)
    return df_loc.groupby(cot_pgd, as_index=False).agg(
        du_no=(cot_tdn, "sum"),
        so_kh=(cot_ma_kh, "nunique"),
        nqh=(cot_dqh, "sum"),
    )


def tinh_co_cau_ct(
    df: pd.DataFrame,
    col_khoanh: str,
    col_gn: str,
    cols_tn_key: str,
    cot_ten_ct: str,
    cot_tdn: str,
    cot_dqh: str,
    cot_dnk: str,
    cot_nv: str,
    cot_ma_kh: str,
) -> pd.DataFrame:
    df_loc = df[df[cot_tdn].fillna(0) > 0].copy()
    cols_to_sum = [cot_tdn, cot_dqh, cot_dnk, cot_nv]
    if col_khoanh:
        cols_to_sum.append(col_khoanh)
    if col_gn:
        cols_to_sum.append(col_gn)
    cols_to_sum = list(set(cols_to_sum))
    for c in cols_to_sum:
        if c in df_loc.columns:
            df_loc[c] = pd.to_numeric(df_loc[c].astype(object), errors="coerce").fillna(0)

    if cot_nv in df_loc.columns:
        nv = pd.to_numeric(df_loc[cot_nv], errors="coerce")
    else:
        nv = pd.Series(0, index=df_loc.index, dtype="float64")
    du_no_tw = df_loc[nv == 1].groupby(cot_ten_ct)[cot_tdn].sum()
    du_no_dp = df_loc[nv == 2].groupby(cot_ten_ct)[cot_tdn].sum()

    so_kh_by_ct = df.groupby(cot_ten_ct)[cot_ma_kh].nunique()
    so_mon_by_ct = df.groupby(cot_ten_ct)[cot_ma_kh].count()

    df_ct = (
        df_loc.groupby(cot_ten_ct)
        .agg(du_no=(cot_tdn, "sum"), _tmp=(cot_tdn, "count"))
        .sort_values("du_no", ascending=False)
        .reset_index()
    )
    df_ct.columns = ["ten_ct", "du_no", "_tmp"]
    df_ct["so_kh"] = df_ct["ten_ct"].map(so_kh_by_ct).fillna(0).astype(int)
    df_ct["so_mon"] = df_ct["ten_ct"].map(so_mon_by_ct).fillna(0).astype(int)
    df_ct.drop(columns=["_tmp"], inplace=True)

    tong = df_ct["du_no"].sum()
    df_ct["ty_trong"] = (df_ct["du_no"] / tong * 100).round(1) if tong > 0 else 0.0
    df_ct["du_no_tw"] = df_ct["ten_ct"].map(du_no_tw).fillna(0)
    df_ct["du_no_dp"] = df_ct["ten_ct"].map(du_no_dp).fillna(0)

    if cot_dqh in df.columns:
        qh = (
            df_loc.groupby(cot_ten_ct)[cot_dqh]
            .apply(lambda x: pd.to_numeric(x, errors="coerce").sum())
            .reset_index()
        )
        qh = qh.rename(columns={qh.columns[0]: "ten_ct", qh.columns[1]: "du_no_qh"})
        df_ct = df_ct.merge(qh, on="ten_ct", how="left")
    else:
        df_ct["du_no_qh"] = 0.0
    df_ct["du_no_qh"] = df_ct["du_no_qh"].fillna(0)

    if col_khoanh and col_khoanh in df.columns:
        nk = (
            df_loc.groupby(cot_ten_ct)[col_khoanh]
            .apply(lambda x: pd.to_numeric(x, errors="coerce").sum())
            .reset_index()
        )
        nk = nk.rename(columns={nk.columns[0]: "ten_ct", nk.columns[1]: "du_no_khoanh"})
        df_ct = df_ct.merge(nk, on="ten_ct", how="left")
    else:
        df_ct["du_no_khoanh"] = 0.0
    df_ct["du_no_khoanh"] = df_ct["du_no_khoanh"].fillna(0)

    if col_gn and col_gn in df.columns:
        gn = (
            df_loc.groupby(cot_ten_ct)[col_gn]
            .apply(lambda x: pd.to_numeric(x, errors="coerce").sum())
            .reset_index()
        )
        gn = gn.rename(columns={gn.columns[0]: "ten_ct", gn.columns[1]: "gn_nam"})
        df_ct = df_ct.merge(gn, on="ten_ct", how="left")
    else:
        df_ct["gn_nam"] = 0.0
    df_ct["gn_nam"] = df_ct["gn_nam"].fillna(0).replace([float("inf"), float("-inf")], 0)

    cols_tn = [c for c in cols_tn_key.split(",") if c and c in df.columns]
    if cols_tn:
        tn = (
            df_loc.groupby(cot_ten_ct)[cols_tn]
            .apply(lambda x: pd.to_numeric(x.stack(), errors="coerce").sum())
            .reset_index()
        )
        tn = tn.rename(columns={tn.columns[0]: "ten_ct", tn.columns[1]: "tn_nam"})
        df_ct = df_ct.merge(tn, on="ten_ct", how="left")
    else:
        df_ct["tn_nam"] = 0.0
    df_ct["tn_nam"] = df_ct["tn_nam"].fillna(0).replace([float("inf"), float("-inf")], 0)
    return df_ct


def tinh_tqpgd_extended(
    df: pd.DataFrame,
    col_khoanh: str,
    col_cv: str,
    cols_thu_key: str,
    nam_ht: str,
    cot_pgd: str,
    cot_tdn: str,
    cot_dqh: str,
    cot_lai_ton: str,
    cot_ngay_dh: str,
    cot_ma_kh: str,
    cot_so_ku: str,
) -> pd.DataFrame:
    cols_to_sum = [cot_tdn, cot_dqh, cot_lai_ton]
    if col_khoanh:
        cols_to_sum.append(col_khoanh)
    if col_cv:
        cols_to_sum.append(col_cv)

    df_loc = df
    for c in set(cols_to_sum):
        if c in df_loc.columns:
            df_loc = df_loc.copy()
            break
    for c in set(cols_to_sum):
        if c in df_loc.columns:
            df_loc[c] = pd.to_numeric(df_loc[c], errors="coerce").fillna(0)

    df_pgd = df_loc.groupby(cot_pgd, as_index=False).agg(
        du_no=(cot_tdn, "sum"),
        so_kh=(cot_ma_kh, "nunique"),
        so_mon=(cot_so_ku, "nunique"),
        nqh=(cot_dqh, "sum"),
    )

    # Gộp khoanh + lãi tồn + nợ ĐH + DS cho vay thành 1 lần tính toán
    agg_extra = {}

    if col_khoanh and col_khoanh in df_loc.columns:
        agg_extra[col_khoanh] = (col_khoanh, "sum")
    else:
        agg_extra["khoanh_tmp"] = (cot_ma_kh, "size")
        col_khoanh = "khoanh_tmp"

    if cot_lai_ton in df_loc.columns:
        agg_extra[cot_lai_ton] = (cot_lai_ton, "sum")

    if col_cv and col_cv in df_loc.columns:
        agg_extra[col_cv] = (col_cv, "sum")

    # Tính 1 lần cho các cột trên
    if agg_extra:
        extra_data = df_loc.groupby(cot_pgd, as_index=False).agg(**agg_extra)
        df_pgd = df_pgd.merge(extra_data, on=cot_pgd, how="left")

    df_pgd = df_pgd.rename(columns={col_khoanh: "du_no_khoanh"})
    df_pgd["du_no_khoanh"] = pd.to_numeric(df_pgd.get("du_no_khoanh", 0), errors="coerce").fillna(0)

    if cot_lai_ton in df_loc.columns:
        df_pgd["lai_ton"] = pd.to_numeric(df_pgd.get(cot_lai_ton, 0), errors="coerce").fillna(0)
        df_pgd = df_pgd.drop(columns=[cot_lai_ton])
    else:
        df_pgd["lai_ton"] = 0.0

    if col_cv and col_cv in df_loc.columns:
        df_pgd["ds_cho_vay"] = pd.to_numeric(df_pgd.get(col_cv, 0), errors="coerce").fillna(0)
        df_pgd = df_pgd.drop(columns=[col_cv])
    else:
        df_pgd["ds_cho_vay"] = 0.0

    # Nợ ĐH năm — tính riêng nếu có ngày đh
    if cot_ngay_dh in df_loc.columns:
        ngay_dh = df_loc[cot_ngay_dh]
        if not pd.api.types.is_datetime64_any_dtype(ngay_dh):
            ngay_dh = pd.to_datetime(ngay_dh, dayfirst=True, errors="coerce")
        tdn_num = pd.to_numeric(df_loc[cot_tdn], errors="coerce").fillna(0)
        mask = ngay_dh.dt.year == int(nam_ht)
        dh = (
            pd.DataFrame({cot_pgd: df_loc.loc[mask, cot_pgd].values, "no_dh_nam": tdn_num[mask].values})
            .groupby(cot_pgd, as_index=False)["no_dh_nam"]
            .sum()
        )
        df_pgd = df_pgd.merge(dh, on=cot_pgd, how="left")
        df_pgd["no_dh_nam"] = df_pgd["no_dh_nam"].fillna(0)
    else:
        df_pgd["no_dh_nam"] = 0.0

    # DS Thu nợ — vectorized thay vì apply(lambda stack)
    cols_thu = [c for c in cols_thu_key.split(",") if c and c in df_loc.columns]
    if cols_thu:
        thu_num = df_loc[[cot_pgd] + cols_thu].copy()
        for c in cols_thu:
            thu_num[c] = pd.to_numeric(thu_num[c], errors="coerce").fillna(0)
        thu_num["ds_thu_no"] = thu_num[cols_thu].sum(axis=1)
        thu = thu_num.groupby(cot_pgd, as_index=False)["ds_thu_no"].sum()
        df_pgd = df_pgd.merge(thu, on=cot_pgd, how="left")
        df_pgd["ds_thu_no"] = df_pgd["ds_thu_no"].fillna(0)
    else:
        df_pgd["ds_thu_no"] = 0.0

    return df_pgd


def chuan_hoa_ngay(
    df: pd.DataFrame,
    cot_ngay: str,
    *,
    dayfirst: bool = True,
) -> pd.DataFrame:
    if cot_ngay not in df.columns:
        return df
    df_loc = df.copy()
    df_loc[cot_ngay] = pd.to_datetime(df_loc[cot_ngay], dayfirst=dayfirst, errors="coerce")
    return df_loc


def ap_dung_loc_ket_hop(
    df: pd.DataFrame,
    *,
    cot_pgd: str,
    cot_ct: str,
    cot_xa: str | None,
    loc_pgd: list | None,
    loc_ct: list | None,
    loc_xa: list | None,
) -> pd.DataFrame:
    df_loc = df
    if loc_pgd and cot_pgd in df_loc.columns:
        df_loc = df_loc[df_loc[cot_pgd].isin(loc_pgd)]
    if loc_ct and cot_ct in df_loc.columns:
        df_loc = df_loc[df_loc[cot_ct].isin(loc_ct)]
    if loc_xa and cot_xa and cot_xa in df_loc.columns:
        df_loc = df_loc[df_loc[cot_xa].isin(loc_xa)]
    return df_loc


def ap_dung_loc_den_han_tab(
    df: pd.DataFrame,
    *,
    cot_xa: str | None,
    cot_nv: str | None,
    loc_xa: list | None,
    loc_nv: int | list[int] | None,
    cot_tdn: str,
    range_no_trieu: tuple | None,
) -> pd.DataFrame:
    df_loc = df
    if loc_xa and cot_xa and cot_xa in df_loc.columns:
        df_loc = df_loc[df_loc[cot_xa].isin(loc_xa)]
    if loc_nv is not None and cot_nv and cot_nv in df_loc.columns:
        nv_num = pd.to_numeric(df_loc[cot_nv], errors="coerce").fillna(0).astype(int)
        if isinstance(loc_nv, list):
            df_loc = df_loc[nv_num.isin(loc_nv)]
        else:
            df_loc = df_loc[nv_num == loc_nv]
    if range_no_trieu and cot_tdn in df_loc.columns:
        lo = range_no_trieu[0] * 1_000_000
        hi = range_no_trieu[1] * 1_000_000
        df_loc = df_loc[(df_loc[cot_tdn] >= lo) & (df_loc[cot_tdn] <= hi)]
    return df_loc


def loc_du_no_duong(df: pd.DataFrame, cot_tdn: str) -> pd.DataFrame:
    if cot_tdn not in df.columns:
        return df
    return df[df[cot_tdn].fillna(0) > 0]


def loc_den_han(
    df: pd.DataFrame,
    *,
    cot_ngay_dh: str,
    tu_ngay,
    den_ngay,
) -> pd.DataFrame:
    if cot_ngay_dh not in df.columns:
        return df.iloc[0:0].copy()
    return df[(df[cot_ngay_dh] >= tu_ngay) & (df[cot_ngay_dh] <= den_ngay)]


def tong_chi_tieu_den_han(
    df: pd.DataFrame,
    *,
    cot_tdn: str,
    cot_so_ku: str,
    cot_ma_kh: str,
) -> dict:
    tong_no = df[cot_tdn].fillna(0).sum() if cot_tdn in df.columns else 0
    tong_mon = df[cot_so_ku].nunique() if cot_so_ku in df.columns else 0
    tong_kh = df[cot_ma_kh].nunique() if cot_ma_kh in df.columns else 0
    return {"tong_no": tong_no, "tong_mon": tong_mon, "tong_kh": tong_kh}


def tong_hop_den_han(
    df: pd.DataFrame,
    *,
    group_cols: list[str],
    cot_so_ku: str,
    cot_ma_kh: str,
    cot_tdn: str,
) -> pd.DataFrame:
    cols_ok = [c for c in group_cols if c in df.columns]
    if not cols_ok:
        return pd.DataFrame(columns=["_mon", "_kh", "_no"])
    if cot_so_ku not in df.columns or cot_ma_kh not in df.columns or cot_tdn not in df.columns:
        return pd.DataFrame(columns=[*cols_ok, "_mon", "_kh", "_no"])
    tg = (
        df.groupby(cols_ok)
        .agg(
            _mon=(cot_so_ku, "nunique"),
            _kh=(cot_ma_kh, "nunique"),
            _no=(cot_tdn, "sum"),
        )
        .reset_index()
    )
    return tg


def tong_hop_den_han_theo_thang(
    df: pd.DataFrame,
    *,
    cot_ngay_dh: str,
    cot_so_ku: str,
    cot_ma_kh: str,
    cot_tdn: str,
) -> pd.DataFrame:
    """Tổng hợp dư nợ đến hạn theo từng tháng — dùng cho biểu đồ timeline."""
    _empty = pd.DataFrame(columns=["nam_thang_label", "_mon", "_kh", "_no"])
    if cot_ngay_dh not in df.columns:
        return _empty
    df2 = df.copy()
    df2["_ym"] = pd.to_datetime(df2[cot_ngay_dh], errors="coerce").dt.to_period("M")
    df2 = df2.dropna(subset=["_ym"])
    if df2.empty or cot_so_ku not in df2.columns or cot_tdn not in df2.columns:
        return _empty
    tg = (
        df2.groupby("_ym")
        .agg(_mon=(cot_so_ku, "nunique"), _kh=(cot_ma_kh, "nunique"), _no=(cot_tdn, "sum"))
        .reset_index()
    )
    tg["nam_thang_label"] = tg["_ym"].apply(lambda p: f"{p.month:02d}/{p.year}")
    return tg.sort_values("_ym").drop(columns=["_ym"]).reset_index(drop=True)


def xuat_excel_tqpgd(df: pd.DataFrame, ten_file: str) -> bytes:
    """Xuất df_show TQPGD ra Excel với định dạng đẹp."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _ = ten_file
    df_xuat = df.copy()
    cols_pct = [c for c in df_xuat.columns if "%" in str(c)]
    for c in cols_pct:
        df_xuat[c] = pd.to_numeric(df_xuat[c], errors="coerce") / 100.0

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_xuat.to_excel(writer, sheet_name="Tổng quan PGD", index=False)
        ws = writer.sheets["Tổng quan PGD"]

        header_fill = PatternFill("solid", fgColor="003D7A")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center")
        right = Alignment(horizontal="right", vertical="center")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        COT_SO = [
            "Số KH",
            "Dư nợ (triệu đồng)",
            "QH (triệu đồng)",
            "TL QH %",
            "Khoanh (triệu đồng)",
            "TL Khoanh %",
            "Nợ xấu (triệu đồng)",
            "TL NPL %",
            "Lãi tồn (triệu đồng)",
            "Nợ ĐH năm (triệu đồng)",
            "DS Cho vay (triệu đồng)",
            "DS Thu nợ (triệu đồng)",
            "Tổng Tổ",
            "Tốt",
            "Khá",
            "TB",
            "Yếu",
        ]
        col_names = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                col_name = col_names[cell.column - 1]
                cell.border = border
                if col_name == df_xuat.columns[0]:
                    cell.alignment = left
                    cell.font = Font(bold=True, size=10)
                elif col_name in COT_SO:
                    cell.alignment = right
                    cell.font = Font(size=10)
                    if "%" in str(col_name):
                        cell.number_format = "0.00%"
                    elif col_name == "Số KH" or col_name in ["Tổng Tổ", "Tốt", "Khá", "TB", "Yếu"]:
                        cell.number_format = "#,##0"
                    else:
                        cell.number_format = "#,##0.000"
                else:
                    cell.alignment = center
                    cell.font = Font(size=10)

        alt_fill = PatternFill("solid", fgColor="EEF4FB")
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
            if i % 2 == 0:
                for cell in row:
                    if not cell.fill or cell.fill.fgColor.rgb == "00000000":
                        cell.fill = alt_fill

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

        ws.freeze_panes = "B2"

    return buf.getvalue()


def tinh_card_pgd(
    df: pd.DataFrame,
    *,
    cot_pgd: str,
    cot_tdn: str,
    cot_dqh: str,
    cot_ma_kh: str,
    cot_so_ku: str,
    cot_ngay_dh: str,
    ds_don_vi: list[str],
) -> pd.DataFrame:
    """Tính chỉ số tóm tắt cho từng PGD (dùng cho tab card 22 PGD).

    Trả về DataFrame gồm:
      ten_pgd, du_no, nqh, ty_le_nqh, so_kh, so_mon, no_den_han_thang, dn_binh_quan_ho
    Các PGD không có dữ liệu vẫn xuất hiện với giá trị 0.
    """
    import datetime as _dt

    for c in [cot_tdn, cot_dqh]:
        if c in df.columns:
            df = df.copy()
            break
    for c in [cot_tdn, cot_dqh]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    df_pgd = (
        df.groupby(cot_pgd, as_index=False)
        .agg(
            du_no=(cot_tdn, "sum"),
            nqh=(cot_dqh, "sum"),
            so_kh=(cot_ma_kh, "nunique"),
            so_mon=(cot_so_ku, "nunique"),
        )
    )

    # Đến hạn trong tháng này
    today = _dt.date.today()
    first_of_month = today.replace(day=1)
    last_of_month = (today.replace(day=28) + _dt.timedelta(days=4)).replace(day=1) - _dt.timedelta(days=1)

    if cot_ngay_dh in df.columns:
        ngay_dh = pd.to_datetime(df[cot_ngay_dh], dayfirst=True, errors="coerce")
        mask_thang = (ngay_dh.dt.date >= first_of_month) & (ngay_dh.dt.date <= last_of_month)
        df_dh = (
            df.loc[mask_thang].groupby(cot_pgd, as_index=False)[cot_tdn]
            .sum()
            .rename(columns={cot_tdn: "no_den_han_thang"})
        )
        df_pgd = df_pgd.merge(df_dh, on=cot_pgd, how="left")
    else:
        df_pgd["no_den_han_thang"] = 0.0

    df_pgd["no_den_han_thang"] = pd.to_numeric(
        df_pgd["no_den_han_thang"] if "no_den_han_thang" in df_pgd.columns else 0,
        errors="coerce",
    ).fillna(0)
    df_pgd["ty_le_nqh"] = (df_pgd["nqh"] / df_pgd["du_no"] * 100).where(df_pgd["du_no"] > 0, 0).round(2)

    # Đảm bảo tất cả đơn vị trong ds_don_vi đều có row (điền 0 nếu không có dữ liệu)
    df_base = pd.DataFrame({"ten_pgd": ds_don_vi})
    df_pgd = df_pgd.rename(columns={cot_pgd: "ten_pgd"})
    df_pgd = df_base.merge(df_pgd, on="ten_pgd", how="left").fillna(0)
    for c in ["so_kh", "so_mon"]:
        df_pgd[c] = df_pgd[c].astype(int)

    # Dư nợ bình quân hộ (đơn vị: đồng)
    df_pgd["dn_binh_quan_ho"] = (
        df_pgd["du_no"] / df_pgd["so_kh"].replace(0, pd.NA)
    ).fillna(0).round(0)

    return df_pgd.reset_index(drop=True)
