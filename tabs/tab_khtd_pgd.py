"""Tab Kế hoạch Tín dụng cấp PGD — Hỗ trợ địa bàn.

Đọc dữ liệu từ khtd_xa ({ten_xa|ma_ct: gia_tri_dong}), hiển thị bảng
tổng hợp hàng=xã, cột=chương trình cho PGD được chọn, có dòng tổng và xuất Excel.
"""


from __future__ import annotations

from logger import get_logger
logger = get_logger(__name__)

import re
from io import BytesIO
from pathlib import Path
from datetime import datetime
from typing import TYPE_CHECKING, Any

import streamlit as st
import pandas as pd
from openpyxl.styles import Font, PatternFill

import db
from auth import la_phan_he_cn, la_executive, normalize_role
from utils import hien_thi_dataframe_phan_trang
from services.khtd_nhap_service import (
    format_kich_thuoc as _svc_format_kich_thuoc,
    doc_meta_qd as _svc_doc_meta_qd,
    luu_meta_qd as _svc_luu_meta_qd,
    luu_file_qd as _svc_luu_file_qd,
)

from config import (

    CHUONG_TRINH_KHTD, TEN_CHINH_THUC_CT,
    COT_TEN_PGD, COT_TEN_CT, COT_TEN_XA, COT_MA_CHUONG_TRINH, COT_NGUON_VON, COT_TONG_DU_NO, COT_DU_NO_TH,
    COT_MA_NHA_DAU_TU,
    DS_PGD, PGD_XA_MAP,
)

if TYPE_CHECKING:
    from streamlit.delta_generator import DeltaGenerator


# ── Hằng số ──────────────────────────────────────────────────────────────────
KV_KEY_XA   = "khtd_xa"
DS_MA_CT    = [row[0] for row in CHUONG_TRINH_KHTD]
NGUON_VON_MA = {mk: nv for mk, _, _, nv, _ in CHUONG_TRINH_KHTD}
MA_CT_BY_MAKEY = {mk: int(ma_ct) for mk, ma_ct, _, _, _ in CHUONG_TRINH_KHTD}
TEN_BASE_BY_MACT: dict[int, str] = {}
for mk, ma_ct, ten, _nv, _ in CHUONG_TRINH_KHTD:
    TEN_BASE_BY_MACT.setdefault(int(ma_ct), str(ten))

# Thư mục lưu văn bản QĐ cấp PGD
DATA_DIR     = Path(__file__).parent.parent / "data"
PGD_DATA_DIR = DATA_DIR / "pgd_data"

_SHORT_CT_PGD: dict[str, str] = {
    "1_TW": "Hộ nghèo",       "2_TW": "HSSV",           "3_TW_NHCSXH": "GQVL HĐ",
    "3_TW_NSNN": "GQVL NS",   "4_TW": "XKLĐ",            "6_TW": "NS sạch",
    "7_TW": "Nhà ở HN",       "9_TW": "Mới TN",           "10_TW": "SXKD KK",
    "12_TW": "Nhà ở XH",      "15_TW": "TN KK",           "17_TW": "DTTS 755",
    "19_TW": "Cận nghèo",     "21_TW": "DTTS 2085",       "25_TW": "Vùng DTTS",
    "26_TW": "Chấp hành",     "99_TW": "Khác TW",
    "1_DP": "HN (ĐP)",        "2_DP": "SV (ĐP)",          "3_DP_TINH": "GQVL tỉnh",
    "3_DP_XA": "GQVL xã",    "6_DP": "NS ĐP",             "9_DP": "Mới TN ĐP",
    "10_DP": "SXKD KK ĐP",   "12_DP": "NOXH ĐP",         "15_DP": "TN KK ĐP",
    "17_DP": "DTTS 755 ĐP",  "19_DP": "Cận nghèo ĐP",    "21_DP": "DTTS 2085 ĐP",
    "25_DP": "Vùng DTTS ĐP", "26_DP": "Chấp hành ĐP",    "99_DP": "Khác ĐP",
}

_LOOKUP_XA_CT: dict[tuple[int, int], str] = {}
for _mk, _ma_ct, _, _nv, _ in CHUONG_TRINH_KHTD:
    _LOOKUP_XA_CT.setdefault((int(_ma_ct), 1 if _nv == "TW" else 2), _mk)


def _ten_ngan_pgd(ma_key: str) -> str:
    return _SHORT_CT_PGD.get(ma_key, ma_key)


def _tinh_th_xa_ct(df: "pd.DataFrame") -> dict[tuple, float]:
    """Returns {(ten_xa, ma_key): vnd_sum} từ HSTD DataFrame."""
    if df is None or df.empty:
        return {}
    required = (COT_TEN_XA, COT_MA_CHUONG_TRINH, COT_NGUON_VON)
    if not all(c in df.columns for c in required):
        return {}
    col_th = COT_TONG_DU_NO if COT_TONG_DU_NO in df.columns else (
        COT_DU_NO_TH if COT_DU_NO_TH in df.columns else None
    )
    if not col_th:
        return {}
    tmp = df[[COT_TEN_XA, COT_MA_CHUONG_TRINH, COT_NGUON_VON, col_th]].copy()
    tmp[COT_MA_CHUONG_TRINH] = pd.to_numeric(tmp[COT_MA_CHUONG_TRINH], errors="coerce").fillna(0).astype(int)
    tmp[COT_NGUON_VON] = pd.to_numeric(tmp[COT_NGUON_VON], errors="coerce").fillna(0).astype(int)
    tmp[col_th] = pd.to_numeric(tmp[col_th], errors="coerce").fillna(0)
    tmp = tmp[(tmp[COT_MA_CHUONG_TRINH] > 0) & (tmp[COT_NGUON_VON].isin([1, 2]))]
    out: dict[tuple, float] = {}
    for (xa, ma_ct, nv), s in tmp.groupby([COT_TEN_XA, COT_MA_CHUONG_TRINH, COT_NGUON_VON])[col_th].sum().items():
        mk = _LOOKUP_XA_CT.get((int(ma_ct), int(nv)))
        if mk:
            out[(str(xa), mk)] = float(s)
    return out


def _html_bdd_pgd_table(pgd: str, kh_xa: dict, th_xa_ct: dict, nguon: str) -> str:
    """HTML pivot table dạng BĐD cho PGD: hàng = xã, cột = CT × (KH | TH | %)."""
    ds_xa = PGD_XA_MAP.get(pgd, [])
    if not ds_xa:
        return f"<p style='color:#94A3B8;padding:12px'>⚠️ Chưa có danh sách xã cho {pgd}.</p>"

    active_mk: list[str] = []
    mk_set: set[str] = set()
    for mk, _, _, nv, _ in CHUONG_TRINH_KHTD:
        if nv != nguon or mk in mk_set:
            continue
        has_kh = any(float(kh_xa.get(f"{xa}|{mk}", 0)) > 0 for xa in ds_xa)
        has_th = any(th_xa_ct.get((xa, mk), 0) > 0 for xa in ds_xa)
        if has_kh or has_th:
            active_mk.append(mk)
            mk_set.add(mk)

    if not active_mk:
        return f"<p style='color:#94A3B8;padding:12px'>Chưa có dữ liệu nguồn {nguon}.</p>"

    def fn(v: float) -> str:
        return "" if v == 0 else f"{v:,.1f}".replace(",", ".")

    def fp(v) -> str:
        return "" if v is None else f"{v:.1f}".replace(".", ",") + "%"

    td = "border:1px solid #e2e8f0;padding:3px 7px"
    th_hdr = "border:1px solid #1e40af;padding:5px 8px;background:rgba(30,64,175,0.85);color:#fff;text-align:center;font-size:11px;white-space:nowrap"
    th_sub = "border:1px solid #2563eb;padding:3px 6px;background:rgba(37,99,235,0.75);color:#fff;text-align:center;font-size:10px;white-space:nowrap"
    th_base = "border:1px solid #e2e8f0;padding:4px 8px;background:rgba(239,246,255,0.9);text-align:center;font-size:11px;font-weight:600;white-space:nowrap"

    html: list[str] = [
        "<div style='overflow-x:auto;font-size:12px;margin-top:8px'>",
        "<table style='border-collapse:collapse;white-space:nowrap'>",
        "<thead>",
        "<tr>",
        f'<th rowspan="2" style="{th_base}">STT</th>',
        f'<th rowspan="2" style="{th_base}">Xã / Phường</th>',
    ]
    for mk in active_mk:
        html.append(f'<th colspan="3" style="{th_hdr}">{_ten_ngan_pgd(mk)}</th>')
    html += ["</tr>", "<tr>"]
    for _ in active_mk:
        html += [
            f'<th style="{th_sub}">KH (tr.đ)</th>',
            f'<th style="{th_sub}">TH (tr.đ)</th>',
            f'<th style="{th_sub}">% TH/KH</th>',
        ]
    html += ["</tr>", "</thead>", "<tbody>"]

    tong_kh = {mk: 0.0 for mk in active_mk}
    tong_th_mk = {mk: 0.0 for mk in active_mk}

    for stt, xa in enumerate(ds_xa, 1):
        html.append("<tr>")
        html.append(f'<td style="{td};text-align:center;opacity:.55">{stt}</td>')
        html.append(f'<td style="{td}">{xa}</td>')
        for mk in active_mk:
            kh = float(kh_xa.get(f"{xa}|{mk}", 0)) / 1e6
            th = th_xa_ct.get((xa, mk), 0.0) / 1e6
            pct = (th / kh * 100) if kh > 0 else None
            tong_kh[mk] += kh
            tong_th_mk[mk] += th
            if pct is None:
                ps = ""
            elif pct >= 100:
                ps = "color:#81C784;font-weight:500"
            elif pct >= 80:
                ps = "color:#FFB74D"
            else:
                ps = "color:#EF9A9A"
            html += [
                f'<td style="{td};text-align:right">{fn(round(kh, 1))}</td>',
                f'<td style="{td};text-align:right">{fn(round(th, 1))}</td>',
                f'<td style="{td};text-align:right;{ps}">{fp(pct)}</td>',
            ]
        html.append("</tr>")

    html.append("<tr style='background:rgba(30,45,35,0.8);font-weight:700'>")
    html.append(f'<td style="{td}" colspan="2">Tổng cộng</td>')
    for mk in active_mk:
        kh = round(tong_kh[mk], 1)
        th = round(tong_th_mk[mk], 1)
        pct = (th / kh * 100) if kh > 0 else None
        if pct is None:
            ps = ""
        elif pct >= 100:
            ps = "color:#81C784"
        elif pct >= 80:
            ps = "color:#FFB74D"
        else:
            ps = "color:#EF9A9A"
        html += [
            f'<td style="{td};text-align:right">{fn(kh)}</td>',
            f'<td style="{td};text-align:right">{fn(th)}</td>',
            f'<td style="{td};text-align:right;{ps}">{fp(pct)}</td>',
        ]
    html += [
        "</tr>", "</tbody>", "</table>",
        "<p style='font-size:11px;margin-top:4px;color:#94A3B8'>Đơn vị: triệu đồng  |  TH = Tổng dư nợ (HSTD)</p>",
        "</div>",
    ]
    return "\n".join(html)


# ── Trợ lý văn bản QĐ ────────────────────────────────────────────────────────
def _pgd_slug(ten_pgd: str) -> str:
    """Chuyển tên PGD thành chuỗi an toàn dùng trong tên file/thư mục."""
    slug = re.sub(r"[^a-zA-Z0-9]", "_", ten_pgd)
    return re.sub(r"_+", "_", slug).strip("_").lower()


def _format_kich_thuoc(byte_count: int) -> str:
    return _svc_format_kich_thuoc(byte_count)


def _doc_meta_qd(kv_key: str) -> list[dict]:
    return _svc_doc_meta_qd(kv_key)


def _luu_meta_qd(kv_key: str, danh_sach: list[dict], username: str) -> None:
    try:
        _svc_luu_meta_qd(kv_key, danh_sach, username)
    except Exception as e:
        logger.error("Lỗi trong khối except: %s", e, exc_info=True)
        st.error(f"Lỗi lưu metadata file QĐ: {e}")


def _luu_file_qd(uploaded, thu_muc: Path, kv_key: str, username: str) -> Path:
    return _svc_luu_file_qd(uploaded, thu_muc, kv_key, username)


def _hien_thi_lich_su_qd(kv_key: str, nhan: str, role: str, username: str) -> None:
    """Hiển thị bảng lịch sử file QĐ, nút tải xuống và (admin) nút xóa."""
    danh_sach = _doc_meta_qd(kv_key)
    st.markdown(f"**{nhan}**")
    if not danh_sach:
        st.info("📭 Chưa có file nào được upload.")
        return

    h1, h2, h3, h4, h5, h6 = st.columns([3, 2, 2, 1.5, 1, 1])
    h1.markdown("**Tên file**")
    h2.markdown("**Ngày upload**")
    h3.markdown("**Người upload**")
    h4.markdown("**Dung lượng**")

    for idx_rev, meta in enumerate(reversed(danh_sach)):
        idx_thuc = len(danh_sach) - 1 - idx_rev
        c1, c2, c3, c4, c5, c6 = st.columns([3, 2, 2, 1.5, 1, 1])
        c1.text(meta.get("ten_file", "—"))
        c2.text(meta.get("ngay_upload", "—"))
        c3.text(meta.get("nguoi_upload", "—"))
        c4.text(_format_kich_thuoc(meta.get("kich_thuoc", 0)))

        duong_dan = Path(meta.get("duong_dan", ""))
        if duong_dan.exists():
            c5.download_button(
                "⬇",
                data=duong_dan.read_bytes(),
                file_name=meta.get("ten_file", duong_dan.name),
                key=f"dl_{kv_key}_{idx_rev}",
                help="Tải xuống",
            )
        else:
            c5.markdown("⚠️")

        if normalize_role(role) in ("admin_cn", "admin"):
            if c6.button("🗑", key=f"del_{kv_key}_{idx_rev}", help="Xóa file"):
                if duong_dan.exists():
                    duong_dan.unlink()
                danh_sach.pop(idx_thuc)
                _luu_meta_qd(kv_key, danh_sach, username)
                db.ghi_audit(
                    username, "xoa_van_ban_qd",
                    f"PGD · File: {meta.get('ten_file')} · key: {kv_key}",
                )
                st.rerun()


def _section_van_ban_qd_pgd(pgd: str, role: str, username: str) -> None:
    """Upload / hiển thị lịch sử văn bản QĐ theo PGD."""
    slug = _pgd_slug(pgd)
    kv_hdqt_tinh = f"qd_files_{slug}_hdqt_tinh"
    kv_hdqt_xa   = f"qd_files_{slug}_hdqt_xa"
    thu_muc_tinh = PGD_DATA_DIR / "qd" / slug / "hdqt_tinh"
    thu_muc_xa   = PGD_DATA_DIR / "qd" / slug / "hdqt_xa"

    with st.expander("📎 Văn bản QĐ", expanded=True):
        # ── Lịch sử từng loại ────────────────────────────────────────────
        col_hist1, col_hist2 = st.columns(2)
        with col_hist1:
            _hien_thi_lich_su_qd(kv_hdqt_tinh, "QĐ HĐQT tỉnh", role, username)
        with col_hist2:
            _hien_thi_lich_su_qd(kv_hdqt_xa, "QĐ HĐQT xã", role, username)

        if not la_phan_he_cn(role) or normalize_role(role) == "executive":
            st.caption("🔒 Chỉ Admin / Manager mới được upload văn bản QĐ.")
            return

        # ── Upload file mới ───────────────────────────────────────────────
        st.divider()
        col_u1, col_u2 = st.columns(2)
        with col_u1:
            f_tinh = st.file_uploader(
                "Upload QĐ HĐQT tỉnh",
                type=["pdf", "xlsx", "xls"],
                key=f"vb_hdqt_tinh_{slug}",
            )
            if f_tinh:
                _id = f"qd_done_tinh_{slug}_{f_tinh.name}_{f_tinh.size}"
                if not st.session_state.get(_id):
                    st.session_state[_id] = True
                    try:
                        dp = _luu_file_qd(f_tinh, thu_muc_tinh, kv_hdqt_tinh, username)
                        db.ghi_audit(username, "upload_vb_hdqt_tinh_pgd",
                                     f"PGD: {pgd} · File: {dp.name}")
                        st.success(f"✅ Đã lưu: `{dp.name}`")
                        st.rerun()
                    except Exception as e:
                        logger.error("Lỗi trong khối except: %s", e, exc_info=True)
                        st.session_state.pop(_id, None)
                        st.error(f"Lỗi lưu file QĐ HĐQT tỉnh: {e}")
        with col_u2:
            f_xa = st.file_uploader(
                "Upload QĐ HĐQT xã",
                type=["pdf", "xlsx", "xls"],
                key=f"vb_hdqt_xa_{slug}",
            )
            if f_xa:
                _id = f"qd_done_xa_{slug}_{f_xa.name}_{f_xa.size}"
                if not st.session_state.get(_id):
                    st.session_state[_id] = True
                    try:
                        dp = _luu_file_qd(f_xa, thu_muc_xa, kv_hdqt_xa, username)
                        db.ghi_audit(username, "upload_vb_hdqt_xa_pgd",
                                     f"PGD: {pgd} · File: {dp.name}")
                        st.success(f"✅ Đã lưu: `{dp.name}`")
                        st.rerun()
                    except Exception as e:
                        logger.error("Lỗi trong khối except: %s", e, exc_info=True)
                        st.session_state.pop(_id, None)
                        st.error(f"Lỗi lưu file QĐ HĐQT xã: {e}")


# ── Đọc kv_store ──────────────────────────────────────────────────────────────
def _tao_column_config_khtd_pgd(cot_so: list[str]) -> dict[str, st.column_config.Column]:
    """
    Tạo column_config cho bảng KHTD PGD.
    
    Args:
        cot_so: Danh sách tên cột số cần format
    
    Returns:
        Dict cấu hình column cho st.dataframe
    """
    config: dict[str, st.column_config.Column] = {
        "STT": st.column_config.TextColumn("STT", width="small"),
        "Chỉ Tiêu": st.column_config.TextColumn("Chỉ Tiêu", width="large"),
        "Cộng": st.column_config.NumberColumn(
            "Cộng",
            format=",.0f",
            help="Đơn vị: triệu đồng"
        ),
    }
    for col in cot_so:
        if col not in ("STT", "Chỉ Tiêu", "Cộng"):
            config[col] = st.column_config.NumberColumn(
                col,
                format=",.0f",
                help="Đơn vị: triệu đồng"
            )
    return config


def _tao_column_config_ss() -> dict[str, st.column_config.Column]:
    """Tạo column_config cho bảng so sánh KH vs TH."""
    trieu_cols = [
        "KH TW (triệu)", "KH ĐP (triệu)", "TH TW (triệu)", "TH ĐP (triệu)",
        "KH Tổng (triệu)", "TH Tổng (triệu)", "Còn lại (triệu)"
    ]
    config: dict[str, st.column_config.Column] = {
        "Chương trình": st.column_config.TextColumn("Chương trình", width="large"),
    }
    for col in trieu_cols:
        config[col] = st.column_config.NumberColumn(
            col,
            format=",.0f",
            help="Đơn vị: triệu đồng"
        )
    config["Tỷ lệ TH %"] = st.column_config.NumberColumn(
        "Tỷ lệ TH %",
        format=".1%",
        help="Tỷ lệ thực hiện %"
    )
    return config


def _doc_khtd_xa() -> dict[str, Any]:
    """
    Đọc toàn bộ kế hoạch xã từ kv_store (key=khtd_xa).
    
    Returns:
        Dict chứa dữ liệu kế hoạch xã
    """
    val = db.doc_kv(KV_KEY_XA)
    return val if isinstance(val, dict) else {}


def _fmt_vn(x, d: int = 1) -> str:
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return "—"
        v = float(x)
        s = f"{v:,.{d}f}"
        return s.replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception as e:
        logger.error("Lỗi trong khối except: %s", e, exc_info=True)
        return "—"


def _fmt_vn_signed(x, d: int = 1) -> str:
    try:
        v = float(x)
        s = _fmt_vn(v, d)
        return f"+{s}" if v > 0 else s
    except Exception as e:
        logger.error("Lỗi trong khối except: %s", e, exc_info=True)
        return "—"

def _nv_int_tu_ma_key(ma_key: str) -> int | None:
    if ma_key.endswith("_TW"):
        return 1
    if ma_key.endswith("_DP"):
        return 2
    if "|" in ma_key:
        try:
            return int(ma_key.split("|", 1)[1])
        except Exception as e:
            logger.error("Lỗi trong khối except: %s", e, exc_info=True)
            return None
    nv = NGUON_VON_MA.get(ma_key)
    if nv == "TW":
        return 1
    if nv == "DP":
        return 2
    return None


def _ma_ct_tu_ma_key(ma_key: str) -> int | None:
    if "|" in ma_key:
        try:
            return int(ma_key.split("|", 1)[0])
        except Exception as e:
            logger.error("Lỗi trong khối except: %s", e, exc_info=True)
            return None
    if ma_key.endswith("_TW") or ma_key.endswith("_DP"):
        try:
            return int(ma_key.rsplit("_", 1)[0])
        except Exception as e:
            logger.error("Lỗi trong khối except: %s", e, exc_info=True)
            return None
    return MA_CT_BY_MAKEY.get(ma_key)


def _bang_so_sanh_kh_th_pivot(
    kh_theo_ct: dict[str, float],
    th_theo_ct: dict[str, float],
    ten_ct_map: dict[str, str],
) -> pd.DataFrame:
    keys = set(kh_theo_ct.keys()) | set(th_theo_ct.keys())
    ma_cts = sorted({m for m in (_ma_ct_tu_ma_key(k) for k in keys) if m is not None})
    rows: list[dict] = []
    for ma_ct in ma_cts:
        mk_tw = f"{ma_ct}|1"
        mk_dp = f"{ma_ct}|2"
        mk_tw2 = f"{ma_ct}_TW"
        mk_dp2 = f"{ma_ct}_DP"
        kh_tw = float(kh_theo_ct.get(mk_tw, kh_theo_ct.get(mk_tw2, 0.0))) / 1_000_000
        kh_dp = float(kh_theo_ct.get(mk_dp, kh_theo_ct.get(mk_dp2, 0.0))) / 1_000_000
        th_tw = float(th_theo_ct.get(mk_tw, th_theo_ct.get(mk_tw2, 0.0))) / 1_000_000
        th_dp = float(th_theo_ct.get(mk_dp, th_theo_ct.get(mk_dp2, 0.0))) / 1_000_000
        kh_tong = kh_tw + kh_dp
        th_tong = th_tw + th_dp
        if kh_tong == 0 and th_tong == 0:
            continue
        ten = TEN_BASE_BY_MACT.get(ma_ct)
        if not ten:
            ten = ten_ct_map.get(mk_tw2) or ten_ct_map.get(mk_dp2) or ten_ct_map.get(mk_tw) or ten_ct_map.get(mk_dp) or str(ma_ct)
        rows.append({
            "Chương trình": ten,
            "KH TW (triệu)": round(kh_tw, 1),
            "KH ĐP (triệu)": round(kh_dp, 1),
            "TH TW (triệu)": round(th_tw, 1),
            "TH ĐP (triệu)": round(th_dp, 1),
            "KH Tổng (triệu)": round(kh_tong, 1),
            "TH Tổng (triệu)": round(th_tong, 1),
            "Còn lại (triệu)": round(kh_tong - th_tong, 1),
            "Tỷ lệ TH %": round(th_tong / kh_tong * 100, 1) if kh_tong > 0 else None,
        })
    return pd.DataFrame(rows)


def _ten_ct_map_tu_df(df: "pd.DataFrame | None") -> dict[str, str]:
    if df is None or df.empty:
        return {}
    if COT_MA_CHUONG_TRINH not in df.columns or COT_NGUON_VON not in df.columns:
        return {}
    if COT_TEN_CT not in df.columns:
        return {}

    lookup: dict[tuple[int, int], str] = {}
    for ma_key, ma_ct, _, nguon_von, _ in CHUONG_TRINH_KHTD:
        nv_int = 1 if nguon_von == "TW" else 2
        if (int(ma_ct), nv_int) not in lookup:
            lookup[(int(ma_ct), nv_int)] = ma_key

    ma_ct_s = pd.to_numeric(df[COT_MA_CHUONG_TRINH], errors="coerce").fillna(0).astype(int)
    nv_s = pd.to_numeric(df[COT_NGUON_VON], errors="coerce").fillna(0).astype(int)
    ten_s = df[COT_TEN_CT].astype(str).fillna("").str.strip()

    out: dict[str, str] = {}
    for i in range(len(ma_ct_s)):
        ma_ct = int(ma_ct_s.iat[i])
        nv_int = int(nv_s.iat[i])
        if ma_ct <= 0 or nv_int not in (1, 2):
            continue
        mk = lookup.get((ma_ct, nv_int), f"{ma_ct}|{nv_int}")
        if mk in out:
            continue
        ten = str(ten_s.iat[i]).strip()
        out[mk] = TEN_CHINH_THUC_CT.get(mk, ten) or mk
    return out


def _tinh_thuc_hien_theo_ct(df: "pd.DataFrame") -> dict[str, float]:
    if df is None or df.empty:
        return {}
    if COT_MA_CHUONG_TRINH not in df.columns or COT_NGUON_VON not in df.columns:
        return {}
    col_th = COT_TONG_DU_NO if COT_TONG_DU_NO in df.columns else (
        COT_DU_NO_TH if COT_DU_NO_TH in df.columns else None
    )
    if not col_th:
        return {}

    lookup: dict[tuple[int, int], str] = {}
    for ma_key, ma_ct, _, nguon_von, _ in CHUONG_TRINH_KHTD:
        nv_int = 1 if nguon_von == "TW" else 2
        if (int(ma_ct), nv_int) not in lookup:
            lookup[(int(ma_ct), nv_int)] = ma_key

    ma_ct_s = pd.to_numeric(df[COT_MA_CHUONG_TRINH], errors="coerce").fillna(0).astype(int)
    nv_s = pd.to_numeric(df[COT_NGUON_VON], errors="coerce").fillna(0).astype(int)
    th_s = pd.to_numeric(df[col_th], errors="coerce").fillna(0).astype(float)

    tmp = pd.DataFrame({"ma_ct": ma_ct_s, "nv": nv_s, "th": th_s})
    tmp = tmp[(tmp["ma_ct"] > 0) & (tmp["nv"].isin([1, 2])) & (tmp["th"] != 0)]
    if tmp.empty:
        return {}

    g = tmp.groupby(["ma_ct", "nv"])["th"].sum()
    out: dict[str, float] = {}
    for (ma_ct, nv_int), val in g.items():
        mk = lookup.get((int(ma_ct), int(nv_int)), f"{int(ma_ct)}|{int(nv_int)}")
        out[mk] = float(val)

    # Phân tầng GQVL ĐP (ma_ct=3, nv=2) và NSVSMT ĐP (ma_ct=6, nv=2) theo Mã NĐT thực tế.
    # first-win lookup ở trên chỉ gán (3,2)→"3_DP_TINH", (6,2)→"6_DP" — ghi đè bằng split đúng.
    if COT_MA_NHA_DAU_TU in df.columns:
        import db as _db
        ma_ndt_arr = df[COT_MA_NHA_DAU_TU].astype(str).fillna("").str.strip().values
        ma_ct_arr = ma_ct_s.values
        nv_arr = nv_s.values
        th_arr = th_s.values

        # GQVL ĐP → 3_DP_TINH / 3_DP_XA
        mask3 = (ma_ct_arr == 3) & (nv_arr == 2) & (th_arr != 0)
        if mask3.any():
            th3 = th_arr[mask3]
            cap3 = [_db.phan_loai_ndt_dp_cap(3, ma) for ma in ma_ndt_arr[mask3]]
            is_tinh3 = [c == "tinh" for c in cap3]
            out["3_DP_TINH"] = float(sum(v for v, t in zip(th3, is_tinh3) if t))
            out["3_DP_XA"] = float(sum(v for v, t in zip(th3, is_tinh3) if not t))

        # NSVSMT ĐP → 6_DP_TINH / 6_DP_XA (xóa key tổng "6_DP" cũ)
        mask6 = (ma_ct_arr == 6) & (nv_arr == 2) & (th_arr != 0)
        if mask6.any():
            th6 = th_arr[mask6]
            cap6 = [_db.phan_loai_ndt_dp_cap(6, ma) for ma in ma_ndt_arr[mask6]]
            is_tinh6 = [c == "tinh" for c in cap6]
            out["6_DP_TINH"] = float(sum(v for v, t in zip(th6, is_tinh6) if t))
            out["6_DP_XA"] = float(sum(v for v, t in zip(th6, is_tinh6) if not t))
            out.pop("6_DP", None)

    return out


def _tinh_ke_hoach_pgd_theo_ct(
    pgd: str,
    kh_xa: dict,
    ds_ct_loc: list[tuple[str, str]],
) -> dict[str, float]:
    ds_xa = PGD_XA_MAP.get(pgd, [])
    if not ds_xa:
        return {}
    out: dict[str, float] = {}
    for ma_key, _ in ds_ct_loc:
        tong = 0.0
        for xa in ds_xa:
            tong += float(kh_xa.get(f"{xa}|{ma_key}", 0.0))
        out[ma_key] = tong
    return out


def _bang_so_sanh_kh_th(
    ds_ct_loc: list[tuple[str, str]],
    kh_theo_ct: dict[str, float],
    th_theo_ct: dict[str, float],
) -> pd.DataFrame:
    ten_ct_map = {mk: ten for mk, ten in ds_ct_loc}
    return _bang_so_sanh_kh_th_pivot(kh_theo_ct, th_theo_ct, ten_ct_map)


# ── Xây dựng DataFrame tổng hợp cho một PGD ──────────────────────────────────
def _xay_dung_bang_pgd(pgd: str, kh_xa: dict, ds_ct_loc: list[tuple[str, str]]) -> pd.DataFrame:
    """Tạo DataFrame ma trận (hàng=chỉ tiêu, cột=xã) cho PGD."""
    danh_sach_xa = PGD_XA_MAP.get(pgd, [])
    if not danh_sach_xa:
        return pd.DataFrame()

    ds_loc_set = {mk for mk, _ in ds_ct_loc}
    hierarchy_rows: list[dict] = [
        {"stt": "A", "chi_tieu": "KẾ HOẠCH", "nhom": "A", "ma_key": None},
        {"stt": "I", "chi_tieu": "Nguồn vốn Trung ương", "nhom": "I", "ma_key": None},
    ]
    stt_tw = 1
    for mk, _ma_ct, ten, nv, *_ in CHUONG_TRINH_KHTD:
        if nv != "TW" or mk not in ds_loc_set:
            continue
        hierarchy_rows.append({
            "stt": str(stt_tw),
            "chi_tieu": f"  {ten}",
            "nhom": "con",
            "ma_key": mk,
        })
        stt_tw += 1
    hierarchy_rows.append({"stt": "II", "chi_tieu": "Nguồn vốn Địa phương", "nhom": "II", "ma_key": None})
    stt_dp = 1
    for mk, _ma_ct, ten, nv, *_ in CHUONG_TRINH_KHTD:
        if nv != "DP" or mk not in ds_loc_set:
            continue
        hierarchy_rows.append({
            "stt": str(stt_dp),
            "chi_tieu": f"  {ten}",
            "nhom": "con",
            "ma_key": mk,
        })
        stt_dp += 1

    rows: list[dict] = []
    for row in hierarchy_rows:
        hang = {"STT": row["stt"], "Chỉ Tiêu": row["chi_tieu"]}
        if row["ma_key"]:
            tong = 0.0
            for xa in danh_sach_xa:
                val = round(float(kh_xa.get(f"{xa}|{row['ma_key']}", 0.0)) / 1_000_000, 1)
                hang[xa] = val
                tong += val
            hang["Cộng"] = round(tong, 1)
        else:
            for xa in danh_sach_xa:
                hang[xa] = None
            hang["Cộng"] = None
        hang["_nhom"] = row["nhom"]
        rows.append(hang)

    # Dòng tổng PGD ở cuối bảng
    tong_hang = {"STT": "", "Chỉ Tiêu": "TỔNG PGD", "_nhom": "tong"}
    for xa in danh_sach_xa:
        tong_hang[xa] = round(
            sum(
                float(kh_xa.get(f"{xa}|{mk}", 0.0)) / 1_000_000
                for mk, _ma_ct, _ten, _nv, *_ in CHUONG_TRINH_KHTD
                if mk in ds_loc_set
            ),
            1,
        )
    tong_hang["Cộng"] = round(sum(v for k, v in tong_hang.items() if k in danh_sach_xa), 1)
    rows.append(tong_hang)

    return pd.DataFrame(rows)


# ── Entry point ───────────────────────────────────────────────────────────────
def render(tab: DeltaGenerator | None = None, **kwargs: dict) -> None:
    """
    Render tab KHTD cấp PGD.
    
    Args:
        tab: Streamlit DeltaGenerator cho tab này
        **kwargs: Chứa role, username, pgd_user, df_full, df
    """
    role_raw = str(kwargs.get("role", "user") or "user")
    role = normalize_role(role_raw)
    username = kwargs.get("username", "unknown")
    pgd_user = kwargs.get("pgd_user", "")
    df = kwargs.get("df_full", kwargs.get("df"))

    _tab_ctx = tab if tab is not None else __import__('streamlit').container()
    with _tab_ctx:
        st.title("🏦 Kế hoạch Tín dụng — Cấp PGD")
        st.caption("Hỗ trợ địa bàn · Xem tổng hợp KH theo Xã × Chương trình")

        # ── Xác định PGD hiển thị ─────────────────────────────────────────
        if la_phan_he_cn(role) and not la_executive(role):
            pgd_hien_tai: str = st.selectbox(
                "Chọn PGD", DS_PGD, key="khtd_pgd_sel_admin"
            )
        else:
            if not pgd_user:
                st.warning("⚠️ Chưa gán PGD cho tài khoản này. Liên hệ Admin.")
                return
            pgd_hien_tai = pgd_user
            st.info(f"📍 PGD của bạn: **{pgd_hien_tai}**")

        # ── Văn bản QĐ theo PGD ──────────────────────────────────────────
        _section_van_ban_qd_pgd(pgd_hien_tai, role, username)

        # ── Đọc dữ liệu ──────────────────────────────────────────────────
        kh_xa = _doc_khtd_xa()

        df_th_pgd = df
        if df_th_pgd is not None and not df_th_pgd.empty:
            if COT_TEN_PGD in df_th_pgd.columns:
                df_th_pgd = df_th_pgd[df_th_pgd[COT_TEN_PGD] == pgd_hien_tai]
        th_theo_ct = _tinh_thuc_hien_theo_ct(df_th_pgd) if df_th_pgd is not None else {}
        th_theo_ct = {k: v for k, v in th_theo_ct.items() if float(v) > 0}
        th_xa_ct = _tinh_th_xa_ct(df_th_pgd) if df_th_pgd is not None else {}
        ten_ct_map = _ten_ct_map_tu_df(df_th_pgd)

        ds_xa_set = set(PGD_XA_MAP.get(pgd_hien_tai, []))
        keys_kh_pgd: set[str] = set()
        for khoa in kh_xa.keys():
            try:
                xa, ma_key = khoa.split("|", 1)
            except ValueError:
                continue
            if xa in ds_xa_set:
                keys_kh_pgd.add(ma_key)

        if not PGD_XA_MAP.get(pgd_hien_tai):
            st.warning(
                f"Chưa có danh sách xã cho **{pgd_hien_tai}**. "
                "Kiểm tra cấu hình `PGD_XA_MAP` trong config.py."
            )
            return
        if not keys_kh_pgd and not th_xa_ct:
            st.info(
                f"Chưa có dữ liệu kế hoạch xã cho **{pgd_hien_tai}**. "
                "Yêu cầu Phòng KH-NV nhập dữ liệu trên tab **KHTD theo Xã**."
            )

        keys_show = set(keys_kh_pgd) | set(th_theo_ct.keys())
        base_all = [(mk, ten) for mk, _, ten, _, _ in CHUONG_TRINH_KHTD]
        if not keys_show:
            keys_show = {mk for mk, _ in base_all}

        base_keys = [mk for mk, _ in base_all if mk in keys_show]
        base_key_set = {mk for mk, _ in base_all}
        extra_keys = sorted(
            (k for k in keys_show if k not in base_key_set),
            key=lambda k: (
                _nv_int_tu_ma_key(k) or 9,
                int(k.split("_", 1)[0]) if "_" in k and k.split("_", 1)[0].isdigit()
                else (int(k.split("|", 1)[0]) if "|" in k and k.split("|", 1)[0].isdigit() else 9_999),
                k,
            ),
        )
        ds_ct_loc: list[tuple[str, str]] = []
        for mk in base_keys + extra_keys:
            ds_ct_loc.append((mk, TEN_CHINH_THUC_CT.get(mk, ten_ct_map.get(mk, mk)) or mk))

        df_pgd = _xay_dung_bang_pgd(pgd_hien_tai, kh_xa, ds_ct_loc)
        kh_theo_ct = _tinh_ke_hoach_pgd_theo_ct(pgd_hien_tai, kh_xa, ds_ct_loc)
        df_ss = _bang_so_sanh_kh_th(ds_ct_loc, kh_theo_ct, th_theo_ct)

        # ── Metrics ──────────────────────────────────────────────────────
        tong_pgd = sum(float(v) for v in kh_theo_ct.values()) / 1_000_000
        tong_th = sum(float(v) for v in th_theo_ct.values()) / 1_000_000
        tl_all = (tong_th / tong_pgd * 100) if tong_pgd > 0 else None
        so_xa = len(PGD_XA_MAP.get(pgd_hien_tai, []))

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Tổng KH PGD (triệu đồng)", _fmt_vn(tong_pgd, 1))
        m2.metric("Tổng TH PGD (triệu đồng)", _fmt_vn(tong_th, 1))
        m3.metric("Tỷ lệ TH", f"{_fmt_vn(tl_all, 1)}%" if tl_all is not None else "—")
        m4.metric("Số xã có kế hoạch", str(so_xa))

        st.divider()

        # ── Bảng BĐD — hàng = xã, cột = nhóm CT ─────────────────────────
        t_tw, t_dp = st.tabs(["🏦 Nguồn TW", "🏙️ Nguồn ĐP"])
        with t_tw:
            st.html(_html_bdd_pgd_table(pgd_hien_tai, kh_xa, th_xa_ct, "TW"))
        with t_dp:
            st.html(_html_bdd_pgd_table(pgd_hien_tai, kh_xa, th_xa_ct, "DP"))

        # ── Xuất Excel ────────────────────────────────────────────────────
        st.divider()
        ten_slug = _pgd_slug(pgd_hien_tai)

        buf = BytesIO()
        df_pgd_xuat = df_pgd.drop(columns=["_nhom"]).copy()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_pgd_xuat.to_excel(writer, index=False, sheet_name="KH_Xa")
            if not df_ss.empty:
                df_ss.to_excel(writer, index=False, sheet_name="KH_vs_TH")
            ws = writer.book["KH_Xa"]
            fill_nhom = PatternFill(fill_type="solid", fgColor="F0F4FA")
            fill_tong = PatternFill(fill_type="solid", fgColor="E8F4FD")
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for ridx, nhom in enumerate(df_pgd["_nhom"].tolist(), start=2):
                if nhom in ("A", "I", "II", "tong"):
                    for cidx in range(1, len(df_pgd_xuat.columns) + 1):
                        cell = ws.cell(row=ridx, column=cidx)
                        cell.font = Font(bold=True)
                        cell.fill = fill_tong if nhom == "tong" else fill_nhom
            for cidx in range(1, len(df_pgd_xuat.columns) + 1):
                ws.column_dimensions[ws.cell(row=1, column=cidx).column_letter].width = 16
            ws.column_dimensions["B"].width = 55
        st.download_button(
            label=f"📥 Xuất Excel — {pgd_hien_tai}",
            data=buf.getvalue(),
            file_name=f"KHTD_{ten_slug}_{datetime.today().strftime('%d%m%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_khtd_pgd_{ten_slug}",
        )
        db.ghi_audit(username, "xuat_excel_khtd_pgd", f"PGD: {pgd_hien_tai}")
