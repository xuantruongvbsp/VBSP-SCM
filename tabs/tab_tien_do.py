from __future__ import annotations

import json
from datetime import datetime, date
from io import BytesIO

import pandas as pd
import plotly.express as px
import streamlit as st

import db
from auth import normalize_role, la_admin_cn
from config import DS_PGD, DON_VI_CHI_NHANH, PGD_XA_MAP, ROLES_PHAN_HE_CN
from utils import fmt_ngay, lazy_tabs

DS_PGD_ALL = [DON_VI_CHI_NHANH] + DS_PGD
_PGD_BIEN_HOA = "Địa bàn Biên Hòa"

LOAI_TASK = {
    "chung":            "📋 Công việc chung",
    "chi_tieu_khtd":    "🎯 Chỉ tiêu KHTD",
    "ho_so_rui_ro":     "🗂️ Hồ sơ rủi ro",
    "khao_sat_nhu_cau": "📊 Khảo sát nhu cầu vay vốn",
    "bao_cao":          "📄 Báo cáo",
    "tap_huan":         "🎓 Tập huấn",
    "giao_dich_xa":     "📅 Giao dịch xã",
    "uy_thac":          "🤝 Hoạt động ủy thác",
    "nguon_von":        "💰 Nguồn vốn",
    "ban_dai_dien":     "📑 Ban đại diện HĐQT",
    "khac":             "Khác",
}
UU_TIEN = {
    "khan_cap":    "🔴 Khẩn cấp",
    "quan_trong":  "🟡 Quan trọng",
    "binh_thuong": "🟢 Bình thường",
}
TS_KQ_LABEL = {
    "chua_thuc_hien": "⬜ Chưa",
    "da_hoan_thanh":  "✅ Xong",
    "khong_ap_dung":  "➖ N/A",
}
UU_TIEN_ORDER = {"khan_cap": 0, "quan_trong": 1, "binh_thuong": 2}


def _doc_tasks(chi_dang_theo_doi: bool = True) -> list[dict]:
    with db.get_conn() as conn:
        sql = "SELECT * FROM tien_do_task"
        if chi_dang_theo_doi:
            sql += " WHERE trang_thai = 'dang_theo_doi'"
        sql += " ORDER BY ngay_deadline ASC"
        return [dict(r) for r in conn.execute(sql).fetchall()]


def _doc_ketqua_task(task_id: int) -> list[dict]:
    with db.get_conn() as conn:
        return [dict(r) for r in conn.execute(
            "SELECT * FROM tien_do_ketqua WHERE task_id=? ORDER BY pgd, ten_xa",
            (task_id,),
        ).fetchall()]


def _khoi_tao_ketqua_task(task_id: int, ds_pgd_task: list[str],
                           cap_theo_doi: str = "xa",
                           loai_noi_dung: str = "chi_tiet_xa") -> None:
    rows = []
    if cap_theo_doi == "pgd":
        for pgd in ds_pgd_task:
            rows.append((task_id, pgd, pgd, loai_noi_dung))
    else:
        for pgd in ds_pgd_task:
            for xa in PGD_XA_MAP.get(pgd, []):
                rows.append((task_id, pgd, xa, loai_noi_dung))
    with db.get_conn() as conn:
        conn.executemany(
            """INSERT OR IGNORE INTO tien_do_ketqua
               (task_id, pgd, ten_xa, trang_thai, loai_noi_dung)
               VALUES (?, ?, ?, 'chua_thuc_hien', ?)""",
            rows,
        )
        conn.commit()


def _sync_bien_hoa_ketqua(task_id: int, cbtd_bien_hoa: str, loai_noi_dung: str) -> None:
    val = str(cbtd_bien_hoa or "").strip()
    with db.get_conn() as conn:
        if val:
            conn.execute(
                """INSERT OR IGNORE INTO tien_do_ketqua
                   (task_id, pgd, ten_xa, trang_thai, loai_noi_dung)
                   VALUES (?, ?, ?, 'chua_thuc_hien', ?)""",
                (task_id, _PGD_BIEN_HOA, _PGD_BIEN_HOA, loai_noi_dung),
            )
        else:
            conn.execute(
                "DELETE FROM tien_do_ketqua WHERE task_id=? AND ten_xa=?",
                (task_id, _PGD_BIEN_HOA),
            )
        conn.commit()


def _upsert_ketqua_xa(
    task_id: int, ten_xa: str, pgd: str,
    trang_thai: str, ngay_ht: str | None,
    ghi_chu: str | None, username: str,
) -> None:
    now = datetime.now().isoformat()
    with db.get_conn() as conn:
        conn.execute(
            """INSERT INTO tien_do_ketqua
               (task_id, pgd, ten_xa, trang_thai, ngay_hoan_thanh,
                ghi_chu, nguoi_nhap, ngay_nhap)
               VALUES (?,?,?,?,?,?,?,?)
               ON CONFLICT(task_id, ten_xa) DO UPDATE SET
                 trang_thai      = excluded.trang_thai,
                 ngay_hoan_thanh = excluded.ngay_hoan_thanh,
                 ghi_chu         = excluded.ghi_chu,
                 nguoi_nhap      = excluded.nguoi_nhap,
                 ngay_nhap       = excluded.ngay_nhap""",
            (task_id, pgd, ten_xa, trang_thai,
             ngay_ht, ghi_chu, username, now),
        )
        conn.commit()


def _render_tong_quan(tab, **kwargs):
    import streamlit as _st
    _tab_ctx = tab if tab is not None else _st.container()
    with _tab_ctx:
        st.subheader("📊 Tổng quan tiến độ")

        c1, c2, c3 = st.columns([2, 1, 1])
        with c1:
            ngay_loc = st.date_input("Thời hạn đến ngày",
                                     value=date.today(), format="DD/MM/YYYY", key="td_ngay")
        with c2:
            loai_loc = st.selectbox("Lọc loại", ["Tất cả"] + list(LOAI_TASK.values()),
                                    key="td_loai")
        with c3:
            nd_loc = st.selectbox("Loại nhiệm vụ",
                                  ["Tất cả", "Chung PGD", "Chi tiết xã"],
                                  key="td_nd")

        hom_nay = date.today().isoformat()
        ds_task = _doc_tasks()
        ds_task = [t for t in ds_task
                   if t["ngay_deadline"] <= ngay_loc.isoformat()]
        if loai_loc != "Tất cả":
            loai_key = next((k for k, v in LOAI_TASK.items() if v == loai_loc), None)
            ds_task = [t for t in ds_task if loai_key and t["loai"] == loai_key]
        if nd_loc != "Tất cả":
            nd_map = {"Chung PGD": "pgd", "Chi tiết xã": "xa"}
            ds_task = [t for t in ds_task if t.get("cap_theo_doi") == nd_map.get(nd_loc)]

        if not ds_task:
            st.info("Không có đầu việc nào trong khoảng thời gian đã chọn.")
            return

        all_kq = {t["id"]: _doc_ketqua_task(t["id"]) for t in ds_task}
        tong_xa = sum(len(v) for v in all_kq.values())
        tong_xong = sum(
            sum(1 for r in v if r["trang_thai"] == "da_hoan_thanh")
            for v in all_kq.values()
        )
        tong_tre = 0
        for t in ds_task:
            kq = all_kq.get(t["id"], [])
            for r in kq:
                if r["trang_thai"] == "chua_thuc_hien" and t["ngay_deadline"] < hom_nay:
                    tong_tre += 1

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Đầu việc", len(ds_task))
        c2.metric("✅ Hoàn thành",
                  f"{tong_xong}/{tong_xa}",
                  f"{round(tong_xong/tong_xa*100) if tong_xa else 0}%")
        c3.metric("🔴 Trễ hạn", tong_tre, delta_color="inverse")
        c4.metric("⬜ Chưa báo cáo", tong_xa - tong_xong - tong_tre)

        st.markdown("#### 📋 Danh sách đầu việc")
        df_task_list = pd.DataFrame([
            {
                "Đầu việc": t.get("tieu_de", ""),
                "Thời hạn": fmt_ngay(t.get("ngay_deadline", "")),
                "Loại": LOAI_TASK.get(t.get("loai", ""), t.get("loai", "")),
                "Người phụ trách": t.get("nguoi_phu_trach") or "",
                "CB KH-NV phụ trách": t.get("nguoi_thuc_hien_cn") or "",
                "Hội sở CN tỉnh": t.get("cbtd_bien_hoa") or "",
                "Ưu tiên": UU_TIEN.get(t.get("uu_tien", ""), t.get("uu_tien", "")),
                "Theo dõi": "Chung PGD" if t.get("cap_theo_doi") == "pgd" else "Chi tiết xã",
            }
            for t in ds_task
        ])
        st.dataframe(df_task_list, use_container_width=True, hide_index=True, height=340)

        st.divider()

        st.markdown("#### 🗺️ Bảng tiến độ theo PGD")
        st.caption("Ô trong bảng = Hoàn thành / Tổng số. 🔴 = có đơn vị trễ hạn.")

        bang_rows = []
        for pgd in DS_PGD_ALL:
            row = {"Đơn vị": pgd}
            for t in ds_task:
                kq_pgd = [r for r in all_kq.get(t["id"], []) if r["pgd"] == pgd]
                if not kq_pgd:
                    row[t["tieu_de"][:18]] = "—"
                    continue
                xong = sum(1 for r in kq_pgd if r["trang_thai"] == "da_hoan_thanh")
                tre = sum(1 for r in kq_pgd
                          if r["trang_thai"] == "chua_thuc_hien"
                          and t["ngay_deadline"] < hom_nay)
                tong = len(kq_pgd)
                ky_hieu = "🔴 " if tre > 0 else ("✅ " if xong == tong else "")
                row[t["tieu_de"][:18]] = f"{ky_hieu}{xong}/{tong}"
            bang_rows.append(row)

        df_pgd = pd.DataFrame(bang_rows)
        st.dataframe(df_pgd, use_container_width=True, hide_index=True, height=620)

        st.divider()

        st.markdown("#### 📈 Tỷ lệ hoàn thành theo đầu việc")
        chart_rows = []
        for t in ds_task:
            kq = all_kq.get(t["id"], [])
            xong = sum(1 for r in kq if r["trang_thai"] == "da_hoan_thanh")
            tong = len(kq)
            chart_rows.append({
                "Đầu việc": t["tieu_de"][:35],
                "% Hoàn thành": round(xong / tong * 100) if tong else 0,
                "Ưu tiên": UU_TIEN.get(t["uu_tien"], ""),
                "Thời hạn": fmt_ngay(t["ngay_deadline"]),
            })
        df_chart = pd.DataFrame(chart_rows)
        color_map = {
            "🔴 Khẩn cấp": "#ef4444",
            "🟡 Quan trọng": "#f59e0b",
            "🟢 Bình thường": "#22c55e",
        }
        fig = px.bar(
            df_chart, x="% Hoàn thành", y="Đầu việc",
            orientation="h", color="Ưu tiên",
            color_discrete_map=color_map,
            text="% Hoàn thành", range_x=[0, 100],
            height=max(300, len(ds_task) * 50),
        )
        fig.update_traces(texttemplate="%{text}%", textposition="outside")
        fig.update_layout(margin=dict(l=10, r=40, t=30, b=10))
        st.plotly_chart(fig, use_container_width=True, key="td_chart_tongquan")

        st.divider()

        st.markdown("#### 🔍 Drill-down theo PGD")
        _DD_TAT_CA = "— Tất cả —"
        col_pgd, col_task = st.columns(2)
        with col_pgd:
            pgd_dd = st.selectbox("Chọn đơn vị", [_DD_TAT_CA] + DS_PGD_ALL, key="td_dd_pgd")
        with col_task:
            task_dd_options = {t["id"]: t["tieu_de"] for t in ds_task}
            task_dd_id = st.selectbox(
                "Chọn đầu việc",
                options=list(task_dd_options.keys()),
                format_func=lambda x: task_dd_options[x],
                key="td_dd_task",
            )

        if pgd_dd and task_dd_id:
            _tat_ca = pgd_dd == _DD_TAT_CA
            kq_xa = (
                all_kq.get(task_dd_id, [])
                if _tat_ca
                else [r for r in all_kq.get(task_dd_id, []) if r["pgd"] == pgd_dd]
            )
            task_sel = next((t for t in ds_task if t["id"] == task_dd_id), None)
            _label = "Tất cả đơn vị" if _tat_ca else pgd_dd

            if not kq_xa or task_sel is None:
                st.info(f"{_label} không có xã nào trong đầu việc này.")
            else:
                xong = sum(1 for r in kq_xa if r["trang_thai"] == "da_hoan_thanh")
                st.caption(
                    f"**{_label}** — {task_sel['tieu_de']} · "
                    f"Hoàn thành: **{xong}/{len(kq_xa)}** · "
                    f"Thời hạn: {fmt_ngay(task_sel['ngay_deadline'])}"
                )
                _cols = (
                    ["Đơn vị", "Xã / Phường", "Trạng thái", "Ngày HT", "Ghi chú", "Người nhập"]
                    if _tat_ca
                    else ["Xã / Phường", "Trạng thái", "Ngày HT", "Ghi chú", "Người nhập"]
                )
                df_xa = pd.DataFrame([
                    {
                        **({"Đơn vị": r["pgd"]} if _tat_ca else {}),
                        "Xã / Phường": r["ten_xa"],
                        "Trạng thái": TS_KQ_LABEL.get(r["trang_thai"], r["trang_thai"]),
                        "Ngày HT": fmt_ngay(r.get("ngay_hoan_thanh")) or "—",
                        "Ghi chú": r.get("ghi_chu") or "",
                        "Người nhập": r.get("nguoi_nhap") or "",
                    }
                    for r in kq_xa
                ])
                st.dataframe(df_xa[_cols], use_container_width=True, hide_index=True)


def _render_tao_task(tab, **kwargs):
    username = kwargs.get("username", "")
    _tab_ctx = tab if tab is not None else __import__('streamlit').container()
    with _tab_ctx:
        st.subheader("➕ Tạo đầu việc mới")

        with st.expander("📖 Hướng dẫn tạo đầu việc", expanded=False):
            st.markdown("""
**1. Tên đầu việc** — Đặt tên ngắn gọn, rõ ràng.  
VD: *Nộp hồ sơ rủi ro tháng 5/2026*, *Khảo sát nhu cầu vay vốn Q2*

**2. Mô tả / Hướng dẫn** — Ghi rõ nội dung cần thực hiện, tài liệu tham khảo,
lưu ý đặc biệt để PGD/CBTD biết cần làm gì.

**3. Loại** — Phân loại đầu việc:
- 📋 Công việc chung — việc hành chính, tổng hợp
- 🎯 Chỉ tiêu KHTD — giao/điều chỉnh chỉ tiêu tín dụng cho PGD/xã
- 🗂️ Hồ sơ rủi ro — liên quan nợ xấu, NQH, xử lý rủi ro
- 📊 Khảo sát nhu cầu — điều tra nhu cầu vay vốn
- 📄 Báo cáo — các loại báo cáo định kỳ, thống kê
- 🎓 Tập huấn — đào tạo, hướng dẫn nghiệp vụ
- 📅 Giao dịch xã — tổ chức, kiểm tra hoạt động giao dịch xã
- 🤝 Hoạt động ủy thác — kiểm tra, giám sát 4 tổ chức CT-XH
- 💰 Nguồn vốn — theo dõi quỹ, điện báo xin vốn, huy động
- 📑 Ban đại diện HĐQT — phiên họp, nghị quyết, kiểm tra giám sát

**4. Loại theo dõi** — Quan trọng!
- 📍 **Chi tiết từng xã** — hệ thống tạo 1 dòng theo dõi cho mỗi xã/phường  
  → Dùng khi cần biết xã nào đã làm, xã nào chưa
- 🏢 **Chung PGD** — hệ thống tạo 1 dòng theo dõi cho mỗi PGD  
  → Dùng khi chỉ cần biết PGD đã hoàn thành chưa

**5. Ưu tiên** — 🔴 Khẩn cấp / 🟡 Quan trọng / 🟢 Bình thường  
Ảnh hưởng màu sắc hiển thị trong biểu đồ tổng quan.

**6. Ngày bắt đầu & Hạn hoàn thành** — Xác định khung thời gian.  
Sau thời hạn hệ thống tự đánh dấu 🔴 trễ hạn.

**7. Áp dụng cho đơn vị** — Mặc định tất cả 22 đơn vị.  
Bỏ chọn nếu chỉ áp dụng cho một số PGD cụ thể.

**8. Người phụ trách** — Ghi tên cán bộ chịu trách nhiệm theo dõi đầu việc này.
            """)

        with st.form("form_tao_task", clear_on_submit=False):
            tieu_de = st.text_area("Tên đầu việc *",
                                   placeholder="VD: Nộp hồ sơ rủi ro tháng 5/2026",
                                   height=68,
                                   key="tao_task_tieu_de")
            mo_ta = st.text_area("Mô tả / Hướng dẫn", key="tao_task_mo_ta")

            c1, c2 = st.columns(2)
            with c1:
                loai = st.selectbox("Loại", list(LOAI_TASK.keys()),
                                    format_func=lambda x: LOAI_TASK[x],
                                    key="tao_task_loai")
                loai_theo_doi = st.radio(
                    "Loại theo dõi",
                    options=["xa", "pgd"],
                    format_func=lambda x: "📍 Chi tiết từng xã" if x == "xa" else "🏢 Chung PGD",
                    horizontal=True,
                    key="tao_task_loai_theo_doi",
                )
                uu_tien = st.selectbox("Ưu tiên", list(UU_TIEN.keys()),
                                       format_func=lambda x: UU_TIEN[x], index=2,
                                       key="tao_task_uu_tien")
                nguoi_phu_trach = st.text_area(
                    "Người phụ trách",
                    placeholder="Tên người phụ trách chính",
                    height=68,
                    key="tao_task_nguoi_pt",
                )
                nguoi_thuc_hien_cn = st.text_input(
                    "👤 Cán bộ phòng KH-NV phụ trách",
                    placeholder="Họ tên cán bộ KH-NV phụ trách nội dung này...",
                    key="tao_task_nguoi_thuc_hien_cn",
                )
            with c2:
                deadline = st.date_input("Ngày kết thúc *", value=date.today(),
                                          format="DD/MM/YYYY", key="tao_task_deadline")
                ngay_bat_dau = st.date_input(
                    "Ngày bắt đầu",
                    value=date.today(),
                    format="DD/MM/YYYY",
                    key="tao_task_ngay_bat_dau",
                )

            st.markdown("**📋 Áp dụng cho đơn vị**")
            st.caption(
                "Chọn đơn vị chịu trách nhiệm thực hiện đầu việc này. "
                "Mỗi đơn vị được chọn sẽ có 1 dòng cập nhật tiến độ riêng. "
                "Có thể áp dụng đồng thời cho PGD huyện và Hội sở CN tỉnh."
            )
            st.markdown("**🏢 Phòng giao dịch trực thuộc**")
            _cc1, _cc2, _cc3 = st.columns(3)
            pgd_chon = []
            for _j, _pgd in enumerate(DS_PGD):
                _i = _j + 1
                with [_cc1, _cc2, _cc3][_j % 3]:
                    if st.checkbox(_pgd, value=True, key=f"tao_task_pgd_{_i}"):
                        pgd_chon.append(_pgd)

            ds_preview = pgd_chon or DS_PGD
            _so_chon = len(pgd_chon)
            _so_bo = len(DS_PGD) - _so_chon
            _thong_ke = f"✅ **{_so_chon}** đơn vị được chọn"
            if _so_bo > 0:
                _khong_chon = [p for p in DS_PGD if p not in pgd_chon]
                _thong_ke += f" · ❌ **{_so_bo}** không chọn: {', '.join(_khong_chon)}"
            st.caption(_thong_ke)
            st.caption(
                "Tick chọn PGD nào phải thực hiện. "
                "Trưởng/Phó PGD sẽ cập nhật tiến độ cho đơn vị mình."
            )

            st.divider()

            st.markdown("**🏛️ Hội sở CN tỉnh**")
            cbtd_bien_hoa = st.text_input(
                "Cán bộ KH-NV phụ trách",
                placeholder="Họ tên CBTD Hội sở phụ trách địa bàn Biên Hòa...",
                key="tao_task_cbtd_bien_hoa",
            )
            st.caption(
                "Địa bàn TP. Biên Hòa không có PGD riêng — CBTD tại Hội sở CN tỉnh "
                "trực tiếp quản lý. Điền tên cán bộ KH-NV phụ trách nếu đầu việc "
                "này áp dụng cho địa bàn Biên Hòa. Để trống nếu không áp dụng."
            )
            if loai_theo_doi == "pgd":
                st.caption(f"🏢 {len(ds_preview)} đơn vị — theo dõi cấp PGD")
            else:
                tong_xa_preview = sum(
                    len(PGD_XA_MAP.get(p, [])) for p in ds_preview
                )
                st.caption(
                    f"📍 {len(ds_preview)} đơn vị · {tong_xa_preview} xã/phường"
                )
            submitted = st.form_submit_button("💾 Tạo đầu việc", type="primary")

        if submitted:
            if not tieu_de.strip():
                st.error("Vui lòng nhập tên đầu việc.")
                return
            pgd_luu = pgd_chon or DS_PGD
            ds_pgd_json = json.dumps(pgd_luu, ensure_ascii=False)
            try:
                with db.get_conn() as conn:
                    cur = conn.execute(
                        """INSERT INTO tien_do_task
                           (tieu_de, mo_ta, ngay_deadline, ds_pgd, loai,
                            uu_tien, nguoi_tao, ngay_tao, trang_thai, ghi_chu,
                            cap_theo_doi, ngay_bat_dau, nguoi_phu_trach, nguoi_thuc_hien_cn,
                            cbtd_bien_hoa)
                           VALUES (?,?,?,?,?,?,?,?,'dang_theo_doi',?,?,?,?,?,?)""",
                        (tieu_de.strip(), mo_ta.strip() or None,
                         deadline.isoformat(), ds_pgd_json,
                         loai, uu_tien, username,
                         datetime.now().isoformat(),
                         None,
                         loai_theo_doi,
                         ngay_bat_dau.isoformat(),
                         nguoi_phu_trach.strip() or None,
                         str(nguoi_thuc_hien_cn).strip(),
                         str(cbtd_bien_hoa).strip()),
                    )
                    task_id = cur.lastrowid
                    conn.commit()

                loai_noi_dung = "chung_pgd" if loai_theo_doi == "pgd" else "chi_tiet_xa"
                _khoi_tao_ketqua_task(task_id, pgd_luu, loai_theo_doi, loai_noi_dung)
                _sync_bien_hoa_ketqua(task_id, cbtd_bien_hoa, loai_noi_dung)

                so_xa = (
                    sum(len(PGD_XA_MAP.get(p, [])) for p in pgd_luu)
                    if loai_theo_doi == "xa" else len(pgd_luu)
                )
                db.ghi_audit(username, "tien_do_tao_task",
                             f"'{tieu_de}' · thời hạn={deadline} · "
                             f"{len(pgd_luu)} PGD · "
                             f"{so_xa} đơn vị {loai_theo_doi} · "
                             f"cap_theo_doi={loai_theo_doi} · "
                             f"cbtd_bien_hoa={str(cbtd_bien_hoa).strip()}")
                st.toast(f"✅ Đã tạo: {tieu_de}")
                st.rerun()
            except Exception as e:
                st.error(f"Lỗi: {e}")


def _render_quan_ly_task(tab, **kwargs):
    username = kwargs.get("username", "")
    role_raw = str(kwargs.get("role", "user") or "user")
    role = normalize_role(role_raw)

    _tab_ctx = tab if tab is not None else __import__("streamlit").container()
    with _tab_ctx:
        st.subheader("✏️ Chỉnh sửa & Xóa đầu việc")

        ds_task = _doc_tasks(chi_dang_theo_doi=False)
        if not ds_task:
            st.info("Chưa có đầu việc nào.")
            return

        task_map = {t["id"]: t for t in ds_task}

        def _fmt_task(task_id: int) -> str:
            t = task_map.get(task_id) or {}
            if t.get("trang_thai") == "dang_theo_doi":
                try:
                    dl = date.fromisoformat(t.get("ngay_deadline", ""))
                    stt = "⚠️ Đã hết hạn" if dl < date.today() else "🟢 Đang thực hiện"
                except Exception:
                    stt = "🟢 Đang thực hiện"
            else:
                stt = "🔒 Đã đóng"
            parts = [stt, t.get("tieu_de", "")]
            ngay_bd = t.get("ngay_bat_dau")
            if ngay_bd:
                parts.append(f"BĐ: {fmt_ngay(ngay_bd)}")
            parts.append(f"KT: {fmt_ngay(t.get('ngay_deadline', ''))}")
            return " · ".join(parts)

        task_id = st.selectbox(
            "Chọn đầu việc",
            options=list(task_map.keys()),
            format_func=_fmt_task,
            key="td_ql_task",
        )
        task = task_map.get(task_id)
        if not task:
            return

        try:
            deadline_default = date.fromisoformat(str(task.get("ngay_deadline") or date.today().isoformat()))
        except Exception:
            deadline_default = date.today()

        loai_keys = list(LOAI_TASK.keys())
        uu_tien_keys = list(UU_TIEN.keys())

        try:
            loai_index = loai_keys.index(task.get("loai")) if task.get("loai") in loai_keys else 0
        except Exception:
            loai_index = 0

        try:
            uu_tien_index = uu_tien_keys.index(task.get("uu_tien")) if task.get("uu_tien") in uu_tien_keys else 2
        except Exception:
            uu_tien_index = 2

        with st.form("form_sua_task"):
            tieu_de = st.text_area(
                "Tên đầu việc *",
                value=str(task.get("tieu_de") or ""),
                height=68,
                key=f"td_sua_tieu_de_{task_id}",
            )
            mo_ta = st.text_area(
                "Mô tả / Hướng dẫn",
                value=str(task.get("mo_ta") or ""),
                key=f"td_sua_mo_ta_{task_id}",
            )
            loai = st.selectbox(
                "Loại",
                options=loai_keys,
                format_func=lambda x: LOAI_TASK.get(x, x),
                index=loai_index,
                key=f"td_sua_loai_{task_id}",
            )
            uu_tien = st.selectbox(
                "Ưu tiên",
                options=uu_tien_keys,
                format_func=lambda x: UU_TIEN.get(x, x),
                index=uu_tien_index,
                key=f"td_sua_uu_tien_{task_id}",
            )
            deadline = st.date_input(
                "Ngày kết thúc *",
                value=deadline_default,
                format="DD/MM/YYYY",
                key=f"td_sua_deadline_{task_id}",
            )
            ghi_chu = st.text_area(
                "Ghi chú thêm",
                value=str(task.get("ghi_chu") or ""),
                height=68,
                key=f"td_sua_ghi_chu_{task_id}",
            )
            nguoi_phu_trach = st.text_area(
                "Người phụ trách",
                value=str(task.get("nguoi_phu_trach") or ""),
                height=68,
                key=f"td_sua_nguoi_pt_{task_id}",
            )
            nguoi_thuc_hien_cn = st.text_input(
                "👤 Cán bộ phòng KH-NV phụ trách",
                value=str(task.get("nguoi_thuc_hien_cn") or ""),
                placeholder="Họ tên cán bộ KH-NV phụ trách nội dung này...",
                key=f"td_sua_nguoi_thuc_hien_cn_{task_id}",
            )
            ngay_bat_dau_val = date.today()
            try:
                ngay_bat_dau_val = date.fromisoformat(
                    str(task.get("ngay_bat_dau") or date.today().isoformat())
                )
            except Exception:
                pass
            ngay_bat_dau = st.date_input(
                "Ngày bắt đầu",
                value=ngay_bat_dau_val,
                format="DD/MM/YYYY",
                key=f"td_sua_ngay_bat_dau_{task_id}",
            )
            cap_theo_doi_keys = ["xa", "pgd"]
            cap_theo_doi_idx = cap_theo_doi_keys.index(
                task.get("cap_theo_doi", "xa")
            ) if task.get("cap_theo_doi") in cap_theo_doi_keys else 0
            cap_theo_doi = st.radio(
                "Loại theo dõi",
                options=cap_theo_doi_keys,
                format_func=lambda x: "📍 Chi tiết từng xã" if x == "xa" else "🏢 Chung PGD",
                index=cap_theo_doi_idx,
                horizontal=True,
                key=f"td_sua_cap_theo_doi_{task_id}",
            )

            st.markdown("**📋 Áp dụng cho đơn vị**")
            st.caption(
                "Chọn đơn vị chịu trách nhiệm thực hiện đầu việc này. "
                "Mỗi đơn vị được chọn sẽ có 1 dòng cập nhật tiến độ riêng. "
                "Có thể áp dụng đồng thời cho PGD huyện và Hội sở CN tỉnh."
            )
            st.markdown("**🏢 Phòng giao dịch trực thuộc**")
            ds_pgd_task = json.loads(task.get("ds_pgd") or "[]") or DS_PGD
            _c1, _c2, _c3 = st.columns(3)
            for _j, _pgd in enumerate(DS_PGD):
                _i = _j + 1
                with [_c1, _c2, _c3][_j % 3]:
                    st.checkbox(
                        _pgd,
                        value=_pgd in ds_pgd_task,
                        key=f"td_sua_pgd_{task_id}_{_i}",
                        disabled=True,
                    )
            st.caption(
                "Tick chọn PGD nào phải thực hiện. "
                "Trưởng/Phó PGD sẽ cập nhật tiến độ cho đơn vị mình. "
                "Danh sách PGD không thể thay đổi sau khi tạo — nếu cần, hãy đóng đầu việc này và tạo mới."
            )

            st.divider()

            st.markdown("**🏛️ Hội sở CN tỉnh**")
            cbtd_bien_hoa = st.text_input(
                "Cán bộ KH-NV phụ trách",
                value=str(task.get("cbtd_bien_hoa") or ""),
                placeholder="Họ tên CBTD Hội sở phụ trách địa bàn Biên Hòa...",
                key=f"td_sua_cbtd_bien_hoa_{task_id}",
            )
            st.caption(
                "Địa bàn TP. Biên Hòa không có PGD riêng — CBTD tại Hội sở CN tỉnh "
                "trực tiếp quản lý. Điền tên cán bộ KH-NV phụ trách nếu đầu việc "
                "này áp dụng cho địa bàn Biên Hòa. Để trống nếu không áp dụng."
            )

            submitted = st.form_submit_button("💾 Lưu thay đổi", type="primary")

        if submitted:
            if not str(tieu_de or "").strip():
                st.error("Vui lòng nhập tên đầu việc.")
                return
            try:
                with db.get_conn() as conn:
                    conn.execute(
                        """UPDATE tien_do_task
                           SET tieu_de=?, mo_ta=?, ngay_deadline=?, loai=?,
                               uu_tien=?, ghi_chu=?,
                               cap_theo_doi=?, ngay_bat_dau=?, nguoi_phu_trach=?,
                               nguoi_thuc_hien_cn=?,
                               cbtd_bien_hoa=?
                           WHERE id=?""",
                        (
                            str(tieu_de).strip(),
                            str(mo_ta).strip() or None,
                            deadline.isoformat(),
                            loai,
                            uu_tien,
                            str(ghi_chu).strip() or None,
                            cap_theo_doi,
                            ngay_bat_dau.isoformat(),
                            str(nguoi_phu_trach).strip() or None,
                            str(nguoi_thuc_hien_cn).strip(),
                            str(cbtd_bien_hoa).strip(),
                            task_id,
                        ),
                    )
                    conn.commit()
                loai_noi_dung = "chung_pgd" if cap_theo_doi == "pgd" else "chi_tiet_xa"
                _sync_bien_hoa_ketqua(task_id, cbtd_bien_hoa, loai_noi_dung)
                db.ghi_audit(
                    username,
                    "tien_do_sua_task",
                    f"ID={task_id} · '{str(tieu_de).strip()}' · thời hạn={deadline}",
                )
                db.ghi_audit(
                    username,
                    "sua_task",
                    f"Task #{task_id}: cbtd_bien_hoa={str(cbtd_bien_hoa).strip()}",
                )
                st.toast("✅ Đã lưu thay đổi.")
                st.rerun()
            except Exception as e:
                st.error(f"Lỗi: {e}")

        c1, c2, c3 = st.columns([1.2, 1.2, 1.6])
        with c1:
            if task.get("trang_thai") == "dang_theo_doi":
                if st.button("🔒 Đóng đầu việc", key=f"td_dong_{task_id}", use_container_width=True):
                    try:
                        with db.get_conn() as conn:
                            conn.execute(
                                "UPDATE tien_do_task SET trang_thai=? WHERE id=?",
                                ("da_dong", task_id),
                            )
                            conn.commit()
                        db.ghi_audit(
                            username,
                            "tien_do_doi_trang_thai",
                            f"ID={task_id} · '{task.get('tieu_de')}' → da_dong",
                        )
                        st.toast("✅ Đã cập nhật trạng thái.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Lỗi: {e}")
            else:
                if st.button("🔓 Mở lại", key=f"td_mo_lai_{task_id}", use_container_width=True):
                    try:
                        with db.get_conn() as conn:
                            conn.execute(
                                "UPDATE tien_do_task SET trang_thai=? WHERE id=?",
                                ("dang_theo_doi", task_id),
                            )
                            conn.commit()
                        db.ghi_audit(
                            username,
                            "tien_do_doi_trang_thai",
                            f"ID={task_id} · '{task.get('tieu_de')}' → dang_theo_doi",
                        )
                        st.toast("✅ Đã cập nhật trạng thái.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Lỗi: {e}")

        if la_admin_cn(role):
            confirm_key = f"_td_xoa_confirm_{task_id}"
            with c3:
                if not st.session_state.get(confirm_key):
                    if st.button("🗑️ Xóa vĩnh viễn", key=f"td_xoa_{task_id}", use_container_width=True):
                        st.session_state[confirm_key] = True
                        st.rerun()
                else:
                    st.warning("⚠️ Hành động này không thể hoàn tác.")
                    if st.button("Xác nhận xóa", key=f"td_xoa_ok_{task_id}", type="primary", use_container_width=True):
                        try:
                            with db.get_conn() as conn:
                                conn.execute("DELETE FROM tien_do_ketqua WHERE task_id=?", (task_id,))
                                conn.execute("DELETE FROM tien_do_task WHERE id=?", (task_id,))
                                conn.commit()
                            db.ghi_audit(
                                username,
                                "tien_do_xoa_task",
                                f"ID={task_id} · '{task.get('tieu_de')}'",
                            )
                            st.session_state.pop(confirm_key, None)
                            st.toast("🗑️ Đã xóa đầu việc.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Lỗi: {e}")


def _render_cap_nhat(tab, **kwargs):
    username = kwargs.get("username", "")
    role = kwargs.get("role", "user")
    pgd_user = kwargs.get("pgd_user") or ""

    _tab_ctx = tab if tab is not None else __import__('streamlit').container()
    with _tab_ctx:
        ds_task = _doc_tasks()
        if not ds_task:
            st.info("Chưa có đầu việc nào đang theo dõi.")
            return

        # ── Khối 1: CHỌN ĐẦU VIỆC ────────────────────────────────────────
        with st.container(border=True):
            st.markdown("**① CHỌN ĐẦU VIỆC**")

            hom_nay = date.today()
            def _fmt_cap_nhat_opt(t):
                try:
                    dl = date.fromisoformat(t["ngay_deadline"])
                    stt = "⚠️ Đã hết hạn" if dl < hom_nay else "🟢 Đang thực hiện"
                except Exception:
                    stt = "🟢 Đang thực hiện"
                ngay_bd = t.get("ngay_bat_dau")
                parts = [stt, t["tieu_de"]]
                if ngay_bd:
                    parts.append(f"BĐ: {fmt_ngay(ngay_bd)}")
                parts.append(f"⏰ KT: {fmt_ngay(t['ngay_deadline'])}")
                return " · ".join(parts)

            task_opts = {t["id"]: _fmt_cap_nhat_opt(t) for t in ds_task}
            task_id = st.selectbox(
                "Đầu việc",
                list(task_opts.keys()),
                format_func=lambda x: task_opts[x],
                key="td_cu_task",
            )
            task = next((t for t in ds_task if t["id"] == task_id), None)
            if task is None:
                return

            cap_theo_doi = task.get("cap_theo_doi", "xa")
            tag_nd = "🏢 Chung PGD" if cap_theo_doi == "pgd" else "🏘️ Chi tiết xã"
            nguoi_pt = task.get("nguoi_phu_trach") or ""
            nguoi_thuc_hien_cn = task.get("nguoi_thuc_hien_cn") or ""
            cbtd_bien_hoa = task.get("cbtd_bien_hoa") or ""

            try:
                deadline_date = date.fromisoformat(task["ngay_deadline"])
                if deadline_date < hom_nay:
                    badge = "⚠️ Đã hết hạn"
                elif (deadline_date - hom_nay).days <= 3:
                    badge = "🟠 Sắp hết hạn (≤ 3 ngày)"
                else:
                    badge = "🟢 Đang thực hiện"
            except Exception:
                badge = ""

            st.caption(
                f"**{task['tieu_de']}** · "
                f"{badge} · "
                f"{tag_nd} · "
                f"{LOAI_TASK.get(task['loai'], task['loai'])} · "
                f"{UU_TIEN.get(task['uu_tien'], '')}"
            )
            c1, c2, c3 = st.columns(3)
            with c1:
                st.caption(f"📅 Ngày bắt đầu: **{fmt_ngay(task.get('ngay_bat_dau') or '—')}**")
            with c2:
                st.caption(f"⏰ Ngày kết thúc: **{fmt_ngay(task['ngay_deadline'])}**")
            with c3:
                if nguoi_pt:
                    st.caption(f"👤 Người PT: **{nguoi_pt}**")
            if nguoi_thuc_hien_cn:
                st.caption(f"👤 Cán bộ KH-NV phụ trách: {nguoi_thuc_hien_cn}")
            if cbtd_bien_hoa:
                st.caption(f"🏛️ Hội sở CN tỉnh — CB KH-NV: {cbtd_bien_hoa}")
            if badge:
                st.info(badge)
            if task.get("mo_ta"):
                st.info(task["mo_ta"])

        pgd_user_val = kwargs.get("pgd_user") or ""
        la_pgd_role = bool(pgd_user_val) and role not in ROLES_PHAN_HE_CN

        # ── Khối 2: CHỌN PHẠM VI ─────────────────────────────────────────
        if cap_theo_doi == "xa":
            if la_pgd_role:
                pgd_sel = pgd_user_val
            else:
                with st.container(border=True):
                    st.markdown("**② CHỌN PHẠM VI**")
                    ds_pgd_task = json.loads(task.get("ds_pgd") or "[]") or DS_PGD
                    if cbtd_bien_hoa:
                        ds_pgd_task = list(ds_pgd_task) + [_PGD_BIEN_HOA]
                    pgd_sel = st.selectbox(
                        "Đơn vị PGD",
                        options=ds_pgd_task,
                        key=f"td_cap_nhat_pgd_{task_id}",
                    )
                    if pgd_sel:
                        st.caption(f"🏘️ Đang xem: **{pgd_sel}**")
            if not pgd_sel:
                st.warning("Tài khoản chưa được gán đơn vị.")
                return
            kq_list = [r for r in _doc_ketqua_task(task_id) if r["pgd"] == pgd_sel]
            ten_cot = "Xã / Phường"
            label_dv = "xã"
        else:
            kq_list = _doc_ketqua_task(task_id)
            if la_pgd_role:
                kq_list = [r for r in kq_list if r["pgd"] == pgd_user_val]
            pgd_sel = "__ALL__"
            ten_cot = "Đơn vị PGD"
            label_dv = "PGD"
            st.caption(f"🏢 Theo dõi chung **{len(kq_list)}** {label_dv}")

        if not kq_list:
            st.warning(f"Không tìm thấy dữ liệu {label_dv}. Thử tạo lại đầu việc.")
            return

        xong = sum(1 for r in kq_list if r["trang_thai"] == "da_hoan_thanh")
        tong = len(kq_list)
        pct = round(xong / tong * 100) if tong else 0

        # ── Khối 3: CẬP NHẬT TIẾN ĐỘ ─────────────────────────────────────
        with st.container(border=True):
            st.markdown("**③ CẬP NHẬT TIẾN ĐỘ**")

            # Progress + thống kê
            st.progress(
                xong / tong,
                text=f"✅ **{xong}/{tong}** {label_dv} hoàn thành  ·  "
                     f"⬜ **{tong - xong}** chưa thực hiện  ·  **{pct}%**",
            )

            confirm_key = f"_td_confirm_save_{task_id}_{pgd_sel}"
            editor_key = f"td_editor_{task_id}_{pgd_sel}"

            an_da_ht = st.toggle("Ẩn đơn vị đã hoàn thành", key=f"td_an_ht_{task_id}_{pgd_sel}")

            editor_key = f"td_editor_{task_id}_{pgd_sel}_{an_da_ht}"

            def _parse_date(val):
                if not val:
                    return None
                try:
                    return pd.to_datetime(val).date()
                except Exception:
                    return None

            kq_hien_thi = kq_list
            if an_da_ht:
                kq_hien_thi = [r for r in kq_list if r["trang_thai"] != "da_hoan_thanh"]

            df_edit = pd.DataFrame([
                {
                    ten_cot: r["ten_xa"],
                    "Trạng thái": TS_KQ_LABEL.get(r["trang_thai"], r["trang_thai"]),
                    "Hoàn thành": r["trang_thai"] == "da_hoan_thanh",
                    "Ngày hoàn thành": _parse_date(r.get("ngay_hoan_thanh")),
                    "Ghi chú": r.get("ghi_chu") or "",
                }
                for r in kq_hien_thi
            ])

            # Action bar: Lưu + Hoàn tác + thống kê
            c_save, c_reset, c_info = st.columns([1.5, 1, 4])
            with c_save:
                if st.session_state.get(confirm_key):
                    st.warning("⚠️ Xác nhận lưu?")
                    c_ok, c_huy = st.columns(2)
                    with c_ok:
                        if st.button("✅ Xác nhận", type="primary", use_container_width=True,
                                     key=f"{confirm_key}_ok"):
                            state = st.session_state.get(editor_key, {})
                            edited = df_edit.copy()
                            if isinstance(state, dict):
                                for row_idx, changes in state.get("edited_rows", {}).items():
                                    for col, val in changes.items():
                                        edited.at[int(row_idx), col] = val
                            elif isinstance(state, pd.DataFrame):
                                edited = state
                            if edited.empty:
                                st.session_state.pop(confirm_key, None)
                                st.rerun()
                            count = 0
                            for i in range(len(edited)):
                                ten_xa_dv = edited.iloc[i][ten_cot]
                                trang_thai = "da_hoan_thanh" if edited.iloc[i]["Hoàn thành"] else "chua_thuc_hien"
                                ngay_val = edited.iloc[i]["Ngày hoàn thành"]
                                ngay_ht = None
                                if pd.notna(ngay_val) and ngay_val:
                                    try:
                                        ngay_ht = pd.to_datetime(ngay_val).date().isoformat()
                                    except Exception:
                                        ngay_ht = None
                                pgd_val = ten_xa_dv if cap_theo_doi == "pgd" else pgd_sel
                                try:
                                    _upsert_ketqua_xa(
                                        task_id, ten_xa_dv, pgd_val,
                                        trang_thai, ngay_ht,
                                        str(edited.iloc[i]["Ghi chú"]).strip() or None,
                                        username,
                                    )
                                    count += 1
                                except Exception as e:
                                    st.warning(f"Lỗi {ten_xa_dv}: {e}")
                            db.ghi_audit(username, "tien_do_cap_nhat_xa",
                                         f"Task '{task['tieu_de']}' · {count} {label_dv}")
                            st.session_state.pop(editor_key, None)
                            st.session_state.pop(confirm_key, None)
                            st.session_state.pop(f"{editor_key}_data", None)
                            st.toast(f"✅ Đã lưu {count} {label_dv}.")
                            st.rerun()
                    with c_huy:
                        if st.button("❌ Hủy", use_container_width=True,
                                     key=f"{confirm_key}_cancel"):
                            st.session_state.pop(confirm_key, None)
                            st.rerun()
                else:
                    if st.button("💾 Lưu thay đổi", type="primary", use_container_width=True,
                                 key=f"{confirm_key}_btn"):
                        st.session_state[confirm_key] = True
                        st.rerun()
            with c_reset:
                if st.button("↩️ Hoàn tác", use_container_width=True,
                             key=f"_td_undo_{task_id}_{pgd_sel}"):
                    base = f"td_editor_{task_id}_{pgd_sel}"
                    for k in list(st.session_state.keys()):
                        if k.startswith(base) or k.startswith(confirm_key):
                            st.session_state.pop(k, None)
                    st.rerun()
            with c_info:
                ht = tong - xong
                _thong_ke = f"✅ {xong} · ⬜ {ht} · {pct}%"
                if an_da_ht:
                    _thong_ke += f" · 👁 Đang ẩn {xong} đã hoàn thành"
                _thong_ke += " · Tick ☑ để đánh dấu hoàn thành"
                st.caption(_thong_ke)

            # Data editor — cache dữ liệu vào session_state để nút lưu dùng được
            edited = st.data_editor(
                df_edit,
                column_config={
                    ten_cot: st.column_config.TextColumn(ten_cot, disabled=True),
                    "Trạng thái": st.column_config.TextColumn(
                        "Trạng thái", disabled=True, width="small",
                    ),
                    "Hoàn thành": st.column_config.CheckboxColumn("✅ Hoàn thành"),
                    "Ngày hoàn thành": st.column_config.DateColumn(
                        "Ngày hoàn thành", format="DD/MM/YYYY", default=None,
                    ),
                    "Ghi chú": st.column_config.TextColumn(width="large"),
                },
                hide_index=True,
                use_container_width=True,
                key=editor_key,
            )
            st.session_state[f"{editor_key}_data"] = edited

def _xuat_excel_tien_do(df_tonghop, df_matran, df_ct) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_tonghop.to_excel(writer, sheet_name="Tổng hợp", index=False)
        df_matran.to_excel(writer, sheet_name="Ma trận PGD", index=False)
        df_ct.to_excel(writer, sheet_name="Chi tiết xã", index=False)

        header_fill = PatternFill("solid", fgColor="003D7A")
        header_font = Font(bold=True, color="FFFFFF", size=10, name="Times New Roman")
        body_font = Font(size=10, name="Times New Roman")
        body_bold = Font(bold=True, size=10, name="Times New Roman")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="BDBDBD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill = PatternFill("solid", fgColor="EEF4FB")

        def _style_sheet(ws, col_left, col_right, col_center, col_pct):
            col_names = [cell.value for cell in ws[1]]

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cn = col_names[cell.column - 1]
                    cell.border = border
                    cell.font = body_font
                    if cn in col_left:
                        cell.alignment = left
                    elif cn in col_right:
                        cell.alignment = right
                        cell.number_format = "#,##0"
                    elif cn in col_center:
                        cell.alignment = center
                    elif cn in col_pct:
                        cell.alignment = center
                        cell.number_format = "0.0\"%\""

            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
                if i % 2 == 0:
                    for cell in row:
                        if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "003D7A"):
                            cell.fill = alt_fill

            for col_idx in range(1, len(col_names) + 1):
                max_len = max(
                    (len(str(cell.value or "")) for cell in ws[ws.min_row:ws.max_row][:, col_idx - 1]),
                    default=15,
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

        # Sheet 1: Tổng hợp
        ws_th = writer.sheets["Tổng hợp"]
        _style_sheet(
            ws_th,
            col_left={"Đầu việc", "Loại", "Người phụ trách", "CB KH-NV phụ trách", "Hội sở CN tỉnh"},
            col_right={"Số PGD", "Tổng xã", "Đã hoàn thành", "Chưa thực hiện", "Trễ hạn", "N/A"},
            col_center={"STT", "Ưu tiên", "Thời hạn", "Loại theo dõi", "Ngày bắt đầu"},
            col_pct={"Tỷ lệ HT%"},
        )

        # Sheet 2: Ma trận PGD
        ws_mt = writer.sheets["Ma trận PGD"]
        col_names_mt = [cell.value for cell in ws_mt[1]]
        for cell in ws_mt[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        for row in ws_mt.iter_rows(min_row=2, max_row=ws_mt.max_row):
            for cell in row:
                cn = col_names_mt[cell.column - 1]
                cell.border = border
                cell.font = body_font
                if cn == "Đơn vị":
                    cell.alignment = left
                    cell.font = body_bold
                else:
                    cell.alignment = center
        for i, row in enumerate(ws_mt.iter_rows(min_row=2, max_row=ws_mt.max_row), start=1):
            if i % 2 == 0:
                for cell in row:
                    if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "003D7A"):
                        cell.fill = alt_fill
        for col_idx in range(1, len(col_names_mt) + 1):
            max_len = max(
                (len(str(cell.value or "")) for cell in ws_mt[ws_mt.min_row:ws_mt.max_row][:, col_idx - 1]),
                default=12,
            )
            ws_mt.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 28)

        # Sheet 3: Chi tiết xã
        ws_ct = writer.sheets["Chi tiết xã"]
        _style_sheet(
            ws_ct,
            col_left={"Đầu việc", "PGD", "Xã / Phường", "Ghi chú"},
            col_right=set(),
            col_center={"Task ID", "Thời hạn", "Trạng thái", "Ngày hoàn thành"},
            col_pct=set(),
        )

    buf.seek(0)
    return buf.getvalue()


def _render_xuat(tab, **kwargs):
    SS_KEY = "_td_xuat_excel"
    username = kwargs.get("username", "")
    _tab_ctx = tab if tab is not None else __import__('streamlit').container()
    with _tab_ctx:
        st.subheader("📤 Xuất báo cáo tiến độ")

        c1, c2 = st.columns(2)
        with c1:
            tu_ngay = st.date_input("Thời hạn từ", value=date.today(), format="DD/MM/YYYY", key="td_x1")
        with c2:
            den_ngay = st.date_input("Thời hạn đến", value=date.today(), format="DD/MM/YYYY", key="td_x2")

        SS_PDF_BAOCAO = "_td_pdf_baocao"

        if st.button("📄 Xuất PDF báo cáo tiến độ", type="primary", key="td_btn_pdf_baocao"):
            ds_task = _doc_tasks(chi_dang_theo_doi=False)
            ds_task = [t for t in ds_task
                       if tu_ngay.isoformat() <= t["ngay_deadline"] <= den_ngay.isoformat()]
            if not ds_task:
                st.info("Không có đầu việc trong khoảng thời gian đã chọn.")
            else:
                pdf_buf = _xuat_pdf_bao_cao_tien_do(ds_task, username)
                if pdf_buf:
                    st.session_state[SS_PDF_BAOCAO] = {
                        "data": pdf_buf.getvalue(),
                        "filename": f"baocao_tiendo_{datetime.now().strftime('%Y%m%d')}.pdf",
                    }
                    db.ghi_audit(username, "tien_do_xuat_pdf_baocao",
                                 f"{len(ds_task)} đầu việc")
                    st.rerun()

        if st.session_state.get(SS_PDF_BAOCAO):
            _p = st.session_state[SS_PDF_BAOCAO]
            st.download_button(
                "⬇ Tải PDF báo cáo tiến độ",
                data=_p["data"],
                file_name=_p["filename"],
                mime="application/pdf",
                key="td_pdf_baocao_dl",
                use_container_width=True,
            )

        if st.button("📥 Tạo Excel", type="primary", key="td_btn_tao"):
            ds_task = _doc_tasks(chi_dang_theo_doi=False)
            ds_task = [t for t in ds_task
                       if tu_ngay.isoformat() <= t["ngay_deadline"] <= den_ngay.isoformat()]

            if not ds_task:
                st.session_state.pop(SS_KEY, None)
                st.info("Không có đầu việc trong khoảng thời gian đã chọn.")
                st.stop()

            hom_nay = date.today().isoformat()

            # Sheet 0: Tổng hợp đầu việc
            summary_rows = []
            for i, t in enumerate(ds_task, 1):
                kq = _doc_ketqua_task(t["id"])
                xong  = sum(1 for r in kq if r["trang_thai"] == "da_hoan_thanh")
                chua  = sum(1 for r in kq if r["trang_thai"] == "chua_thuc_hien")
                tre   = sum(1 for r in kq
                            if r["trang_thai"] == "chua_thuc_hien"
                            and t["ngay_deadline"] < hom_nay)
                na    = sum(1 for r in kq if r["trang_thai"] == "khong_ap_dung")
                tong  = len(kq)
                ds_p  = json.loads(t.get("ds_pgd") or "[]")
                summary_rows.append({
                    "STT":           i,
                    "Đầu việc":      t["tieu_de"],
                    "Loại":          LOAI_TASK.get(t["loai"], t["loai"]),
                    "Ưu tiên":       UU_TIEN.get(t["uu_tien"], t["uu_tien"]).replace("🔴","").replace("🟡","").replace("🟢","").strip(),
                    "Thời hạn":      t["ngay_deadline"],
                    "Số PGD":        len(ds_p),
                    "Tổng xã":       tong,
                    "Đã hoàn thành": xong,
                    "Chưa thực hiện": chua - tre,
                    "Trễ hạn":       tre,
                    "N/A":           na,
                    "Tỷ lệ HT%":    round(xong / tong * 100, 1) if tong else 0,
                    "Loại theo dõi": "Chung PGD" if t.get("cap_theo_doi") == "pgd" else "Chi tiết xã",
                    "Ngày bắt đầu": t.get("ngay_bat_dau") or "",
                    "Người phụ trách": t.get("nguoi_phu_trach") or "",
                    "CB KH-NV phụ trách": t.get("nguoi_thuc_hien_cn") or "",
                    "Hội sở CN tỉnh": t.get("cbtd_bien_hoa") or "",
                })
            df_tonghop = pd.DataFrame(summary_rows)

            # Sheet 1: Ma trận PGD × đầu việc
            bang_rows = []
            for pgd in DS_PGD_ALL:
                row = {"Đơn vị": pgd}
                for t in ds_task:
                    kq = [r for r in _doc_ketqua_task(t["id"]) if r["pgd"] == pgd]
                    xong = sum(1 for r in kq if r["trang_thai"] == "da_hoan_thanh")
                    tre = sum(1 for r in kq
                              if r["trang_thai"] == "chua_thuc_hien"
                              and t["ngay_deadline"] < hom_nay)
                    row[f"{t['tieu_de'][:18]} (#{t['id']})"] = f"{xong}/{len(kq)}" + ("🔴" if tre else "")
                bang_rows.append(row)
            df_matran = pd.DataFrame(bang_rows)

            # Sheet 2: Chi tiết theo xã
            rows_ct = []
            for t in ds_task:
                kq = _doc_ketqua_task(t["id"])
                for r in kq:
                    rows_ct.append({
                        "Task ID": t["id"],
                        "Đầu việc": t["tieu_de"],
                        "Thời hạn": t["ngay_deadline"],
                        "PGD": r["pgd"],
                        "Xã / Phường": r["ten_xa"],
                        "Trạng thái": TS_KQ_LABEL.get(r["trang_thai"], r["trang_thai"]),
                        "Ngày hoàn thành": r.get("ngay_hoan_thanh") or "",
                        "Ghi chú": r.get("ghi_chu") or "",
                    })
            df_ct = pd.DataFrame(rows_ct)

            excel_data = _xuat_excel_tien_do(df_tonghop, df_matran, df_ct)

            st.session_state[SS_KEY] = {
                "data": excel_data,
                "filename": f"TienDoCongViec_{date.today().isoformat()}.xlsx",
                "n_task": len(ds_task),
                "n_ct": len(df_ct),
                "n_th": len(df_tonghop),
            }
            st.rerun()

        if SS_KEY in st.session_state:
            payload = st.session_state[SS_KEY]
            col_dl, col_clear = st.columns([4, 1])
            with col_dl:
                st.download_button(
                    label="⬇ Tải Excel báo cáo tiến độ",
                    data=payload["data"],
                    file_name=payload["filename"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="td_xuat_dl",
                )
            with col_clear:
                if st.button("✕", key="td_xuat_clear", help="Tạo lại"):
                    del st.session_state[SS_KEY]
                    st.rerun()
            st.success(f"Đã xuất {payload['n_task']} đầu việc · "
                       f"{payload['n_ct']} dòng chi tiết · "
                       f"{payload['n_th']} đầu việc tổng hợp.")

        st.divider()
        st.markdown("### 📄 Xuất PDF tiến độ theo task")

        ds_task_pdf = _doc_tasks(chi_dang_theo_doi=False)
        if not ds_task_pdf:
            st.info("Chưa có đầu việc nào.")
        else:
            task_map_pdf = {t["id"]: t for t in ds_task_pdf}

            def _fmt_task_pdf(task_id):
                t = task_map_pdf.get(task_id) or {}
                if t.get("trang_thai") == "dang_theo_doi":
                    try:
                        dl = date.fromisoformat(t.get("ngay_deadline", ""))
                        stt = "⚠️ Đã hết hạn" if dl < date.today() else "🟢 Đang thực hiện"
                    except Exception:
                        stt = "🟢 Đang thực hiện"
                else:
                    stt = "🔒 Đã đóng"
                parts = [stt, t.get("tieu_de", "")]
                ngay_bd = t.get("ngay_bat_dau")
                if ngay_bd:
                    parts.append(f"BĐ: {fmt_ngay(ngay_bd)}")
                parts.append(f"KT: {fmt_ngay(t.get('ngay_deadline', ''))}")
                return " · ".join(parts)

            task_id_pdf = st.selectbox(
                "Chọn đầu việc",
                options=list(task_map_pdf.keys()),
                format_func=_fmt_task_pdf,
                key="td_xuat_pdf_task",
            )
            task_pdf = task_map_pdf.get(task_id_pdf)

            if task_pdf and st.button(
                "📄 Xuất PDF tiến độ",
                type="primary",
                key="td_btn_xuat_pdf"
            ):
                kq_all = _doc_ketqua_task(task_id_pdf)
                pdf_bytes = _xuat_pdf_tien_do(task_pdf, kq_all, username)
                if pdf_bytes:
                    ten_file = (
                        f"TienDo_{task_pdf['tieu_de'][:20].replace(' ', '_')}_"
                        f"{date.today().isoformat()}.pdf"
                    )
                    st.session_state["_td_pdf_bytes"] = pdf_bytes
                    st.session_state["_td_pdf_file"] = ten_file
                    db.ghi_audit(
                        username, "tien_do_xuat_pdf",
                        f"Task '{task_pdf['tieu_de']}'"
                    )
                    st.rerun()

            if st.session_state.get("_td_pdf_bytes"):
                st.download_button(
                    "⬇ Tải PDF",
                    data=st.session_state["_td_pdf_bytes"],
                    file_name=st.session_state.get("_td_pdf_file", "TienDo.pdf"),
                    mime="application/pdf",
                    key="td_pdf_dl",
                    use_container_width=True,
                )




def _xuat_pdf_bao_cao_tien_do(ds_task, username):
    try:
        from pdf_service import _dang_ky_font, VBSP_GREEN, ROW_ALT, BORDER_COLOR, _REPORTLAB_READY
        if not _REPORTLAB_READY:
            st.error("Chưa cài thư viện reportlab. Chạy: pip install reportlab")
            return None
        _dang_ky_font()

        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable, PageBreak,
        )
        from io import BytesIO
        from datetime import date, datetime
        import json

        FONT_NORMAL = "TNR"
        FONT_BOLD = "TNR-Bold"
        FONT_FALLBACK = "Helvetica"

        fn = FONT_NORMAL if FONT_NORMAL else FONT_FALLBACK
        fb = FONT_BOLD if FONT_BOLD else FONT_FALLBACK

        hom_nay = date.today().isoformat()
        ngay_str = datetime.now().strftime("%d/%m/%Y")
        buf = BytesIO()
        margin = 1.5 * cm
        page_size = A4
        usable_w = page_size[0] - 2 * margin

        doc = SimpleDocTemplate(
            buf, pagesize=page_size,
            leftMargin=margin, rightMargin=margin,
            topMargin=margin, bottomMargin=1.5 * cm,
            title=f"Báo cáo tiến độ công việc {ngay_str}",
            author="VBSP-SCM",
        )

        story = []

        # ── Header ─────────────────────────────────────────────────────
        story.append(Paragraph(
            "NGÂN HÀNG CHÍNH SÁCH XÃ HỘI VIỆT NAM",
            ParagraphStyle("bank", fontName=fb, fontSize=13,
                           alignment=TA_CENTER, spaceAfter=2)
        ))
        story.append(Paragraph(
            "CHI NHÁNH TỈNH ĐỒNG NAI",
            ParagraphStyle("branch", fontName=fn, fontSize=11,
                           alignment=TA_CENTER, spaceAfter=4)
        ))
        story.append(HRFlowable(width="100%", thickness=1.5,
                                color=colors.HexColor("#2E7D32"), spaceAfter=6))
        story.append(Paragraph(
            f"BÁO CÁO TIẾN ĐỘ CÔNG VIỆC — {ngay_str}",
            ParagraphStyle("title", fontName=fb, fontSize=14,
                           alignment=TA_CENTER, spaceAfter=4,
                           textColor=colors.HexColor("#003D7A"))
        ))
        story.append(Paragraph(
            f"Người xuất: {username}  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            ParagraphStyle("meta", fontName=fn, fontSize=9,
                           alignment=TA_CENTER, spaceAfter=10,
                           textColor=colors.grey)
        ))

        # ── Phần 1: Tiến độ theo đầu việc ─────────────────────────────
        story.append(Paragraph(
            "PHẦN 1: TIẾN ĐỘ THEO ĐẦU VIỆC",
            ParagraphStyle("p1_title", fontName=fb, fontSize=11,
                           alignment=TA_LEFT, spaceBefore=4, spaceAfter=6,
                           textColor=colors.HexColor("#2E7D32"))
        ))

        task_header_s = ParagraphStyle("task_h", fontName=fb, fontSize=13,
                                       leading=17, spaceBefore=10, spaceAfter=3,
                                       textColor=colors.HexColor("#003D7A"))
        task_meta_s = ParagraphStyle("task_m", fontName=fn, fontSize=11,
                                     leading=15, spaceAfter=4, textColor=colors.grey)
        pgd_group_s = ParagraphStyle("pgd_g", fontName=fb, fontSize=10,
                                     leading=14, spaceBefore=4, spaceAfter=1,
                                     textColor=colors.HexColor("#2E7D32"))
        xa_line_s = ParagraphStyle("xa_l", fontName=fn, fontSize=9,
                                   leading=12, leftIndent=12)
        th_s = ParagraphStyle("th", fontName=fb, fontSize=11,
                              alignment=TA_CENTER, textColor=colors.white,
                              leading=14)
        td_s = ParagraphStyle("td", fontName=fn, fontSize=11,
                              leading=14, wordWrap="CJK")
        td_c = ParagraphStyle("td_c", fontName=fn, fontSize=11,
                              alignment=TA_CENTER, leading=14)
        td_green = ParagraphStyle("td_green", fontName=fn, fontSize=11,
                                  alignment=TA_CENTER, leading=14,
                                  textColor=colors.HexColor("#2E7D32"))
        td_red = ParagraphStyle("td_red", fontName=fn, fontSize=11,
                                alignment=TA_CENTER, leading=14,
                                textColor=colors.HexColor("#C62828"))
        td_r = ParagraphStyle("td_r", fontName=fn, fontSize=11,
                              alignment=TA_RIGHT, leading=14)

        for i, t in enumerate(ds_task, 1):
            kq = _doc_ketqua_task(t["id"])
            cbtd_bien_hoa = str(t.get("cbtd_bien_hoa") or "").strip()
            if cbtd_bien_hoa and not any(r.get("ten_xa") == _PGD_BIEN_HOA for r in kq):
                kq = list(kq or [])
                kq.append({
                    "pgd": _PGD_BIEN_HOA,
                    "ten_xa": _PGD_BIEN_HOA,
                    "trang_thai": "chua_thuc_hien",
                    "ngay_hoan_thanh": None,
                    "ghi_chu": "",
                })
            xong = sum(1 for r in kq if r["trang_thai"] == "da_hoan_thanh")
            tong = len(kq)
            pct = round(xong / tong * 100) if tong else 0

            loai_label = LOAI_TASK.get(t["loai"], t["loai"])
            nguoi_pt = t.get("nguoi_phu_trach") or "—"
            cap = t.get("cap_theo_doi", "xa")

            tag_cap = "🏢 Theo dõi chung PGD" if cap == "pgd" else "📍 Theo dõi chi tiết xã"

            if pct == 100:
                css_cls = "✅ Hoàn thành"
            elif t["ngay_deadline"] < hom_nay:
                css_cls = "🔴 Trễ hạn"
            else:
                css_cls = "⏳ Đang thực hiện"

            story.append(Paragraph(
                f"{i}. {t['tieu_de']}",
                task_header_s
            ))
            story.append(Paragraph(
                f"Thời hạn cuối cùng: {fmt_ngay(t['ngay_deadline'])} | Tiến độ hoàn thành: {pct}% | "
                f"{css_cls} | Loại: {loai_label} | {tag_cap} | "
                f"Người PT: {nguoi_pt}",
                task_meta_s
            ))

            if cap == "pgd":
                pgd_order = sorted(set(r["pgd"] for r in kq), key=lambda x: (x == _PGD_BIEN_HOA, x))
                tbl_data = [[
                    Paragraph("STT", th_s),
                    Paragraph("PGD", th_s),
                    Paragraph("Trạng thái", th_s),
                    Paragraph("Ghi chú", th_s),
                ]]
                for j, pgd in enumerate(pgd_order, 1):
                    r_pgd = next((r for r in kq if r["pgd"] == pgd), {})
                    tt = r_pgd.get("trang_thai", "chua_thuc_hien")
                    if tt == "da_hoan_thanh":
                        tt_txt = "✅ Hoàn thành"
                        stt_style = td_green
                    elif tt == "khong_ap_dung":
                        tt_txt = "➖ N/A"
                        stt_style = td_c
                    else:
                        tt_txt = "⬜ Chưa thực hiện"
                        stt_style = td_red
                    gc = (r_pgd.get("ghi_chu") or "").strip()
                    tbl_data.append([
                        Paragraph(str(j), td_c),
                        Paragraph(pgd, td_s),
                        Paragraph(tt_txt, stt_style),
                        Paragraph(gc, td_s),
                    ])
                cw = [1.5 * cm, 7.0 * cm, 3.5 * cm, 6.0 * cm]
                tbl = Table(tbl_data, colWidths=cw, repeatRows=1)
                tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BDBDBD")),
                ]))
                for r_idx in range(1, len(tbl_data)):
                    if r_idx % 2 == 0:
                        tbl.setStyle(TableStyle([
                            ("BACKGROUND", (0, r_idx), (-1, r_idx),
                             colors.HexColor("#F5F5F5"))
                        ]))
                story.append(tbl)
            else:
                tbl_data = [[
                    Paragraph("STT", th_s),
                    Paragraph("PGD", th_s),
                    Paragraph("Xã / Phường", th_s),
                    Paragraph("Trạng thái", th_s),
                    Paragraph("Ghi chú", th_s),
                ]]
                for j, r in enumerate(
                    sorted(kq, key=lambda x: (x["pgd"], x["ten_xa"])), 1
                ):
                    tt = r["trang_thai"]
                    if tt == "da_hoan_thanh":
                        tt_txt = "✅ Hoàn thành"
                        stt_style = td_green
                        ngay_ht = r.get("ngay_hoan_thanh") or ""
                        try:
                            ngay_ht = datetime.strptime(
                                ngay_ht[:10], "%Y-%m-%d"
                            ).strftime("%d/%m/%Y")
                        except Exception:
                            pass
                        gc = ngay_ht
                    elif tt == "khong_ap_dung":
                        tt_txt = "➖ N/A"
                        stt_style = td_c
                        gc = ""
                    else:
                        tt_txt = "⬜ Chưa thực hiện"
                        stt_style = td_red
                        gc = ""
                    tbl_data.append([
                        Paragraph(str(j), td_c),
                        Paragraph(r["pgd"], td_s),
                        Paragraph(r["ten_xa"] or "", td_s),
                        Paragraph(tt_txt, stt_style),
                        Paragraph(gc, td_s),
                    ])
                cw = [1.2 * cm, 5.2 * cm, 4.3 * cm, 3.3 * cm, 4.0 * cm]
                tbl = Table(tbl_data, colWidths=cw, repeatRows=1)
                tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BDBDBD")),
                ]))
                for r_idx in range(1, len(tbl_data)):
                    if r_idx % 2 == 0:
                        tbl.setStyle(TableStyle([
                            ("BACKGROUND", (0, r_idx), (-1, r_idx),
                             colors.HexColor("#F5F5F5"))
                        ]))
                story.append(tbl)

            story.append(Spacer(1, 0.3 * cm))
            story.append(HRFlowable(width="60%", thickness=0.3,
                                    color=colors.HexColor("#E0E0E0"),
                                    spaceAfter=2))

        # ── Page break → Phần 2 ───────────────────────────────────────
        story.append(PageBreak())

        story.append(Paragraph(
            "PHẦN 2: BÁO CÁO TRỄ HẠN",
            ParagraphStyle("p2_title", fontName=fb, fontSize=11,
                           alignment=TA_LEFT, spaceBefore=4, spaceAfter=6,
                           textColor=colors.HexColor("#C62828"))
        ))

        p2_th = ParagraphStyle("p2_th", fontName=fb, fontSize=11,
                                alignment=TA_CENTER, textColor=colors.white,
                                leading=14)
        p2_td = ParagraphStyle("p2_td", fontName=fn, fontSize=11,
                               leading=14, wordWrap="CJK")
        p2_td_c = ParagraphStyle("p2_td_c", fontName=fn, fontSize=11,
                                 alignment=TA_CENTER, leading=14)
        p2_td_r = ParagraphStyle("p2_td_r", fontName=fb, fontSize=11,
                                 alignment=TA_CENTER, leading=14,
                                 textColor=colors.HexColor("#C62828"))

        p2_count = 0
        for t in ds_task:
            if t["ngay_deadline"] >= hom_nay:
                continue
            kq = _doc_ketqua_task(t["id"])
            tre_list = [r for r in kq if r["trang_thai"] == "chua_thuc_hien"]
            if not tre_list:
                continue

            so_ngay_tre = (date.today() -
                           date.fromisoformat(t["ngay_deadline"])).days
            cap = t.get("cap_theo_doi", "xa")

            story.append(Paragraph(
                f"▪ {t['tieu_de']} — Trễ {so_ngay_tre} ngày (thời hạn: {fmt_ngay(t['ngay_deadline'])})",
                ParagraphStyle("p2_task", fontName=fb, fontSize=11,
                               leading=15, spaceBefore=8, spaceAfter=3,
                               textColor=colors.HexColor("#C62828"))
            ))

            if cap == "pgd":
                cols = ["STT", "PGD", "Số ngày trễ"]
                tbl_data = [[Paragraph(c, p2_th) for c in cols]]
                for j, r in enumerate(tre_list, 1):
                    p2_count += 1
                    tbl_data.append([
                        Paragraph(str(j), p2_td_c),
                        Paragraph(r["pgd"], p2_td),
                        Paragraph(f"{so_ngay_tre} ngày", p2_td_r),
                    ])
                cw2 = [1.5 * cm, 10.0 * cm, 6.5 * cm]
                tbl2 = Table(tbl_data, colWidths=cw2, repeatRows=1)
                tbl2.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C62828")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BDBDBD")),
                ]))
                for r_idx in range(1, len(tbl_data)):
                    if r_idx % 2 == 0:
                        tbl2.setStyle(TableStyle([
                            ("BACKGROUND", (0, r_idx), (-1, r_idx),
                             colors.HexColor("#FFEBEE"))
                        ]))
                story.append(tbl2)
            else:
                cols = ["STT", "PGD", "Xã / Phường", "Số ngày trễ"]
                tbl_data = [[Paragraph(c, p2_th) for c in cols]]
                idx2 = 0
                kq_by_pgd = {}
                for r in tre_list:
                    kq_by_pgd.setdefault(r["pgd"], []).append(r)
                for pgd, items in sorted(kq_by_pgd.items()):
                    for r in items:
                        idx2 += 1
                        p2_count += 1
                        tbl_data.append([
                            Paragraph(str(idx2), p2_td_c),
                            Paragraph(pgd, p2_td),
                            Paragraph(r["ten_xa"], p2_td),
                            Paragraph(f"{so_ngay_tre} ngày", p2_td_r),
                        ])
                cw2 = [1.2 * cm, 5.5 * cm, 6.5 * cm, 4.8 * cm]
                tbl2 = Table(tbl_data, colWidths=cw2, repeatRows=1)
                tbl2.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C62828")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BDBDBD")),
                ]))
                for r_idx in range(1, len(tbl_data)):
                    if r_idx % 2 == 0:
                        tbl2.setStyle(TableStyle([
                            ("BACKGROUND", (0, r_idx), (-1, r_idx),
                             colors.HexColor("#FFEBEE"))
                        ]))
                story.append(tbl2)

        if p2_count > 0:
            story.append(Spacer(1, 0.5 * cm))
            story.append(Paragraph(
                f"Tổng số: {p2_count} đơn vị trễ hạn",
                ParagraphStyle("sum", fontName=fb, fontSize=10,
                               alignment=TA_CENTER, spaceBefore=8,
                               textColor=colors.HexColor("#C62828"))
            ))
        else:
            story.append(Paragraph(
                "✅ Không có đơn vị nào trễ hạn.",
                ParagraphStyle("ok", fontName=fn, fontSize=10,
                               alignment=TA_CENTER, spaceAfter=6,
                               textColor=colors.HexColor("#2E7D32"))
            ))

        # ── Chữ ký ───────────────────────────────────────────────────
        story.append(Spacer(1, 0.6 * cm))
        story.append(HRFlowable(width="100%", thickness=0.5,
                                color=colors.HexColor("#BDBDBD"), spaceAfter=4))
        sig_style = ParagraphStyle("sig", fontName=fn, fontSize=11,
                                   alignment=TA_CENTER, leading=15)
        sig_bold = ParagraphStyle("sig_b", fontName=fb, fontSize=12,
                                  alignment=TA_CENTER, leading=17)
        sig_data = [[
            Paragraph("Người lập", sig_bold),
            Paragraph("Kiểm soát", sig_bold),
            Paragraph("Giám đốc", sig_bold),
        ], [
            Paragraph("(Ký, ghi rõ họ tên)", sig_style),
            Paragraph("(Ký, ghi rõ họ tên)", sig_style),
            Paragraph("(Ký, ghi rõ họ tên, đóng dấu)", sig_style),
        ]]
        sig_tbl = Table(sig_data, colWidths=[5.5*cm, 5.5*cm, 5.5*cm])
        sig_tbl.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(sig_tbl)

        # ── Footer ────────────────────────────────────────────────────
        def _on_page(canvas, _doc):
            canvas.saveState()
            canvas.setFont(fn if FONT_NORMAL else FONT_FALLBACK, 8)
            canvas.setFillColor(colors.grey)
            canvas.drawRightString(
                page_size[0] - margin,
                0.6 * cm,
                f"Trang {_doc.page}  |  VBSP-SCM  |  {ngay_str}"
            )
            canvas.restoreState()

        doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
        buf.seek(0)
        return buf

    except Exception as e:
        st.error(f"Lỗi tạo PDF báo cáo tiến độ: {e}")
        return None


def _xuat_pdf_tien_do(task, ds_kq, username):
    try:
        from itertools import groupby
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, HRFlowable,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import os

        FONT_NORMAL = "Helvetica"
        FONT_BOLD = "Helvetica-Bold"

        arial_path = "C:/Windows/Fonts/arial.ttf"
        arialbd_path = "C:/Windows/Fonts/arialbd.ttf"
        if os.path.exists(arial_path):
            pdfmetrics.registerFont(TTFont("ArialUni", arial_path))
            FONT_NORMAL = "ArialUni"
        if os.path.exists(arialbd_path):
            pdfmetrics.registerFont(TTFont("ArialUni-Bold", arialbd_path))
            FONT_BOLD = "ArialUni-Bold"

        ds_kq = list(ds_kq or [])
        cbtd_bien_hoa = str(task.get("cbtd_bien_hoa") or "").strip()
        if cbtd_bien_hoa and not any(r.get("ten_xa") == _PGD_BIEN_HOA for r in ds_kq):
            ds_kq.append({
                "pgd": _PGD_BIEN_HOA,
                "ten_xa": _PGD_BIEN_HOA,
                "trang_thai": "chua_thuc_hien",
                "ngay_hoan_thanh": None,
                "ghi_chu": "",
            })

        kq_theo_pgd = {}
        for r in sorted(ds_kq, key=lambda x: x["pgd"]):
            kq_theo_pgd.setdefault(r["pgd"], []).append(r)

        pgd_order = sorted(kq_theo_pgd.keys(), key=lambda x: (x == _PGD_BIEN_HOA, x))

        TS_LABEL = {
            "chua_thuc_hien": "○",
            "da_hoan_thanh": "✓",
            "khong_ap_dung": "—",
        }

        buf = BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=28, rightMargin=28,
            topMargin=20, bottomMargin=20,
        )

        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "Title_VN", parent=styles["Title"],
            fontName=FONT_BOLD, fontSize=14,
            alignment=TA_CENTER, spaceAfter=2, leading=18,
        )
        subtitle_style = ParagraphStyle(
            "Sub_VN", parent=styles["Normal"],
            fontName=FONT_BOLD, fontSize=12,
            alignment=TA_CENTER, spaceAfter=4, leading=16,
        )
        header_style = ParagraphStyle(
            "Header_VN", parent=styles["Normal"],
            fontName=FONT_BOLD, fontSize=11,
            alignment=TA_CENTER, spaceAfter=6, leading=15,
        )
        normal_style = ParagraphStyle(
            "Normal_VN", parent=styles["Normal"],
            fontName=FONT_NORMAL, fontSize=9, leading=13,
        )
        pgd_header_style = ParagraphStyle(
            "PGDHeader", parent=styles["Normal"],
            fontName=FONT_BOLD, fontSize=10, leading=14,
            spaceBefore=6, spaceAfter=2,
        )
        xa_style = ParagraphStyle(
            "XaStyle", parent=styles["Normal"],
            fontName=FONT_NORMAL, fontSize=9, leading=12, leftIndent=12,
        )
        summary_style = ParagraphStyle(
            "Summary", parent=styles["Normal"],
            fontName=FONT_BOLD, fontSize=10, leading=14,
            spaceBefore=8, spaceAfter=4,
        )

        story.append(Paragraph(
            "NGÂN HÀNG CHÍNH SÁCH XÃ HỘI VIỆT NAM", title_style))
        story.append(Paragraph("CHI NHÁNH TỈNH ĐỒNG NAI", subtitle_style))
        story.append(Spacer(1, 6))
        story.append(Paragraph("BÁO CÁO TIẾN ĐỘ CÔNG VIỆC", header_style))
        story.append(Paragraph(task.get("tieu_de", ""), header_style))
        story.append(Spacer(1, 4))

        loai = LOAI_TASK.get(task.get("loai", ""), task.get("loai", ""))
        uu_tien = UU_TIEN.get(task.get("uu_tien", ""), task.get("uu_tien", ""))
        deadline = task.get("ngay_deadline", "")
        now = datetime.now()
        ngay_xuat = now.strftime("%d/%m/%Y %H:%M")

        story.append(Paragraph(
            f"Loại: {loai} | Ưu tiên: {uu_tien} | "
            f"Người phụ trách: {task.get('nguoi_phu_trach') or '—'} | "
            f"Từ ngày: {fmt_ngay(task.get('ngay_bat_dau')) or '—'} → Thời hạn: {fmt_ngay(deadline)}",
            normal_style,
        ))
        story.append(Paragraph(
            f"Ngày xuất: {ngay_xuat} | Người xuất: {username}",
            normal_style,
        ))

        story.append(HRFlowable(
            width="100%", thickness=0.5, color=colors.grey))
        story.append(Spacer(1, 4))

        tong_dv = 0
        tong_dv_ht = 0
        tong_xa_all = 0
        tong_xa_xong_all = 0

        for pgd in pgd_order:
            items = kq_theo_pgd[pgd]
            tong = len(items)
            xong = sum(1 for r in items if r["trang_thai"] == "da_hoan_thanh")
            tong_xa_all += tong
            tong_xa_xong_all += xong
            pct = round(xong / tong * 100) if tong > 0 else 0

            story.append(Paragraph(
                f"<b>{pgd}</b>  ({xong}/{tong} xã — {pct}%)",
                pgd_header_style,
            ))

            for r in items:
                symbol = TS_LABEL.get(r["trang_thai"], "?")
                ten_xa = r["ten_xa"]
                ngay_ht = r.get("ngay_hoan_thanh") or ""

                if r["trang_thai"] == "da_hoan_thanh":
                    try:
                        ngay_ht = datetime.strptime(
                            ngay_ht[:10], "%Y-%m-%d"
                        ).strftime("%d/%m/%Y")
                    except Exception:
                        pass
                    status_text = ngay_ht
                elif r["trang_thai"] == "khong_ap_dung":
                    status_text = "N/A"
                else:
                    status_text = "Chưa thực hiện"

                dots = "." * max(2, 45 - len(ten_xa) - len(status_text))
                line = f"{symbol} {ten_xa} {dots} {status_text}"
                story.append(Paragraph(line, xa_style))

            story.append(Spacer(1, 2))

            tong_dv += 1
            if xong == tong and tong > 0:
                tong_dv_ht += 1

        story.append(HRFlowable(
            width="100%", thickness=0.5, color=colors.grey))
        story.append(Spacer(1, 6))

        tong_pct = (
            round(tong_xa_xong_all / tong_xa_all * 100)
            if tong_xa_all > 0 else 0
        )
        story.append(Paragraph(
            f"TỔNG KẾT: {tong_dv_ht}/{tong_dv} đơn vị hoàn thành"
            f" | {tong_pct}% toàn Chi nhánh",
            summary_style,
        ))

        doc.build(story)
        buf.seek(0)
        return buf

    except Exception as e:
        st.error(f"Lỗi tạo PDF: {e}")
        return None




from tabs.base_tab import TabContext


def render(tab, **kwargs):
    ctx = TabContext(tab, **kwargs)
    with ctx:
        st.subheader("📅 Tiến độ Công việc Hàng ngày")

        if ctx.is_exec:
            _render_tong_quan(tab, **kwargs)
            return

        if ctx.is_pgd:
            lazy_tabs(
                ["📊 Tổng quan", "📋 Cập nhật tiến độ"],
                [
                    lambda: _render_tong_quan(st.container(), **kwargs),
                    lambda: _render_cap_nhat(st.container(), **kwargs),
                ],
                key="td_pgd",
            )
            return

        if not ctx.is_cn:
            _render_tong_quan(tab, **kwargs)
            return

        lazy_tabs(
            ["📊 Tổng quan", "➕ Tạo đầu việc mới", "✏️ Chỉnh sửa & Xóa đầu việc",
             "📋 Cập nhật tiến độ", "📤 Xuất báo cáo"],
            [
                lambda: _render_tong_quan(st.container(), **kwargs),
                lambda: _render_tao_task(st.container(), **kwargs),
                lambda: _render_quan_ly_task(st.container(), **kwargs),
                lambda: _render_cap_nhat(st.container(), **kwargs),
                lambda: _render_xuat(st.container(), **kwargs),
            ],
            key="td_cn",
        )


def render_tong_quan_only(tab, **kwargs):
    ctx = TabContext(tab, **kwargs)
    with ctx:
        st.subheader("📅 Tiến độ Công việc Hàng ngày")
        _render_tong_quan(tab, **kwargs)
