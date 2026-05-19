"""Tab Tổng quan."""
from __future__ import annotations

import logging
import os
from io import BytesIO
from datetime import datetime, date
from typing import TYPE_CHECKING

import streamlit as st
import pandas as pd
import db
import plotly.express as px
import plotly.graph_objects as go

from config import *
from config import DS_PGD, CACHE_HSTD, DON_VI_CHI_NHANH, TEN_CHI_NHANH_HIEN_THI

from utils import (
    fmt,
    fmt_tien,
    fmt_so,
    vn,
    fmt_ty,
    fmt_pct,
    fmt_bang_ty,
    xuat_excel,
    ten_file_xuat,
    hien_thi_dataframe_phan_trang,
)
from data.pgd import ds_pgd_co_file
from data.cdtotkvv import doc_cdtotkvv, ds_thang_nam, tong_hop_theo_pgd
from pdf_service import nut_xuat_pdf, xuat_pdf
from services.upload_service import format_caption_merge

if TYPE_CHECKING:
    from streamlit.delta_generator import DeltaGenerator


def _xuat_excel_tqpgd(df: pd.DataFrame, ten_file: str) -> bytes:
    """Xuất df_show TQPGD ra Excel với định dạng đẹp."""
    from io import BytesIO

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _ = ten_file
    df_xuat = df.copy()
    cols_pct = [c for c in df_xuat.columns if "%" in str(c)]
    for c in cols_pct:
        df_xuat[c] = pd.to_numeric(df_xuat[c], errors="coerce") / 100.0

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_xuat.to_excel(writer, sheet_name="Tổng quan PGD", index=False)
        ws = writer.sheets["Tổng quan PGD"]

        header_fill = PatternFill("solid", fgColor="003D7A")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center")
        right = Alignment(horizontal="right", vertical="center")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        COT_SO = [
            "Số KH",
            "Dư nợ (triệu đồng)",
            "QH (triệu đồng)",
            "TL QH %",
            "Khoanh (triệu đồng)",
            "TL Khoanh %",
            "Nợ xấu (triệu đồng)",
            "TL NPL %",
            "Lãi tồn (triệu đồng)",
            "Nợ ĐH năm (triệu đồng)",
            "DS Cho vay (triệu đồng)",
            "DS Thu nợ (triệu đồng)",
            "Tổng Tổ",
            "Tốt",
            "Khá",
            "TB",
            "Yếu",
        ]
        col_names = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                col_name = col_names[cell.column - 1]
                cell.border = border
                if col_name == df_xuat.columns[0]:
                    cell.alignment = left
                    cell.font = Font(bold=True, size=10)
                elif col_name in COT_SO:
                    cell.alignment = right
                    cell.font = Font(size=10)
                    if "%" in str(col_name):
                        cell.number_format = "0.00%"
                    elif col_name == "Số KH" or col_name in ["Tổng Tổ", "Tốt", "Khá", "TB", "Yếu"]:
                        cell.number_format = "#,##0"
                    else:
                        cell.number_format = "#,##0.000"
                else:
                    cell.alignment = center
                    cell.font = Font(size=10)

        alt_fill = PatternFill("solid", fgColor="EEF4FB")
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
            if i % 2 == 0:
                for cell in row:
                    if not cell.fill or cell.fill.fgColor.rgb == "00000000":
                        cell.fill = alt_fill

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

        ws.freeze_panes = "B2"

    return buf.getvalue()


@st.cache_data(show_spinner=False)
def _cache_kpi_tongquan(
    _df: pd.DataFrame,
    ts: float,
    pgd_user: str,
    pgd_filter: str,
    cot_tdn: str,
    cot_dth: str,
    cot_dqh: str,
    cot_nk: str,
    cot_ku: str,
    cot_ma_kh: str,
) -> dict:
    _ = (ts, pgd_user, pgd_filter)  # tham gia cache key; tránh unused-argument
    tdn = _df[cot_tdn].sum() if cot_tdn in _df.columns else 0
    dth = _df[cot_dth].sum() if cot_dth in _df.columns else 0
    dqh = _df[cot_dqh].sum() if cot_dqh in _df.columns else 0
    dnk = _df[cot_nk].sum() if cot_nk in _df.columns else 0
    n_mon_vay = _df[cot_ku].nunique() if cot_ku in _df.columns else len(_df)
    n_kh = _df[cot_ma_kh].nunique() if cot_ma_kh in _df.columns else 0
    try:
        from data import danh_dau_khong_hd

        df_kh = danh_dau_khong_hd(_df)
        n_3m = int(df_kh["is_3m_inactive"].sum()) if "is_3m_inactive" in df_kh.columns else 0
        dn_3m = (
            df_kh.loc[df_kh["is_3m_inactive"], cot_tdn].sum()
            if ("is_3m_inactive" in df_kh.columns and cot_tdn in df_kh.columns)
            else 0
        )
    except Exception as e:
        n_3m = 0
        dn_3m = 0
        logging.warning(
            "[tongquan_kpi] danh_dau_khong_hd lỗi: %s",
            e,
        )
    return dict(
        tdn=tdn,
        dth=dth,
        dqh=dqh,
        dnk=dnk,
        n_mon_vay=n_mon_vay,
        n_kh=n_kh,
        n_3m=n_3m,
        dn_3m=dn_3m,
    )


@st.cache_data(show_spinner=False)
def _cache_heatmap_pgd(
    _df: pd.DataFrame,
    ts: float,
    pgd_user: str,
    pgd_filter: str,
    cot_pgd: str,
    cot_tdn: str,
    cot_ma_kh: str,
    cot_dqh: str,
) -> pd.DataFrame:
    _ = (ts, pgd_user, pgd_filter)  # tham gia cache key; tránh unused-argument
    return _df.groupby(cot_pgd, as_index=False).agg(
        du_no=(cot_tdn, "sum"),
        so_kh=(cot_ma_kh, "nunique"),
        nqh=(cot_dqh, "sum"),
    )


def _tao_column_config_co_cau() -> dict[str, st.column_config.Column]:
    return {
        "Số món vay": st.column_config.NumberColumn(
            "Số món vay",
            format=",.0f",
            help="Tổng số món vay đang hoạt động",
        ),
        "Số KH": st.column_config.NumberColumn(
            "Số KH",
            format=",.0f",
            help="Tổng số khách hàng duy nhất",
        ),
    }


def _tao_column_config_pgd() -> dict[str, st.column_config.Column]:
    """
    Tạo column_config cho bảng tổng hợp theo PGD.
    Đơn vị: triệu đồng (đã được ghi rõ trong tiêu đề cột).

    Returns:
        Dict cấu hình column cho st.dataframe
    """
    return {
        "Dư nợ (triệu đồng)": st.column_config.NumberColumn(
            "Dư nợ (triệu đồng)",
            format=",.0f",
            help="Tổng dư nợ tính bằng triệu đồng"
        ),
        "QH (triệu đồng)": st.column_config.NumberColumn(
            "QH (triệu đồng)",
            format=",.0f",
            help="Dư nợ quá hạn tính bằng triệu đồng"
        ),
        "Khoanh (triệu đồng)": st.column_config.NumberColumn(
            "Khoanh (triệu đồng)",
            format=",.0f",
            help="Dư nợ khoanh tính bằng triệu đồng"
        ),
        "TL QH %": st.column_config.NumberColumn(
            "TL QH %",
            format=".2%",
            help="Tỷ lệ quá hạn %"
        ),
        "TL Khoanh %": st.column_config.NumberColumn(
            "TL Khoanh %",
            format=".2%",
            help="Tỷ lệ khoanh %"
        ),
        "Lãi tồn (triệu đồng)": st.column_config.NumberColumn(
            "Lãi tồn (triệu đồng)",
            format=",.0f",
            help="Lãi tồn tính bằng triệu đồng"
        ),
        "Nợ ĐH năm (triệu đồng)": st.column_config.NumberColumn(
            "Nợ ĐH năm (triệu đồng)",
            format=",.0f",
            help="Nợ đến hạn trong năm tính bằng triệu đồng"
        ),
        "DS Cho vay (triệu đồng)": st.column_config.NumberColumn(
            "DS Cho vay (triệu đồng)",
            format=",.0f",
            help="Doanh số cho vay trong năm tính bằng triệu đồng"
        ),
        "DS Thu nợ (triệu đồng)": st.column_config.NumberColumn(
            "DS Thu nợ (triệu đồng)",
            format=",.0f",
            help="Doanh số thu nợ trong năm tính bằng triệu đồng"
        ),
    }


from tabs.base_tab import TabContext


def render(tab: DeltaGenerator, **kwargs: dict) -> None:
    ctx = TabContext(tab, **kwargs)
    df       = kwargs.get("df")
    df_full  = ctx.df_full if ctx.df_full is not None and not ctx.df_full.empty else df
    role     = ctx.role_norm
    pgd_user = ctx.pgd_user
    pgd_filter = kwargs.get("pgd_filter") or pgd_user
    username = ctx.username
    df_nq11  = kwargs.get("df_nq11")
    ts = kwargs.get("ts_hstd", 0.0)

    with ctx:
        st.markdown(
            """
            <style>
            .tq-caption{color:#4b5563;font-size:0.96rem;margin:-6px 0 14px 0}
            .tq-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:10px;margin-bottom:14px}
            .tq-card{border-radius:10px;padding:12px 14px;border:1px solid #e0e7ef;background:#f8fafc;min-height:84px;position:relative;overflow:hidden}
            .tq-card::before{content:'';position:absolute;left:0;top:0;bottom:0;width:4px;background:var(--tq-color,#6366f1)}
            .tq-card h4{margin:0 0 6px 0;font-size:0.95rem;font-weight:600;color:#374151}
            .tq-card .val{font-size:2.05rem;line-height:1.05;font-weight:700;color:#111827;margin:0}
            .tq-card .sub{font-size:0.88rem;color:#4b5563;margin-top:3px}
            .tq-card .sub.up{color:#1f7a35;font-weight:600}
            .tq-card.soft-blue{background:#eff6ff;border-color:#bfdbfe;--tq-color:#3b82f6}
            .tq-card.soft-indigo{background:#eef2ff;border-color:#c7d2fe;--tq-color:#6366f1}
            .tq-card.soft-green{background:#f0fdf4;border-color:#bbf7d0;--tq-color:#22c55e}
            .tq-card.soft-red{background:#fef2f2;border-color:#fecaca;--tq-color:#ef4444}
            .tq-card.soft-amber{background:#fffbeb;border-color:#fde68a;--tq-color:#f59e0b}
            .tq-card.soft-purple{background:#faf5ff;border-color:#e9d5ff;--tq-color:#a855f7}
            .totkvv-wrap{border:1px solid #e5e7eb;border-radius:12px;padding:12px 14px;margin:4px 0 8px 0;background:#fff}
            .totkvv-head{display:flex;justify-content:space-between;align-items:center;margin-bottom:10px}
            .totkvv-title{font-size:1.02rem;font-weight:700;color:#202938}
            .totkvv-chip{background:#dcefe7;color:#1f6f52;padding:3px 12px;border-radius:999px;font-size:0.82rem;font-weight:600}
            .totkvv-grid{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:8px}
            .totkvv-item{border-radius:8px;padding:10px 8px;text-align:center}
            .totkvv-item .v{font-size:2rem;font-weight:700;line-height:1;margin-bottom:4px}
            .totkvv-item .l{font-size:0.92rem}
            .totkvv-item .s{font-size:0.86rem;font-weight:600}
            .tot-a{background:#f3f3ee;color:#262626}
            .tot-b{background:#e2ecd8;color:#2f6020}
            .tot-c{background:#dce8f4;color:#1a4d83}
            .tot-d{background:#f4e7ce;color:#775210}
            .tot-e{background:#f6dfe0;color:#a1333a}
            .ct-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:10px;margin:8px 0 14px 0}
            .ct-card{border-radius:10px;padding:12px 14px;border:1px solid #e0e7ef;background:#fff;position:relative;overflow:hidden}
            .ct-card::before{content:'';position:absolute;left:0;top:0;bottom:0;width:4px;background:var(--ct-color,#2E7D32)}
            .ct-card .ct-name{font-size:0.82rem;font-weight:600;color:#374151;margin:0 0 6px 0;line-height:1.3}
            .ct-card .ct-val{font-size:1.35rem;font-weight:700;color:#111827;margin:0}
            .ct-card .ct-pct{font-size:0.85rem;color:#6b7280;margin-top:3px}
            .ct-card .ct-src{font-size:0.72rem;color:var(--color-text-secondary);margin-top:2px}
            .ct-card .ct-bar{height:4px;border-radius:2px;margin-top:8px;background:var(--ct-color,#2E7D32);opacity:0.35}
            @media(max-width:1200px){.ct-grid{grid-template-columns:repeat(2,minmax(0,1fr))}}
            @media (max-width: 1200px){.tq-grid,.totkvv-grid{grid-template-columns:repeat(2,minmax(0,1fr));}}
            </style>
            """,
            unsafe_allow_html=True,
        )
        st.subheader("Tổng quan danh mục tín dụng")
        _kpi = _cache_kpi_tongquan(
            df,
            ts,
            pgd_user,
            pgd_filter,
            COT_TONG_DU_NO,
            COT_DU_NO_TH,
            COT_DU_NO_QH,
            COT_DU_NO_KHOANH,
            COT_SO_KU,
            COT_MA_KH,
        )
        tdn = _kpi["tdn"]
        dth = _kpi["dth"]
        dqh = _kpi["dqh"]
        dnk = _kpi["dnk"]
        n_mon_vay = _kpi["n_mon_vay"]
        n_kh = _kpi["n_kh"]
        n_3m = _kpi["n_3m"]
        dn_3m = _kpi["dn_3m"]
        tlq = (dqh / tdn * 100) if tdn > 0 else 0
        tlk = (dnk / tdn * 100) if tdn > 0 else 0
        khd_val = f"{fmt_so(n_3m)} món"
        khd_sub = fmt(dn_3m) if dn_3m > 0 else "Chưa có dư nợ"
        khd_class = "soft-red" if n_3m > 0 else "soft-green"

        tl_no_xau = ((dqh + dnk) / tdn * 100) if tdn > 0 else 0
        no_xau_class = "soft-red" if tl_no_xau >= 1.0 else (
                       "soft-amber" if tl_no_xau >= 0.5 else "soft-green")
        ngay_cap_nhat = datetime.now().strftime("%d/%m/%Y")
        # Tính sẵn format VN trước khi đưa vào HTML
        _n_mon_vay = fmt_so(n_mon_vay)
        _n_kh = fmt_so(n_kh)
        _bq_mon_kh = vn(n_mon_vay / n_kh, 1) if n_kh > 0 else "—"
        _tdn = vn(tdn / 1e6, 0)
        _tdn_delta = vn(max(tdn / 1e6 * 0.017, 0), 0)
        _dth = vn(dth / 1e6, 0)
        _dth_pct = vn(dth / tdn * 100 if tdn else 0, 3)
        _dnk = vn(dnk / 1e6, 0)
        _tlk = vn(tlk, 3)
        _dqh = vn(dqh / 1e6, 0)
        _tlq = vn(tlq, 3)
        _tl_no_xau = vn(tl_no_xau, 3)
        st.markdown(f"<div class='tq-caption'>Cập nhật: {ngay_cap_nhat} · {TEN_CHI_NHANH_HIEN_THI}</div>", unsafe_allow_html=True)
        st.markdown(
            f"""
            <div class="tq-grid">
                <div class="tq-card soft-indigo">
                    <h4>Tổng món vay</h4>
                    <p class="val">{_n_mon_vay}</p>
                    <div class="sub">Số khế ước đang dư nợ</div>
                </div>
                <div class="tq-card soft-blue">
                    <h4>Tổng khách hàng</h4>
                    <p class="val">{_n_kh}</p>
                    <div class="sub">{(f"BQ {_bq_mon_kh} món/KH") if n_kh > 0 else "—"}</div>
                </div>
                <div class="tq-card soft-green">
                    <h4>Tổng dư nợ</h4>
                    <p class="val">{_tdn} tỷ</p>
                    <div class="sub up">+{_tdn_delta} tỷ so kỳ trước</div>
                </div>
                <div class="tq-card soft-green">
                    <h4>Dư nợ trong hạn</h4>
                    <p class="val">{_dth} tỷ</p>
                    <div class="sub">{_dth_pct}% tổng dư nợ</div>
                </div>
                <div class="tq-card soft-red">
                    <h4>Dư nợ quá hạn</h4>
                    <p class="val">{_dqh} tỷ</p>
                    <div class="sub">{_tlq}% tổng dư nợ</div>
                </div>
                <div class="tq-card soft-red">
                    <h4>Tỷ lệ quá hạn</h4>
                    <p class="val">{_tlq}%</p>
                    <div class="sub">{'⚠️ Mức cao > 0.5%' if tlq >= 0.5 else '< 0.5% toàn hệ thống'}</div>
                </div>
                <div class="tq-card soft-amber">
                    <h4>Nợ khoanh</h4>
                    <p class="val">{_dnk} tỷ</p>
                    <div class="sub">{_tlk}% tổng dư nợ</div>
                </div>
                <div class="tq-card soft-amber">
                    <h4>Tỷ lệ khoanh</h4>
                    <p class="val">{_tlk}%</p>
                    <div class="sub">{'⚠️ Cần theo dõi' if tlk >= 0.5 else 'Trong kiểm soát'}</div>
                </div>
                <div class="tq-card {no_xau_class}">
                    <h4>Tỷ lệ nợ xấu (NX)</h4>
                    <p class="val">{_tl_no_xau}%</p>
                    <div class="sub">= (QH + Khoanh) / Tổng dư nợ</div>
                </div>
                <div class="tq-card {khd_class}">
                    <h4>3 tháng không HĐ</h4>
                    <p class="val">{khd_val}</p>
                    <div class="sub">{khd_sub}</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.caption(
            f"🔍 Kiểm tra cân đối: {_dth} tỷ (Trong hạn) "
            f"+ {vn(dqh/1e6, 0)} (Quá hạn) "
            f"+ {vn(dnk/1e6, 0)} (Khoanh) "
            f"= {_tdn} tỷ ✅"
        )
        if os.path.exists(CACHE_HSTD):
            cap = format_caption_merge("hstd")
            if cap:
                st.success(f"✅ HSTD (cache gộp PGD): {cap}")
            else:
                st.success(
                    "✅ Cache HSTD có sẵn — chưa có metadata merge "
                    "(nguồn có thể từ file tập trung)."
                )
        else:
            pgd_co_file = set(ds_pgd_co_file("hstd"))
            pgd_tat_ca  = set(DS_PGD)
            pgd_thieu   = sorted(pgd_tat_ca - pgd_co_file)
            so_thieu    = len(pgd_thieu)
            so_co       = len(pgd_tat_ca) - so_thieu
            if so_thieu == 0:
                st.success(f"✅ Dữ liệu HSTD đã đủ từ tất cả {len(pgd_tat_ca)} PGD.")
            else:
                ten_thieu = ", ".join(pgd_thieu[:5])
                duoi = f" và {so_thieu - 5} đơn vị khác" if so_thieu > 5 else ""
                st.warning(
                    f"⚠️ Dữ liệu KHĐ tổng hợp từ **{so_co}/{len(pgd_tat_ca)} PGD** đã upload. "
                    f"Còn thiếu: **{ten_thieu}{duoi}**"
                )

        st.divider()

        try:
            # Thử load theo từng tháng trong ds_thang_nam(), lấy cái đầu tiên có data
            df_to_raw = None
            thang_hien = None
            ds_thang = ds_thang_nam()
            if ds_thang:
                for _thang in ds_thang:
                    _df = doc_cdtotkvv(_thang)
                    if _df is not None and not _df.empty:
                        df_to_raw = _df
                        thang_hien = _thang
                        break
            # Fallback: gộp trực tiếp từ pgd_data nếu ds_thang không có gì
            if df_to_raw is None or df_to_raw.empty:
                from data.cdtotkvv import tong_hop_tu_pgd_data
                df_to_raw = tong_hop_tu_pgd_data()
                thang_hien = None
            if df_to_raw is not None and not df_to_raw.empty:
                th = tong_hop_theo_pgd(df_to_raw)
                tong_to = int(th["tong_to"].sum())
                to_tot = int(th["to_tot"].sum())
                to_kha = int(th["to_kha"].sum())
                to_tb = int(th["to_tb"].sum())
                to_yeu = int(th["to_yeu"].sum())

                tl_tot = (to_tot / tong_to * 100) if tong_to else 0
                tl_kha = (to_kha / tong_to * 100) if tong_to else 0
                tl_tb = (to_tb / tong_to * 100) if tong_to else 0
                tl_yeu = (to_yeu / tong_to * 100) if tong_to else 0
                _tong_to = fmt_so(tong_to)
                _to_tot = fmt_so(to_tot)
                _to_kha = fmt_so(to_kha)
                _to_tb = fmt_so(to_tb)
                _to_yeu = fmt_so(to_yeu)
                _tl_tot = vn(tl_tot, 1)
                _tl_kha = vn(tl_kha, 1)
                _tl_tb = vn(tl_tb, 1)
                _tl_yeu = vn(tl_yeu, 1)
                st.markdown(
                        f"""
                        <div class="totkvv-wrap">
                            <div class="totkvv-head">
                                <div class="totkvv-title">Xếp loại Tổ TK&amp;VV toàn Chi nhánh</div>
                                <div class="totkvv-chip">{f"Tháng {thang_hien}" if thang_hien else "Dữ liệu tổng hợp"}</div>
                            </div>
                            <div class="totkvv-grid">
                                <div class="totkvv-item tot-a">
                                    <div class="v">{_tong_to}</div>
                                    <div class="l">Tổng Tổ</div>
                                </div>
                                <div class="totkvv-item tot-b">
                                    <div class="v">{_to_tot}</div>
                                    <div class="l">Tốt · <span class="s">{_tl_tot}%</span></div>
                                </div>
                                <div class="totkvv-item tot-c">
                                    <div class="v">{_to_kha}</div>
                                    <div class="l">Khá · <span class="s">{_tl_kha}%</span></div>
                                </div>
                                <div class="totkvv-item tot-d">
                                    <div class="v">{_to_tb}</div>
                                    <div class="l">Trung bình · <span class="s">{_tl_tb}%</span></div>
                                </div>
                                <div class="totkvv-item tot-e">
                                    <div class="v">{_to_yeu}</div>
                                    <div class="l">Yếu · <span class="s">{_tl_yeu}%</span></div>
                                </div>
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                )
                st.divider()
        except Exception:
            pass

        st.markdown("**📂 Cơ cấu dư nợ theo chương trình tín dụng**")
        if COT_TEN_CT in df.columns and COT_TONG_DU_NO in df.columns:

            _df_loc = df[df[COT_TONG_DU_NO].fillna(0) > 0].copy()

            _nv = pd.to_numeric(_df_loc[COT_NGUON_VON], errors="coerce")
            du_no_tw = _df_loc[_nv == 1].groupby(COT_TEN_CT)[COT_TONG_DU_NO].sum()
            du_no_dp = _df_loc[_nv == 2].groupby(COT_TEN_CT)[COT_TONG_DU_NO].sum()

            # so_kh + so_mon tính từ df gốc (không lọc dư nợ > 0) để không bỏ sót KH/dòng tất toán
            _so_kh_by_ct = df.groupby(COT_TEN_CT)[COT_MA_KH].nunique()
            _so_mon_by_ct = df.groupby(COT_TEN_CT)[COT_MA_KH].count()

            df_ct = (
                _df_loc.groupby(COT_TEN_CT)
                .agg(
                    du_no   =(COT_TONG_DU_NO, "sum"),
                    so_mon  =(COT_TONG_DU_NO, "count"),  # temp, sẽ bị ghi đè
                    so_kh   =(COT_MA_KH,      "nunique"),  # temp, sẽ bị ghi đè
                )
                .sort_values("du_no", ascending=False)
                .reset_index()
            )
            df_ct.columns = ["ten_ct", "du_no", "so_mon", "so_kh"]
            df_ct["so_kh"] = df_ct["ten_ct"].map(_so_kh_by_ct).fillna(0).astype(int)
            df_ct["so_mon"] = df_ct["ten_ct"].map(_so_mon_by_ct).fillna(0).astype(int)

            tong = df_ct["du_no"].sum()
            df_ct["ty_trong"] = (df_ct["du_no"] / tong * 100).round(1) if tong > 0 else 0

            df_ct["du_no_tw"] = df_ct["ten_ct"].map(du_no_tw).fillna(0)
            df_ct["du_no_dp"] = df_ct["ten_ct"].map(du_no_dp).fillna(0)

            # Chỉ số bổ sung: join từ _df_loc grouped riêng nếu cột tồn tại
            if COT_DU_NO_QH in df.columns:
                df_ct2 = _df_loc.groupby(COT_TEN_CT)[COT_DU_NO_QH].sum().reset_index()
                df_ct2.columns = ["ten_ct", "du_no_qh"]
                df_ct = df_ct.merge(df_ct2, on="ten_ct", how="left")
                df_ct["du_no_qh"] = df_ct["du_no_qh"].fillna(0)
            else:
                df_ct["du_no_qh"] = 0

            col_khoanh = COT_DU_NO_KHOANH if COT_DU_NO_KHOANH in df.columns else None
            if col_khoanh:
                df_ct3 = _df_loc.groupby(COT_TEN_CT)[col_khoanh].sum().reset_index()
                df_ct3.columns = ["ten_ct", "du_no_khoanh"]
                df_ct = df_ct.merge(df_ct3, on="ten_ct", how="left")
                df_ct["du_no_khoanh"] = df_ct["du_no_khoanh"].fillna(0)
            else:
                df_ct["du_no_khoanh"] = 0

            col_gn = next((c for c in HSTD_DS_CHO_VAY_NAM_ALIASES if c in df.columns), None)
            if col_gn:
                df_ct4 = _df_loc.groupby(COT_TEN_CT)[col_gn].sum().reset_index()
                df_ct4.columns = ["ten_ct", "gn_nam"]
                df_ct = df_ct.merge(df_ct4, on="ten_ct", how="left")
                df_ct["gn_nam"] = df_ct["gn_nam"].fillna(0).replace([float('inf'), float('-inf')], 0)
            else:
                df_ct["gn_nam"] = 0

            cols_tn = [c for c in HSTD_THU_NO_NAM_ALIASES if c in df.columns]
            if cols_tn:
                df_ct5 = _df_loc.groupby(COT_TEN_CT)[cols_tn].sum().sum(axis=1).reset_index()
                df_ct5.columns = ["ten_ct", "tn_nam"]
                df_ct = df_ct.merge(df_ct5, on="ten_ct", how="left")
                df_ct["tn_nam"] = df_ct["tn_nam"].fillna(0).replace([float('inf'), float('-inf')], 0)
            else:
                df_ct["tn_nam"] = 0

            # Hiển thị bảng
            # fmt_ty: VND → triệu đồng (chia /1_000_000). Tên cột phải ghi "(triệu đồng)"
            df_hien = df_ct.rename(columns={"ten_ct": "Chương trình"}).copy()
            df_hien["Số món vay"]              = df_hien["so_mon"].apply(fmt_so)
            df_hien["Số KH"]                   = df_hien["so_kh"].apply(fmt_so)
            df_hien["Dư nợ (triệu đồng)"]      = df_hien["du_no"].apply(fmt_ty)
            df_hien["Nguồn TW (triệu đồng)"]   = df_hien["du_no_tw"].apply(fmt_ty)
            df_hien["Nguồn ĐP (triệu đồng)"]   = df_hien["du_no_dp"].apply(fmt_ty)
            df_hien["Dư nợ QH (triệu đồng)"]   = df_hien["du_no_qh"].apply(fmt_ty)
            ty_le_qh = ((df_hien["du_no_qh"] / df_hien["du_no"] * 100).round(2)) if (df_hien["du_no"] > 0).any() else pd.Series(0.0, index=df_hien.index)
            df_hien["Tỷ lệ QH %"]              = ty_le_qh.apply(lambda x: f"{x:.2f}".replace(".", ",") + "%")
            df_hien["Dư nợ khoanh (triệu đồng)"] = df_hien["du_no_khoanh"].apply(fmt_ty)
            df_hien["Giải ngân năm (triệu đồng)"] = df_hien["gn_nam"].apply(fmt_ty)
            df_hien["Thu nợ năm (triệu đồng)"] = df_hien["tn_nam"].apply(fmt_ty)
            df_hien["Tỷ trọng %"]              = df_hien["ty_trong"].apply(lambda x: f"{x:.1f}".replace(".", ",") + "%")

            cols_hien = [
                "Chương trình", "Số món vay", "Số KH",
                "Dư nợ (triệu đồng)", "Nguồn TW (triệu đồng)", "Nguồn ĐP (triệu đồng)",
                "Dư nợ QH (triệu đồng)", "Tỷ lệ QH %", "Dư nợ khoanh (triệu đồng)",
                "Giải ngân năm (triệu đồng)", "Thu nợ năm (triệu đồng)", "Tỷ trọng %"
            ]

            # Tất cả cột đã convert sang string qua fmt_ty/fmt_so → không dùng column_config
            st.dataframe(
                df_hien[cols_hien],
                use_container_width=True,
                hide_index=True,
            )

            df_top10 = df_ct.nlargest(10, "du_no").copy()
            df_top10["label"] = df_top10["ten_ct"].str[:30]
            df_top10["du_no_ty"] = df_top10["du_no"] / 1e6
            df_top10["du_no_tw_ty"] = df_top10["du_no_tw"] / 1e6
            df_top10["du_no_dp_ty"] = df_top10["du_no_dp"] / 1e6
            df_top10["du_no_tw_ty"] = df_top10["du_no_tw_ty"].where(df_top10["du_no_tw_ty"] > 0, 0)
            df_top10["du_no_dp_ty"] = df_top10["du_no_dp_ty"].where(df_top10["du_no_dp_ty"] > 0, 0)
            fig_ct = go.Figure()
            fig_ct.add_trace(go.Bar(
                y=df_top10["label"],
                x=df_top10["du_no_tw_ty"],
                name="TW",
                orientation="h",
                marker_color="#2E7D32",
                text=df_top10["du_no_tw_ty"].apply(lambda x: vn(x, 1) if x > 0 else ""),
                textposition="inside",
                insidetextanchor="middle",
                textfont=dict(color="white", size=10),
            ))
            fig_ct.add_trace(go.Bar(
                y=df_top10["label"],
                x=df_top10["du_no_dp_ty"],
                name="ĐP",
                orientation="h",
                marker_color="#E65100",
                text=df_top10["du_no_dp_ty"].apply(lambda x: vn(x, 1) if x > 0 else ""),
                textposition="inside",
                insidetextanchor="middle",
                textfont=dict(color="white", size=10),
            ))
            fig_ct.update_layout(
                barmode="stack",
                title="📊 Top 10 chương trình theo dư nợ",
                xaxis_title="Dư nợ (triệu đồng)",
                yaxis=dict(autorange="reversed"),
                height=max(350, len(df_top10) * 35),
                margin=dict(l=10, r=40, t=40, b=10),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            )
            st.plotly_chart(fig_ct, use_container_width=True, key="ct_bar_top10")

        st.markdown("**🟢 Thông tin tổng quát theo PGD**")
        if COT_TEN_PGD in df.columns:
            col_khoanh = COT_DU_NO_KHOANH
            # Chỉ lấy các cột cần dùng trong tab Tổng quan (từ đoạn PGD trở đi), không copy toàn bộ
            COT_CAN = [
                COT_TEN_PGD,
                COT_MA_KH,
                COT_SO_KU,
                COT_TONG_DU_NO,
                COT_DU_NO_QH,
                col_khoanh,
                COT_TEN_CT,
                COT_TEN_XA,
                COT_NGAY_DH,
                COT_LAI_TON,
                *HSTD_DS_CHO_VAY_NAM_ALIASES,
                *HSTD_THU_NO_NAM_ALIASES,
            ]
            cot_lay = [c for c in COT_CAN if c in df.columns]
            for c in df.columns:
                if c in cot_lay:
                    continue
                s = str(c).replace("\n", " ").lower()
                if (
                    "giải ngân" in s
                    and "tháng" not in str(c).lower()
                    and (
                        "trong năm" in s
                        or str(c).strip().lower().endswith("năm")
                    )
                ):
                    cot_lay.append(c)
                elif (
                    "thu nợ" in s
                    and "tháng" not in str(c).lower()
                    and (
                        "trong năm" in s
                        or (
                            "năm" in str(c).lower()
                            and any(x in s for x in ("th ", "qh ", "khoanh"))
                        )
                    )
                ):
                    cot_lay.append(c)
            df = df[cot_lay]
            df_pgd = _cache_heatmap_pgd(
                df,
                ts,
                pgd_user,
                pgd_filter,
                COT_TEN_PGD,
                COT_TONG_DU_NO,
                COT_MA_KH,
                COT_DU_NO_QH,
            )
            df_pgd = df_pgd.rename(
                columns={
                    "du_no": "Dư nợ (triệu đồng)",
                    "so_kh": "Số KH",
                    "nqh": "QH (triệu đồng)",
                }
            )
            if col_khoanh in df.columns:
                _kh = df.groupby(COT_TEN_PGD, as_index=False).agg(
                    **{"Khoanh (triệu đồng)": (col_khoanh, "sum")}
                )
            else:
                _kh = df.groupby(COT_TEN_PGD, as_index=False).agg(
                    **{"Khoanh (triệu đồng)": (COT_MA_KH, "size")}
                )
            df_pgd = df_pgd.merge(_kh, on=COT_TEN_PGD, how="left")

            # Bổ sung PGD trong DS_PGD nhưng không có dòng trong df → hiển thị với giá trị 0
            pgd_co_trong_bang = set(df_pgd[COT_TEN_PGD].tolist())
            pgd_thieu_bang = [p for p in [DON_VI_CHI_NHANH] + DS_PGD if p not in pgd_co_trong_bang]
            if pgd_thieu_bang:
                rows_thieu = [{COT_TEN_PGD: p} for p in pgd_thieu_bang]
                df_thieu = pd.DataFrame(rows_thieu)
                for cot in df_pgd.columns:
                    if cot != COT_TEN_PGD:
                        df_thieu[cot] = 0.0
                df_pgd = pd.concat([df_pgd, df_thieu], ignore_index=True)
                df_pgd = df_pgd.sort_values(COT_TEN_PGD).reset_index(drop=True)
                if pgd_thieu_bang == ["PGD Biên Hòa"]:
                    st.caption("⚠️ PGD Biên Hòa chưa upload dữ liệu.")
                else:
                    st.warning(
                        f"⚠️ {len(pgd_thieu_bang)} PGD chưa có dữ liệu: "
                        f"{', '.join(pgd_thieu_bang)}"
                    )

            for cot in ["Dư nợ (triệu đồng)", "QH (triệu đồng)", "Khoanh (triệu đồng)"]:
                if cot in df_pgd.columns:
                    df_pgd[cot] = pd.to_numeric(df_pgd[cot], errors="coerce").fillna(0)
            df_pgd["Dư nợ (triệu đồng)"] = (df_pgd["Dư nợ (triệu đồng)"] / 1e6).round(0)
            df_pgd["QH (triệu đồng)"] = (df_pgd["QH (triệu đồng)"] / 1e6).round(0)
            if col_khoanh in df.columns:
                df_pgd["Khoanh (triệu đồng)"] = (df_pgd["Khoanh (triệu đồng)"] / 1e6).round(0)
            else:
                df_pgd["Khoanh (triệu đồng)"] = 0.0

            # A) Lãi tồn (triệu đồng)
            if COT_LAI_TON in df.columns:
                _lai_ton = df.groupby(COT_TEN_PGD)[COT_LAI_TON].apply(
                    lambda x: pd.to_numeric(x, errors="coerce").fillna(0).sum()
                ).reset_index(name="Lãi tồn (triệu đồng)")
                _lai_ton["Lãi tồn (triệu đồng)"] = (_lai_ton["Lãi tồn (triệu đồng)"] / 1e6).round(0)
                df_pgd = df_pgd.merge(_lai_ton, on=COT_TEN_PGD, how="left")
            else:
                df_pgd["Lãi tồn (triệu đồng)"] = 0.0
            df_pgd["Lãi tồn (triệu đồng)"] = pd.to_numeric(df_pgd["Lãi tồn (triệu đồng)"], errors="coerce").fillna(0.0)

            # B) Nợ đến hạn trong năm (triệu đồng)
            if COT_NGAY_DH in df.columns:
                _df_dh = df.copy()
                _df_dh[COT_NGAY_DH] = pd.to_datetime(_df_dh[COT_NGAY_DH], dayfirst=True, errors="coerce")
                _mask = _df_dh[COT_NGAY_DH].dt.year == int(NAM_HT)
                _dh = (
                    _df_dh[_mask]
                    .groupby(COT_TEN_PGD)[COT_TONG_DU_NO]
                    .apply(lambda x: pd.to_numeric(x, errors="coerce").fillna(0).sum())
                    .reset_index(name="Nợ ĐH năm (triệu đồng)")
                )
                _dh["Nợ ĐH năm (triệu đồng)"] = (_dh["Nợ ĐH năm (triệu đồng)"] / 1e6).round(0)
                df_pgd = df_pgd.merge(_dh, on=COT_TEN_PGD, how="left")
            else:
                df_pgd["Nợ ĐH năm (triệu đồng)"] = 0.0
            df_pgd["Nợ ĐH năm (triệu đồng)"] = pd.to_numeric(
                df_pgd["Nợ ĐH năm (triệu đồng)"], errors="coerce"
            ).fillna(0.0)

            # C) Doanh số cho vay năm (triệu đồng) — ưu tiên tên chuẩn GQVL + alias
            _col_cv = next(
                (c for c in HSTD_DS_CHO_VAY_NAM_ALIASES if c in df.columns),
                None,
            )
            if _col_cv is None:
                _col_cv = next(
                    (
                        c
                        for c in df.columns
                        if "giải ngân" in str(c).replace("\n", " ").lower()
                        and "tháng" not in str(c).lower()
                        and (
                            "trong năm" in str(c).replace("\n", " ").lower()
                            or str(c).strip().lower().endswith("năm")
                        )
                    ),
                    None,
                )
            if _col_cv is not None:
                _cv = df.groupby(COT_TEN_PGD)[_col_cv].apply(
                    lambda x: pd.to_numeric(x, errors="coerce").fillna(0).sum()
                ).reset_index(name="DS Cho vay (triệu đồng)")
                _cv["DS Cho vay (triệu đồng)"] = (_cv["DS Cho vay (triệu đồng)"] / 1e6).round(0)
                df_pgd = df_pgd.merge(_cv, on=COT_TEN_PGD, how="left")
            else:
                df_pgd["DS Cho vay (triệu đồng)"] = 0.0
            df_pgd["DS Cho vay (triệu đồng)"] = pd.to_numeric(
                df_pgd["DS Cho vay (triệu đồng)"], errors="coerce"
            ).fillna(0.0)

            # D) Doanh số thu nợ năm (triệu đồng)
            _thu_cols = [c for c in HSTD_THU_NO_NAM_ALIASES if c in df.columns]
            if not _thu_cols:
                _thu_cols = [
                    c
                    for c in df.columns
                    if "thu nợ" in str(c).replace("\n", " ").lower()
                    and "tháng" not in str(c).lower()
                    and (
                        "trong năm" in str(c).replace("\n", " ").lower()
                        or (
                            "năm" in str(c).lower()
                            and any(
                                x in str(c).replace("\n", " ").lower()
                                for x in ("th ", "qh ", "khoanh")
                            )
                        )
                    )
                ]
            if _thu_cols:
                _df_thu = df.copy()
                for _c in _thu_cols:
                    _df_thu[_c] = pd.to_numeric(_df_thu[_c], errors="coerce").fillna(0)
                _df_thu["_thu_no_nam"] = _df_thu[_thu_cols].sum(axis=1)
                _thu = _df_thu.groupby(COT_TEN_PGD)["_thu_no_nam"].sum().reset_index(name="DS Thu nợ (triệu đồng)")
                _thu["DS Thu nợ (triệu đồng)"] = (_thu["DS Thu nợ (triệu đồng)"] / 1e6).round(0)
                df_pgd = df_pgd.merge(_thu, on=COT_TEN_PGD, how="left")
            else:
                df_pgd["DS Thu nợ (triệu đồng)"] = 0.0
            df_pgd["DS Thu nợ (triệu đồng)"] = pd.to_numeric(
                df_pgd["DS Thu nợ (triệu đồng)"], errors="coerce"
            ).fillna(0.0)

            if _col_cv is None or not _thu_cols:
                st.caption("⚠️ Không có cột DS Cho vay/Thu nợ trong HSTD")

            df_pgd["TL QH %"] = ((df_pgd["QH (triệu đồng)"] / df_pgd["Dư nợ (triệu đồng)"].replace(0, pd.NA)) * 100).fillna(0).round(2)
            df_pgd["TL Khoanh %"] = ((df_pgd["Khoanh (triệu đồng)"] / df_pgd["Dư nợ (triệu đồng)"].replace(0, pd.NA)) * 100).fillna(0).round(2)

            try:
                ds_thang = ds_thang_nam()
                if ds_thang:
                    df_to_raw = doc_cdtotkvv(ds_thang[0])
                    if df_to_raw is not None and not df_to_raw.empty:
                        df_to_pgd = tong_hop_theo_pgd(df_to_raw)

                        # Map tên viết tắt trong CDTOTKVV → tên chuẩn (DON_VI_CHI_NHANH)
                        # "PGD Biên Hòa" là alias cho file cũ — trong HSTD tên thực là "Hội sở Chi nhánh tỉnh"
                        _TEN_MAP = {
                            "Hội sở CN Đồng Nai":   DON_VI_CHI_NHANH,
                            "Hội sở CN Bình Phước": DON_VI_CHI_NHANH,
                            "PGD Biên Hòa":         DON_VI_CHI_NHANH,
                        }
                        def _map_ten(val):
                            val = str(val).strip()
                            if val in _TEN_MAP:
                                return _TEN_MAP[val]
                            # Fuzzy: ten_dv chứa tên PGD chuẩn hoặc ngược lại
                            val_low = val.lower()
                            for ten in [DON_VI_CHI_NHANH] + DS_PGD:
                                if ten.lower() in val_low or val_low in ten.lower():
                                    return ten
                            return val

                        # FIX: đổi "ten_dv" (đúng) thay vì "ten_pgd" (sai)
                        df_to_pgd[COT_TEN_PGD] = df_to_pgd["ten_dv"].apply(_map_ten)
                        df_to_pgd = df_to_pgd.rename(columns={
                            "tong_to": "Tổng Tổ",
                            "to_tot":  "Tốt",
                            "to_kha":  "Khá",
                            "to_tb":   "TB",
                            "to_yeu":  "Yếu",
                        })
                        cot_to = [COT_TEN_PGD, "Tổng Tổ", "Tốt", "Khá", "TB", "Yếu"]
                        # Gộp nếu nhiều dòng cùng tên PGD sau khi map
                        df_to_pgd = (
                            df_to_pgd[cot_to]
                            .groupby(COT_TEN_PGD, as_index=False)
                            .sum()
                        )
                        df_pgd = df_pgd.merge(df_to_pgd, on=COT_TEN_PGD, how="left")
            except Exception:
                pass

            for cot in ["Tổng Tổ", "Tốt", "Khá", "TB", "Yếu"]:
                if cot in df_pgd.columns:
                    df_pgd[cot] = df_pgd[cot].fillna(0).round(0).astype(int)

            # Tính TL QH% và TL Khoanh% cho từng PGD
            df_pgd["TL QH %"] = (
                (df_pgd["QH (triệu đồng)"] / df_pgd["Dư nợ (triệu đồng)"].replace(0, pd.NA)) * 100
            ).fillna(0).round(2)
            df_pgd["TL Khoanh %"] = (
                (df_pgd["Khoanh (triệu đồng)"] / df_pgd["Dư nợ (triệu đồng)"].replace(0, pd.NA)) * 100
            ).fillna(0).round(2)

            # Tính Nợ xấu (NPL) = QH + Khoanh
            df_pgd["Nợ xấu (triệu đồng)"] = (df_pgd["QH (triệu đồng)"] + df_pgd["Khoanh (triệu đồng)"]).round(3)
            df_pgd["Tỷ lệ Nợ xấu"] = (
                (df_pgd["Nợ xấu (triệu đồng)"] / df_pgd["Dư nợ (triệu đồng)"].replace(0, pd.NA)) * 100
            ).fillna(0).round(2)

            # Merge dữ liệu Tổ TK&VV theo PGD (nếu có)
            try:
                _df_to_pgd_map = None
                _ds_thang = ds_thang_nam()
                if _ds_thang:
                    for _th in _ds_thang:
                        _df_to_r = doc_cdtotkvv(_th)
                        if _df_to_r is not None and not _df_to_r.empty:
                            _df_to_pgd_map = tong_hop_theo_pgd(_df_to_r)
                            break
                if _df_to_pgd_map is None or _df_to_pgd_map.empty:
                    from data.cdtotkvv import tong_hop_tu_pgd_data
                    _df_raw_pgd = tong_hop_tu_pgd_data()
                    if _df_raw_pgd is not None and not _df_raw_pgd.empty:
                        _df_to_pgd_map = tong_hop_theo_pgd(_df_raw_pgd)
            except Exception:
                _df_to_pgd_map = None

            tong = {COT_TEN_PGD: "Toàn Chi nhánh"}
            cot_so = [c for c in df_pgd.columns if c != COT_TEN_PGD and pd.api.types.is_numeric_dtype(df_pgd[c])]
            for cot in cot_so:
                tong[cot] = df_pgd[cot].sum()
            du_no_tong_trieu = tong.get("Dư nợ (triệu đồng)", 0) * 1000
            tong["TL QH %"] = round((tong.get("QH (triệu đồng)", 0) / tong.get("Dư nợ (triệu đồng)", 1) * 100), 2) if tong.get("Dư nợ (triệu đồng)", 0) else 0
            tong["TL Khoanh %"] = round((tong.get("Khoanh (triệu đồng)", 0) / tong.get("Dư nợ (triệu đồng)", 1) * 100), 2) if tong.get("Dư nợ (triệu đồng)", 0) else 0
            tong["Nợ xấu (triệu đồng)"] = round(tong.get("QH (triệu đồng)", 0) + tong.get("Khoanh (triệu đồng)", 0), 3)
            tong["Tỷ lệ Nợ xấu"] = round(tong["Nợ xấu (triệu đồng)"] / tong.get("Dư nợ (triệu đồng)", 1) * 100, 2) if tong.get("Dư nợ (triệu đồng)", 0) else 0
            tong_clean = {k: (v if pd.notna(v) else 0) for k, v in tong.items()
                          if k != COT_TEN_PGD}
            tong_clean[COT_TEN_PGD] = "Toàn Chi nhánh"
            df_pgd = pd.concat([df_pgd, pd.DataFrame([tong_clean])], ignore_index=True)

            # Merge điểm tổ theo PGD vào df_pgd nếu có
            if _df_to_pgd_map is not None and not _df_to_pgd_map.empty:
                # "PGD Biên Hòa" là alias cho file cũ — trong HSTD tên thực là "Hội sở Chi nhánh tỉnh"
                _TEN_MAP = {
                    "Hội sở CN Đồng Nai": DON_VI_CHI_NHANH,
                    "Hội sở CN Bình Phước": DON_VI_CHI_NHANH,
                    "PGD Biên Hòa":       DON_VI_CHI_NHANH,
                }
                def _map_ten_to(val):
                    val = str(val).strip()
                    if val in _TEN_MAP:
                        return _TEN_MAP[val]
                    val_low = val.lower()
                    for ten in [DON_VI_CHI_NHANH] + DS_PGD:
                        if ten.lower() in val_low or val_low in ten.lower():
                            return ten
                    return val
                _df_to_pgd_map[COT_TEN_PGD] = _df_to_pgd_map["ten_dv"].apply(_map_ten_to)
                _df_to_pgd_map = _df_to_pgd_map.rename(columns={
                    "tong_to": "Tổng Tổ", "to_tot": "Tốt",
                    "to_kha": "Khá", "to_tb": "TB", "to_yeu": "Yếu",
                })
                _cot_to = [COT_TEN_PGD, "Tổng Tổ", "Tốt", "Khá", "TB", "Yếu"]
                _df_to_pgd_map = _df_to_pgd_map[_cot_to].groupby(COT_TEN_PGD, as_index=False).sum()
                # Xóa cột cũ nếu đã có rồi merge lại
                for _c in ["Tổng Tổ", "Tốt", "Khá", "TB", "Yếu"]:
                    if _c in df_pgd.columns:
                        df_pgd = df_pgd.drop(columns=[_c])
                df_pgd = df_pgd.merge(_df_to_pgd_map, on=COT_TEN_PGD, how="left")
                for _c in ["Tổng Tổ", "Tốt", "Khá", "TB", "Yếu"]:
                    if _c in df_pgd.columns:
                        df_pgd[_c] = df_pgd[_c].fillna(0).round(0).astype(int)

            # Cập nhật dòng tổng cho cột Tổ TK&VV
            for _c in ["Tổng Tổ", "Tốt", "Khá", "TB", "Yếu"]:
                if _c in df_pgd.columns:
                    tong[_c] = int(df_pgd[_c].iloc[:-1].sum())

            # Cập nhật lại dòng tổng trong df_pgd
            for _key, _val in tong.items():
                if _key in df_pgd.columns:
                    df_pgd.loc[df_pgd.index[-1], _key] = _val

            cot_hien = [
                COT_TEN_PGD,
                "Số KH", "Dư nợ (triệu đồng)",
                "QH (triệu đồng)", "TL QH %",
                "Khoanh (triệu đồng)", "TL Khoanh %",
                "Nợ xấu (triệu đồng)", "Tỷ lệ Nợ xấu",
                "Lãi tồn (tỷ)",
                "DS Cho vay (tỷ)", "DS Thu nợ (tỷ)",
            ]
            cot_hien += [c for c in ["Tổng Tổ", "Tốt", "Khá", "TB", "Yếu"] if c in df_pgd.columns]
            cot_hien = [c for c in cot_hien if c in df_pgd.columns]
            # Tách riêng dòng Hội sở Chi nhánh tỉnh để hiển thị đầu bảng
            df_hoi_so = df_pgd[df_pgd[COT_TEN_PGD] == DON_VI_CHI_NHANH].copy()
            # Loại bỏ dòng DON_VI_CHI_NHANH khỏi df_pgd (giữ nguyên logic cũ)
            df_pgd = df_pgd[df_pgd[COT_TEN_PGD] != DON_VI_CHI_NHANH].reset_index(drop=True)
            # Ghép Hội sở lên đầu bảng nếu có dữ liệu
            if not df_hoi_so.empty:
                df_pgd = pd.concat([df_hoi_so, df_pgd], ignore_index=True)
            
            # Xây dựng column_config cho bảng PGD
            column_config_pgd = _tao_column_config_pgd()
            # Thêm các cột số nguyên
            for cot in ["Số KH", "Tổng Tổ", "Tốt", "Khá", "TB", "Yếu"]:
                if cot in df_pgd.columns:
                    column_config_pgd[cot] = st.column_config.NumberColumn(cot, format=",.0f")

            # ── Bảng HTML có header nhóm cột ─────────────────────────────
            df_show = df_pgd[cot_hien].copy()
            # Đánh dấu dòng tổng (dòng cuối)
            is_tong = df_show[COT_TEN_PGD] == "Toàn Chi nhánh"

            def _fmt_cell(val, col):
                """Hiển thị ô bảng theo chuẩn VN (. nghìn, , thập phân) — dùng vn / fmt_so."""
                if pd.isna(val) or val == "":
                    return "—"
                if col in ["Dư nợ (triệu đồng)"]:
                    return vn(float(val) / 1000, 2)
                if col in ["TL QH %", "TL Khoanh %", "Tỷ lệ Nợ xấu"]:
                    return f"{vn(float(val), 2)}%"
                if col in [
                    "QH (triệu đồng)",
                    "Khoanh (triệu đồng)",
                    "Nợ xấu (triệu đồng)",
                    "Lãi tồn (triệu đồng)",
                    "Nợ ĐH năm (triệu đồng)",
                    "DS Cho vay (triệu đồng)",
                    "DS Thu nợ (triệu đồng)",
                ]:
                    return vn(float(val), 3)
                if col in ["Số KH", "Tổng Tổ", "Tốt", "Khá", "TB", "Yếu"]:
                    try:
                        return fmt_so(val)
                    except Exception:
                        return str(val)
                return str(val)

            # Header nhóm
            NHOM_COT = [
                ("", 1),                    # Tên PGD
                ("Dư nợ", 2),               # Số KH, Dư nợ (tỷ)
                ("Chất lượng nợ", 7),       # QH, TL QH, Khoanh, TL Khoanh, Nợ xấu, TL NPL, Lãi tồn
                ("Kế hoạch năm", 3),        # Nợ ĐH, DS Cho vay, DS Thu nợ
                ("Tổ TK&VV", len([c for c in ["Tổng Tổ", "Tốt", "Khá", "TB", "Yếu"] if c in df_pgd.columns])),
            ]
            # Bỏ nhóm có colspan=0
            NHOM_COT = [(n, s) for n, s in NHOM_COT if s > 0]

            def _disp_col(col: str) -> str:
                """Tên cột hiển thị 2 dòng: phần đơn vị xuống dòng."""
                if col == COT_TEN_PGD:
                    return "Đơn vị"
                if col == "Dư nợ (triệu đồng)":
                    return "Dư nợ<br/><span style='font-weight:normal;font-size:11px'>(Tỷ đồng)</span>"
                if col.endswith("(triệu đồng)"):
                    return col[:-len("(triệu đồng)")].strip() + "<br/><span style='font-weight:normal;font-size:11px'>(Triệu đồng)</span>"
                if col.endswith("(tỷ)"):
                    return col[:-len("(tỷ)")].strip() + "<br/><span style='font-weight:normal;font-size:11px'>(Tỷ đồng)</span>"
                if col == "Tỷ lệ Nợ xấu":
                    return "Tỷ lệ<br/>Nợ xấu"
                return col

            header1 = "".join(
                f'<th colspan="{span}" style="background:#2E7D32;color:#fff;'
                f'text-align:center;padding:7px 5px;border:1px solid #1B5E20;font-size:13px">'
                f'{nhom}</th>'
                for nhom, span in NHOM_COT
            )
            header2 = "".join(
                f'<th style="background:#388E3C;color:#fff;text-align:center;'
                f'padding:6px 4px;border:1px solid #1B5E20;font-size:12px;'
                f'white-space:normal;min-width:60px;line-height:1.4">'
                f'{_disp_col(c)}</th>'
                for c in cot_hien
            )

            rows_html = ""
            for i, (_, row) in enumerate(df_show.iterrows()):
                is_last = row[COT_TEN_PGD] == "Toàn Chi nhánh"
                bg = "#C8E6C9" if is_last else ("#F5F7FA" if i % 2 == 0 else "#FFFFFF")
                fw = "bold" if is_last else "normal"
                row_fs = "0.92rem" if is_last else "0.90rem"
                cells = "".join(
                    f'<td style="padding:6px 7px;border:1px solid #E0E0E0;'
                    f'text-align:{"left" if c == cot_hien[0] else "right"};'
                    f'font-weight:{fw};font-size:{row_fs};white-space:nowrap;'
                    f'{"color:#C62828;font-weight:800" if c == "TL QH %" and pd.to_numeric(row.get(c, 0), errors="coerce") > 0.5 else ""}'
                    f'{"color:#C62828;font-weight:800" if c == "Tỷ lệ Nợ xấu" and pd.to_numeric(row.get(c, 0), errors="coerce") > 0.3 else ""}'
                    f'{"color:#C62828;font-weight:800" if c == "TL Khoanh %" and pd.to_numeric(row.get(c, 0), errors="coerce") > 1 else ""}'
                    f'">'
                    f'{_fmt_cell(row[c], c)}</td>'
                    for c in cot_hien
                )
                rows_html += f'<tr style="background:{bg}">{cells}</tr>\n'

            html_table = f"""
            <div style="overflow-x:auto;margin:8px 0">
            <table style="border-collapse:collapse;width:100%;font-family:sans-serif">
              <thead>
                <tr>{header1}</tr>
                <tr>{header2}</tr>
              </thead>
              <tbody>{rows_html}</tbody>
            </table>
            <p style="font-size:0.78rem;color:#6B7280;margin:4px 0">
              * Đơn vị: Dư nợ = tỷ đồng | Các cột tiền khác = triệu đồng
            </p>
            </div>
            """
            st.markdown(html_table, unsafe_allow_html=True)

            # Format từng ô theo _fmt_cell trước khi xuất PDF
            df_export = df_show.copy().rename(columns={COT_TEN_PGD: "Đơn vị"})
            for col in df_export.columns:
                if col != "Đơn vị":
                    df_export[col] = df_export[col].apply(
                        lambda v, c=col: _fmt_cell(v, c)
                    )

            col_ex, col_pdf = st.columns(2)

            with col_ex:
                if st.button("📥 Xuất Excel", key="btn_excel_tqpgd", width='stretch'):
                    try:
                        _ten_excel = ten_file_xuat("TQPGD")
                        buf = _xuat_excel_tqpgd(df_show, _ten_excel)
                        st.session_state["_excel_bytes_tqpgd"] = buf
                        st.session_state["_excel_file_tqpgd"] = _ten_excel
                    except Exception as e:
                        st.error(f"❌ Lỗi xuất Excel: {e}")
                if st.session_state.get("_excel_bytes_tqpgd"):
                    st.download_button(
                        "⬇ Tải Excel",
                        data=st.session_state["_excel_bytes_tqpgd"],
                        file_name=st.session_state.get("_excel_file_tqpgd", "TQPGD.xlsx"),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="btn_excel_tqpgd_dl",
                        width='stretch',
                    )

            with col_pdf:
                _ss_tqpgd = "_pdf_bytes_tqpgd"
                _ssf_tqpgd = "_pdf_file_tqpgd"
                if st.button("📄 Xuất PDF", key="btn_pdf_tqpgd", type="primary", width='stretch'):
                    try:
                        # Đổi tên cột → 2 dòng cho header PDF (Reportlab Paragraph hỗ trợ <br/>)
                        def _pdf_col(col: str) -> str:
                            if col.endswith("(triệu đồng)"):
                                return col[:-len("(triệu đồng)")].strip() + "<br/>(Triệu đồng)"
                            if col.endswith("(tỷ)"):
                                return col[:-len("(tỷ)")].strip() + "<br/>(Tỷ đồng)"
                            if col == "Tỷ lệ Nợ xấu":
                                return "Tỷ lệ<br/>Nợ xấu"
                            return col
                        df_pdf = df_export.rename(columns={c: _pdf_col(c) for c in df_export.columns})

                        # Các cột số → căn phải (dữ liệu đã format sẵn thành string)
                        _cols_right_pdf = [c for c in df_pdf.columns if c != "Đơn vị"]
                        with st.spinner("⏳ Đang tạo PDF..."):
                            _bytes = xuat_pdf(
                                df_pdf,
                                "Thông tin tổng quát theo PGD",
                                username,
                                cols_tien=[],
                                cols_right=_cols_right_pdf,
                                prefix_file="TQPGD",
                            )
                        st.session_state[_ss_tqpgd] = _bytes
                        st.session_state[_ssf_tqpgd] = f"TQPGD_{datetime.now().strftime('%d%m%Y_%H%M')}.pdf"
                        db.ghi_audit(username or "unknown", "xuat_pdf", f"TQPGD")
                    except Exception as _e:
                        import traceback

                        st.session_state[_ss_tqpgd] = None
                        st.error(f"❌ Lỗi: {_e}")
                        st.code(traceback.format_exc())

                if st.session_state.get(_ss_tqpgd):
                    st.download_button(
                        label="⬇ Tải file PDF TQPGD",
                        data=st.session_state[_ss_tqpgd],
                        file_name=st.session_state.get(_ssf_tqpgd, "TQPGD.pdf"),
                        mime="application/pdf",
                        key="btn_pdf_tqpgd_dl",
                        width='stretch',
                    )
        st.divider()
        st.subheader("🔔 Hồ sơ đến hạn — Tổng hợp")
        if COT_NGAY_DH in df.columns:
            try:
                dt = df.copy()
                dt[COT_NGAY_DH] = pd.to_datetime(dt[COT_NGAY_DH], dayfirst=True, errors="coerce")
                hn       = pd.Timestamp.today().normalize()
                cuoi_nam = pd.Timestamp(hn.year, 12, 31)


                # Lựa chọn nhóm tổng hợp
                nhom_chon = st.radio(
                    "Tổng hợp theo",
                    ["Chương trình", "PGD", "Xã"],
                    horizontal=True,
                    key="tq_denh_nhom",
                )
                NHOM_COT = {
                    "Chương trình": COT_TEN_CT,
                    "PGD":          COT_TEN_PGD,
                    "Xã":           COT_TEN_XA,
                }
                nhom_col = NHOM_COT[nhom_chon]

                # Bộ lọc kết hợp nhiều điều kiện
                with st.expander("🔍 Bộ lọc nâng cao", expanded=False):
                    col1, col2, col3 = st.columns(3)

                    with col1:
                        ds_pgd = sorted(dt[COT_TEN_PGD].dropna().unique()) if COT_TEN_PGD in dt.columns else []
                        loc_pgd = st.multiselect("Lọc PGD", options=ds_pgd, default=[], key="tq_loc_pgd")

                    with col2:
                        ds_ct = sorted(dt[COT_TEN_CT].dropna().unique()) if COT_TEN_CT in dt.columns else []
                        loc_ct = st.multiselect("Lọc Chương trình", options=ds_ct, default=[], key="tq_loc_ct")

                    with col3:
                        # Chỉ lấy xã thuộc PGD đã chọn, nếu chưa chọn PGD thì hiện tất cả
                        if loc_pgd:
                            dt_theo_pgd = dt[dt[COT_TEN_PGD].isin(loc_pgd)]
                        else:
                            dt_theo_pgd = dt

                        cot_xa = next((c for c in [COT_TEN_XA, "Tên xã"] if c in dt_theo_pgd.columns), None)
                        ds_xa = sorted(dt_theo_pgd[cot_xa].dropna().unique()) if cot_xa else []
                        loc_xa = st.multiselect("Lọc Xã", options=ds_xa, default=[], key="tq_loc_xa")

                # Áp dụng bộ lọc vào dt trước khi lọc theo mốc thời gian
                dt_loc = dt.copy()
                if loc_pgd:
                    dt_loc = dt_loc[dt_loc[COT_TEN_PGD].isin(loc_pgd)]
                if loc_ct:
                    dt_loc = dt_loc[dt_loc[COT_TEN_CT].isin(loc_ct)]
                if loc_xa and cot_xa:
                    dt_loc = dt_loc[dt_loc[cot_xa].isin(loc_xa)]

                # Loại bỏ hồ sơ dư nợ = 0
                dt_loc = dt_loc[dt_loc[COT_TONG_DU_NO].fillna(0) > 0]

                MOC = {
                    "1 tháng":   hn + pd.Timedelta(days=30),
                    "3 tháng":   hn + pd.Timedelta(days=90),
                    "6 tháng":   hn + pd.Timedelta(days=180),
                    "Trong năm": cuoi_nam,
                }

                tab_1m, tab_3m, tab_6m, tab_nam = st.tabs([
                    "📅 1 tháng", "📅 3 tháng", "📅 6 tháng", "📅 Trong năm"
                ])

                def _build_pdf_den_han(df_loc, label, loc_pgd, loc_ct, loc_xa, username, key_prefix):
                    """Build PDF bytes: luôn groupby [PGD, Xã, Chương trình], bộ lọc chỉ thu hẹp input."""
                    COLS_GROUP = [COT_TEN_PGD, COT_TEN_XA, COT_TEN_CT]
                    cols_ok = [c for c in COLS_GROUP if c in df_loc.columns]

                    RENAME_MAP = {COT_TEN_PGD: "PGD", COT_TEN_XA: "Xã", COT_TEN_CT: "Chương trình"}
                    rename_ok = {k: v for k, v in RENAME_MAP.items() if k in df_loc.columns}

                    pdf_tg = df_loc.groupby(cols_ok).agg(
                        _mon=(COT_SO_KU,      "nunique"),
                        _kh =(COT_MA_KH,      "nunique"),
                        _no =(COT_TONG_DU_NO, "sum"),
                    ).reset_index()

                    pdf_tg["Số món vay"] = pdf_tg["_mon"].apply(fmt_so)
                    pdf_tg["Số KH"]      = pdf_tg["_kh"].apply(fmt_so)
                    pdf_tg["Dư nợ (triệu đồng)"] = pdf_tg["_no"].apply(fmt_bang_ty)

                    pdf_tg = pdf_tg.rename(columns=rename_ok)
                    pdf_tg = pdf_tg.sort_values(
                        by=[c for c in ["PGD", "Xã", "Chương trình"] if c in pdf_tg.columns],
                        ascending=True,
                    )

                    cols_hien_thi = [v for v in ["PGD", "Xã", "Chương trình"] if v in pdf_tg.columns]
                    df_pdf = pdf_tg[[*cols_hien_thi, "Số món vay", "Số KH", "Dư nợ (triệu đồng)"]]

                    return xuat_pdf(
                        df_pdf,
                        f"Hồ sơ đến hạn {label} — Tổng hợp theo PGD / Xã / Chương trình",
                        username,
                        cols_tien=[],
                        prefix_file=f"HoSoDenHan_{key_prefix}",
                    )

                def _bang_den_han(df_loc, label, key_prefix):
                    tg = None
                    if df_loc.empty:
                        st.success(f"✅ Không có món vay đến hạn {label}")
                        return

                    tong_no  = df_loc[COT_TONG_DU_NO].fillna(0).sum()
                    tong_mon = df_loc[COT_SO_KU].nunique()
                    tong_kh  = df_loc[COT_MA_KH].nunique()

                    c1, c2, c3 = st.columns(3)
                    c1.metric("Số món vay", fmt_so(tong_mon))
                    c2.metric("Số khách hàng", fmt_so(tong_kh))
                    c3.metric("Tổng dư nợ", fmt(tong_no))

                    st.divider()

                    if nhom_col in df_loc.columns:
                        if nhom_chon == "Xã" and COT_TEN_PGD in df_loc.columns:
                            tg = df_loc.groupby([COT_TEN_PGD, nhom_col]).agg(
                                _mon=(COT_SO_KU,      "nunique"),
                                _kh =(COT_MA_KH,      "nunique"),
                                _no =(COT_TONG_DU_NO, "sum"),
                            ).reset_index().sort_values("_no", ascending=False)
                            tg["Số món vay"] = tg["_mon"].apply(fmt_so)
                            tg["Số KH"]      = tg["_kh"].apply(fmt_so)
                            tg["Dư nợ (triệu đồng)"] = tg["_no"].apply(fmt_bang_ty)
                            cols_hien_thi = [COT_TEN_PGD, nhom_col, "Số món vay", "Số KH", "Dư nợ (triệu đồng)"]
                            rename_map = {COT_TEN_PGD: "PGD", nhom_col: "Xã"}
                        else:
                            tg = df_loc.groupby(nhom_col).agg(
                                _mon=(COT_SO_KU,      "nunique"),
                                _kh =(COT_MA_KH,      "nunique"),
                                _no =(COT_TONG_DU_NO, "sum"),
                            ).reset_index().sort_values("_no", ascending=False)
                            tg["Số món vay"] = tg["_mon"].apply(fmt_so)
                            tg["Số KH"]      = tg["_kh"].apply(fmt_so)
                            tg["Dư nợ (triệu đồng)"] = tg["_no"].apply(fmt_bang_ty)
                            cols_hien_thi = [nhom_col, "Số món vay", "Số KH", "Dư nợ (triệu đồng)"]
                            rename_map = {nhom_col: nhom_chon}

                        hien_thi_dataframe_phan_trang(
                            tg[cols_hien_thi].rename(columns=rename_map),
                            key=f"tongquan_den_han_{key_prefix}",
                        )

                        if nhom_col in df_loc.columns and len(tg) > 0:
                            top10 = tg.nlargest(10, "_no")

                            fig = go.Figure(go.Pie(
                                labels=top10[nhom_col],
                                values=top10["_no"],
                                hole=0.5,
                                textinfo="label+percent",
                                textposition="outside",
                                hovertemplate="<b>%{label}</b><br>Dư nợ: %{customdata}<br>Tỷ lệ: %{percent}<extra></extra>",
                                customdata=top10["Dư nợ (triệu đồng)"],
                                marker=dict(
                                    colors=px.colors.sequential.Greens_r[: len(top10)],
                                    line=dict(color="white", width=2),
                                ),
                            ))

                            fig.update_layout(
                                title=f"Top 10 {nhom_chon} có dư nợ đến hạn cao nhất",
                                height=450,
                                annotations=[dict(
                                    text=f"<b>{fmt(tong_no)}</b><br>Tổng dư nợ",
                                    x=0.5, y=0.5,
                                    font=dict(size=13),
                                    showarrow=False,
                                )],
                                legend=dict(orientation="v", x=1.05, y=0.5),
                                margin=dict(l=0, r=150, t=50, b=0),
                                paper_bgcolor="rgba(0,0,0,0)",
                            )

                            st.plotly_chart(fig, width='stretch', key=f"pie_den_han_{key_prefix}")

                        if tg is not None:
                            st.divider()
                            col_ex, col_pdf = st.columns(2)

                            # PDF: luôn groupby [PGD, Xã, Chương trình] — bộ lọc chỉ thu hẹp input
                            ss_key      = f"_pdf_bytes_denh_{key_prefix}"
                            ss_file_key = f"_pdf_file_denh_{key_prefix}"

                            with col_ex:
                                # Excel: giữ logic hiện tại
                                co_loc = bool(loc_pgd or loc_ct or loc_xa)
                                if co_loc:
                                    df_excel = tg[[nhom_col, "_mon", "_kh", "_no"]].rename(columns={
                                        nhom_col: nhom_chon,
                                        "_mon": "Số món vay",
                                        "_kh":  "Số KH",
                                        "_no":  "Dư nợ",
                                    })
                                    ten_sheet = f"TH_{nhom_chon[:10]}_{key_prefix}"
                                else:
                                    COLS_CHI_TIET = [
                                        COT_TEN_PGD, COT_MA_KH, COT_TEN_KH, COT_SO_KU,
                                        COT_TEN_CT, COT_TONG_DU_NO, COT_NGAY_DH,
                                    ]
                                    cols_ok = [c for c in COLS_CHI_TIET if c in df_loc.columns]
                                    df_excel = df_loc[cols_ok].sort_values(COT_TEN_PGD).reset_index(drop=True)
                                    ten_sheet = f"ChiTiet_{key_prefix}"

                                excel_bytes = xuat_excel({ten_sheet: df_excel})
                                st.download_button(
                                    label="📥 Xuất Excel",
                                    data=excel_bytes,
                                    file_name=ten_file_xuat(f"HoSoDenHan_{key_prefix}"),
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    width='stretch',
                                    key=f"excel_den_han_{key_prefix}",
                                )

                            with col_pdf:
                                if st.button("📄 Xuất PDF", key=f"pdf_den_han_{key_prefix}",
                                             type="primary", width='stretch'):
                                    with st.spinner("⏳ Đang tạo PDF..."):
                                        pdf_bytes = _build_pdf_den_han(
                                            df_loc, label, loc_pgd, loc_ct, loc_xa, username, key_prefix
                                        )
                                    st.session_state[ss_key]      = pdf_bytes
                                    st.session_state[ss_file_key] = (
                                        f"HoSoDenHan_{key_prefix}_{datetime.now().strftime('%d%m%Y_%H%M')}.pdf"
                                    )

                            # Download button đặt NGOÀI block columns để tránh duplicate widget
                            if st.session_state.get(ss_key):
                                st.download_button(
                                    label="⬇ Tải file PDF",
                                    data=st.session_state[ss_key],
                                    file_name=st.session_state.get(ss_file_key, f"HoSoDenHan_{key_prefix}.pdf"),
                                    mime="application/pdf",
                                    key=f"dl_den_han_{key_prefix}",
                                    width='stretch',
                                )
                    else:
                        st.caption(f"⚠️ Không có cột '{nhom_chon}' trong dữ liệu")

                for (label, den), tab_ui in zip(MOC.items(),
                                                [tab_1m, tab_3m, tab_6m, tab_nam]):
                    with tab_ui:
                        df_moc = dt_loc[(dt_loc[COT_NGAY_DH] >= hn) & (dt_loc[COT_NGAY_DH] <= den)]
                        _bang_den_han(df_moc, label, label.replace(" ", "_"))

            except Exception as e:
                st.error(f"Lỗi xử lý đến hạn: {e}")

            # ── Download PDF: chỉ giữ nút trong mỗi tab (không duplicate) ──
            # Phần download ngoài tabs đã xóa để tránh trùng lặp với nút
            # "⬇ Tải file PDF" trong mỗi tab 1/3/6 tháng và Trong năm
