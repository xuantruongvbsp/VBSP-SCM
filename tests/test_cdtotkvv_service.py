"""Unit tests cho services/cdtotkvv_service.py — Chấm điểm Tổ TK&VV."""
from __future__ import annotations

from io import BytesIO

import openpyxl
import pandas as pd

import services.cdtotkvv_service as cdtotkvv_service
from config import (
    COT_NGAY_SINH,
    COT_TEN_KH,
    COT_TEN_PGD,
    COT_TEN_XA,
    DON_VI_CHI_NHANH,
)
from data.cdtotkvv import doc_cdtotkvv_path, doc_thang_tu_cdto_toan_cn, tach_file_cdto_toan_cn
from services.cdtotkvv_service import (
    loc_df,
    cdtotkvv_ten_sheet_excel,
    fmt_xuat_to_khong_dat_vn,
)


class TestLocDf:
    def test_df_None(self):
        df = loc_df(None, "cn", "")
        assert df is None

    def test_df_trong(self):
        df = loc_df(pd.DataFrame(), "cn", "")
        assert df.empty

    def test_mode_cn_tra_ve_toan_bo(self):
        df_in = pd.DataFrame({"ten_dv": ["PGD A", "PGD B"]})
        df_out = loc_df(df_in, "cn", "PGD A")
        assert len(df_out) == 2

    def test_mode_pgd_loc_theo_ten(self):
        df_in = pd.DataFrame({"ten_dv": ["PGD A", "PGD B"]})
        df_out = loc_df(df_in, "pgd", "PGD B")
        assert len(df_out) == 1
        assert df_out.iloc[0]["ten_dv"] == "PGD B"

    def test_mode_pgd_khong_khop(self):
        df_in = pd.DataFrame({"ten_dv": ["PGD A"]})
        df_out = loc_df(df_in, "pgd", "PGD X")
        assert df_out.empty

    def test_mode_pgd_loc_theo_ma_khi_khong_co_ten(self):
        df_in = pd.DataFrame({"ma_dv": ["pgd_a", "pgd_b"]})
        df_out = loc_df(df_in, "pgd", "pgd_a")
        assert len(df_out) == 1

    def test_mode_pgd_khong_pgd_user(self):
        df_in = pd.DataFrame({"ten_dv": ["PGD A"]})
        df_out = loc_df(df_in, "pgd", "")
        assert len(df_out) == 1


class TestCdtotkvvTenSheetExcel:
    def test_ten_don_gian(self):
        da_dung = set()
        ten = cdtotkvv_ten_sheet_excel("PGD A - Tổ 1", da_dung)
        assert len(ten) <= 31
        assert ten in da_dung

    def test_ky_tu_cam(self):
        da_dung = set()
        ten = cdtotkvv_ten_sheet_excel("Tổ [1]: A/B\\C", da_dung)
        assert "[" not in ten
        assert ":" not in ten
        assert "/" not in ten
        assert "\\" not in ten

    def test_trung_thi_danh_so(self):
        da_dung = set()
        t1 = cdtotkvv_ten_sheet_excel("PGD A", da_dung)
        t2 = cdtotkvv_ten_sheet_excel("PGD A", da_dung)
        assert t1 != t2
        assert "_1" in t2 or t1.endswith("_")

    def test_ten_dai_cat_bot(self):
        da_dung = set()
        ten = cdtotkvv_ten_sheet_excel("A" * 40, da_dung)
        assert len(ten) <= 31

    def test_ten_toan_ky_tu_cam(self):
        da_dung = set()
        ten = cdtotkvv_ten_sheet_excel("[/:*]", da_dung)
        assert len(ten) > 0
        assert ten not in {"", "/", "\\"}


class TestFmtXuatToKhongDatVn:
    def test_co_du_no_va_diem(self):
        df_in = pd.DataFrame({
            "Dư nợ": [1234567, 0],
            "Số dư TK": [500000, 100000],
            "Điểm đạt được": [85, 90],
            "Điểm tối đa": [100, 100],
            "Tổng điểm": [85, 90],
        })
        df_out = fmt_xuat_to_khong_dat_vn(df_in)
        assert df_out["Dư nợ"].iloc[0] == "1.234.567"
        assert df_out["Điểm đạt được"].iloc[0] == "85"

    def test_cot_ko_co_van_chay(self):
        df_in = pd.DataFrame({"A": [1]})
        df_out = fmt_xuat_to_khong_dat_vn(df_in)
        assert len(df_out) == 1


class TestEnrichTuoiToTruongFallback:
    def test_method_c_fill_dung_mask_khi_df_giu_index_cu(self, tmp_path, monkeypatch):
        hstd_path = tmp_path / "hstd.parquet"
        pd.DataFrame({
            COT_TEN_PGD: ["PGD A", "PGD A", "PGD A"],
            COT_TEN_XA: ["Xã A", "Xã A", "Xã A"],
            COT_TEN_KH: ["Nguyễn Văn A", "Trần Thị B", "Người Khác"],
            COT_NGAY_SINH: ["01/01/1950", "01/01/1980", "01/01/1990"],
        }).to_parquet(hstd_path)
        monkeypatch.setattr(cdtotkvv_service, "CACHE_HSTD", str(hstd_path))

        df_cdto = pd.DataFrame({
            "ten_dv": ["PGD A", "PGD A", "PGD A"],
            "ten_xa": ["Xã A", "Xã A", "Xã A"],
            "ten_to_truong": ["Nguyễn Văn A", "Trần Thị B", "Không Có"],
        }, index=[10, 20, 30])

        enriched, source_msg, so_fill = cdtotkvv_service.enrich_tuoi_to_truong_fallback_tu_hstd(df_cdto)

        assert so_fill == 3
        assert "Tên tổ trưởng trùng KH HSTD" in source_msg
        assert "2 tổ" in source_msg

        by_name = enriched.set_index("ten_to_truong")
        assert int(by_name.loc["Nguyễn Văn A", "tuoi_to_truong"]) >= 70
        assert 18 <= int(by_name.loc["Không Có", "tuoi_to_truong"]) <= 100
        assert "Tên tổ trùng KH" in by_name.loc["Nguyễn Văn A", "_nguon_tuoi_est"]
        assert "TB tuổi KH xã" in by_name.loc["Không Có", "_nguon_tuoi_est"]
        assert int(by_name.loc["Nguyễn Văn A", "_co_vay_von"]) == 1
        assert int(by_name.loc["Không Có", "_co_vay_von"]) == 0

    def test_upload_duoi_30_phan_tram_giu_na_va_nguon_chi_tiet(self, tmp_path, monkeypatch):
        hstd_path = tmp_path / "hstd.parquet"
        pd.DataFrame({
            COT_TEN_PGD: ["PGD A", "PGD A"],
            COT_TEN_XA: ["Xã A", "Xã A"],
            COT_TEN_KH: ["Trần Thị B", "Người Khác"],
            COT_NGAY_SINH: ["01/01/1980", "01/01/1990"],
        }).to_parquet(hstd_path)
        monkeypatch.setattr(cdtotkvv_service, "CACHE_HSTD", str(hstd_path))

        df_cdto = pd.DataFrame({
            "ten_dv": ["PGD A", "PGD A", "PGD A", "PGD A"],
            "ten_xa": ["Xã A", "Xã A", "Xã A", "Xã A"],
            "ten_to_truong": ["Nguyễn Văn A", "Trần Thị B", "Không Có 1", "Không Có 2"],
            "tuoi_to_truong": [66, pd.NA, pd.NA, pd.NA],
        })

        enriched, source_msg, so_fill = cdtotkvv_service.enrich_tuoi_to_truong_fallback_tu_hstd(df_cdto)

        assert so_fill == 3
        assert "CDTOTKVV upload cho phần còn lại" in source_msg
        by_name = enriched.set_index("ten_to_truong")
        assert int(by_name.loc["Nguyễn Văn A", "tuoi_to_truong"]) == 66
        assert pd.isna(by_name.loc["Nguyễn Văn A", "_co_vay_von"])
        assert "Upload thật" in by_name.loc["Nguyễn Văn A", "_nguon_chi_tiet"]
        assert int(by_name.loc["Trần Thị B", "_co_vay_von"]) == 1
        assert int(by_name.loc["Không Có 1", "_co_vay_von"]) == 0


def _build_cdto_toan_cn_bytes(leading_blank: bool) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["BÁO CÁO CDTOTKVV"])
    ws.append(["Kỳ chấm điểm tháng 05/2026"])
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])
    header = [
        "STT",
        "Mã PGD",
        "Tên PGD",
        "Mã xã",
        "Tên xã",
        "Mã tổ",
        "Tên tổ trưởng",
        "Loại tổ",
        "ĐVUT",
        "Dư nợ",
        "Tham gia GDX",
        "TL thu nợ gốc",
        "TL thu lãi",
        "TG Tổ TKVV",
        "TL nợ quá hạn",
        "Tổng điểm",
        "Xếp loại",
        "NGAYBC",
    ]
    row = [
        1,
        "004601",
        DON_VI_CHI_NHANH,
        "460001",
        "Xã A",
        "T01",
        "Nguyễn Văn A",
        "Tổ tốt",
        "Hội Phụ nữ",
        100_000_000,
        10,
        10,
        10,
        10,
        10,
        95,
        "Tốt",
        "31/05/2026",
    ]
    if leading_blank:
        header = [None] + header
        row = [None] + row
    ws.append(header)
    ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestCdtotkvvToanCnParser:
    def test_tach_file_nhan_dien_duoc_khi_khong_co_cot_trong_dau(self):
        file_bytes = _build_cdto_toan_cn_bytes(leading_blank=False)

        pgd_map = tach_file_cdto_toan_cn(file_bytes)

        assert DON_VI_CHI_NHANH in pgd_map
        assert isinstance(pgd_map[DON_VI_CHI_NHANH], bytes)

    def test_doc_thang_tu_file_toan_cn_khi_lech_cot(self):
        file_bytes = _build_cdto_toan_cn_bytes(leading_blank=False)

        thang = doc_thang_tu_cdto_toan_cn(file_bytes)

        assert thang == "05/2026"

    def test_tach_file_khong_bat_nham_cot_ma_don_vi_khac(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["BÁO CÁO CDTOTKVV"])
        ws.append(["Kỳ chấm điểm tháng 05/2026"])
        for _ in range(5):
            ws.append([])
        ws.append([
            "STT",
            "Mã đơn vị",
            "Tên đơn vị",
            "Mã PGD",
            "Tên PGD",
            "Mã xã",
            "Tên xã",
            "Mã tổ",
            "Tên tổ trưởng",
            "Loại tổ",
            "ĐVUT",
            "Dư nợ",
            "Tổng điểm",
            "Xếp loại",
            "NGAYBC",
        ])
        ws.append([1, "999999", "ĐVUT A", "004601", DON_VI_CHI_NHANH, "001", "Xã A", "T01", "A", "Tổ tốt", "Hội PN", 1, 90, "Tốt", "31/05/2026"])
        ws.append([2, "999999", "ĐVUT A", "004602", "PGD Long Thành", "002", "Xã B", "T02", "B", "Tổ tốt", "Hội ND", 1, 91, "Tốt", "31/05/2026"])

        buf = BytesIO()
        wb.save(buf)

        pgd_map = tach_file_cdto_toan_cn(buf.getvalue())

        assert DON_VI_CHI_NHANH in pgd_map
        assert "PGD Long Thành" in pgd_map
        assert len(pgd_map) == 2

    def test_tach_file_van_tach_don_vi_theo_ten_khi_ma_pgd_trong(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["BÁO CÁO CDTOTKVV"])
        ws.append(["Kỳ chấm điểm tháng 05/2026"])
        for _ in range(5):
            ws.append([])
        ws.append([
            "STT",
            "Mã PGD",
            "Tên PGD",
            "Mã xã",
            "Tên xã",
            "Mã tổ",
            "Tên tổ trưởng",
            "Loại tổ",
            "ĐVUT",
            "Dư nợ",
            "Tổng điểm",
            "Xếp loại",
            "NGAYBC",
        ])
        ws.append([1, None, DON_VI_CHI_NHANH, "001", "Xã A", "T01", "A", "Tổ tốt", "Hội PN", 1, 90, "Tốt", "31/05/2026"])
        ws.append([2, None, "PGD Long Thành", "002", "Xã B", "T02", "B", "Tổ tốt", "Hội ND", 1, 91, "Tốt", "31/05/2026"])

        buf = BytesIO()
        wb.save(buf)

        pgd_map = tach_file_cdto_toan_cn(buf.getvalue())

        assert DON_VI_CHI_NHANH in pgd_map
        assert "PGD Long Thành" in pgd_map
        assert len(pgd_map) == 2

    def test_tach_file_ke_thua_don_vi_theo_block_khi_ma_va_ten_trong(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["BÁO CÁO CDTOTKVV"])
        ws.append(["Kỳ chấm điểm tháng 05/2026"])
        for _ in range(5):
            ws.append([])
        ws.append([
            "STT",
            "Mã PGD",
            "Tên PGD",
            "Mã xã",
            "Tên xã",
            "Mã tổ",
            "Tên tổ trưởng",
            "Loại tổ",
            "ĐVUT",
            "Dư nợ",
            "Tổng điểm",
            "Xếp loại",
            "NGAYBC",
        ])
        ws.append([1, None, DON_VI_CHI_NHANH, "001", "Xã A", "T01", "A", "Tổ tốt", "Hội PN", 1, 90, "Tốt", "31/05/2026"])
        ws.append([2, None, None, "001", "Xã A", "T02", "B", "Tổ tốt", "Hội PN", 1, 91, "Tốt", "31/05/2026"])
        ws.append([3, None, "PGD Long Thành", "002", "Xã B", "T03", "C", "Tổ tốt", "Hội ND", 1, 92, "Tốt", "31/05/2026"])
        ws.append([4, None, None, "002", "Xã B", "T04", "D", "Tổ tốt", "Hội ND", 1, 93, "Tốt", "31/05/2026"])

        buf = BytesIO()
        wb.save(buf)

        pgd_map = tach_file_cdto_toan_cn(buf.getvalue())
        hoi_so_path = tmp_path / "hoi_so.xlsx"
        long_thanh_path = tmp_path / "long_thanh.xlsx"
        hoi_so_path.write_bytes(pgd_map[DON_VI_CHI_NHANH])
        long_thanh_path.write_bytes(pgd_map["PGD Long Thành"])

        df_hoi_so = doc_cdtotkvv_path(str(hoi_so_path), 1)
        df_long_thanh = doc_cdtotkvv_path(str(long_thanh_path), 1)

        assert df_hoi_so is not None
        assert df_long_thanh is not None
        assert len(df_hoi_so) == 2
        assert len(df_long_thanh) == 2
        assert set(df_hoi_so["ten_dv"]) == {DON_VI_CHI_NHANH}
        assert set(df_long_thanh["ma_dv"]) == {"004602"}

    def test_tach_file_van_map_dung_khi_header_khong_co_ma_pgd(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["BÁO CÁO CDTOTKVV"])
        ws.append(["Kỳ chấm điểm tháng 05/2026"])
        for _ in range(5):
            ws.append([])
        ws.append([
            "STT",
            "Tên đơn vị",
            "Mã xã",
            "Tên xã",
            "Mã tổ",
            "Tên tổ trưởng",
            "Loại tổ",
            "ĐVUT",
            "Dư nợ",
            "Tổng điểm",
            "Xếp loại",
            "NGAYBC",
        ])
        ws.append([1, "Hội sở CN Đồng Nai", "001", "Xã A", "T01", "A", "Tổ tốt", "Hội PN", 1, 90, "Tốt", "31/05/2026"])
        ws.append([2, "Long Thành", "002", "Xã B", "T02", "B", "Tổ tốt", "Hội ND", 1, 91, "Tốt", "31/05/2026"])

        buf = BytesIO()
        wb.save(buf)

        pgd_map = tach_file_cdto_toan_cn(buf.getvalue())
        hoi_so_path = tmp_path / "hoi_so_no_ma.xlsx"
        long_thanh_path = tmp_path / "long_thanh_no_ma.xlsx"
        hoi_so_path.write_bytes(pgd_map[DON_VI_CHI_NHANH])
        long_thanh_path.write_bytes(pgd_map["PGD Long Thành"])

        df_hoi_so = doc_cdtotkvv_path(str(hoi_so_path), 1)
        df_long_thanh = doc_cdtotkvv_path(str(long_thanh_path), 1)

        assert set(pgd_map) == {DON_VI_CHI_NHANH, "PGD Long Thành"}
        assert df_hoi_so is not None
        assert df_long_thanh is not None
        assert df_hoi_so.iloc[0]["ma_dv"] == "004601"
        assert df_hoi_so.iloc[0]["ma_xa"] == "000001"
        assert df_long_thanh.iloc[0]["ma_dv"] == "004602"
        assert df_long_thanh.iloc[0]["ten_dv"] == "PGD Long Thành"
        assert df_long_thanh.iloc[0]["ma_xa"] == "000002"
        assert df_long_thanh.iloc[0]["ma_to"] == "T02"

    def test_tach_file_ghi_ma_dv_theo_cot_duoc_chon_tot_nhat(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["BÁO CÁO CDTOTKVV"])
        ws.append(["Kỳ chấm điểm tháng 05/2026"])
        for _ in range(5):
            ws.append([])
        ws.append([
            "STT",
            "Mã PGD",
            "Tên đơn vị",
            "Mã phòng giao dịch",
            "Tên PGD",
            "Mã xã",
            "Tên xã",
            "Mã tổ",
            "Tên tổ trưởng",
            "Loại tổ",
            "ĐVUT",
            "Dư nợ",
            "Tổng điểm",
            "Xếp loại",
            "NGAYBC",
        ])
        ws.append([
            1, "004601", "Cột nhiễu", "004601", DON_VI_CHI_NHANH,
            "460001", "Xã A", "T01", "A", "Tổ tốt", "Hội PN", 1,
            90, "Tốt", "31/05/2026",
        ])
        ws.append([
            2, "004601", "Cột nhiễu", "004602", "PGD Long Thành",
            "460002", "Xã B", "T02", "B", "Tổ tốt", "Hội ND", 1,
            91, "Tốt", "31/05/2026",
        ])

        buf = BytesIO()
        wb.save(buf)
        pgd_map = tach_file_cdto_toan_cn(buf.getvalue())
        path = tmp_path / "long_thanh.xlsx"
        path.write_bytes(pgd_map["PGD Long Thành"])

        df = doc_cdtotkvv_path(str(path), 1)

        assert df is not None
        assert df.iloc[0]["ma_dv"] == "004602"

    def test_tach_file_doc_lai_dung_cot_cho_ca_layout_cu_va_moi(self, tmp_path):
        for leading_blank in (False, True):
            file_bytes = _build_cdto_toan_cn_bytes(leading_blank=leading_blank)
            pgd_map = tach_file_cdto_toan_cn(file_bytes)
            path = tmp_path / f"cdtotkvv_{leading_blank}.xlsx"
            path.write_bytes(pgd_map[DON_VI_CHI_NHANH])

            df = doc_cdtotkvv_path(str(path), 1)

            assert df is not None
            assert len(df) == 1
            row = df.iloc[0]
            assert row["ma_dv"] == "004601"
            assert row["ten_dv"] == DON_VI_CHI_NHANH
            assert row["ma_xa"] == "460001"
            assert row["ten_xa"] == "Xã A"
            assert row["ma_to"] == "T01"
            assert row["ten_to_truong"] == "Nguyễn Văn A"
            assert row["dvut"] == "Hội Phụ nữ"
            assert row["loai_to"] == "Tổ tốt"
            assert row["du_no"] == 100_000_000
            assert row["tong_diem"] == 95
            assert row["xep_loai"] == "Tốt"
