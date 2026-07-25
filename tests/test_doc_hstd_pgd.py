"""tests/test_doc_hstd_pgd.py — Test doc_hstd_pgd() và luu_file_he_thong()."""
from __future__ import annotations

import os
import tempfile
import time
from pathlib import Path
from unittest.mock import patch, MagicMock

import pandas as pd
import pytest
from openpyxl import Workbook

from data.pgd import doc_hstd_pgd, duong_dan_hstd_hien_hanh, pgd_slug


# ── Helper ────────────────────────────────────────────────────────────────────

def _tao_hstd_excel(path, ten_pgd="PGD Test"):
    """Tạo file HSTD giả với sheet BCQUERY."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BCQUERY"
    # Header at row 5 (pandas header=4 → 0-indexed row 4 = Excel row 5)
    headers = ["BoQua", "Số khế ước", "Mã KH", "Tên PGD", "Tên xã",
               "Dư nợ trong hạn", "Dư nợ quá hạn", "Tổng dư nợ", "Nguồn vốn"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
    # Data rows
    ws.cell(row=6, column=1, value="x")
    ws.cell(row=6, column=2, value="KU001")
    ws.cell(row=6, column=3, value="KH001")
    ws.cell(row=6, column=4, value=ten_pgd)
    ws.cell(row=6, column=5, value="Xã Test")
    ws.cell(row=6, column=6, value=1000)
    ws.cell(row=6, column=7, value=0)
    ws.cell(row=6, column=8, value=1000)
    ws.cell(row=6, column=9, value=1)
    wb.save(str(path))


# ══════════════════════════════════════════════════════════════════════════════
# TestDocHstdPgd
# ══════════════════════════════════════════════════════════════════════════════

class TestDocHstdPgd:

    def test_file_khong_ton_tai_tra_ve_none(self):
        """Không có file HSTD nào → trả về None."""
        with tempfile.TemporaryDirectory() as tmp:
            with patch("data.pgd.PGD_DATA_DIR", tmp):
                result = doc_hstd_pgd("PGD Khong Ton Tai", file_mtime=0.0)
        assert result is None

    def test_chon_file_moi_hon(self):
        """Khi cả hstd_latest.xlsx và hstd_khnv.xlsx tồn tại → chọn file mới hơn."""
        with tempfile.TemporaryDirectory() as tmp:
            slug = pgd_slug("PGD Long Thanh")
            pgd_dir = Path(tmp) / slug
            pgd_dir.mkdir(parents=True)

            # Tạo file hstd_latest.xlsx (cũ hơn) với dữ liệu "CU"
            path_latest = pgd_dir / "hstd_latest.xlsx"
            _tao_hstd_excel(path_latest, ten_pgd="PGD CU")

            # Đợi 1 chút để mtime khác nhau
            time.sleep(0.05)

            # Tạo file hstd_khnv.xlsx (mới hơn) với dữ liệu "MOI"
            path_khnv = pgd_dir / "hstd_khnv.xlsx"
            _tao_hstd_excel(path_khnv, ten_pgd="PGD MOI")

            with patch("data.pgd.PGD_DATA_DIR", tmp):
                # duong_dan_hstd_hien_hanh phải chọn file mới hơn (hstd_khnv.xlsx)
                path_chon = duong_dan_hstd_hien_hanh("PGD Long Thanh")
                assert path_chon.endswith("hstd_khnv.xlsx")

                # doc_hstd_pgd phải đọc dữ liệu từ file mới hơn
                mtime = os.path.getmtime(path_chon)
                df = doc_hstd_pgd("PGD Long Thanh", file_mtime=mtime)

            assert df is not None
            assert isinstance(df, pd.DataFrame)
            # Cột "Tên PGD" phải chứa giá trị từ file mới
            assert "PGD MOI" in df["Tên PGD"].values

    def test_doc_duoc_dataframe(self):
        """Tạo file HSTD hợp lệ → trả về DataFrame với cột mong đợi."""
        with tempfile.TemporaryDirectory() as tmp:
            slug = pgd_slug("PGD Test Doc")
            pgd_dir = Path(tmp) / slug
            pgd_dir.mkdir(parents=True)

            path = pgd_dir / "hstd_latest.xlsx"
            _tao_hstd_excel(path, ten_pgd="PGD Test Doc")

            with patch("data.pgd.PGD_DATA_DIR", tmp):
                mtime = os.path.getmtime(path)
                df = doc_hstd_pgd("PGD Test Doc", file_mtime=mtime)

            assert df is not None
            assert isinstance(df, pd.DataFrame)
            assert len(df) == 1
            # doc_file drops first column (BoQua) via iloc[:, 1:]
            assert "Số khế ước" in df.columns
            assert "Mã KH" in df.columns
            assert "Tên PGD" in df.columns
            assert "Tổng dư nợ" in df.columns
            assert df.iloc[0]["Mã KH"] == "KH001"
            assert df.iloc[0]["Tổng dư nợ"] == 1000


# ══════════════════════════════════════════════════════════════════════════════
# TestLuuFileHeThong
# ══════════════════════════════════════════════════════════════════════════════

class TestLuuFileHeThong:

    def _file_bytes_hop_le(self):
        """Tạo bytes file xlsx hợp lệ (> 1KB)."""
        wb = Workbook()
        ws = wb.active
        ws.title = "BCQUERY"
        ws.cell(row=1, column=1, value="test")
        # Thêm dữ liệu để file > 1KB
        for i in range(2, 100):
            ws.cell(row=i, column=1, value=f"data_{i}")
            ws.cell(row=i, column=2, value=i)
        import io
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_luu_thanh_cong(self):
        """File hợp lệ → lưu thành công, trả về KetQuaUpload.thanh_cong=True."""
        from services.upload_service import luu_file_he_thong

        file_bytes = self._file_bytes_hop_le()

        with tempfile.TemporaryDirectory() as tmp:
            fake_path = os.path.join(tmp, "test_file.XLSX")
            fake_files_he_thong = {
                "test_file.XLSX": {
                    "mo_ta": "Test file",
                    "path": fake_path,
                }
            }
            with patch("services.upload_service.FILES_HE_THONG", fake_files_he_thong), \
                 patch("services.upload_service.st.session_state", {"username": "test_user"}), \
                 patch("services.upload_service.db.ghi_audit"):
                kq = luu_file_he_thong("test_file.XLSX", file_bytes)

            assert kq.thanh_cong is True
            assert os.path.exists(fake_path)
            assert "test_file.XLSX" in kq.thong_bao

    def test_ten_file_khong_hop_le(self):
        """Tên file không nằm trong FILES_HE_THONG → thất bại."""
        from services.upload_service import luu_file_he_thong

        file_bytes = self._file_bytes_hop_le()

        with patch("services.upload_service.FILES_HE_THONG", {"valid.XLSX": {"mo_ta": "x", "path": "/tmp/x"}}):
            kq = luu_file_he_thong("invalid_file.XLSX", file_bytes)

        assert kq.thanh_cong is False
        assert "không hợp lệ" in kq.thong_bao

    def test_ghi_audit(self):
        """Sau khi lưu thành công → db.ghi_audit được gọi."""
        from services.upload_service import luu_file_he_thong

        file_bytes = self._file_bytes_hop_le()

        with tempfile.TemporaryDirectory() as tmp:
            fake_path = os.path.join(tmp, "audit_test.XLSX")
            fake_files_he_thong = {
                "audit_test.XLSX": {
                    "mo_ta": "Audit test",
                    "path": fake_path,
                }
            }
            mock_audit = MagicMock()
            with patch("services.upload_service.FILES_HE_THONG", fake_files_he_thong), \
                 patch("services.upload_service.st.session_state", {"username": "audit_user"}), \
                 patch("services.upload_service.db.ghi_audit", mock_audit):
                kq = luu_file_he_thong("audit_test.XLSX", file_bytes)

            assert kq.thanh_cong is True
            mock_audit.assert_called_once()
            args = mock_audit.call_args[0]
            assert args[0] == "audit_user"
            assert args[1] == "upload_he_thong"
