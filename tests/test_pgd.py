"""tests/test_pgd.py — data/pgd.py: pgd_slug(), duong_dan_pgd()"""
from __future__ import annotations
import tempfile
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from data.pgd import pgd_slug, duong_dan_pgd, thu_muc_pgd, _doc_ngay_so_lieu


class TestPgdSlug:

    def test_ten_tieng_viet_co_ban(self):
        assert pgd_slug("PGD Biên Hòa 1") == "pgd_bien_hoa_1"

    def test_chu_d_viet_duoc_xu_ly(self):
        result = pgd_slug("PGD Đồng Nai")
        assert "d" in result and "ong" in result
        assert "đ" not in result

    def test_chu_d_hoa_duoc_xu_ly(self):
        result = pgd_slug("PGD Điện Biên")
        assert "d" in result
        assert "Đ" not in result

    def test_khoang_trang_dau_cuoi_bi_strip(self):
        assert pgd_slug("  PGD ABC  ") == "pgd_abc"

    def test_so_trong_ten_duoc_giu(self):
        result = pgd_slug("PGD Biên Hòa 1")
        assert "1" in result

    def test_dau_gach_ngang_thanh_underscore(self):
        result = pgd_slug("PGD A-B")
        assert "-" not in result
        assert "_" in result

    def test_ket_qua_luon_lowercase(self):
        result = pgd_slug("PGD LONG THÀNH")
        assert result == result.lower()

    def test_hai_ten_pgd_khac_nhau_cho_slug_khac(self):
        assert pgd_slug("PGD Long Thành") != pgd_slug("PGD Nhơn Trạch")

    def test_hoi_so_khong_crash(self):
        result = pgd_slug("Hội sở Chi nhánh tỉnh")
        assert isinstance(result, str)
        assert len(result) > 0

    def test_khong_co_ky_tu_dac_biet(self):
        result = pgd_slug("PGD Xuân Lộc")
        assert all(c.isalnum() or c == "_" for c in result)


class TestDuongDanPgd:

    def test_hstd_ket_thuc_dung_ten_file(self):
        assert duong_dan_pgd("PGD Long Thành", "hstd").endswith("hstd_latest.xlsx")

    def test_nq11(self):
        assert duong_dan_pgd("PGD Long Thành", "nq11").endswith("nq11_latest.xlsx")

    def test_gqvl(self):
        assert duong_dan_pgd("PGD Long Thành", "gqvl").endswith("gqvl_latest.xlsx")

    def test_cdtotkvv(self):
        assert duong_dan_pgd("PGD Long Thành", "cdtotkvv").endswith("cdtotkvv_latest.xlsx")

    def test_dienbao_ht(self):
        assert duong_dan_pgd("PGD Long Thành", "dienbao_ht").endswith("dienbao_ht.xlsx")

    def test_dienbao_prev(self):
        assert duong_dan_pgd("PGD Long Thành", "dienbao_prev").endswith("dienbao_prev.xlsx")

    def test_slug_co_trong_path(self):
        path = duong_dan_pgd("PGD Long Thành", "hstd")
        assert "pgd_long_thanh" in path

    def test_hai_pgd_co_path_khac_nhau(self):
        p1 = duong_dan_pgd("PGD Long Thành", "hstd")
        p2 = duong_dan_pgd("PGD Nhơn Trạch", "hstd")
        assert p1 != p2

    def test_path_la_chuoi(self):
        result = duong_dan_pgd("PGD Long Thành", "hstd")
        assert isinstance(result, str)

    def test_hstd_khnv_ket_thuc_dung_ten_file(self):
        assert duong_dan_pgd("PGD Long Thành", "hstd_khnv").endswith("hstd_khnv.xlsx")

    def test_hstd_khnv_slug_co_trong_path(self):
        path = duong_dan_pgd("PGD Long Thành", "hstd_khnv")
        assert "pgd_long_thanh" in path


class TestThuMucPgdTaoDirectory:

    def test_hstd_tao_thu_muc(self):
        """duong_dan_pgd('hstd') tạo thư mục PGD qua thu_muc_pgd."""
        with tempfile.TemporaryDirectory() as tmp:
            with patch("data.pgd.PGD_DATA_DIR", tmp):
                path = duong_dan_pgd("PGD Long Thành", "hstd")
                assert Path(path).parent.exists()
                assert Path(path).parent.is_dir()

    def test_nq11_tao_thu_muc(self):
        """duong_dan_pgd('nq11') tạo thư mục PGD qua thu_muc_pgd."""
        with tempfile.TemporaryDirectory() as tmp:
            with patch("data.pgd.PGD_DATA_DIR", tmp):
                path = duong_dan_pgd("PGD Nhơn Trạch", "nq11")
                assert Path(path).parent.exists()
                assert Path(path).parent.is_dir()

    def test_gqvl_tao_thu_muc(self):
        """duong_dan_pgd('gqvl') tạo thư mục PGD qua thu_muc_pgd."""
        with tempfile.TemporaryDirectory() as tmp:
            with patch("data.pgd.PGD_DATA_DIR", tmp):
                path = duong_dan_pgd("PGD Xuân Lộc", "gqvl")
                assert Path(path).parent.exists()
                assert Path(path).parent.is_dir()


class TestDocNgaySoLieuHstd:

    def _make_hstd_file(self, tmp_path, ngay_col: int):
        fp = tmp_path / f"hstd_col_{ngay_col}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "BCQUERY"
        ws.cell(row=5, column=4, value="Tên PGD")
        ws.cell(row=6, column=4, value="PGD Test")
        ws.cell(row=5, column=ngay_col, value="Ngày số liệu")
        ws.cell(row=6, column=ngay_col, value="30/06/2026")
        wb.save(fp)
        return fp

    def test_doc_ngay_so_lieu_hstd_layout_cu_fs(self, tmp_path):
        fp = self._make_hstd_file(tmp_path, 175)

        assert _doc_ngay_so_lieu(fp, "hstd", fp.stat().st_mtime) == datetime(2026, 6, 30)

    def test_doc_ngay_so_lieu_hstd_layout_moi_ft(self, tmp_path):
        fp = self._make_hstd_file(tmp_path, 176)

        assert _doc_ngay_so_lieu(fp, "hstd", fp.stat().st_mtime) == datetime(2026, 6, 30)
