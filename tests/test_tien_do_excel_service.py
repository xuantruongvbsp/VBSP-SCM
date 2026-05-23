"""
Tests for services/tien_do_excel_service.py
Pure function: xuat_excel_tien_do → bytes with 3 styled sheets.
"""
from __future__ import annotations

from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from services.tien_do_excel_service import xuat_excel_tien_do


# ── fixtures ─────────────────────────────────────────────────────────────────

def _df_tonghop():
    return pd.DataFrame({
        "STT":                [1, 2],
        "Đầu việc":           ["Việc A", "Việc B"],
        "Loại":               ["chung", "cbtd"],
        "Người phụ trách":    ["NV 1", "NV 2"],
        "CB KH-NV phụ trách": ["CB 1", "CB 2"],
        "Ưu tiên":            ["cao", "binh_thuong"],
        "Thời hạn":           ["2026-05-31", "2026-06-30"],
        "Số PGD":             [21, 21],
        "Tổng xã":            [95, 95],
        "Đã hoàn thành":      [10, 5],
        "Chưa thực hiện":     [85, 90],
        "Trễ hạn":            [0, 0],
        "N/A":                [0, 0],
        "Tỷ lệ HT%":          [10.5, 5.3],
    })


def _df_matran():
    return pd.DataFrame({
        "Đơn vị":        ["PGD Long Thành", "PGD Trảng Bom"],
        "Việc A":        ["✅", "⏳"],
        "Việc B":        ["⏳", "✅"],
    })


def _df_ct():
    return pd.DataFrame({
        "Task ID":          [1, 1],
        "Đầu việc":         ["Việc A", "Việc A"],
        "PGD":              ["PGD Long Thành", "PGD Long Thành"],
        "Xã / Phường":      ["Xã A", "Xã B"],
        "Trạng thái":       ["da_hoan_thanh", "chua_thuc_hien"],
        "Thời hạn":         ["2026-05-31", "2026-05-31"],
        "Ngày hoàn thành":  ["2026-05-20", None],
        "Ghi chú":          ["OK", ""],
    })


# ── basic output ──────────────────────────────────────────────────────────────

def test_xuat_returns_bytes():
    result = xuat_excel_tien_do(_df_tonghop(), _df_matran(), _df_ct())
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_xuat_produces_valid_xlsx():
    result = xuat_excel_tien_do(_df_tonghop(), _df_matran(), _df_ct())
    wb = load_workbook(BytesIO(result))
    assert wb is not None


def test_xuat_three_sheets():
    result = xuat_excel_tien_do(_df_tonghop(), _df_matran(), _df_ct())
    wb = load_workbook(BytesIO(result))
    assert len(wb.sheetnames) == 3


def test_xuat_correct_sheet_names():
    result = xuat_excel_tien_do(_df_tonghop(), _df_matran(), _df_ct())
    wb = load_workbook(BytesIO(result))
    assert "Tổng hợp" in wb.sheetnames
    assert "Ma trận PGD" in wb.sheetnames
    assert "Chi tiết xã" in wb.sheetnames


def test_xuat_tonghop_has_data():
    result = xuat_excel_tien_do(_df_tonghop(), _df_matran(), _df_ct())
    wb = load_workbook(BytesIO(result))
    ws = wb["Tổng hợp"]
    # Row 1 = header, rows 2+ = data
    assert ws.max_row >= 3


def test_xuat_matran_has_correct_data():
    result = xuat_excel_tien_do(_df_tonghop(), _df_matran(), _df_ct())
    wb = load_workbook(BytesIO(result))
    ws = wb["Ma trận PGD"]
    header = [cell.value for cell in ws[1]]
    assert "Đơn vị" in header
    # Data rows should have PGD names
    col_dvien_idx = header.index("Đơn vị") + 1
    dvien_values = [ws.cell(row=r, column=col_dvien_idx).value for r in range(2, ws.max_row + 1)]
    assert "PGD Long Thành" in dvien_values


def test_xuat_ct_has_correct_columns():
    result = xuat_excel_tien_do(_df_tonghop(), _df_matran(), _df_ct())
    wb = load_workbook(BytesIO(result))
    ws = wb["Chi tiết xã"]
    header = [cell.value for cell in ws[1]]
    assert "Đầu việc" in header
    assert "PGD" in header
    assert "Xã / Phường" in header


def test_xuat_empty_dataframes_no_crash():
    empty = pd.DataFrame()
    result = xuat_excel_tien_do(empty, empty, empty)
    assert isinstance(result, bytes)
    assert len(result) > 0
