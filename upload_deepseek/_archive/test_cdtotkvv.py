"""Unit test cho data/cdtotkvv.py — Chấm điểm Tổ TK&VV."""
import os
import unittest
from datetime import datetime
from io import BytesIO

import pandas as pd


class TestCdtotkvvDocThangNam(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        from data import cdtotkvv
        cls.cdtotkvv = cdtotkvv

    def _tao_file_voi_noi_dung(self, text: str) -> bytes:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=text)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_doc_thang_nam_tu_file_ngay_viet(self) -> None:
        content = "ngày 08 tháng 05 năm 2026"
        result = self.cdtotkvv.doc_thang_nam_tu_file(self._tao_file_voi_noi_dung(content))
        self.assertEqual(result, "05/2026")

    def test_doc_thang_nam_tu_file_thang_slash_nam(self) -> None:
        content = "BÁO CÁO tháng 06/2026"
        result = self.cdtotkvv.doc_thang_nam_tu_file(self._tao_file_voi_noi_dung(content))
        self.assertEqual(result, "06/2026")

    def test_doc_thang_nam_tu_file_ddmmyyyy(self) -> None:
        content = "08/05/2026"
        result = self.cdtotkvv.doc_thang_nam_tu_file(self._tao_file_voi_noi_dung(content))
        self.assertEqual(result, "05/2026")

    def test_doc_thang_nam_tu_file_dd_mm_yyyy(self) -> None:
        content = "08-05-2026"
        result = self.cdtotkvv.doc_thang_nam_tu_file(self._tao_file_voi_noi_dung(content))
        self.assertEqual(result, "05/2026")

    def test_doc_thang_nam_tu_file_thang_word(self) -> None:
        content = "tháng 3 / 2026"
        result = self.cdtotkvv.doc_thang_nam_tu_file(self._tao_file_voi_noi_dung(content))
        self.assertEqual(result, "03/2026")

    def test_doc_thang_nam_tu_file_khong_tim_thay(self) -> None:
        content = "Không có thông tin ngày tháng"
        result = self.cdtotkvv.doc_thang_nam_tu_file(self._tao_file_voi_noi_dung(content))
        self.assertIsNone(result)

    def test_doc_thang_nam_tu_file_file_rong(self) -> None:
        result = self.cdtotkvv.doc_thang_nam_tu_file(b"")
        self.assertIsNone(result)

    def test_doc_thang_nam_tu_file_datetime_object(self) -> None:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=datetime(2026, 5, 8))
        buf = BytesIO()
        wb.save(buf)
        result = self.cdtotkvv.doc_thang_nam_tu_file(buf.getvalue())
        self.assertEqual(result, "05/2026")


class TestCdtotkvvTongHop(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        from data import cdtotkvv
        cls.cdtotkvv = cdtotkvv

    def test_tong_hop_theo_pgd_co_du_lieu(self) -> None:
        df = pd.DataFrame({
            "stt": [1, 2, 3, 4],
            "ma_dv": ["DV001", "DV001", "DV002", "DV002"],
            "ten_dv": ["Đơn vị A", "Đơn vị A", "Đơn vị B", "Đơn vị B"],
            "ma_xa": ["X001", "X001", "X002", "X002"],
            "ma_to": ["T001", "T002", "T003", "T004"],
            "tong_diem": [85, 70, 55, 40],
            "xep_loai": ["Tốt", "Khá", "Trung bình", "Yếu"],
            "tinh_trang": ["A", "B", "B", "C"],
            "du_no": [10_000_000, 5_000_000, 3_000_000, 1_000_000],
            "so_du_tk": [1_000_000, 500_000, 200_000, 0],
        })
        result = self.cdtotkvv.tong_hop_theo_pgd(df)
        self.assertIsInstance(result, pd.DataFrame)
        self.assertIn("tong_to", result.columns)
        self.assertIn("tong_diem_tb", result.columns)
        self.assertIn("to_tot", result.columns)
        self.assertIn("to_kha", result.columns)
        self.assertIn("to_tb", result.columns)
        self.assertIn("to_yeu", result.columns)
        self.assertIn("to_tinh_trang_a", result.columns)
        self.assertIn("to_tinh_trang_b", result.columns)
        self.assertIn("to_tinh_trang_c", result.columns)

    def test_tong_hop_theo_pgd_so_luong_to(self) -> None:
        df = pd.DataFrame({
            "stt": [1, 2],
            "ma_dv": ["DV001", "DV001"],
            "ten_dv": ["Đơn vị A", "Đơn vị A"],
            "ma_xa": ["X001", "X001"],
            "ma_to": ["T001", "T002"],
            "tong_diem": [80, 75],
            "xep_loai": ["Tốt", "Khá"],
            "tinh_trang": ["A", "B"],
            "du_no": [10_000_000, 5_000_000],
            "so_du_tk": [1_000_000, 500_000],
        })
        result = self.cdtotkvv.tong_hop_theo_pgd(df)
        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]["tong_to"], 2)
        self.assertEqual(result.iloc[0]["to_tot"], 1)
        self.assertEqual(result.iloc[0]["to_kha"], 1)
        self.assertEqual(result.iloc[0]["to_tinh_trang_a"], 1)
        self.assertEqual(result.iloc[0]["to_tinh_trang_b"], 1)

    def test_tong_hop_theo_pgd_df_rong(self) -> None:
        df = pd.DataFrame({"ma_dv": [], "ten_dv": [], "stt": [], "tong_diem": [], "xep_loai": [], "tinh_trang": []})
        result = self.cdtotkvv.tong_hop_theo_pgd(df)
        self.assertIsInstance(result, pd.DataFrame)
        self.assertTrue(result.empty)

    def test_tong_hop_theo_pgd_nhieu_don_vi(self) -> None:
        df = pd.DataFrame({
            "stt": [1, 2],
            "ma_dv": ["DV001", "DV002"],
            "ten_dv": ["Đơn vị A", "Đơn vị B"],
            "ma_xa": ["X001", "X002"],
            "ma_to": ["T001", "T002"],
            "tong_diem": [90, 45],
            "xep_loai": ["Tốt", "Yếu"],
            "tinh_trang": ["A", "C"],
            "du_no": [10_000_000, 2_000_000],
            "so_du_tk": [1_000_000, 0],
        })
        result = self.cdtotkvv.tong_hop_theo_pgd(df)
        self.assertEqual(len(result), 2)

    def test_tong_hop_theo_pgd_diem_trung_binh(self) -> None:
        df = pd.DataFrame({
            "stt": [1, 2],
            "ma_dv": ["DV001", "DV001"],
            "ten_dv": ["Đơn vị A", "Đơn vị A"],
            "ma_xa": ["X001", "X001"],
            "ma_to": ["T001", "T002"],
            "tong_diem": [80, 90],
            "xep_loai": ["Khá", "Tốt"],
            "tinh_trang": ["B", "A"],
            "du_no": [5_000_000, 10_000_000],
            "so_du_tk": [500_000, 1_000_000],
        })
        result = self.cdtotkvv.tong_hop_theo_pgd(df)
        self.assertEqual(result.iloc[0]["tong_diem_tb"], 85.0)


class TestCdtotkvvDocPath(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        from data import cdtotkvv
        cls.cdtotkvv = cdtotkvv

    def test_doc_cdtotkvv_path_duong_dan_khong_ton_tai(self) -> None:
        result = self.cdtotkvv.doc_cdtotkvv_path("/nonexistent/file.xlsx", 0)
        self.assertIsNone(result)

    def test_doc_cdtotkvv_thang_nam_sai(self) -> None:
        result = self.cdtotkvv.doc_cdtotkvv("invalid")
        self.assertIsNone(result)


if __name__ == "__main__":
    unittest.main()
