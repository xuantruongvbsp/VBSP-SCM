"""Dịch vụ xuất Excel chuyên nghiệp (chuẩn VSPPRO) — KPI sheet + Detail sheet + Bìa."""

from io import BytesIO
from datetime import datetime
from typing import Optional
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    raise ImportError("Cần cài: pip install openpyxl")

from config import COT_SO_KU, COT_MA_KH, COT_TEN_KH, COT_TEN_PGD, COT_TEN_CT
from config import COT_TONG_DU_NO, COT_DU_NO_TH, COT_DU_NO_QH
from config import COT_NGAY_VAY, COT_NGAY_DH, COT_THOI_HAN, COT_MUC_VAY

# ── Styles ───────────────────────────────────────────────────────────────────
FONT_TITLE   = Font(name="Times New Roman", size=18, bold=True, color="2E75B6")
FONT_HEADER  = Font(name="Times New Roman", size=14, bold=True)
FONT_NORMAL  = Font(name="Times New Roman", size=12)
FONT_SMALL   = Font(name="Times New Roman", size=10, italic=True, color="666666")
FONT_TABLE_H = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
FONT_TABLE   = Font(name="Times New Roman", size=11)

FILL_HEADER  = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
FILL_ALT     = PatternFill(start_color="F2F7F2", end_color="F2F7F2", fill_type="solid")
FILL_BIA_LOGO_AREA = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")
ALIGN_WRAP   = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

NUM_FMT_TIEN = '#,##0'
NUM_FMT_PCT  = '0.00%'
NUM_FMT_SO   = '#,##0'

# ── Danh sách cột mặc định cho sheet chi tiết ──────────────────────────────
LOAN_COLUMNS = [
    ("Tên PGD",              COT_TEN_PGD,    "text"),
    ("Số khế ước",            COT_SO_KU,      "text"),
    ("Mã KH",                 COT_MA_KH,      "text"),
    ("Tên KH",                COT_TEN_KH,     "text"),
    ("Chương trình",          COT_TEN_CT,     "text"),
    ("Ngày vay",              COT_NGAY_VAY,   "date"),
    ("Ngày ĐH",               COT_NGAY_DH,    "date"),
    ("Thời hạn (tháng)",      COT_THOI_HAN,   "so"),
    ("Mức vay",               COT_MUC_VAY,    "tien"),
    ("Dư nợ trong hạn",       COT_DU_NO_TH,   "tien"),
    ("Dư nợ quá hạn",         COT_DU_NO_QH,   "tien"),
    ("Tổng dư nợ",            COT_TONG_DU_NO, "tien"),
]


def _fmt_date(v):
    if pd.isna(v):
        return ""
    try:
        dt = pd.to_datetime(v)
        return dt.strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return str(v)


def _fmt_tien(v):
    if pd.isna(v) or v == 0:
        return 0
    try:
        return round(float(v))
    except (ValueError, TypeError):
        return v


def _safe_sheet_name(name: str) -> str:
    cleaned = name.replace("\\", " ").replace("/", " ").replace("?", " ")
    cleaned = cleaned.replace("*", " ").replace("[", " ").replace("]", " ").replace(":", " ")
    return cleaned.strip()[:31] or "Sheet"


# ═══════════════════════════════════════════════════════════════════════════════
# COVER SHEET — Bìa báo cáo
# ═══════════════════════════════════════════════════════════════════════════════
def _build_cover_sheet(wb: Workbook, title: str, subtitle: str = "",
                       nguoi_xuat: str = "") -> None:
    ws = wb.active
    ws.title = "Bìa"

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 20

    row = 3
    cell = ws.cell(row=row, column=2, value="NGÂN HÀNG CHÍNH SÁCH XÃ HỘI VIỆT NAM")
    cell.font = FONT_HEADER
    cell.alignment = ALIGN_CENTER

    row += 1
    cell = ws.cell(row=row, column=2, value="CHI NHÁNH TỈNH")
    cell.font = FONT_NORMAL
    cell.alignment = ALIGN_CENTER

    row += 3
    cell = ws.cell(row=row, column=2, value=title.upper())
    cell.font = FONT_TITLE
    cell.alignment = ALIGN_CENTER

    if subtitle:
        row += 1
        cell = ws.cell(row=row, column=2, value=subtitle)
        cell.font = Font(name="Times New Roman", size=13, italic=True, color="555555")
        cell.alignment = ALIGN_CENTER

    row += 3
    now = datetime.now()
    info_items = [
        f"Ngày xuất: {now.strftime('%d/%m/%Y')}",
        f"Giờ xuất: {now.strftime('%H:%M:%S')}",
    ]
    if nguoi_xuat:
        info_items.insert(1, f"Người xuất: {nguoi_xuat}")

    for item in info_items:
        cell = ws.cell(row=row, column=2, value=item)
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_LEFT
        row += 1

    row += 2
    cell = ws.cell(row=row, column=2,
                   value="Báo cáo được tạo tự động bởi Hệ thống VBSP-SCM")
    cell.font = FONT_SMALL
    cell.alignment = ALIGN_CENTER

    ws.sheet_properties.tabColor = "2E7D32"


# ═══════════════════════════════════════════════════════════════════════════════
# KPI SUMMARY SHEET
# ═══════════════════════════════════════════════════════════════════════════════
def _build_kpi_sheet(wb: Workbook, kpi_items: list[tuple[str, object, str]],
                     subtitle: str = "") -> None:
    ws = wb.create_sheet("Tóm tắt KPI")

    for c, w in enumerate([35, 22, 28], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    if subtitle:
        ws.cell(row=1, column=1, value=subtitle).font = FONT_NORMAL
        ws.merge_cells("A1:C1")
        start = 3
    else:
        start = 1

    headers = ["Chỉ tiêu", "Giá trị", "Ghi chú"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.font = FONT_TABLE_H
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws.freeze_panes = ws.cell(row=start + 1, column=1)

    for i, (label, value, note) in enumerate(kpi_items):
        r = start + 1 + i
        cell_a = ws.cell(row=r, column=1, value=label)
        cell_a.font = FONT_TABLE
        cell_a.alignment = ALIGN_LEFT
        cell_a.border = THIN_BORDER

        cell_b = ws.cell(row=r, column=2, value=value)
        cell_b.font = FONT_TABLE
        cell_b.alignment = ALIGN_RIGHT
        cell_b.border = THIN_BORDER

        cell_c = ws.cell(row=r, column=3, value=note)
        cell_c.font = Font(name="Times New Roman", size=10, color="666666")
        cell_c.alignment = ALIGN_LEFT
        cell_c.border = THIN_BORDER

        if i % 2 == 1:
            for cc in range(1, 4):
                ws.cell(row=r, column=cc).fill = FILL_ALT

    ws.sheet_properties.tabColor = "2E7D32"


# ═══════════════════════════════════════════════════════════════════════════════
# DETAIL SHEET — Chi tiết khế ước
# ═══════════════════════════════════════════════════════════════════════════════
def _build_detail_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame,
                        columns: Optional[list] = None) -> None:
    ws = wb.create_sheet(_safe_sheet_name(sheet_name))

    cols = columns or LOAN_COLUMNS
    headers = [c[0] for c in cols]
    col_keys = [c[1] for c in cols]
    col_types = [c[2] for c in cols]

    has_data = df is not None and not df.empty

    if has_data:
        avail = {k for k in col_keys if k in df.columns}
    else:
        avail = set()

    visible_headers = []
    visible_keys = []
    visible_types = []
    for h, k, t in zip(headers, col_keys, col_types):
        if k in avail:
            visible_headers.append(h)
            visible_keys.append(k)
            visible_types.append(t)

    if not visible_headers:
        ws.cell(row=1, column=1, value="Không có dữ liệu").font = FONT_NORMAL
        return

    # Header row
    for c, h in enumerate(visible_headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = FONT_TABLE_H
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_WRAP
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Data rows
    for r_idx in range(len(df)):
        r = r_idx + 2
        row_data = df.iloc[r_idx]
        for c_idx, (key, typ) in enumerate(zip(visible_keys, visible_types), 1):
            val = row_data.get(key) if hasattr(row_data, "get") else row_data[key]
            cell = ws.cell(row=r, column=c_idx)

            if typ == "tien":
                cell.value = _fmt_tien(val)
                cell.number_format = NUM_FMT_TIEN
                cell.alignment = ALIGN_RIGHT
            elif typ == "date":
                cell.value = _fmt_date(val)
                cell.alignment = ALIGN_CENTER
            elif typ == "so":
                try:
                    cell.value = round(float(val)) if pd.notna(val) else 0
                    cell.number_format = NUM_FMT_SO
                    cell.alignment = ALIGN_RIGHT
                except (ValueError, TypeError):
                    cell.value = str(val) if pd.notna(val) else ""
            else:
                cell.value = str(val) if pd.notna(val) else ""

            cell.font = FONT_TABLE
            cell.border = THIN_BORDER

        if r_idx % 2 == 1:
            for cc in range(1, len(visible_headers) + 1):
                ws.cell(row=r, column=cc).fill = FILL_ALT

    # Auto column width
    for c_idx, h in enumerate(visible_headers, 1):
        max_len = len(h)
        col_data = df.iloc[:5] if len(df) > 5 else df
        for key in [visible_keys[c_idx - 1]]:
            samples = col_data[key].dropna().astype(str).tolist() if key in col_data.columns else []
            for s in samples:
                max_len = max(max_len, min(len(str(s)), 40))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 3, 50)

    ws.sheet_properties.tabColor = "2E7D32"


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN EXPORT FUNCTION
# ═══════════════════════════════════════════════════════════════════════════════

class ExcelReport:
    """Xây dựng báo cáo Excel nhiều sheet với định dạng chuẩn VBSP.

    Usage:
        rpt = ExcelReport("Báo cáo dư nợ", nguoi_xuat="Nguyễn Văn A")
        rpt.add_kpi("Số khế ước", 1234)
        rpt.add_kpi("Tổng dư nợ", "1.234.567.000", "tỷ đồng")
        rpt.add_detail("Chi tiết PGD", df_pgd)
        rpt.add_sheet("Bảng phụ", df_phu)
        buf = rpt.build()
        st.download_button(..., data=buf)
    """

    def __init__(self, title: str, subtitle: str = "",
                 nguoi_xuat: str = "",
                 detail_columns: Optional[list] = None):
        self.title = title
        self.subtitle = subtitle
        self.nguoi_xuat = nguoi_xuat
        self.detail_columns = detail_columns or LOAN_COLUMNS
        self._kpi_items: list[tuple[str, object, str]] = []
        self._detail_sheets: list[tuple[str, pd.DataFrame]] = []
        self._extra_sheets: list[tuple[str, pd.DataFrame]] = []

    def add_kpi(self, label: str, value: object,
                note: str = "") -> "ExcelReport":
        self._kpi_items.append((label, value, note))
        return self

    def add_detail(self, name: str,
                   df: pd.DataFrame) -> "ExcelReport":
        if df is not None and not df.empty:
            self._detail_sheets.append((name, df))
        return self

    def add_sheet(self, name: str,
                  df: pd.DataFrame) -> "ExcelReport":
        if df is not None and not df.empty:
            self._extra_sheets.append((name, df))
        return self

    def build(self) -> bytes:
        wb = Workbook()

        _build_cover_sheet(wb, self.title, self.subtitle, self.nguoi_xuat)

        if self._kpi_items:
            _build_kpi_sheet(wb, self._kpi_items, self.subtitle)

        for name, df in self._detail_sheets:
            _build_detail_sheet(wb, name, df, self.detail_columns)

        for name, df in self._extra_sheets:
            _build_detail_sheet(wb, name, df)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# CONVENIENCE FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def xuat_excel_chuyen_nghiep(
    df: pd.DataFrame,
    title: str,
    nguoi_xuat: str = "",
    subtitle: str = "",
    columns: Optional[list] = None,
    kpi_items: Optional[list[tuple[str, object, str]]] = None,
    extra_sheets: Optional[list[tuple[str, pd.DataFrame]]] = None,
) -> bytes:
    """Xuất 1 DataFrame thành Excel chuyên nghiệp (1 lần gọn nhẹ).

    Args:
        df: DataFrame chính (sheet chi tiết)
        title: Tiêu đề báo cáo (hiển thị trên Bìa)
        nguoi_xuat: Tên người xuất
        subtitle: Phụ đề (hiển thị trên Bìa + KPI sheet)
        columns: Danh sách cột xuất, None = LOAN_COLUMNS mặc định
        kpi_items: List (label, value, note) cho sheet KPI
        extra_sheets: List (name, df) cho các sheet bổ sung

    Returns:
        bytes — nội dung file .xlsx
    """
    rpt = ExcelReport(title, subtitle, nguoi_xuat, columns)

    if kpi_items:
        for label, value, note in kpi_items:
            rpt.add_kpi(label, value, note)

    sheet_name = _safe_sheet_name(title.split("—")[-1].strip() if "—" in title else title)
    rpt.add_detail(sheet_name, df)

    if extra_sheets:
        for name, extra_df in extra_sheets:
            rpt.add_sheet(name, extra_df)

    return rpt.build()


def ten_file_xuat(prefix: str, ext: str = "xlsx") -> str:
    now = datetime.now()
    return f"{prefix}_{now.strftime('%d%m%Y')}_{now.strftime('%H%M')}.{ext}"
